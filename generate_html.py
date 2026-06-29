#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_html.py
================
直接讀取 *_Dividend_Analysis.xlsx 總覽分頁，
生成最新的 index.html 上傳到 GitHub。

使用方法：
    python generate_html.py

每日執行順序：
    1. python daily_importer_global_v5.py
    2. python generate_html.py
    3. 上傳 index.html 到 GitHub

每週有新股時：
    1. python screener_global.py
    2. 揀股 → python add_to_tracking.py
    3. python batch_importer_global_v5.py
    4. python daily_importer_global_v5.py
    5. python generate_html.py
    6. 上傳 index.html 到 GitHub
"""

import datetime
import os
import json
import openpyxl

# ── 設定 ─────────────────────────────────────────────────
EXCEL_FILES = {
    "UK": r"D:\finance_project\UK_Dividend_Analysis.xlsx",
    "HK": r"D:\finance_project\HK_Dividend_Analysis.xlsx",
    "US": r"D:\finance_project\US_Dividend_Analysis.xlsx",
    "CN": r"D:\finance_project\CN_Dividend_Analysis.xlsx",
}
OUTPUT_BASE  = r"D:\finance_project\web_site"
OUTPUT_FILE  = os.path.join(OUTPUT_BASE, datetime.date.today().strftime("%Y%m%d"), "index.html")
TOP_N        = 10    # 柱狀圖顯示前N名
PICKS_N      = 12    # 推介卡片數量
MIN_SCORE    = 50    # 入圍門檻（顯示用）
RATING_STRONG = 75
RATING_WATCH  = 60

# ── 讀取 Excel 總覽 ───────────────────────────────────────
def load_excel_data():
    """讀取三個市場的總覽分頁，合併成一個股票列表"""
    all_stocks = []

    for mkt, filepath in EXCEL_FILES.items():
        if not os.path.exists(filepath):
            print(f"  ⚠️  找不到 {filepath}，跳過")
            continue

        wb = openpyxl.load_workbook(filepath, data_only=True)
        summary_name = f"{mkt} 總覽"
        if summary_name not in wb.sheetnames:
            print(f"  ⚠️  找不到分頁「{summary_name}」，跳過")
            continue

        ws = wb[summary_name]
        headers = [str(ws.cell(1, c).value or "").strip()
                   for c in range(1, ws.max_column + 1)]

        def ci(name):
            try: return headers.index(name)
            except ValueError: return None

        i_ticker = ci("股票代號")
        i_name   = ci("公司名稱")
        i_price  = ci("現價")
        i_yield  = ci("最新股息率")
        i_pe     = ci("最新 PE")
        i_pb     = ci("P/B")
        i_score  = ci("📊 總分_100")
        i_diag   = ci("📊 綜合診斷")
        i_sq     = ci("S_股息質量")
        i_sv     = ci("S_估值")
        i_sf     = ci("S_財務健康")
        i_sg     = ci("S_增長")
        i_st     = ci("S_技術")

        if i_ticker is None or i_score is None:
            print(f"  ⚠️  {summary_name} 欄位不符，跳過")
            continue

        for row in range(2, ws.max_row + 1):
            ticker = str(ws.cell(row, i_ticker + 1).value or "").strip()
            score  = ws.cell(row, i_score + 1).value if i_score is not None else None
            if not ticker or score is None:
                continue
            try:
                score = float(score)
            except:
                continue

            all_stocks.append({
                "ticker": ticker,
                "name":   str(ws.cell(row, i_name  + 1).value or ticker).strip() if i_name  is not None else ticker,
                "mkt":    mkt,
                "price":  ws.cell(row, i_price + 1).value if i_price is not None else None,
                "yield":  ws.cell(row, i_yield + 1).value if i_yield is not None else None,
                "pe":     ws.cell(row, i_pe    + 1).value if i_pe    is not None else None,
                "pb":     ws.cell(row, i_pb    + 1).value if i_pb    is not None else None,
                "score":  score,
                "diag":   str(ws.cell(row, i_diag + 1).value or "").strip() if i_diag is not None else "",
                "sq":     ws.cell(row, i_sq + 1).value if i_sq is not None else None,
                "sv":     ws.cell(row, i_sv + 1).value if i_sv is not None else None,
                "sf":     ws.cell(row, i_sf + 1).value if i_sf is not None else None,
                "sg":     ws.cell(row, i_sg + 1).value if i_sg is not None else None,
                "st":     ws.cell(row, i_st + 1).value if i_st is not None else None,
            })

        print(f"  ✅ {mkt}: 讀取 {sum(1 for s in all_stocks if s['mkt']==mkt)} 隻")

    all_stocks.sort(key=lambda x: x["score"], reverse=True)
    return all_stocks

def get_rating_key(score):
    if score >= RATING_STRONG: return "strong"
    if score >= RATING_WATCH:  return "watch"
    return "hold"

def get_market_stats(stocks):
    stats = {}
    for mkt in ["US", "HK", "UK", "CN"]:
        ms = [s for s in stocks if s["mkt"] == mkt]
        if not ms:
            stats[mkt] = {"total":0,"strong":0,"watch":0,"hold":0,"avg":0,"max":0}
            continue
        stats[mkt] = {
            "total":  len(ms),
            "strong": sum(1 for s in ms if s["score"] >= RATING_STRONG),
            "watch":  sum(1 for s in ms if RATING_WATCH <= s["score"] < RATING_STRONG),
            "hold":   sum(1 for s in ms if s["score"] < RATING_WATCH),
            "avg":    round(sum(s["score"] for s in ms) / len(ms), 1),
            "max":    int(max(s["score"] for s in ms)),
        }
    return stats

def fmt(v, decimals=1, suffix=""):
    if v is None: return "─"
    try: return f"{float(v):.{decimals}f}{suffix}"
    except: return "─"

# ── 生成 HTML ─────────────────────────────────────────────
SETLANG_JS = """
function setLang(lang) {
  try{localStorage.setItem('hidh_lang',lang);}catch(e){}
  ['#btn-zh-hk','#btn-zh-cn','#btn-en'].forEach(function(id){
    var el=document.querySelector(id);
    if(el){el.style.background='transparent';el.style.color='#666';el.style.borderColor='#e5e5e5';}
  });
  var ab=document.querySelector({'zh-hk':'#btn-zh-hk','zh-cn':'#btn-zh-cn','en':'#btn-en'}[lang]);
  if(ab){ab.style.background='#1D9E75';ab.style.color='#fff';ab.style.borderColor='#1D9E75';}
  var T={
    'zh-hk':{heroTitle:'全球高息股<br>每日精選分析',heroSub:'覆蓋香港、美國、英國及A股四大市場，以系統化評分篩選出具備穩定派息能力的優質股票。',heroTag:'每日更新',stat1:'今日追蹤股票數',stat2:'入圍股票（≥60分）',stat3:'強力買入',stat4:'今日最高分',mktHK:'港股 HK',mktUS:'美股 US',mktUK:'英股 UK',mktCN:'A股 CN',cUS:'美股 US',cHK:'港股 HK',cUK:'英股 UK',cCN:'A股 CN',top10:'前10名評分分佈',sec0:'選股理念與方法',sec1:'評分系統',sec2:'市場概覽',sec3:'最新精選推介',ah:['選股理念','三大市場','系統化篩選','每日更新'],ap:['長期穩定的股息收入是財富增長的重要基石。我們不單看當前息率高低，更重視企業的派息可持續性、財務健康狀況及估值合理性。','同步覆蓋香港、美國及英國市場，以統一標準進行跨市場比較，讓投資者掌握全球高息機會。','每日自動更新數據，以量化評分模型對數百隻股票進行排名，過濾雜訊，聚焦真正值得關注的機會。','每個交易日收市後自動重新評分，確保推介反映最新的估值及財務狀況。'],scoreIntro:'每隻股票以100分制進行綜合評分，涵蓋五個範疇：',th:['評分範疇','滿分','主要考量'],rows:[['股息質量','30分','息率水平、派息穩定性及覆蓋率'],['估值','25分','現價相對歷史息率及市場的吸引程度'],['財務健康','25分','資產負債、現金流及償債能力'],['增長','10分','股息增長趨勢及盈利前景'],['技術走勢','10分','RSI、52週位置等技術指標']],rth:['評級','分數','意義'],rmean:['各方面均表現優秀，值得重點關注','基本面良好，可納入觀察名單','有一定吸引力，但需留意風險'],ll:['強力買入','值得關注','觀望'],unit:'隻',rl:{'strong':'🟢🟢 強力買入','watch':'🟢 值得關注','hold':'⚖️ 觀望'},picksNote:'以下為今日評分最高股票（≥50分），綠框為強力買入。',pros:'✅ 優點',cons:'⚠️ 缺點',score:'評分',yieldLbl:'股息率',discLabel:'免責聲明',disc:'本網站所有內容僅供參考及教育用途，不構成任何投資建議或買賣邀請。投資涉及風險，過往表現不代表未來回報。讀者應自行進行盡職審查，並在作出任何投資決定前諮詢持牌財務顧問。',f2:'每個交易日更新',f3:'資料來源：Yahoo Finance · 僅供參考'},
    'zh-cn':{heroTitle:'全球高息股<br>每日精选分析',heroSub:'覆盖香港、美国、英国及A股四大市场，以系统化评分筛选出具备稳定派息能力的优质股票。',heroTag:'每日更新',stat1:'今日追踪股票数',stat2:'入围股票（≥60分）',stat3:'强力买入',stat4:'今日最高分',mktHK:'港股 HK',mktUS:'美股 US',mktUK:'英股 UK',mktCN:'A股 CN',cUS:'美股 US',cHK:'港股 HK',cUK:'英股 UK',cCN:'A股 CN',top10:'前10名评分分布',sec0:'选股理念与方法',sec1:'评分系统',sec2:'市场概览',sec3:'最新精选推介',ah:['选股理念','三大市场','系统化筛选','每日更新'],ap:['长期稳定的股息收入是财富增长的重要基石。我们不单看当前息率高低，更重视企业的派息可持续性、财务健康状况及估值合理性。','同步覆盖香港、美国及英国市场，以统一标准进行跨市场比较，让投资者掌握全球高息机会。','每日自动更新数据，以量化评分模型对数百只股票进行排名，过滤杂讯，聚焦真正值得关注的机会。','每个交易日收市后自动重新评分，确保推介反映最新的估值及财务状况。'],scoreIntro:'每只股票以100分制进行综合评分，涵盖五个范畴：',th:['评分范畴','满分','主要考量'],rows:[['股息质量','30分','息率水平、派息稳定性及覆盖率'],['估值','25分','现价相对历史息率及市场的吸引程度'],['财务健康','25分','资产负债、现金流及偿债能力'],['增长','10分','股息增长趋势及盈利前景'],['技术走势','10分','RSI、52周位置等技术指标']],rth:['评级','分数','意义'],rmean:['各方面均表现优秀，值得重点关注','基本面良好，可纳入观察名单','有一定吸引力，但需留意风险'],ll:['强力买入','值得关注','观望'],unit:'只',rl:{'strong':'🟢🟢 强力买入','watch':'🟢 值得关注','hold':'⚖️ 观望'},picksNote:'以下为今日评分最高股票（≥50分），绿框为强力买入。',pros:'✅ 优点',cons:'⚠️ 缺点',score:'评分',yieldLbl:'股息率',discLabel:'免责声明',disc:'本网站所有内容仅供参考及教育用途，不构成任何投资建议或买卖邀请。投资涉及风险，过往表现不代表未来回报。读者应自行进行尽职审查，并在作出任何投资决定前咨询持牌财务顾问。',f2:'每个交易日更新',f3:'资料来源：Yahoo Finance · 仅供参考'},
    'en':{heroTitle:'Global Dividend Stocks<br>Daily Analysis',heroSub:'Covering HK, US, UK and China A-shares with systematic scoring to identify quality dividend stocks.',heroTag:'Daily Update',stat1:'Stocks Tracked',stat2:'Qualified (≥60pts)',stat3:'Strong Buy',stat4:"Today\'s High",mktHK:'HK Stocks',mktUS:'US Stocks',mktUK:'UK Stocks',mktCN:'China A-Shares',cUS:'US Stocks',cHK:'HK Stocks',cUK:'UK Stocks',cCN:'China A-Shares',top10:'Top 10 Score Distribution',sec0:'Investment Philosophy',sec1:'Scoring System',sec2:'Market Overview',sec3:'Top Picks',ah:['Philosophy','4 Markets','Systematic Screening','Daily Update'],ap:['We focus not just on yield but on dividend sustainability, financial health and valuation to find quality long-term holdings.','Covering HK, US, UK and China A-shares with a unified scoring framework for cross-market comparison.','Daily automated updates with quantitative scoring to rank hundreds of stocks and surface the best opportunities.','Re-scored every trading day after market close to reflect the latest valuations and conditions.'],scoreIntro:'Each stock is scored on a 100-point scale across five dimensions:',th:['Category','Max','Key Criteria'],rows:[['Dividend Quality','30pts','Yield level, payout stability & coverage'],['Valuation','25pts','Current price vs historical yield & attractiveness'],['Financial Health','25pts','Balance sheet, cash flow & debt coverage'],['Growth','10pts','Dividend growth trend & earnings outlook'],['Technical','10pts','RSI, 52-week position & other indicators']],rth:['Rating','Score','Meaning'],rmean:['Excellent across all dimensions, high priority','Good fundamentals, worth monitoring','Some appeal, monitor risks'],ll:['Strong Buy','Watch','Hold'],unit:'stk',rl:{'strong':'🟢🟢 Strong Buy','watch':'🟢 Watch','hold':'⚖️ Hold'},picksNote:'Top-rated stocks today (≥50pts). Green border = Strong Buy.',pros:'✅ Pro',cons:'⚠️ Con',score:'Score',yieldLbl:'Yield',discLabel:'Disclaimer',disc:'All content is for reference and educational purposes only. Not investment advice. Investing involves risk. Past performance does not guarantee future results.',f2:'Updated every trading day',f3:'Data: Yahoo Finance · For reference only'},
  };
  var t=T[lang]||T['zh-hk'];
  document.title=(lang==='en'?'Global Dividend Analysis':'全球高息股分析')+' | HiDH Dividend Analyst';
  document.querySelectorAll('a[data-zh-hk]').forEach(function(el){el.textContent=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');});
  var h1=document.querySelector('.hero h1');if(h1)h1.innerHTML=t.heroTitle;
  var hs=document.querySelector('.hero-sub');if(hs)hs.textContent=t.heroSub;
  var ht=document.querySelector('.hero-tag');if(ht){var p=ht.textContent.split('·');ht.textContent=t.heroTag+' · '+(p[1]||'').trim();}
  var sl=document.querySelectorAll('.stat-label');[t.stat1,t.stat2,t.stat3,t.stat4].forEach(function(s,i){if(sl[i])sl[i].textContent=s;});
  var hb=document.querySelectorAll('.hero-badges .badge');[t.mktHK,t.mktUS,t.mktUK,t.mktCN].forEach(function(s,i){if(hb[i])hb[i].textContent=s;});
  var ct=document.querySelectorAll('.chart-title');[t.cUS,t.cHK,t.cUK,t.cCN,t.top10].forEach(function(s,i){if(ct[i])ct[i].firstChild.textContent=s;});
  var st=document.querySelectorAll('.section-title');
  if(st[0])st[0].childNodes[0].textContent=t.sec0;
  if(st[1])st[1].childNodes[0].textContent=t.sec1;
  if(st[2]){var d2=st[2].textContent.match(/[0-9]{4}-[0-9]{2}-[0-9]{2}/);st[2].childNodes[0].textContent=t.sec2+(d2?' — '+d2[0]:'');}
  if(st[3]){var d3=st[3].textContent.match(/[0-9]{4}-[0-9]{2}-[0-9]{2}/);st[3].childNodes[0].textContent=t.sec3+(d3?' — '+d3[0]:'');}
  var ag=document.getElementById('about-grid');
  if(ag){ag.querySelectorAll('h3').forEach(function(el,i){if(t.ah[i])el.textContent=t.ah[i];});ag.querySelectorAll('p').forEach(function(el,i){if(t.ap[i])el.textContent=t.ap[i];});}
  var si=document.querySelector('#about > p');if(si)si.textContent=t.scoreIntro;
  var tbls=document.querySelectorAll('.score-table');
  if(tbls[0]){tbls[0].querySelectorAll('th').forEach(function(el,i){if(t.th[i])el.textContent=t.th[i];});tbls[0].querySelectorAll('tbody tr').forEach(function(tr,i){if(t.rows[i]){var tds=tr.querySelectorAll('td');tds.forEach(function(td,j){if(t.rows[i][j])td.textContent=t.rows[i][j];});}});}
  if(tbls[1]){tbls[1].querySelectorAll('th').forEach(function(el,i){if(t.rth[i])el.textContent=t.rth[i];});tbls[1].querySelectorAll('tbody tr').forEach(function(tr,i){var tds=tr.querySelectorAll('td');if(tds[1])tds[1].textContent=tds[1].getAttribute('data-'+lang)||tds[1].getAttribute('data-zh-hk')||tds[1].textContent;if(tds[2]&&t.rmean[i])tds[2].textContent=t.rmean[i];});}
  document.querySelectorAll('[data-rating]').forEach(function(el){var r=el.getAttribute('data-rating');if(t.rl[r])el.textContent=t.rl[r];});
  document.querySelectorAll('.legend span[data-zh-hk]').forEach(function(el){var k=el.getAttribute('data-zh-hk'),strong=el.querySelector('strong'),ldEl=el.querySelector('.ld'),ld=ldEl?ldEl.outerHTML:'',idx=['強力買入','值得關注','觀望'].indexOf(k);if(idx>=0)el.innerHTML=ld+t.ll[idx]+(strong?' '+strong.outerHTML:'');});
  document.querySelectorAll('.chart-sub[data-zh-hk]').forEach(function(el){el.textContent=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');});
  document.querySelectorAll('svg[id]').forEach(function(svg){var texts=svg.querySelectorAll('text'),last=texts[texts.length-1];if(last){var m=last.textContent.match(/^([0-9]+)/);if(m)last.textContent=m[1]+t.unit;}});
  var pn=document.getElementById('picks-note');if(pn)pn.textContent=t.picksNote;
  document.querySelectorAll('.pros-label').forEach(function(el){el.textContent=t.pros;});
  document.querySelectorAll('.cons-label').forEach(function(el){el.textContent=t.cons;});
  document.querySelectorAll('.score-row span:first-child').forEach(function(el){el.textContent=t.score;});
  document.querySelectorAll('.pick-stat-label[data-zh-hk]').forEach(function(el){el.textContent=t.yieldLbl;});
  document.querySelectorAll('.pros-text').forEach(function(el){el.textContent=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');});
  document.querySelectorAll('.cons-text').forEach(function(el){el.textContent=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');});
  var fi=document.querySelectorAll('.footer-inner span');if(fi[1])fi[1].textContent='© 2026 prosynchk.com · '+t.f2;if(fi[2])fi[2].textContent=t.f3;
  var ch={'zh-hk':['合作機會','歡迎合作','如您有興趣進行廣告合作、內容授權、數據合作或其他商業合作，歡迎透過以下方式聯絡我們。'],'zh-cn':['合作机会','欢迎合作','如您有兴趣进行广告合作、内容授权、数据合作或其他商业合作，欢迎通过以下方式联系我们。'],'en':['Partnership','Work With Us','Interested in advertising, content licensing, data collaboration or other business opportunities? Get in touch.']};
  var cl=ch[lang]||ch['zh-hk'];
  var chead=document.getElementById('contact-heading');if(chead)chead.childNodes[0].textContent=cl[0];
  var ctitle=document.getElementById('contact-title');if(ctitle)ctitle.textContent=cl[1];
  var cbody=document.getElementById('contact-body');if(cbody)cbody.textContent=cl[2];
  var disc=document.querySelector('.disclaimer-inner');if(disc)disc.innerHTML='<strong>'+t.discLabel+'：</strong>'+t.disc;
}
"""



def generate_html(stocks, stats, report_date):
    top10  = stocks[:TOP_N]
    picks  = [s for s in stocks if s["score"] >= MIN_SCORE][:PICKS_N]
    total  = sum(s["total"] for s in stats.values())
    n_strong = sum(s["strong"] for s in stats.values())
    max_score  = int(stocks[0]["score"]) if stocks else 0
    max_ticker = stocks[0]["ticker"] if stocks else ""

    top10_js = json.dumps([{
        "label":  s["ticker"],
        "score":  int(s["score"]),
        "mkt":    s["mkt"],
        "rating": get_rating_key(s["score"]),
    } for s in top10])

    def get_pros_cons(s):
        phk,chk,pcn,ccn,pen,cen = [],[],[],[],[],[]
        try:
            yld = float(s["yield"]) if s["yield"] else 0
            pe  = float(s["pe"])    if s["pe"]    else 0
            sq  = float(s["sq"])    if s["sq"]    else 0
            sv  = float(s["sv"])    if s["sv"]    else 0
            sf  = float(s["sf"])    if s["sf"]    else 0
            sg  = float(s["sg"])    if s["sg"]    else 0
            if yld >= 6:   phk.append(f"高息率 {yld:.1f}%，現金回報吸引"); pcn.append(f"高股息率 {yld:.1f}%，现金回报吸引"); pen.append(f"High yield {yld:.1f}%, attractive income")
            elif yld >= 4: phk.append(f"息率 {yld:.1f}%，優於市場平均"); pcn.append(f"股息率 {yld:.1f}%，优于市场平均"); pen.append(f"Yield {yld:.1f}%, above market average")
            if sq >= 22:   phk.append("股息質量高，派息穩定可靠"); pcn.append("股息质量高，派息稳定可靠"); pen.append("High dividend quality, stable payout")
            if sv >= 20:   phk.append("估值具吸引力，現價相對合理"); pcn.append("估值具吸引力，现价相对合理"); pen.append("Attractive valuation at current price")
            if sf >= 20:   phk.append("財務健康，資產負債穩健"); pcn.append("财务健康，资产负债稳健"); pen.append("Strong balance sheet")
            if sg >= 7:    phk.append("股息增長趨勢良好"); pcn.append("股息增长趋势良好"); pen.append("Positive dividend growth trend")
            if 0 < pe < 12:phk.append(f"PE {pe:.1f}x，估值偏低"); pcn.append(f"PE {pe:.1f}x，估值偏低"); pen.append(f"Low PE of {pe:.1f}x")
            if yld < 3:    chk.append("息率偏低，收息吸引力有限"); ccn.append("股息率偏低，收息吸引力有限"); cen.append("Low yield limits income appeal")
            if sq < 15:    chk.append("股息質量一般，需留意派息持續性"); ccn.append("股息质量一般，需留意派息持续性"); cen.append("Moderate dividend quality, sustainability risk")
            if sv < 12:    chk.append("估值偏高，上升空間有限"); ccn.append("估值偏高，上升空间有限"); cen.append("Rich valuation limits upside")
            if sf < 15:    chk.append("財務健康度一般，需關注負債"); ccn.append("财务健康度一般，需关注负债"); cen.append("Moderate financials, watch debt levels")
            if sg < 5:     chk.append("股息增長動力不足"); ccn.append("股息增长动力不足"); cen.append("Limited dividend growth momentum")
            if pe > 25:    chk.append(f"PE {pe:.1f}x，估值偏貴"); ccn.append(f"PE {pe:.1f}x，估值偏贵"); cen.append(f"High PE of {pe:.1f}x")
        except: pass
        return {
            "pro_zh_hk": phk[0] if phk else "評分良好，具備一定投資價值",
            "con_zh_hk": chk[0] if chk else "需留意市場風險及行業波動",
            "pro_zh_cn": pcn[0] if pcn else "评分良好，具备一定投资价值",
            "con_zh_cn": ccn[0] if ccn else "需留意市场风险及行业波动",
            "pro_en":    pen[0] if pen else "Good overall score, investment potential",
            "con_en":    cen[0] if cen else "Monitor market and sector risks",
        }

    picks_js = json.dumps([{
        "ticker": s["ticker"],
        "name":   s["name"],
        "mkt":    s["mkt"],
        "score":  int(s["score"]),
        "rating": get_rating_key(s["score"]),
        "yield_": fmt(s["yield"]),
        "pe":     fmt(s["pe"]),
        "pb":     fmt(s["pb"], 2),
        **get_pros_cons(s),
    } for s in picks])

    html = f"""<!DOCTYPE html>
<html lang="zh-HK">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>全球高息股分析 | HiDH Dividend Analyst</title>
<link rel="icon" type="image/png" href="/icon.png">
<link rel="apple-touch-icon" href="/icon.png">
<script async src="https://pagead2.googlesyndication.com/pagead/js/adsbygoogle.js?client=ca-pub-9533474113956980" crossorigin="anonymous"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:system-ui,-apple-system,sans-serif;background:#f7f7f5;color:#333;line-height:1.7}}
a{{color:#1D9E75;text-decoration:none}}a:hover{{text-decoration:underline}}
.header{{background:#fff;border-bottom:1px solid #e5e5e5;padding:0 1.5rem}}
.header-inner{{max-width:960px;margin:0 auto;display:flex;align-items:center;justify-content:space-between;height:60px}}
.logo{{font-size:18px;font-weight:600;color:#1D9E75}}.logo span{{color:#333}}
.nav{{display:flex;gap:1.5rem;font-size:14px}}.nav a{{color:#555}}.nav a:hover{{color:#1D9E75;text-decoration:none}}
.hero{{background:#fff;border-bottom:1px solid #e5e5e5;padding:2.5rem 1.5rem}}
.hero-inner{{max-width:960px;margin:0 auto;display:grid;grid-template-columns:1fr 1fr;gap:2rem;align-items:center}}
.hero-tag{{font-size:11px;font-weight:600;color:#1D9E75;text-transform:uppercase;letter-spacing:.1em;margin-bottom:.5rem}}
.hero h1{{font-size:26px;font-weight:600;line-height:1.35;margin-bottom:.75rem;color:#222}}
.hero-sub{{font-size:15px;color:#666;margin-bottom:1.25rem}}
.hero-badges{{display:flex;gap:8px;flex-wrap:wrap}}
.badge{{font-size:11px;padding:4px 10px;border-radius:20px;font-weight:500}}
.badge-hk{{background:#E1F5EE;color:#0F6E56}}.badge-us{{background:#E6F1FB;color:#185FA5}}.badge-uk{{background:#FAEEDA;color:#854F0B}}.badge-cn{{background:#FDECEA;color:#B71C1C}}
.hero-stats{{display:grid;grid-template-columns:1fr 1fr;gap:12px}}
.stat-box{{background:#f7f7f5;border-radius:10px;padding:1rem}}
.stat-val{{font-size:24px;font-weight:600;color:#222}}.stat-label{{font-size:12px;color:#888;margin-top:2px}}
.section{{max-width:960px;margin:0 auto;padding:2rem 1.5rem}}
.section-title{{font-size:13px;font-weight:600;color:#999;text-transform:uppercase;letter-spacing:.1em;margin-bottom:1.25rem;display:flex;align-items:center;gap:10px}}
.section-title::after{{content:'';flex:1;height:1px;background:#e5e5e5}}
.about-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:12px;margin-bottom:1.5rem}}
.about-card{{background:#fff;border:1px solid #e5e5e5;border-radius:12px;padding:1.25rem}}
.about-icon{{font-size:20px;margin-bottom:.5rem}}
.about-card h3{{font-size:14px;font-weight:600;color:#222;margin-bottom:.4rem}}
.about-card p{{font-size:13px;color:#666;line-height:1.6}}
.score-table{{width:100%;border-collapse:collapse;font-size:13px;background:#fff;border-radius:12px;overflow:hidden;border:1px solid #e5e5e5}}
.score-table th{{background:#f7f7f5;padding:10px 14px;text-align:left;font-weight:600;color:#555;font-size:12px}}
.score-table td{{padding:10px 14px;border-top:1px solid #f0f0f0;color:#444}}
.r-strong{{background:#E1F5EE;color:#0F6E56;padding:3px 8px;border-radius:4px;font-size:11px;font-weight:600;white-space:nowrap}}
.r-watch{{background:#E6F1FB;color:#185FA5;padding:3px 8px;border-radius:4px;font-size:11px;font-weight:600;white-space:nowrap}}
.r-hold{{background:#FAEEDA;color:#854F0B;padding:3px 8px;border-radius:4px;font-size:11px;font-weight:600;white-space:nowrap}}
.top-row{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px}}
.chart-card{{background:#fff;border:1px solid #e5e5e5;border-radius:12px;padding:16px}}
.chart-title{{font-size:13px;font-weight:600;color:#444;margin-bottom:8px}}
.legend{{display:flex;gap:10px;margin-bottom:8px;font-size:11px;color:#666;flex-wrap:wrap}}
.ld{{width:10px;height:10px;border-radius:2px;display:inline-block;margin-right:3px;vertical-align:2px}}
.chart-sub{{text-align:center;font-size:11px;color:#888;margin-top:6px}}
svg text{{font-family:system-ui,sans-serif}}
.picks-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:12px}}
.pick-card{{background:#fff;border:1px solid #e5e5e5;border-radius:12px;padding:1.1rem}}
.pick-card.top{{border:2px solid #1D9E75}}
.pick-header{{display:block;margin-bottom:.5rem}}
.pick-ticker{{font-size:15px;font-weight:600;color:#222}}
.pick-name{{font-size:12px;color:#888;margin-bottom:.75rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.pick-stats{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:.75rem}}
.pick-stat{{background:#f7f7f5;border-radius:6px;padding:6px 8px}}
.pick-stat-label{{font-size:10px;color:#999}}.pick-stat-val{{font-size:13px;font-weight:600;color:#222}}
.score-track{{height:4px;background:#f0f0f0;border-radius:2px;overflow:hidden;margin-top:.5rem}}
.score-fill{{height:100%;border-radius:2px;background:#1D9E75}}
.score-row{{display:flex;justify-content:space-between;font-size:10px;color:#aaa;margin-top:3px}}
.pros-cons{{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-top:8px;margin-bottom:6px}}
.pros{{background:#F0FBF5;border-radius:6px;padding:6px 8px;font-size:11px;color:#0F6E56}}
.cons{{background:#FEF3F2;border-radius:6px;padding:6px 8px;font-size:11px;color:#B42318}}
.pros-label,.cons-label{{font-weight:600;margin-bottom:2px;font-size:10px}}
.disclaimer{{background:#fff;border-top:1px solid #e5e5e5;padding:1.5rem;margin-top:2rem}}
.disclaimer-inner{{max-width:960px;margin:0 auto;font-size:12px;color:#aaa;line-height:1.7}}
.disclaimer-inner strong{{color:#888}}
.footer{{background:#222;color:#aaa;padding:1.5rem;font-size:12px}}
.footer-inner{{max-width:960px;margin:0 auto;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:.5rem}}
.footer-logo{{color:#1D9E75;font-weight:600;font-size:14px}}
@media(max-width:640px){{.hero-inner{{grid-template-columns:1fr}}.hero-stats{{display:none}}.top-row{{grid-template-columns:1fr}}}}
</style>
</head>
<body>

<div class="header">
  <div class="header-inner">
    <div class="logo">HiDH <span>Dividend Analyst</span></div>
    <nav class="nav">
      <a href="#about">選股方法</a>
      <a href="#charts">市場概覽</a>
      <a href="#picks">最新精選</a>
    </nav>
  </div>
</div>
<div style="display:flex;justify-content:flex-end;gap:6px;padding:.4rem 1.5rem;background:#fff;border-bottom:0.5px solid #e5e5e5">
  <button id="btn-zh-hk" style="font-size:12px;padding:4px 12px;border-radius:20px;border:0.5px solid #1D9E75;background:#1D9E75;color:#fff;cursor:pointer" onclick="setLang('zh-hk')">繁體</button>
  <button id="btn-zh-cn" style="font-size:12px;padding:4px 12px;border-radius:20px;border:0.5px solid #e5e5e5;background:transparent;color:#666;cursor:pointer" onclick="setLang('zh-cn')">简体</button>
  <button id="btn-en"    style="font-size:12px;padding:4px 12px;border-radius:20px;border:0.5px solid #e5e5e5;background:transparent;color:#666;cursor:pointer" onclick="setLang('en')">English</button>
</div>

<div class="hero">
  <div class="hero-inner">
    <div>
      <div class="hero-tag">每日更新 · {report_date}</div>
      <h1>全球高息股<br>每日精選分析</h1>
      <p class="hero-sub">覆蓋香港、美國、英國三大市場，以系統化評分篩選出具備穩定派息能力的優質股票。</p>
      <div class="hero-badges">
        <span class="badge badge-hk">港股 HK</span>
        <span class="badge badge-us">美股 US</span>
        <span class="badge badge-uk">英股 UK</span>
        <span class="badge badge-cn">A股 CN</span>
      </div>
    </div>
    <div class="hero-stats">
      <div class="stat-box"><div class="stat-val">{total}</div><div class="stat-label">今日追蹤股票數</div></div>
      <div class="stat-box"><div class="stat-val">{sum(s["strong"]+s["watch"] for s in stats.values())}</div><div class="stat-label">入圍股票（≥{RATING_WATCH}分）</div></div>
      <div class="stat-box"><div class="stat-val">{n_strong}</div><div class="stat-label">強力買入</div></div>
      <div class="stat-box"><div class="stat-val">{max_score}</div><div class="stat-label">今日最高分 ({max_ticker})</div></div>
    </div>
  </div>
</div>

<div class="section" id="about">
  <div class="section-title">選股理念與方法</div>
  <div class="about-grid">
    <div class="about-card">
      <div class="about-icon">💡</div>
      <h3>選股理念</h3>
      <p>長期穩定的股息收入是財富增長的重要基石。我們不單看當前息率高低，更重視企業的派息可持續性、財務健康狀況及估值合理性。</p>
    </div>
    <div class="about-card">
      <div class="about-icon">🌍</div>
      <h3>三大市場</h3>
      <p>同步覆蓋香港、美國及英國市場，以統一標準進行跨市場比較，讓投資者掌握全球高息機會。</p>
    </div>
    <div class="about-card">
      <div class="about-icon">📊</div>
      <h3>系統化篩選</h3>
      <p>每日自動更新數據，以量化評分模型對數百隻股票進行排名，過濾雜訊，聚焦真正值得關注的機會。</p>
    </div>
    <div class="about-card">
      <div class="about-icon">🔄</div>
      <h3>每日更新</h3>
      <p>每個交易日收市後自動重新評分，確保推介反映最新的估值及財務狀況。</p>
    </div>
  </div>

  <div class="section-title" style="margin-top:2rem">評分系統</div>
  <p style="font-size:14px;color:#666;margin-bottom:1rem">每隻股票以100分制進行綜合評分，涵蓋五個範疇：</p>
  <table class="score-table" style="margin-bottom:1.25rem">
    <thead><tr><th>評分範疇</th><th>滿分</th><th>主要考量</th></tr></thead>
    <tbody>
      <tr><td>股息質量</td><td>30分</td><td>息率水平、派息穩定性及覆蓋率</td></tr>
      <tr><td>估值</td><td>25分</td><td>現價相對歷史息率及市場的吸引程度</td></tr>
      <tr><td>財務健康</td><td>25分</td><td>資產負債、現金流及償債能力</td></tr>
      <tr><td>增長</td><td>10分</td><td>股息增長趨勢及盈利前景</td></tr>
      <tr><td>技術走勢</td><td>10分</td><td>RSI、52週位置等技術指標</td></tr>
    </tbody>
  </table>
  <table class="score-table">
    <thead><tr><th>評級</th><th>分數</th><th>意義</th></tr></thead>
    <tbody>
      <tr><td><span class="r-strong" data-rating="strong">🟢🟢 強力買入</span></td><td>75分以上</td><td>各方面均表現優秀，值得重點關注</td></tr>
      <tr><td><span class="r-watch" data-rating="watch">🟢 值得關注</span></td><td>60–74分</td><td>基本面良好，可納入觀察名單</td></tr>
      <tr><td><span class="r-hold" data-rating="hold">⚖️ 觀望</span></td><td>40–49分</td><td>有一定吸引力，但需留意風險</td></tr>
    </tbody>
  </table>
</div>

<div class="section" id="charts">
  <div class="section-title">市場概覽 — {report_date}</div>
  <div class="top-row">
    <div class="chart-card">
      <div class="chart-title">美股 US</div>
      <div class="legend">
        <span><span class="ld" style="background:#1D9E75"></span>強力買入 <strong>{stats['US']['strong']}</strong></span>
        <span><span class="ld" style="background:#378ADD"></span>值得關注 <strong>{stats['US']['watch']}</strong></span>
        <span><span class="ld" style="background:#EF9F27"></span>觀望 <strong>{stats['US']['hold']}</strong></span>
      </div>
      <svg id="us" viewBox="0 0 180 160" width="100%" height="160"></svg>
      <div class="chart-sub" data-zh-hk="{stats['US']['total']}隻 · 均分 {stats['US']['avg']} · 最高 {stats['US']['max']}" data-zh-cn="{stats['US']['total']}只 · 均分 {stats['US']['avg']} · 最高 {stats['US']['max']}" data-en="{stats['US']['total']} stks · Avg {stats['US']['avg']} · High {stats['US']['max']}">{stats['US']['total']}隻 · 均分 {stats['US']['avg']} · 最高 {stats['US']['max']}</div>
    </div>
    <div class="chart-card">
      <div class="chart-title">港股 HK</div>
      <div class="legend">
        <span><span class="ld" style="background:#1D9E75"></span>強力買入 <strong>{stats['HK']['strong']}</strong></span>
        <span><span class="ld" style="background:#378ADD"></span>值得關注 <strong>{stats['HK']['watch']}</strong></span>
        <span><span class="ld" style="background:#EF9F27"></span>觀望 <strong>{stats['HK']['hold']}</strong></span>
      </div>
      <svg id="hk" viewBox="0 0 180 160" width="100%" height="160"></svg>
      <div class="chart-sub" data-zh-hk="{stats['HK']['total']}隻 · 均分 {stats['HK']['avg']} · 最高 {stats['HK']['max']}" data-zh-cn="{stats['HK']['total']}只 · 均分 {stats['HK']['avg']} · 最高 {stats['HK']['max']}" data-en="{stats['HK']['total']} stks · Avg {stats['HK']['avg']} · High {stats['HK']['max']}">{stats['HK']['total']}隻 · 均分 {stats['HK']['avg']} · 最高 {stats['HK']['max']}</div>
    </div>
    <div class="chart-card">
      <div class="chart-title">英股 UK</div>
      <div class="legend">
        <span><span class="ld" style="background:#1D9E75"></span>強力買入 <strong>{stats['UK']['strong']}</strong></span>
        <span><span class="ld" style="background:#378ADD"></span>值得關注 <strong>{stats['UK']['watch']}</strong></span>
        <span><span class="ld" style="background:#EF9F27"></span>觀望 <strong>{stats['UK']['hold']}</strong></span>
      </div>
      <svg id="uk" viewBox="0 0 180 160" width="100%" height="160"></svg>
      <div class="chart-sub" data-zh-hk="{stats['UK']['total']}隻 · 均分 {stats['UK']['avg']} · 最高 {stats['UK']['max']}" data-zh-cn="{stats['UK']['total']}只 · 均分 {stats['UK']['avg']} · 最高 {stats['UK']['max']}" data-en="{stats['UK']['total']} stks · Avg {stats['UK']['avg']} · High {stats['UK']['max']}">{stats['UK']['total']}隻 · 均分 {stats['UK']['avg']} · 最高 {stats['UK']['max']}</div>
    </div>
    <div class="chart-card">
      <div class="chart-title">A股 CN</div>
      <div class="legend">
        <span><span class="ld" style="background:#1D9E75"></span>強力買入 <strong>{stats['CN']['strong']}</strong></span>
        <span><span class="ld" style="background:#378ADD"></span>值得關注 <strong>{stats['CN']['watch']}</strong></span>
        <span><span class="ld" style="background:#EF9F27"></span>觀望 <strong>{stats['CN']['hold']}</strong></span>
      </div>
      <svg id="cn" viewBox="0 0 180 160" width="100%" height="160"></svg>
      <div class="chart-sub" data-zh-hk="{stats['CN']['total']}隻 · 均分 {stats['CN']['avg']} · 最高 {stats['CN']['max']}" data-zh-cn="{stats['CN']['total']}只 · 均分 {stats['CN']['avg']} · 最高 {stats['CN']['max']}" data-en="{stats['CN']['total']} stks · Avg {stats['CN']['avg']} · High {stats['CN']['max']}">{stats['CN']['total']}隻 · 均分 {stats['CN']['avg']} · 最高 {stats['CN']['max']}</div>
    </div>
  </div>
  <div class="chart-card">
    <div class="chart-title">前{TOP_N}名評分分佈</div>
    <div class="legend">
      <span><span class="ld" style="background:#1D9E75"></span>強力買入</span>
      <span><span class="ld" style="background:#378ADD"></span>值得關注</span>
      <span><span class="ld" style="background:#EF9F27"></span>觀望</span>
    </div>
    <svg id="bar" viewBox="0 0 900 240" width="100%" height="240"></svg>
  </div>
</div>

<div class="section" id="picks">
  <div class="section-title">最新精選推介 — {report_date}</div>
  <p id="picks-note" style="font-size:13px;color:#999;margin-bottom:1.25rem">以下為今日評分最高股票（≥{MIN_SCORE}分），綠框為強力買入。</p>
  <div class="picks-grid" id="picksGrid"></div>
</div>


<div class="section" id="contact">
  <div class="section-title" id="contact-heading">合作機會</div>
  <div style="background:#fff;border-radius:16px;padding:2rem;text-align:center;border:1px solid #e5e5e5">
    <div style="font-size:2rem;margin-bottom:1rem">🤝</div>
    <h3 style="font-size:18px;font-weight:600;color:#222;margin-bottom:.75rem" id="contact-title">歡迎合作</h3>
    <p style="font-size:14px;color:#555;line-height:1.7;margin-bottom:1.5rem" id="contact-body">如您有興趣進行廣告合作、內容授權、數據合作或其他商業合作，歡迎透過以下方式聯絡我們。</p>
    <a href="mailto:prosynchk@gmail.com" style="display:inline-flex;align-items:center;gap:8px;background:#1D9E75;color:#fff;padding:10px 24px;border-radius:8px;text-decoration:none;font-size:14px;font-weight:600">
      <span>✉️</span>
      <span>prosynchk@gmail.com</span>
    </a>
  </div>
</div>
<div class="disclaimer">
  <div class="disclaimer-inner">
    <strong>免責聲明：</strong>本網站所有內容僅供參考及教育用途，不構成任何投資建議或買賣邀請。投資涉及風險，過往表現不代表未來回報。讀者應自行進行盡職審查，並在作出任何投資決定前諮詢持牌財務顧問。本站對因使用本站資料而引起的任何損失概不負責。
  </div>
</div>

<div class="footer">
  <div class="footer-inner">
    <span class="footer-logo">HiDH Dividend Analyst</span>
    <span>
      <a href="about.html" style="color:#aaa;text-decoration:none">關於我們</a> ·
      <a href="privacy.html" style="color:#aaa;text-decoration:none">私隱政策</a>
    </span>
    <span>© {datetime.date.today().year} prosynchk.com · 資料來源：Yahoo Finance · 僅供參考</span>
  </div>
</div>

<script>
const US_STATS = {json.dumps(stats['US'])};
const HK_STATS = {json.dumps(stats['HK'])};
const UK_STATS = {json.dumps(stats['UK'])};
const CN_STATS = {json.dumps(stats['CN'])};
const TOP10    = {top10_js};
const PICKS    = {picks_js};

function drawDonut(svgId, strong, watch, hold) {{
  const svg = document.getElementById(svgId);
  if(!svg) return;
  const values=[strong,watch,hold], colors=['#1D9E75','#378ADD','#EF9F27'];
  const cx=90,cy=80,R=65,r=42;
  const total=values.reduce((a,b)=>a+b,0);
  if(total===0) return;
  let angle=-Math.PI/2;
  values.forEach((v,i)=>{{
    if(v===0) return;
    const sweep=(v/total)*2*Math.PI;
    const x1=cx+R*Math.cos(angle),y1=cy+R*Math.sin(angle);
    const x2=cx+R*Math.cos(angle+sweep),y2=cy+R*Math.sin(angle+sweep);
    const ix1=cx+r*Math.cos(angle),iy1=cy+r*Math.sin(angle);
    const ix2=cx+r*Math.cos(angle+sweep),iy2=cy+r*Math.sin(angle+sweep);
    const large=sweep>Math.PI?1:0;
    const path=document.createElementNS('http://www.w3.org/2000/svg','path');
    path.setAttribute('d',`M ${{x1}} ${{y1}} A ${{R}} ${{R}} 0 ${{large}} 1 ${{x2}} ${{y2}} L ${{ix2}} ${{iy2}} A ${{r}} ${{r}} 0 ${{large}} 0 ${{ix1}} ${{iy1}} Z`);
    path.setAttribute('fill',colors[i]);
    svg.appendChild(path);
    if(v/total>0.03){{
      const mid=angle+sweep/2;
      const lx=cx+(R+r)/2*Math.cos(mid),ly=cy+(R+r)/2*Math.sin(mid);
      const t=document.createElementNS('http://www.w3.org/2000/svg','text');
      t.setAttribute('x',lx);t.setAttribute('y',ly+4);
      t.setAttribute('text-anchor','middle');t.setAttribute('font-size','11');
      t.setAttribute('fill','#fff');t.setAttribute('font-weight','600');
      t.textContent=v;svg.appendChild(t);
    }}
    angle+=sweep;
  }});
  const ct=document.createElementNS('http://www.w3.org/2000/svg','text');
  ct.setAttribute('x',cx);ct.setAttribute('y',cy+5);
  ct.setAttribute('text-anchor','middle');ct.setAttribute('font-size','13');
  ct.setAttribute('fill','#555');ct.setAttribute('font-weight','600');
  ct.textContent=total+'隻';svg.appendChild(ct);
}}

drawDonut('us',US_STATS.strong,US_STATS.watch,US_STATS.hold);
drawDonut('hk',HK_STATS.strong,HK_STATS.watch,HK_STATS.hold);
drawDonut('uk',UK_STATS.strong,UK_STATS.watch,UK_STATS.hold);
drawDonut('cn',CN_STATS.strong,CN_STATS.watch,CN_STATS.hold);

const RATING_COLOR={{strong:'#1D9E75',watch:'#378ADD',hold:'#EF9F27'}};
const MKT_STYLE={{
  US:{{bg:'#E6F1FB',text:'#185FA5'}},
  HK:{{bg:'#E1F5EE',text:'#0F6E56'}},
  UK:{{bg:'#FAEEDA',text:'#854F0B'}},
  CN:{{bg:'#FDECEA',text:'#B71C1C'}},
}};
const svg=document.getElementById('bar');
const W=900,H=240,padL=36,padB=46,padT=10,padR=10;
const chartW=W-padL-padR,chartH=H-padB-padT;
const scores=TOP10.map(s=>s.score);
const minScore=Math.max(40,Math.min(...scores)-5);
const maxScore=Math.min(100,Math.max(...scores)+5);
const barW=chartW/TOP10.length,gap=barW*0.22;

[...Array(6)].map((_,i)=>Math.round(minScore+i*(maxScore-minScore)/5)).forEach(v=>{{
  const y=padT+chartH-((v-minScore)/(maxScore-minScore))*chartH;
  const line=document.createElementNS('http://www.w3.org/2000/svg','line');
  line.setAttribute('x1',padL);line.setAttribute('x2',W-padR);
  line.setAttribute('y1',y);line.setAttribute('y2',y);
  line.setAttribute('stroke',v===Math.round(minScore)?'#ccc':'#eee');line.setAttribute('stroke-width','1');
  svg.appendChild(line);
  const t=document.createElementNS('http://www.w3.org/2000/svg','text');
  t.setAttribute('x',padL-4);t.setAttribute('y',y+4);
  t.setAttribute('text-anchor','end');t.setAttribute('font-size','10');
  t.setAttribute('fill','#999');t.textContent=v;svg.appendChild(t);
}});

TOP10.forEach((s,i)=>{{
  const barH=((s.score-minScore)/(maxScore-minScore))*chartH;
  const x=padL+i*barW+gap/2,y=padT+chartH-barH,w=barW-gap;
  const rect=document.createElementNS('http://www.w3.org/2000/svg','rect');
  rect.setAttribute('x',x);rect.setAttribute('y',y);
  rect.setAttribute('width',w);rect.setAttribute('height',barH);
  rect.setAttribute('fill',RATING_COLOR[s.rating]);rect.setAttribute('rx','3');
  svg.appendChild(rect);
  const ts=document.createElementNS('http://www.w3.org/2000/svg','text');
  ts.setAttribute('x',x+w/2);ts.setAttribute('y',y-4);
  ts.setAttribute('text-anchor','middle');ts.setAttribute('font-size','10');
  ts.setAttribute('fill','#555');ts.setAttribute('font-weight','600');
  ts.textContent=s.score;svg.appendChild(ts);
  const tl=document.createElementNS('http://www.w3.org/2000/svg','text');
  tl.setAttribute('x',x+w/2);tl.setAttribute('y',H-padB+14);
  tl.setAttribute('text-anchor','middle');tl.setAttribute('font-size','10');
  tl.setAttribute('fill','#777');tl.textContent=s.label;svg.appendChild(tl);
  const ms=MKT_STYLE[s.mkt];
  const bw=24,bh=13,bx=x+w/2-bw/2,by=H-padB+18;
  const brect=document.createElementNS('http://www.w3.org/2000/svg','rect');
  brect.setAttribute('x',bx);brect.setAttribute('y',by);
  brect.setAttribute('width',bw);brect.setAttribute('height',bh);
  brect.setAttribute('fill',ms.bg);brect.setAttribute('rx','3');
  svg.appendChild(brect);
  const bt=document.createElementNS('http://www.w3.org/2000/svg','text');
  bt.setAttribute('x',bx+bw/2);bt.setAttribute('y',by+9);
  bt.setAttribute('text-anchor','middle');bt.setAttribute('font-size','9');
  bt.setAttribute('fill',ms.text);bt.setAttribute('font-weight','600');
  bt.textContent=s.mkt;svg.appendChild(bt);
}});

const RATING_LABEL={{strong:'🟢🟢 強力買入',watch:'🟢 值得關注',hold:'⚖️ 觀望'}};
const RATING_CLASS={{strong:'r-strong',watch:'r-watch',hold:'r-hold'}};
const MKT_BADGE={{
  HK:'<span class="badge badge-hk">HK</span>',
  US:'<span class="badge badge-us">US</span>',
  UK:'<span class="badge badge-uk">UK</span>',
  CN:'<span class="badge badge-cn">CN</span>',
}};
document.getElementById('picksGrid').innerHTML=PICKS.map(p=>`
  <div class="pick-card${{p.rating==='strong'?' top':''}}">
    <div class="pick-header">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:4px;margin-bottom:4px">
        <div style="display:flex;align-items:center;gap:5px">
          <span style="font-size:14px;font-weight:600;color:#222;white-space:nowrap">${{p.ticker}}</span>
          ${{MKT_BADGE[p.mkt]}}
        </div>
        <span class="${{RATING_CLASS[p.rating]}}" style="white-space:nowrap;flex-shrink:0" data-rating="${{p.rating}}">${{RATING_LABEL[p.rating]}}</span>
      </div>
      <div class="pick-name">${{p.name}}</div>
    </div>
    <div class="pick-stats">
      <div class="pick-stat"><div class="pick-stat-label" data-zh-hk="股息率" data-zh-cn="股息率" data-en="Yield">股息率</div><div class="pick-stat-val">${{p.yield_}}%</div></div>
      <div class="pick-stat"><div class="pick-stat-label">PE</div><div class="pick-stat-val">${{p.pe}}x</div></div>
      <div class="pick-stat"><div class="pick-stat-label">P/B</div><div class="pick-stat-val">${{p.pb}}</div></div>
    </div>
    <div class="pros-cons">
      <div class="pros"><div class="pros-label">✅ 優點</div><div class="pros-text" data-zh-hk="${{p.pro_zh_hk}}" data-zh-cn="${{p.pro_zh_cn}}" data-en="${{p.pro_en}}">${{p.pro_zh_hk}}</div></div>
      <div class="cons"><div class="cons-label">⚠️ 缺點</div><div class="cons-text" data-zh-hk="${{p.con_zh_hk}}" data-zh-cn="${{p.con_zh_cn}}" data-en="${{p.con_en}}">${{p.con_zh_hk}}</div></div>
    </div>
    <div class="score-track"><div class="score-fill" style="width:${{p.score}}%"></div></div>
    <div class="score-row"><span>評分</span><span>${{p.score}}/100</span></div>
  </div>
`).join('');
</script>
</body>
</html>"""
    html = html.replace("</body>", "<script>" + SETLANG_JS + "</script></body>")
    return html

# ── 主程式 ────────────────────────────────────────────────
def main():
    print("=" * 50)
    print(f"generate_html.py  ({datetime.date.today()})")
    print("=" * 50)

    print("\n📂 讀取 Excel 數據庫...")
    stocks = load_excel_data()

    if not stocks:
        print("❌ 找不到任何股票數據，請先執行 daily_importer_global_v5.py")
        return

    print(f"\n✅ 合計 {len(stocks)} 隻股票")

    stats = get_market_stats(stocks)
    for mkt, s in stats.items():
        print(f"   {mkt}: {s['total']}隻，強力買入 {s['strong']}，值得關注 {s['watch']}，觀望 {s['hold']}，均分 {s['avg']}")

    report_date = str(datetime.date.today())
    html = generate_html(stocks, stats, report_date)

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(html)

    top5 = ', '.join(s['ticker'] for s in stocks[:5])
    print(f"\n✅ 已生成：{OUTPUT_FILE}")
    print(f"   前5名：{top5}")
    print(f"\n👉 下一步：上傳 {OUTPUT_FILE} 到 GitHub")

if __name__ == "__main__":
    main()#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_html.py
================
直接讀取 *_Dividend_Analysis.xlsx 總覽分頁，
生成最新的 index.html 上傳到 GitHub。

使用方法：
    python generate_html.py

每日執行順序：
    1. python daily_importer_global_v5.py
    2. python generate_html.py
    3. 上傳 index.html 到 GitHub

每週有新股時：
    1. python screener_global.py
    2. 揀股 → python add_to_tracking.py
    3. python batch_importer_global_v5.py
    4. python daily_importer_global_v5.py
    5. python generate_html.py
    6. 上傳 index.html 到 GitHub
"""

import datetime
import os
import json
import openpyxl
import pandas as pd

# ── 設定 ─────────────────────────────────────────────────
EXCEL_FILES = {
    "UK": "UK_Dividend_Analysis.xlsx",
    "HK": "HK_Dividend_Analysis.xlsx",
    "US": "US_Dividend_Analysis.xlsx",
    "CN": "CN_Dividend_Analysis.xlsx",
}
OUTPUT_BASE  = "web_site"   # 相對路徑；GitHub Actions 會直接將今日嗰個子資料夾發佈去 Pages，
                            # 唔會將呢個資料夾 commit 落 git（每次都係即跑即棄嘅 build output）
OUTPUT_FILE  = os.path.join(OUTPUT_BASE, datetime.date.today().strftime("%Y%m%d"), "index.html")
HISTORY_FILE = "pick_history.json"  # 記錄每隻股票「最早推介」日期與股價，長期累積、不會被每日的輸出資料夾覆蓋
                                     # （相對路徑＝repo根目錄，會被 commit 落 git 以保持跨日持久化）
TOP_N        = 10    # 柱狀圖顯示前N名
PICKS_N      = 12    # 推介卡片數量
AVOID_N      = 10    # 避開股票卡片數量（評分最低的N隻）
MIN_SCORE    = 50    # 入圍門檻（顯示用）
RATING_STRONG = 75
RATING_WATCH  = 60

# ── 讀取 Excel 總覽 ───────────────────────────────────────
def load_excel_data():
    """讀取三個市場的總覽分頁，合併成一個股票列表"""
    all_stocks = []

    for mkt, filepath in EXCEL_FILES.items():
        if not os.path.exists(filepath):
            print(f"  ⚠️  找不到 {filepath}，跳過")
            continue

        wb = openpyxl.load_workbook(filepath, data_only=True)
        summary_name = f"{mkt} 總覽"
        if summary_name not in wb.sheetnames:
            print(f"  ⚠️  找不到分頁「{summary_name}」，跳過")
            continue

        ws = wb[summary_name]
        headers = [str(ws.cell(1, c).value or "").strip()
                   for c in range(1, ws.max_column + 1)]

        def ci(name):
            try: return headers.index(name)
            except ValueError: return None

        i_ticker = ci("股票代號")
        i_name   = ci("公司名稱")
        i_price  = ci("現價")
        i_yield  = ci("最新股息率")
        i_pe     = ci("最新 PE")
        i_pb     = ci("P/B")
        i_score  = ci("📊 總分_100")
        i_diag   = ci("📊 綜合診斷")
        i_sq     = ci("S_股息質量")
        i_sv     = ci("S_估值")
        i_sf     = ci("S_財務健康")
        i_sg     = ci("S_增長")
        i_st     = ci("S_技術")

        if i_ticker is None or i_score is None:
            print(f"  ⚠️  {summary_name} 欄位不符，跳過")
            continue

        for row in range(2, ws.max_row + 1):
            ticker = str(ws.cell(row, i_ticker + 1).value or "").strip()
            score  = ws.cell(row, i_score + 1).value if i_score is not None else None
            if not ticker or score is None:
                continue
            try:
                score = float(score)
            except:
                continue

            all_stocks.append({
                "ticker": ticker,
                "name":   str(ws.cell(row, i_name  + 1).value or ticker).strip() if i_name  is not None else ticker,
                "mkt":    mkt,
                "price":  ws.cell(row, i_price + 1).value if i_price is not None else None,
                "yield":  ws.cell(row, i_yield + 1).value if i_yield is not None else None,
                "pe":     ws.cell(row, i_pe    + 1).value if i_pe    is not None else None,
                "pb":     ws.cell(row, i_pb    + 1).value if i_pb    is not None else None,
                "score":  score,
                "diag":   str(ws.cell(row, i_diag + 1).value or "").strip() if i_diag is not None else "",
                "sq":     ws.cell(row, i_sq + 1).value if i_sq is not None else None,
                "sv":     ws.cell(row, i_sv + 1).value if i_sv is not None else None,
                "sf":     ws.cell(row, i_sf + 1).value if i_sf is not None else None,
                "sg":     ws.cell(row, i_sg + 1).value if i_sg is not None else None,
                "st":     ws.cell(row, i_st + 1).value if i_st is not None else None,
            })

        print(f"  ✅ {mkt}: 讀取 {sum(1 for s in all_stocks if s['mkt']==mkt)} 隻")

    all_stocks.sort(key=lambda x: x["score"], reverse=True)
    return all_stocks

# ── 讀取各股歷年股息（用於 sparkline）────────────────────────
def _ticker_to_sheet(ticker, mkt):
    """將總覽 ticker 轉換成 Excel sheet name"""
    t = str(ticker)
    if mkt == "HK": return t.replace(".HK", "")
    if mkt == "CN": return t.replace(".", "_")
    if mkt == "UK": return t.replace(".L", "")
    return t  # US: 原樣

def load_dividend_history():
    """從四個 Excel 讀取每隻股票的歷年股息及季度股價，
    回傳 dict: { ticker -> {"annual":{year:amt}, "trend":"grow|cut|flat", "px_pts":[{v:float},...]} }
    以總覽 sheet 的 ticker 做 key，確保同 load_excel_data() 一致。"""
    cur_year = datetime.date.today().year
    result = {}
    for mkt, filepath in EXCEL_FILES.items():
        if not os.path.exists(filepath):
            continue
        try:
            xl = pd.ExcelFile(filepath)
            summary = f"{mkt} 總覽"
            sheet_set = set(xl.sheet_names)

            # 從總覽 sheet 建立 ticker -> sheet_name mapping
            wb_ov = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws_ov = wb_ov[summary]
            ov_headers = [ws_ov.cell(1, c).value for c in range(1, 30)]
            tc = next((i for i, h in enumerate(ov_headers) if h and "代號" in str(h)), None)
            if tc is None:
                wb_ov.close()
                continue
            ticker_sheet_map = {}
            for row in range(2, ws_ov.max_row + 1):
                t = ws_ov.cell(row, tc + 1).value
                if not t:
                    continue
                t = str(t).strip()
                sn = _ticker_to_sheet(t, mkt)
                if sn in sheet_set:
                    ticker_sheet_map[t] = sn
            wb_ov.close()

            before = len(result)
            for ticker, sheet_name in ticker_sheet_map.items():
                try:
                    df = xl.parse(sheet_name, usecols=[0, 1, 3], header=0)
                    df.columns = ["date", "close", "div"]
                    df["date"]  = pd.to_datetime(df["date"], errors="coerce")
                    df["close"] = pd.to_numeric(df["close"], errors="coerce")
                    df["div"]   = pd.to_numeric(df["div"],   errors="coerce")
                    df = df.dropna(subset=["date"])

                    # ── 股息：按年彙總，排除當年 ──
                    ddf = df[df["div"] > 0].copy()
                    ddf["year"] = ddf["date"].dt.year
                    ddf = ddf[ddf["year"] < cur_year]
                    annual = {}
                    if not ddf.empty:
                        annual = {int(y): round(float(v), 4)
                                  for y, v in ddf.groupby("year")["div"].sum().items()}
                    trend = _classify_div_trend(annual)

                    # ── 股價：季末收市，最多保留20個點 ──
                    pdf = df[df["close"].notna()].copy()
                    px_pts = []
                    if not pdf.empty:
                        pdf["qtr"] = pdf["date"].dt.to_period("Q")
                        q_close = pdf.groupby("qtr")["close"].last()
                        # 最近 20 季
                        q_close = q_close.iloc[-20:]
                        px_pts = [{"v": round(float(v), 3)} for v in q_close.values]

                    result[ticker] = {"annual": annual, "trend": trend, "px_pts": px_pts}
                except Exception:
                    pass
            print(f"  📈 {mkt}: {len(result)-before} 隻股息+股價歷史")
        except Exception as e:
            print(f"  ⚠️  {mkt} 歷史讀取失敗：{e}")
    print(f"  ✅ 歷史合計：{len(result)} 隻")
    return result

def _classify_div_trend(annual):
    vals = list(annual.values())
    if len(vals) < 2:
        return "flat"
    cuts  = sum(1 for i in range(1, len(vals)) if vals[i] < vals[i-1] * 0.95)
    grows = sum(1 for i in range(1, len(vals)) if vals[i] > vals[i-1] * 1.01)
    if cuts > 0:
        return "cut"
    if grows >= len(vals) - 1:
        return "grow"
    return "flat"

def get_rating_key(score):
    if score >= RATING_STRONG: return "strong"
    if score >= RATING_WATCH:  return "watch"
    return "hold"

def get_market_stats(stocks):
    stats = {}
    for mkt in ["US", "HK", "UK", "CN"]:
        ms = [s for s in stocks if s["mkt"] == mkt]
        if not ms:
            stats[mkt] = {"total":0,"strong":0,"watch":0,"hold":0,"avg":0,"max":0}
            continue
        stats[mkt] = {
            "total":  len(ms),
            "strong": sum(1 for s in ms if s["score"] >= RATING_STRONG),
            "watch":  sum(1 for s in ms if RATING_WATCH <= s["score"] < RATING_STRONG),
            "hold":   sum(1 for s in ms if s["score"] < RATING_WATCH),
            "avg":    round(sum(s["score"] for s in ms) / len(ms), 1),
            "max":    int(max(s["score"] for s in ms)),
        }
    return stats

def fmt(v, decimals=1, suffix=""):
    if v is None: return "─"
    try: return f"{float(v):.{decimals}f}{suffix}"
    except: return "─"

# ── 推介／避開歷史記錄（最早推介或避開日期／股價）──────────
# 每隻股票的記錄是一個「期間」清單：[{date, price, end_date}, ...]
# end_date 是 None 代表這段期間還在進行中（目前仍在名單裡）；
# 一旦跌出名單，這段期間就會被封存（end_date 填上封存當天），永遠不再更動；
# 之後同一隻股票重新入選/重新被列為避開，會另外開一筆新的期間，舊的那筆完全不受影響。
# key 的格式是 "pick:HK:0001.HK" 或 "avoid:US:XYZ"，用 kind 前綴把「推介」跟「避開」分開，
# 同一隻股票理論上可以同時有自己的推介歷史與避開歷史，互不影響。
def load_history():
    """讀取長期累積的歷史記錄；檔案不存在或損毀時回傳空白記錄。
    也會自動把舊版資料轉成新版格式（單一物件→清單；沒有 kind 前綴的舊 key→視為 pick:）。"""
    if not os.path.exists(HISTORY_FILE):
        return {}
    try:
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception as e:
        print(f"  ⚠️  讀取 {HISTORY_FILE} 失敗（{e}），視為空白記錄重新開始")
        return {}

    migrated = {}
    for key, val in raw.items():
        if isinstance(val, dict):
            val = [{"date": val.get("date"), "price": val.get("price"), "end_date": None}]
        if not (key.startswith("pick:") or key.startswith("avoid:")):
            key = f"pick:{key}"
        migrated[key] = val
    return migrated

def save_history(history):
    d = os.path.dirname(HISTORY_FILE)
    if d:
        os.makedirs(d, exist_ok=True)
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def history_key(s, kind="pick"):
    return f"{kind}:{s['mkt']}:{s['ticker']}"

def get_active_episode(history, key):
    """回傳目前進行中的那一筆期間（end_date 是 None），沒有就回傳 None"""
    eps = history.get(key)
    if eps and eps[-1].get("end_date") is None:
        return eps[-1]
    return None

def reconcile_history(history, items, today_str, kind="pick"):
    """每天執行一次（推介、避開各跑一次，用 kind 區分）：
    1. 之前在追蹤、但今天已經不在名單（評分轉好/轉弱、被剔除）的股票，
       把它目前進行中的那筆記錄封存（end_date=今天），之後永遠不會再更新。
    2. 今天在名單裡的股票：
       - 完全沒記錄過，或上一筆記錄已經封存（代表是重新入選/重新被列為避開）→ 開一筆全新記錄。
       - 上一筆記錄還在進行中 → 完全不動，沿用原本的日期與股價。
    回傳 (新開期間數, 封存期間數)。"""
    prefix = f"{kind}:"
    active_keys = {history_key(s, kind) for s in items}
    opened = closed = 0

    for key, eps in history.items():
        if not key.startswith(prefix) or key in active_keys:
            continue
        if eps and eps[-1].get("end_date") is None:
            eps[-1]["end_date"] = today_str
            closed += 1

    for s in items:
        key = history_key(s, kind)
        eps = history.setdefault(key, [])
        if not eps or eps[-1].get("end_date") is not None:
            try:
                price = float(s["price"])
            except (TypeError, ValueError):
                price = None
            eps.append({"date": today_str, "price": price, "end_date": None})
            opened += 1

    return opened, closed

def calc_change(current, base):
    """計算漲跌幅，回傳顯示文字與顏色"""
    try:
        current = float(current)
        base    = float(base)
        if base == 0:
            return {"text": "─", "color": "#999"}
        pct = (current - base) / base * 100
        sign = "+" if pct > 0 else ""
        color = "#0F6E56" if pct > 0 else ("#B42318" if pct < 0 else "#888")
        return {"text": f"{sign}{pct:.1f}%", "color": color}
    except (TypeError, ValueError):
        return {"text": "─", "color": "#999"}

# ── 生成 HTML ─────────────────────────────────────────────
SETLANG_JS = """
function setLang(lang) {
  try{localStorage.setItem('hidh_lang',lang);}catch(e){}
  ['#btn-zh-hk','#btn-zh-cn','#btn-en'].forEach(function(id){
    var el=document.querySelector(id);
    if(el){el.style.background='transparent';el.style.color='#666';el.style.borderColor='#e5e5e5';}
  });
  var ab=document.querySelector({'zh-hk':'#btn-zh-hk','zh-cn':'#btn-zh-cn','en':'#btn-en'}[lang]);
  if(ab){ab.style.background='#1D9E75';ab.style.color='#fff';ab.style.borderColor='#1D9E75';}
  var T={
    'zh-hk':{heroTitle:'全球高息股<br>每日精選分析',heroSub:'覆蓋香港、美國、英國及A股四大市場，以系統化評分篩選出具備穩定派息能力的優質股票。',heroTag:'每日更新',stat1:'今日追蹤股票數',stat2:'入圍股票（≥60分）',stat3:'強力買入',stat4:'今日最高分',mktHK:'港股 HK',mktUS:'美股 US',mktUK:'英股 UK',mktCN:'A股 CN',cUS:'美股 US',cHK:'港股 HK',cUK:'英股 UK',cCN:'A股 CN',top10:'前10名評分分佈',sec0:'選股理念與方法',sec1:'評分系統',sec2:'市場概覽',sec3:'最新精選推介',sec4:'高危名單',ah:['選股理念','四大市場','系統化篩選','每日更新'],ap:['長期穩定的股息收入是財富增長的重要基石。我們不單看當前息率高低，更重視企業的派息可持續性、財務健康狀況及估值合理性。','同步覆蓋香港、美國、英國及A股四大市場，以統一標準進行跨市場比較，讓投資者掌握全球高息機會。','每日自動更新數據，以量化評分模型對數百隻股票進行排名，過濾雜訊，聚焦真正值得關注的機會。','每個交易日收市後自動重新評分，確保推介反映最新的估值及財務狀況。'],scoreIntro:'每隻股票以100分制進行綜合評分，涵蓋五個範疇：',th:['評分範疇','滿分','主要考量'],rows:[['股息質量','30分','息率水平、派息穩定性及覆蓋率'],['估值','25分','現價相對歷史息率及市場的吸引程度'],['財務健康','25分','資產負債、現金流及償債能力'],['增長','10分','股息增長趨勢及盈利前景'],['技術走勢','10分','RSI、52週位置等技術指標']],rth:['評級','分數','意義'],rmean:['各方面均表現優秀，值得重點關注','基本面良好，可納入觀察名單','有一定吸引力，但需留意風險'],ll:['強力買入','值得關注','觀望'],unit:'隻',rl:{'strong':'🟢🟢 強力買入','watch':'🟢 值得關注','hold':'⚖️ 觀望'},picksNote:'以下為今日評分最高股票（≥50分），綠框為強力買入。',avoidNote:'以下為今日評分最低的股票（最多10隻），紅框代表高危股票，數據與上方推介股一致。',pros:'✅ 優點',cons:'⚠️ 缺點',score:'評分',yieldLbl:'股息率',discLabel:'免責聲明',disc:'本網站所有內容僅供參考及教育用途，不構成任何投資建議或買賣邀請。投資涉及風險，過往表現不代表未來回報。讀者應自行進行盡職審查，並在作出任何投資決定前諮詢持牌財務顧問。',f2:'每個交易日更新',f3:'資料來源：Yahoo Finance · 僅供參考'},
    'zh-cn':{heroTitle:'全球高息股<br>每日精选分析',heroSub:'覆盖香港、美国、英国及A股四大市场，以系统化评分筛选出具备稳定派息能力的优质股票。',heroTag:'每日更新',stat1:'今日追踪股票数',stat2:'入围股票（≥60分）',stat3:'强力买入',stat4:'今日最高分',mktHK:'港股 HK',mktUS:'美股 US',mktUK:'英股 UK',mktCN:'A股 CN',cUS:'美股 US',cHK:'港股 HK',cUK:'英股 UK',cCN:'A股 CN',top10:'前10名评分分布',sec0:'选股理念与方法',sec1:'评分系统',sec2:'市场概览',sec3:'最新精选推介',sec4:'高危名单',ah:['选股理念','四大市场','系统化筛选','每日更新'],ap:['长期稳定的股息收入是财富增长的重要基石。我们不单看当前息率高低，更重视企业的派息可持续性、财务健康状况及估值合理性。','同步覆盖香港、美国、英国及A股四大市场，以统一标准进行跨市场比较，让投资者掌握全球高息机会。','每日自动更新数据，以量化评分模型对数百只股票进行排名，过滤杂讯，聚焦真正值得关注的机会。','每个交易日收市后自动重新评分，确保推介反映最新的估值及财务状况。'],scoreIntro:'每只股票以100分制进行综合评分，涵盖五个范畴：',th:['评分范畴','满分','主要考量'],rows:[['股息质量','30分','息率水平、派息稳定性及覆盖率'],['估值','25分','现价相对历史息率及市场的吸引程度'],['财务健康','25分','资产负债、现金流及偿债能力'],['增长','10分','股息增长趋势及盈利前景'],['技术走势','10分','RSI、52周位置等技术指标']],rth:['评级','分数','意义'],rmean:['各方面均表现优秀，值得重点关注','基本面良好，可纳入观察名单','有一定吸引力，但需留意风险'],ll:['强力买入','值得关注','观望'],unit:'只',rl:{'strong':'🟢🟢 强力买入','watch':'🟢 值得关注','hold':'⚖️ 观望'},picksNote:'以下为今日评分最高股票（≥50分），绿框为强力买入。',avoidNote:'以下为今日评分最低的股票（最多10只），红框代表高危股票，数据与上方推介股一致。',pros:'✅ 优点',cons:'⚠️ 缺点',score:'评分',yieldLbl:'股息率',discLabel:'免责声明',disc:'本网站所有内容仅供参考及教育用途，不构成任何投资建议或买卖邀请。投资涉及风险，过往表现不代表未来回报。读者应自行进行尽职审查，并在作出任何投资决定前咨询持牌财务顾问。',f2:'每个交易日更新',f3:'资料来源：Yahoo Finance · 仅供参考'},
    'en':{heroTitle:'Global Dividend Stocks<br>Daily Analysis',heroSub:'Covering HK, US, UK and China A-shares with systematic scoring to identify quality dividend stocks.',heroTag:'Daily Update',stat1:'Stocks Tracked',stat2:'Qualified (≥60pts)',stat3:'Strong Buy',stat4:"Today\'s High",mktHK:'HK Stocks',mktUS:'US Stocks',mktUK:'UK Stocks',mktCN:'China A-Shares',cUS:'US Stocks',cHK:'HK Stocks',cUK:'UK Stocks',cCN:'China A-Shares',top10:'Top 10 Score Distribution',sec0:'Investment Philosophy',sec1:'Scoring System',sec2:'Market Overview',sec3:'Top Picks',sec4:'High-Risk List',ah:['Philosophy','4 Markets','Systematic Screening','Daily Update'],ap:['We focus not just on yield but on dividend sustainability, financial health and valuation to find quality long-term holdings.','Covering HK, US, UK and China A-shares with a unified scoring framework for cross-market comparison.','Daily automated updates with quantitative scoring to rank hundreds of stocks and surface the best opportunities.','Re-scored every trading day after market close to reflect the latest valuations and conditions.'],scoreIntro:'Each stock is scored on a 100-point scale across five dimensions:',th:['Category','Max','Key Criteria'],rows:[['Dividend Quality','30pts','Yield level, payout stability & coverage'],['Valuation','25pts','Current price vs historical yield & attractiveness'],['Financial Health','25pts','Balance sheet, cash flow & debt coverage'],['Growth','10pts','Dividend growth trend & earnings outlook'],['Technical','10pts','RSI, 52-week position & other indicators']],rth:['Rating','Score','Meaning'],rmean:['Excellent across all dimensions, high priority','Good fundamentals, worth monitoring','Some appeal, monitor risks'],ll:['Strong Buy','Watch','Hold'],unit:'stk',rl:{'strong':'🟢🟢 Strong Buy','watch':'🟢 Watch','hold':'⚖️ Hold'},picksNote:'Top-rated stocks today (≥50pts). Green border = Strong Buy.',avoidNote:'Lowest-scoring stocks today (up to 10). Red border = high risk. Same data fields as the picks above.',pros:'✅ Pro',cons:'⚠️ Con',score:'Score',yieldLbl:'Yield',discLabel:'Disclaimer',disc:'All content is for reference and educational purposes only. Not investment advice. Investing involves risk. Past performance does not guarantee future results.',f2:'Updated every trading day',f3:'Data: Yahoo Finance · For reference only'},
  };
  var t=T[lang]||T['zh-hk'];
  document.title=(lang==='en'?'Global Dividend Analysis':'全球高息股分析')+' | HiDH Dividend Analyst';
  document.querySelectorAll('a[data-zh-hk]').forEach(function(el){el.textContent=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');});
  var h1=document.querySelector('.hero h1');if(h1)h1.innerHTML=t.heroTitle;
  var hs=document.querySelector('.hero-sub');if(hs)hs.textContent=t.heroSub;
  var ht=document.querySelector('.hero-tag');if(ht){var p=ht.textContent.split('·');ht.textContent=t.heroTag+' · '+(p[1]||'').trim();}
  var sl=document.querySelectorAll('.stat-label');[t.stat1,t.stat2,t.stat3,t.stat4].forEach(function(s,i){if(sl[i])sl[i].textContent=s;});
  var hb=document.querySelectorAll('.hero-badges .badge');[t.mktHK,t.mktUS,t.mktUK,t.mktCN].forEach(function(s,i){if(hb[i])hb[i].textContent=s;});
  var ct=document.querySelectorAll('.chart-title');[t.cUS,t.cHK,t.cUK,t.cCN,t.top10].forEach(function(s,i){if(ct[i])ct[i].firstChild.textContent=s;});
  var st=document.querySelectorAll('.section-title');
  if(st[0])st[0].childNodes[0].textContent=t.sec0;
  if(st[1])st[1].childNodes[0].textContent=t.sec1;
  if(st[2]){var d2=st[2].textContent.match(/[0-9]{4}-[0-9]{2}-[0-9]{2}/);st[2].childNodes[0].textContent=t.sec2+(d2?' — '+d2[0]:'');}
  if(st[3]){var d3=st[3].textContent.match(/[0-9]{4}-[0-9]{2}-[0-9]{2}/);st[3].childNodes[0].textContent=t.sec3+(d3?' — '+d3[0]:'');}
  if(st[4]){var d4=st[4].textContent.match(/[0-9]{4}-[0-9]{2}-[0-9]{2}/);st[4].childNodes[0].textContent=t.sec4+(d4?' — '+d4[0]:'');}
  var ag=document.getElementById('about-grid');
  if(ag){ag.querySelectorAll('h3').forEach(function(el,i){if(t.ah[i])el.textContent=t.ah[i];});ag.querySelectorAll('p').forEach(function(el,i){if(t.ap[i])el.textContent=t.ap[i];});}
  var si=document.getElementById('score-intro');if(si)si.textContent=t.scoreIntro;
  var tbls=document.querySelectorAll('.score-table');
  if(tbls[0]){tbls[0].querySelectorAll('th').forEach(function(el,i){if(t.th[i])el.textContent=t.th[i];});tbls[0].querySelectorAll('tbody tr').forEach(function(tr,i){if(t.rows[i]){var tds=tr.querySelectorAll('td');tds.forEach(function(td,j){if(t.rows[i][j])td.textContent=t.rows[i][j];});}});}
  if(tbls[1]){tbls[1].querySelectorAll('th').forEach(function(el,i){if(t.rth[i])el.textContent=t.rth[i];});tbls[1].querySelectorAll('tbody tr').forEach(function(tr,i){var tds=tr.querySelectorAll('td');if(tds[1])tds[1].textContent=tds[1].getAttribute('data-'+lang)||tds[1].getAttribute('data-zh-hk')||tds[1].textContent;if(tds[2]&&t.rmean[i])tds[2].textContent=t.rmean[i];});}
  document.querySelectorAll('[data-rating]').forEach(function(el){var r=el.getAttribute('data-rating');if(t.rl[r])el.textContent=t.rl[r];});
  document.querySelectorAll('.legend span[data-zh-hk]').forEach(function(el){var k=el.getAttribute('data-zh-hk'),strong=el.querySelector('strong'),ldEl=el.querySelector('.ld'),ld=ldEl?ldEl.outerHTML:'',idx=['強力買入','值得關注','觀望'].indexOf(k);if(idx>=0)el.innerHTML=ld+t.ll[idx]+(strong?' '+strong.outerHTML:'');});
  document.querySelectorAll('.chart-sub[data-zh-hk]').forEach(function(el){el.textContent=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');});
  document.querySelectorAll('svg[id]').forEach(function(svg){var texts=svg.querySelectorAll('text'),last=texts[texts.length-1];if(last){var m=last.textContent.match(/^([0-9]+)/);if(m)last.textContent=m[1]+t.unit;}});
  var pn=document.getElementById('picks-note');if(pn)pn.textContent=t.picksNote;
  var an=document.getElementById('avoid-note');if(an)an.textContent=t.avoidNote;
  document.querySelectorAll('.pros-label').forEach(function(el){el.textContent=t.pros;});
  document.querySelectorAll('.cons-label').forEach(function(el){el.textContent=t.cons;});
  document.querySelectorAll('.score-row span:first-child').forEach(function(el){el.textContent=t.score;});
  document.querySelectorAll('.pick-stat-label[data-zh-hk], .pick-track-label[data-zh-hk], .r-avoid[data-zh-hk], .risk-label[data-zh-hk], .risk-text[data-zh-hk]').forEach(function(el){el.textContent=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');});
  document.querySelectorAll('.pros-text').forEach(function(el){el.textContent=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');});
  document.querySelectorAll('.cons-text').forEach(function(el){el.textContent=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');});
  var fi=document.querySelectorAll('.footer-inner span');if(fi[2])fi[2].textContent='© 2026 prosynchk.com · '+t.f2+' · '+t.f3;
  var disc=document.querySelector('.disclaimer-inner');if(disc)disc.innerHTML='<strong>'+t.discLabel+'：</strong>'+t.disc;
  var ch={'zh-hk':['合作機會','歡迎合作','如您有興趣進行廣告合作、內容授權、數據合作或其他商業合作，歡迎透過以下方式聯絡我們。'],'zh-cn':['合作机会','欢迎合作','如您有兴趣进行广告合作、内容授权、数据合作或其他商业合作，欢迎通过以下方式联系我们。'],'en':['Partnership','Work With Us','Interested in advertising, content licensing, data collaboration or other business opportunities? Get in touch.']};
  var cl=ch[lang]||ch['zh-hk'];
  var chead=document.getElementById('contact-heading');if(chead)chead.childNodes[0].textContent=cl[0];
  var ctitle=document.getElementById('contact-title');if(ctitle)ctitle.textContent=cl[1];
  var cbody=document.getElementById('contact-body');if(cbody)cbody.textContent=cl[2];
  document.querySelectorAll('.div-spark-desc-lbl[data-zh-hk]').forEach(function(el){el.textContent=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');});
  document.querySelectorAll('.div-spark-leg[data-zh-hk]').forEach(function(el){var line=el.querySelector('.div-spark-leg-line');var lbl=el.getAttribute('data-'+lang)||el.getAttribute('data-zh-hk');el.innerHTML=(line?line.outerHTML:'')+lbl;});
}
(function(){var l=localStorage.getItem('hidh_lang');if(l&&l!=='zh-hk')setLang(l);})();
"""



def generate_html(stocks, stats, report_date, div_hist=None):
    if div_hist is None:
        div_hist = {}
    top10   = stocks[:TOP_N]
    picks   = [s for s in stocks if s["score"] >= MIN_SCORE][:PICKS_N]
    avoided = sorted(stocks, key=lambda x: x["score"])[:AVOID_N]   # 評分最低的N隻，列入高危名單
    total  = sum(s["total"] for s in stats.values())
    n_strong = sum(s["strong"] for s in stats.values())
    max_score  = int(stocks[0]["score"]) if stocks else 0
    max_ticker = stocks[0]["ticker"] if stocks else ""

    # 載入歷史記錄：推介、避開各自校對一次 —
    # 封存跌出名單的舊期間、為新入選/重新入選的股票開新期間
    history = load_history()
    op, cp = reconcile_history(history, picks,   report_date, kind="pick")
    oa, ca = reconcile_history(history, avoided, report_date, kind="avoid")
    if op or cp or oa or ca:
        save_history(history)
        parts = []
        if op: parts.append(f"推介新開 {op} 筆")
        if cp: parts.append(f"推介封存 {cp} 筆")
        if oa: parts.append(f"避開新開 {oa} 筆")
        if ca: parts.append(f"避開封存 {ca} 筆")
        print(f"  🔄 歷史記錄更新：{'、'.join(parts)}（{HISTORY_FILE}）")

    top10_js = json.dumps([{
        "label":  s["ticker"],
        "score":  int(s["score"]),
        "mkt":    s["mkt"],
        "rating": get_rating_key(s["score"]),
    } for s in top10])

    def get_metric_signals(s):
        """根據各項指標，列出所有適用的優點/缺點原因清單，供 get_pros_cons 與 get_risk_reasons 共用"""
        phk,chk,pcn,ccn,pen,cen = [],[],[],[],[],[]
        try:
            yld = float(s["yield"]) if s["yield"] else 0
            pe  = float(s["pe"])    if s["pe"]    else 0
            sq  = float(s["sq"])    if s["sq"]    else 0
            sv  = float(s["sv"])    if s["sv"]    else 0
            sf  = float(s["sf"])    if s["sf"]    else 0
            sg  = float(s["sg"])    if s["sg"]    else 0
            if yld >= 6:   phk.append(f"高息率 {yld:.1f}%，現金回報吸引"); pcn.append(f"高股息率 {yld:.1f}%，现金回报吸引"); pen.append(f"High yield {yld:.1f}%, attractive income")
            elif yld >= 4: phk.append(f"息率 {yld:.1f}%，優於市場平均"); pcn.append(f"股息率 {yld:.1f}%，优于市场平均"); pen.append(f"Yield {yld:.1f}%, above market average")
            if sq >= 22:   phk.append("股息質量高，派息穩定可靠"); pcn.append("股息质量高，派息稳定可靠"); pen.append("High dividend quality, stable payout")
            if sv >= 20:   phk.append("估值具吸引力，現價相對合理"); pcn.append("估值具吸引力，现价相对合理"); pen.append("Attractive valuation at current price")
            if sf >= 20:   phk.append("財務健康，資產負債穩健"); pcn.append("财务健康，资产负债稳健"); pen.append("Strong balance sheet")
            if sg >= 7:    phk.append("股息增長趨勢良好"); pcn.append("股息增长趋势良好"); pen.append("Positive dividend growth trend")
            if 0 < pe < 12:phk.append(f"PE {pe:.1f}x，估值偏低"); pcn.append(f"PE {pe:.1f}x，估值偏低"); pen.append(f"Low PE of {pe:.1f}x")
            if yld < 3:    chk.append("息率偏低，收息吸引力有限"); ccn.append("股息率偏低，收息吸引力有限"); cen.append("Low yield limits income appeal")
            if sq < 15:    chk.append("股息質量一般，需留意派息持續性"); ccn.append("股息质量一般，需留意派息持续性"); cen.append("Moderate dividend quality, sustainability risk")
            if sv < 12:    chk.append("估值偏高，上升空間有限"); ccn.append("估值偏高，上升空间有限"); cen.append("Rich valuation limits upside")
            if sf < 15:    chk.append("財務健康度一般，需關注負債"); ccn.append("财务健康度一般，需关注负债"); cen.append("Moderate financials, watch debt levels")
            if sg < 5:     chk.append("股息增長動力不足"); ccn.append("股息增长动力不足"); cen.append("Limited dividend growth momentum")
            if pe > 25:    chk.append(f"PE {pe:.1f}x，估值偏貴"); ccn.append(f"PE {pe:.1f}x，估值偏贵"); cen.append(f"High PE of {pe:.1f}x")
        except: pass
        return phk,chk,pcn,ccn,pen,cen

    def get_pros_cons(s):
        phk,chk,pcn,ccn,pen,cen = get_metric_signals(s)
        return {
            "pro_zh_hk": phk[0] if phk else "評分良好，具備一定投資價值",
            "con_zh_hk": chk[0] if chk else "需留意市場風險及行業波動",
            "pro_zh_cn": pcn[0] if pcn else "评分良好，具备一定投资价值",
            "con_zh_cn": ccn[0] if ccn else "需留意市场风险及行业波动",
            "pro_en":    pen[0] if pen else "Good overall score, investment potential",
            "con_en":    cen[0] if cen else "Monitor market and sector risks",
        }

    def get_risk_reasons(s):
        """彙整這隻股票所有偏弱的指標，作為「高危原因」——跟 get_pros_cons 不同，這裡不只取第一項，
        而是把全部適用的弱點都列出來，因為高危卡片不需要優點，只需要把低分原因講清楚。"""
        _,chk,_,ccn,_,cen = get_metric_signals(s)
        if not chk:
            chk, ccn, cen = ["整體評分偏低，建議謹慎評估"], ["整体评分偏低，建议谨慎评估"], ["Overall score is low, evaluate with caution"]
        return {
            "risk_zh_hk": "；".join(chk),
            "risk_zh_cn": "；".join(ccn),
            "risk_en":    "; ".join(cen),
        }

    def get_track_record(s, kind="pick"):
        """根據目前進行中的期間，算出這隻股票的日期、當時股價、現價與漲跌"""
        ep = get_active_episode(history, history_key(s, kind)) or {"date": report_date, "price": s["price"]}
        first_date  = ep.get("date", report_date)
        first_price = ep.get("price")
        chg = calc_change(s["price"], first_price)
        return {
            "first_date":  first_date,
            "first_price": fmt(first_price, 2),
            "price":       fmt(s["price"], 2),
            "chg_text":    chg["text"],
            "chg_color":   chg["color"],
        }

    def get_div_data(s):
        """取出該股的年度股息 sparkline 資料及季度股價"""
        info   = div_hist.get(s["ticker"], {})
        annual = info.get("annual", {})
        trend  = info.get("trend", "flat")
        px_pts = info.get("px_pts", [])
        years  = sorted(annual.keys())[-6:]
        pts    = [{"y": yr, "v": annual[yr]} for yr in years]
        return {"div_pts": pts, "div_trend": trend, "px_pts": px_pts}

    picks_js = json.dumps([{
        "ticker": s["ticker"],
        "name":   s["name"],
        "mkt":    s["mkt"],
        "score":  int(s["score"]),
        "rating": get_rating_key(s["score"]),
        "yield_": fmt(s["yield"]),
        "pe":     fmt(s["pe"]),
        "pb":     fmt(s["pb"], 2),
        **get_pros_cons(s),
        **get_track_record(s, "pick"),
        **get_div_data(s),
    } for s in picks])

    avoid_js = json.dumps([{
        "ticker": s["ticker"],
        "name":   s["name"],
        "mkt":    s["mkt"],
        "score":  int(s["score"]),
        "yield_": fmt(s["yield"]),
        "pe":     fmt(s["pe"]),
        "pb":     fmt(s["pb"], 2),
        **get_risk_reasons(s),
        **get_track_record(s, "avoid"),
        **get_div_data(s),
    } for s in avoided])

    html = f"""<!DOCTYPE html>
<html lang="zh-HK">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>全球高息股分析 | HiDH Dividend Analyst</title>
<link rel="icon" type="image/png" href="/icon.png">
<link rel="apple-touch-icon" href="/icon.png">
<script async src="https://pagead2.googlesyndication.com/pagead/js/adsbygoogle.js?client=ca-pub-9533474113956980" crossorigin="anonymous"></script>
<!-- Google tag (gtag.js) -->
<script async src="https://www.googletagmanager.com/gtag/js?id=G-P1VNXYE3FB"></script>
<script>
  window.dataLayer = window.dataLayer || [];
  function gtag(){{dataLayer.push(arguments);}}
  gtag('js', new Date());
  gtag('config', 'G-P1VNXYE3FB');
</script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:system-ui,-apple-system,sans-serif;background:#f7f7f5;color:#333;line-height:1.7}}
a{{color:#1D9E75;text-decoration:none}}a:hover{{text-decoration:underline}}
.header{{background:#fff;border-bottom:1px solid #e5e5e5;padding:0 1.5rem}}
.header-inner{{max-width:960px;margin:0 auto;display:flex;align-items:center;justify-content:space-between;height:60px}}
.logo{{font-size:18px;font-weight:600;color:#1D9E75}}.logo span{{color:#333}}
.nav{{display:flex;gap:1.5rem;font-size:14px}}.nav a{{color:#555}}.nav a:hover{{color:#1D9E75;text-decoration:none}}
.hero{{background:#fff;border-bottom:1px solid #e5e5e5;padding:2.5rem 1.5rem}}
.hero-inner{{max-width:960px;margin:0 auto;display:grid;grid-template-columns:1fr 1fr;gap:2rem;align-items:center}}
.hero-tag{{font-size:11px;font-weight:600;color:#1D9E75;text-transform:uppercase;letter-spacing:.1em;margin-bottom:.5rem}}
.hero h1{{font-size:26px;font-weight:600;line-height:1.35;margin-bottom:.75rem;color:#222}}
.hero-sub{{font-size:15px;color:#666;margin-bottom:1.25rem}}
.hero-badges{{display:flex;gap:8px;flex-wrap:wrap}}
.badge{{font-size:11px;padding:4px 10px;border-radius:20px;font-weight:500}}
.badge-hk{{background:#E1F5EE;color:#0F6E56}}.badge-us{{background:#E6F1FB;color:#185FA5}}.badge-uk{{background:#FAEEDA;color:#854F0B}}.badge-cn{{background:#FDECEA;color:#B71C1C}}
.hero-stats{{display:grid;grid-template-columns:1fr 1fr;gap:12px}}
.stat-box{{background:#f7f7f5;border-radius:10px;padding:1rem}}
.stat-val{{font-size:24px;font-weight:600;color:#222}}.stat-label{{font-size:12px;color:#888;margin-top:2px}}
.section{{max-width:960px;margin:0 auto;padding:2rem 1.5rem}}
.section-title{{font-size:20px;font-weight:700;color:#222;text-transform:uppercase;letter-spacing:.02em;margin-bottom:1.25rem;display:flex;align-items:center;gap:10px}}
.section-title::after{{content:'';flex:1;height:1px;background:#e5e5e5}}
.about-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:12px;margin-bottom:1.5rem}}
.about-card{{background:#fff;border:1px solid #e5e5e5;border-radius:12px;padding:1.25rem}}
.about-icon{{font-size:20px;margin-bottom:.5rem}}
.about-card h3{{font-size:14px;font-weight:600;color:#222;margin-bottom:.4rem}}
.about-card p{{font-size:13px;color:#666;line-height:1.6}}
.score-table{{width:100%;border-collapse:collapse;font-size:13px;background:#fff;border-radius:12px;overflow:hidden;border:1px solid #e5e5e5}}
.score-table th{{background:#f7f7f5;padding:10px 14px;text-align:left;font-weight:600;color:#555;font-size:12px}}
.score-table td{{padding:10px 14px;border-top:1px solid #f0f0f0;color:#444}}
.r-strong{{background:#E1F5EE;color:#0F6E56;padding:3px 8px;border-radius:4px;font-size:11px;font-weight:600;white-space:nowrap}}
.r-watch{{background:#E6F1FB;color:#185FA5;padding:3px 8px;border-radius:4px;font-size:11px;font-weight:600;white-space:nowrap}}
.r-hold{{background:#FAEEDA;color:#854F0B;padding:3px 8px;border-radius:4px;font-size:11px;font-weight:600;white-space:nowrap}}
.r-avoid{{background:#FDECEA;color:#B71C1C;padding:3px 8px;border-radius:4px;font-size:11px;font-weight:600;white-space:nowrap}}
.top-row{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px}}
.chart-card{{background:#fff;border:1px solid #e5e5e5;border-radius:12px;padding:16px}}
.chart-title{{font-size:13px;font-weight:600;color:#444;margin-bottom:8px}}
.legend{{display:flex;gap:10px;margin-bottom:8px;font-size:11px;color:#666;flex-wrap:wrap}}
.ld{{width:10px;height:10px;border-radius:2px;display:inline-block;margin-right:3px;vertical-align:2px}}
.chart-sub{{text-align:center;font-size:11px;color:#888;margin-top:6px}}
svg text{{font-family:system-ui,sans-serif}}
.picks-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:12px}}
.pick-card{{background:#fff;border:1px solid #e5e5e5;border-radius:12px;padding:1.1rem}}
.pick-card.top{{border:2px solid #1D9E75}}
.pick-card.avoid{{border:2px solid #B71C1C}}
.pick-header{{display:block;margin-bottom:.5rem}}
.pick-ticker{{font-size:15px;font-weight:600;color:#222}}
.pick-name{{font-size:12px;color:#888;margin-bottom:.75rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.pick-stats{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:.75rem}}
.pick-stat{{background:#f7f7f5;border-radius:6px;padding:6px 8px}}
.pick-stat-label{{font-size:10px;color:#999}}.pick-stat-val{{font-size:13px;font-weight:600;color:#222}}
.pick-track{{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin:.75rem 0;padding:.6rem 0;border-top:1px dashed #eee;border-bottom:1px dashed #eee}}
.pick-track-item{{text-align:center}}
.pick-track-label{{font-size:9px;color:#999;margin-bottom:2px}}
.pick-track-val{{font-size:12px;font-weight:600;color:#222;white-space:nowrap}}
.score-track{{height:4px;background:#f0f0f0;border-radius:2px;overflow:hidden;margin-top:.5rem}}
.score-fill{{height:100%;border-radius:2px;background:#1D9E75}}
.score-fill.avoid{{background:#B71C1C}}
.score-row{{display:flex;justify-content:space-between;font-size:10px;color:#aaa;margin-top:3px}}
.pros-cons{{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-top:8px;margin-bottom:6px}}
.pros{{background:#F0FBF5;border-radius:6px;padding:6px 8px;font-size:11px;color:#0F6E56}}
.cons{{background:#FEF3F2;border-radius:6px;padding:6px 8px;font-size:11px;color:#B42318}}
.pros-label,.cons-label{{font-weight:600;margin-bottom:2px;font-size:10px}}
.risk-box{{background:#FEF3F2;border-radius:6px;padding:8px 10px;font-size:11px;color:#B42318;margin-top:8px;margin-bottom:6px}}
.risk-label{{font-weight:600;margin-bottom:3px;font-size:10px}}
.div-spark{{margin-top:8px;margin-bottom:2px}}
.div-spark-row{{display:flex;align-items:center;gap:6px;margin-bottom:4px}}
.div-trend-badge{{font-size:10px;font-weight:600;padding:2px 7px;border-radius:3px;white-space:nowrap}}
.div-trend-grow{{background:#E1F5EE;color:#0F6E56}}
.div-trend-cut{{background:#FEF3F2;color:#B42318}}
.div-trend-flat{{background:#FAEEDA;color:#854F0B}}
.div-spark svg{{display:block;width:100%;overflow:visible}}
.div-spark-legend{{display:flex;gap:10px;margin-bottom:4px;align-items:center}}
.div-spark-leg{{display:flex;align-items:center;gap:4px;font-size:9px;color:#888}}
.div-spark-leg-line{{width:12px;height:2px;border-radius:1px;display:inline-block}}
.div-spark-desc{{display:grid;grid-template-columns:1fr 1fr;gap:4px;margin-top:6px}}
.div-spark-desc-box{{border-radius:4px;padding:4px 7px}}
.div-spark-desc-lbl{{font-size:9px;font-weight:600;margin-bottom:1px}}
.div-spark-desc-val{{font-size:10px}}
.disclaimer{{background:#fff;border-top:1px solid #e5e5e5;padding:1.5rem;margin-top:2rem}}
.disclaimer-inner{{max-width:960px;margin:0 auto;font-size:12px;color:#aaa;line-height:1.7}}
.disclaimer-inner strong{{color:#888}}
.footer{{background:#222;color:#aaa;padding:1.5rem;font-size:12px}}
.footer-inner{{max-width:960px;margin:0 auto;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:.5rem}}
.footer-logo{{color:#1D9E75;font-weight:600;font-size:14px}}
@media(max-width:640px){{.hero-inner{{grid-template-columns:1fr}}.hero-stats{{display:none}}.top-row{{grid-template-columns:1fr}}}}
</style>
</head>
<body>

<div class="header">
  <div class="header-inner">
    <div class="logo">HiDH <span>Dividend Analyst</span></div>
    <nav class="nav">
      <a href="#about" data-zh-hk="選股方法" data-zh-cn="选股方法" data-en="Methodology">選股方法</a>
      <a href="#charts" data-zh-hk="市場概覽" data-zh-cn="市场概览" data-en="Markets">市場概覽</a>
      <a href="#picks" data-zh-hk="最新精選" data-zh-cn="最新精选" data-en="Top Picks">最新精選</a>
      <a href="#avoid" data-zh-hk="高危名單" data-zh-cn="高危名单" data-en="High-Risk List">高危名單</a>
    </nav>
  </div>
</div>
<div style="display:flex;justify-content:flex-end;gap:6px;padding:.4rem 1.5rem;background:#fff;border-bottom:0.5px solid #e5e5e5">
  <button id="btn-zh-hk" style="font-size:12px;padding:4px 12px;border-radius:20px;border:0.5px solid #1D9E75;background:#1D9E75;color:#fff;cursor:pointer" onclick="setLang('zh-hk')">繁體</button>
  <button id="btn-zh-cn" style="font-size:12px;padding:4px 12px;border-radius:20px;border:0.5px solid #e5e5e5;background:transparent;color:#666;cursor:pointer" onclick="setLang('zh-cn')">简体</button>
  <button id="btn-en"    style="font-size:12px;padding:4px 12px;border-radius:20px;border:0.5px solid #e5e5e5;background:transparent;color:#666;cursor:pointer" onclick="setLang('en')">English</button>
</div>

<div class="hero">
  <div class="hero-inner">
    <div>
      <div class="hero-tag">每日更新 · {report_date}</div>
      <h1>全球高息股<br>每日精選分析</h1>
      <p class="hero-sub">覆蓋香港、美國、英國及A股四大市場，以系統化評分篩選出具備穩定派息能力的優質股票。</p>
      <div class="hero-badges">
        <span class="badge badge-hk">港股 HK</span>
        <span class="badge badge-us">美股 US</span>
        <span class="badge badge-uk">英股 UK</span>
        <span class="badge badge-cn">A股 CN</span>
      </div>
    </div>
    <div class="hero-stats">
      <div class="stat-box"><div class="stat-val">{total}</div><div class="stat-label">今日追蹤股票數</div></div>
      <div class="stat-box"><div class="stat-val">{sum(s["strong"]+s["watch"] for s in stats.values())}</div><div class="stat-label">入圍股票（≥{RATING_WATCH}分）</div></div>
      <div class="stat-box"><div class="stat-val">{n_strong}</div><div class="stat-label">強力買入</div></div>
      <div class="stat-box"><div class="stat-val">{max_score}</div><div class="stat-label">今日最高分 ({max_ticker})</div></div>
    </div>
  </div>
</div>

<div class="section" id="about">
  <div class="section-title">選股理念與方法</div>
  <div class="about-grid" id="about-grid">
    <div class="about-card">
      <div class="about-icon">💡</div>
      <h3>選股理念</h3>
      <p>長期穩定的股息收入是財富增長的重要基石。我們不單看當前息率高低，更重視企業的派息可持續性、財務健康狀況及估值合理性。</p>
    </div>
    <div class="about-card">
      <div class="about-icon">🌍</div>
      <h3>四大市場</h3>
      <p>同步覆蓋香港、美國、英國及A股四大市場，以統一標準進行跨市場比較，讓投資者掌握全球高息機會。</p>
    </div>
    <div class="about-card">
      <div class="about-icon">📊</div>
      <h3>系統化篩選</h3>
      <p>每日自動更新數據，以量化評分模型對數百隻股票進行排名，過濾雜訊，聚焦真正值得關注的機會。</p>
    </div>
    <div class="about-card">
      <div class="about-icon">🔄</div>
      <h3>每日更新</h3>
      <p>每個交易日收市後自動重新評分，確保推介反映最新的估值及財務狀況。</p>
    </div>
  </div>

  <div class="section-title" style="margin-top:2rem">評分系統</div>
  <p id="score-intro" style="font-size:14px;color:#666;margin-bottom:1rem">每隻股票以100分制進行綜合評分，涵蓋五個範疇：</p>
  <table class="score-table" style="margin-bottom:1.25rem">
    <thead><tr><th>評分範疇</th><th>滿分</th><th>主要考量</th></tr></thead>
    <tbody>
      <tr><td>股息質量</td><td>30分</td><td>息率水平、派息穩定性及覆蓋率</td></tr>
      <tr><td>估值</td><td>25分</td><td>現價相對歷史息率及市場的吸引程度</td></tr>
      <tr><td>財務健康</td><td>25分</td><td>資產負債、現金流及償債能力</td></tr>
      <tr><td>增長</td><td>10分</td><td>股息增長趨勢及盈利前景</td></tr>
      <tr><td>技術走勢</td><td>10分</td><td>RSI、52週位置等技術指標</td></tr>
    </tbody>
  </table>
  <table class="score-table">
    <thead><tr><th>評級</th><th>分數</th><th>意義</th></tr></thead>
    <tbody>
      <tr><td><span class="r-strong" data-rating="strong">🟢🟢 強力買入</span></td><td data-zh-hk="75分以上" data-zh-cn="75分以上" data-en="75+">75分以上</td><td>各方面均表現優秀，值得重點關注</td></tr>
      <tr><td><span class="r-watch" data-rating="watch">🟢 值得關注</span></td><td data-zh-hk="60–74分" data-zh-cn="60–74分" data-en="60–74">60–74分</td><td>基本面良好，可納入觀察名單</td></tr>
      <tr><td><span class="r-hold" data-rating="hold">⚖️ 觀望</span></td><td data-zh-hk="40–49分" data-zh-cn="40–49分" data-en="40–49">40–49分</td><td>有一定吸引力，但需留意風險</td></tr>
    </tbody>
  </table>
</div>

<div class="section" id="charts">
  <div class="section-title">市場概覽 — {report_date}</div>
  <div class="top-row">
    <div class="chart-card">
      <div class="chart-title">美股 US</div>
      <div class="legend">
        <span data-zh-hk="強力買入"><span class="ld" style="background:#1D9E75"></span>強力買入 <strong>{stats['US']['strong']}</strong></span>
        <span data-zh-hk="值得關注"><span class="ld" style="background:#378ADD"></span>值得關注 <strong>{stats['US']['watch']}</strong></span>
        <span data-zh-hk="觀望"><span class="ld" style="background:#EF9F27"></span>觀望 <strong>{stats['US']['hold']}</strong></span>
      </div>
      <svg id="us" viewBox="0 0 180 160" width="100%" height="160"></svg>
      <div class="chart-sub" data-zh-hk="{stats['US']['total']}隻 · 均分 {stats['US']['avg']} · 最高 {stats['US']['max']}" data-zh-cn="{stats['US']['total']}只 · 均分 {stats['US']['avg']} · 最高 {stats['US']['max']}" data-en="{stats['US']['total']} stks · Avg {stats['US']['avg']} · High {stats['US']['max']}">{stats['US']['total']}隻 · 均分 {stats['US']['avg']} · 最高 {stats['US']['max']}</div>
    </div>
    <div class="chart-card">
      <div class="chart-title">港股 HK</div>
      <div class="legend">
        <span data-zh-hk="強力買入"><span class="ld" style="background:#1D9E75"></span>強力買入 <strong>{stats['HK']['strong']}</strong></span>
        <span data-zh-hk="值得關注"><span class="ld" style="background:#378ADD"></span>值得關注 <strong>{stats['HK']['watch']}</strong></span>
        <span data-zh-hk="觀望"><span class="ld" style="background:#EF9F27"></span>觀望 <strong>{stats['HK']['hold']}</strong></span>
      </div>
      <svg id="hk" viewBox="0 0 180 160" width="100%" height="160"></svg>
      <div class="chart-sub" data-zh-hk="{stats['HK']['total']}隻 · 均分 {stats['HK']['avg']} · 最高 {stats['HK']['max']}" data-zh-cn="{stats['HK']['total']}只 · 均分 {stats['HK']['avg']} · 最高 {stats['HK']['max']}" data-en="{stats['HK']['total']} stks · Avg {stats['HK']['avg']} · High {stats['HK']['max']}">{stats['HK']['total']}隻 · 均分 {stats['HK']['avg']} · 最高 {stats['HK']['max']}</div>
    </div>
    <div class="chart-card">
      <div class="chart-title">英股 UK</div>
      <div class="legend">
        <span data-zh-hk="強力買入"><span class="ld" style="background:#1D9E75"></span>強力買入 <strong>{stats['UK']['strong']}</strong></span>
        <span data-zh-hk="值得關注"><span class="ld" style="background:#378ADD"></span>值得關注 <strong>{stats['UK']['watch']}</strong></span>
        <span data-zh-hk="觀望"><span class="ld" style="background:#EF9F27"></span>觀望 <strong>{stats['UK']['hold']}</strong></span>
      </div>
      <svg id="uk" viewBox="0 0 180 160" width="100%" height="160"></svg>
      <div class="chart-sub" data-zh-hk="{stats['UK']['total']}隻 · 均分 {stats['UK']['avg']} · 最高 {stats['UK']['max']}" data-zh-cn="{stats['UK']['total']}只 · 均分 {stats['UK']['avg']} · 最高 {stats['UK']['max']}" data-en="{stats['UK']['total']} stks · Avg {stats['UK']['avg']} · High {stats['UK']['max']}">{stats['UK']['total']}隻 · 均分 {stats['UK']['avg']} · 最高 {stats['UK']['max']}</div>
    </div>
    <div class="chart-card">
      <div class="chart-title">A股 CN</div>
      <div class="legend">
        <span data-zh-hk="強力買入"><span class="ld" style="background:#1D9E75"></span>強力買入 <strong>{stats['CN']['strong']}</strong></span>
        <span data-zh-hk="值得關注"><span class="ld" style="background:#378ADD"></span>值得關注 <strong>{stats['CN']['watch']}</strong></span>
        <span data-zh-hk="觀望"><span class="ld" style="background:#EF9F27"></span>觀望 <strong>{stats['CN']['hold']}</strong></span>
      </div>
      <svg id="cn" viewBox="0 0 180 160" width="100%" height="160"></svg>
      <div class="chart-sub" data-zh-hk="{stats['CN']['total']}隻 · 均分 {stats['CN']['avg']} · 最高 {stats['CN']['max']}" data-zh-cn="{stats['CN']['total']}只 · 均分 {stats['CN']['avg']} · 最高 {stats['CN']['max']}" data-en="{stats['CN']['total']} stks · Avg {stats['CN']['avg']} · High {stats['CN']['max']}">{stats['CN']['total']}隻 · 均分 {stats['CN']['avg']} · 最高 {stats['CN']['max']}</div>
    </div>
  </div>
  <div class="chart-card">
    <div class="chart-title">前{TOP_N}名評分分佈</div>
    <div class="legend">
      <span data-zh-hk="強力買入"><span class="ld" style="background:#1D9E75"></span>強力買入</span>
      <span data-zh-hk="值得關注"><span class="ld" style="background:#378ADD"></span>值得關注</span>
      <span data-zh-hk="觀望"><span class="ld" style="background:#EF9F27"></span>觀望</span>
    </div>
    <svg id="bar" viewBox="0 0 900 240" width="100%" height="240"></svg>
  </div>
</div>

<div class="section" id="picks">
  <div class="section-title">最新精選推介 — {report_date}</div>
  <p id="picks-note" style="font-size:13px;color:#999;margin-bottom:1.25rem">以下為今日評分最高股票（≥{MIN_SCORE}分），綠框為強力買入。</p>
  <div class="picks-grid" id="picksGrid"></div>
</div>

<div class="section" id="avoid">
  <div class="section-title">高危名單 — {report_date}</div>
  <p id="avoid-note" style="font-size:13px;color:#999;margin-bottom:1.25rem">以下為今日評分最低的股票（最多{AVOID_N}隻），紅框代表高危股票，數據與上方推介股一致。</p>
  <div class="picks-grid" id="avoidGrid"></div>
</div>

<div class="section" id="contact">
  <div class="section-title" id="contact-heading">合作機會</div>
  <div style="background:#fff;border-radius:16px;padding:2rem;text-align:center;border:1px solid #e5e5e5">
    <div style="font-size:2rem;margin-bottom:1rem">🤝</div>
    <h3 style="font-size:18px;font-weight:600;color:#222;margin-bottom:.75rem" id="contact-title">歡迎合作</h3>
    <p style="font-size:14px;color:#555;line-height:1.7;margin-bottom:1.5rem" id="contact-body">如您有興趣進行廣告合作、內容授權、數據合作或其他商業合作，歡迎透過以下方式聯絡我們。</p>
    <a href="mailto:prosynchk@gmail.com" style="display:inline-flex;align-items:center;gap:8px;background:#1D9E75;color:#fff;padding:10px 24px;border-radius:8px;text-decoration:none;font-size:14px;font-weight:600">
      <span>✉️</span>
      <span>prosynchk@gmail.com</span>
    </a>
  </div>
</div>

<div class="disclaimer">
  <div class="disclaimer-inner">
    <strong>免責聲明：</strong>本網站所有內容僅供參考及教育用途，不構成任何投資建議或買賣邀請。投資涉及風險，過往表現不代表未來回報。讀者應自行進行盡職審查，並在作出任何投資決定前諮詢持牌財務顧問。本站對因使用本站資料而引起的任何損失概不負責。
  </div>
</div>

<div class="footer">
  <div class="footer-inner">
    <span class="footer-logo">HiDH Dividend Analyst</span>
    <span>
      <a href="about.html" style="color:#aaa;text-decoration:none" data-zh-hk="關於我們" data-zh-cn="关于我们" data-en="About">關於我們</a> ·
      <a href="privacy.html" style="color:#aaa;text-decoration:none" data-zh-hk="私隱政策" data-zh-cn="隐私政策" data-en="Privacy">私隱政策</a>
    </span>
    <span>© {datetime.date.today().year} prosynchk.com · 每個交易日更新 · 資料來源：Yahoo Finance · 僅供參考</span>
  </div>
</div>

<script>
const US_STATS = {json.dumps(stats['US'])};
const HK_STATS = {json.dumps(stats['HK'])};
const UK_STATS = {json.dumps(stats['UK'])};
const CN_STATS = {json.dumps(stats['CN'])};
const TOP10    = {top10_js};
const PICKS    = {picks_js};
const AVOID    = {avoid_js};

function drawDonut(svgId, strong, watch, hold) {{
  const svg = document.getElementById(svgId);
  if(!svg) return;
  const values=[strong,watch,hold], colors=['#1D9E75','#378ADD','#EF9F27'];
  const cx=90,cy=80,R=65,r=42;
  const total=values.reduce((a,b)=>a+b,0);
  if(total===0) return;
  let angle=-Math.PI/2;
  values.forEach((v,i)=>{{
    if(v===0) return;
    const sweep=(v/total)*2*Math.PI;
    const x1=cx+R*Math.cos(angle),y1=cy+R*Math.sin(angle);
    const x2=cx+R*Math.cos(angle+sweep),y2=cy+R*Math.sin(angle+sweep);
    const ix1=cx+r*Math.cos(angle),iy1=cy+r*Math.sin(angle);
    const ix2=cx+r*Math.cos(angle+sweep),iy2=cy+r*Math.sin(angle+sweep);
    const large=sweep>Math.PI?1:0;
    const path=document.createElementNS('http://www.w3.org/2000/svg','path');
    path.setAttribute('d',`M ${{x1}} ${{y1}} A ${{R}} ${{R}} 0 ${{large}} 1 ${{x2}} ${{y2}} L ${{ix2}} ${{iy2}} A ${{r}} ${{r}} 0 ${{large}} 0 ${{ix1}} ${{iy1}} Z`);
    path.setAttribute('fill',colors[i]);
    svg.appendChild(path);
    if(v/total>0.03){{
      const mid=angle+sweep/2;
      const lx=cx+(R+r)/2*Math.cos(mid),ly=cy+(R+r)/2*Math.sin(mid);
      const t=document.createElementNS('http://www.w3.org/2000/svg','text');
      t.setAttribute('x',lx);t.setAttribute('y',ly+4);
      t.setAttribute('text-anchor','middle');t.setAttribute('font-size','11');
      t.setAttribute('fill','#fff');t.setAttribute('font-weight','600');
      t.textContent=v;svg.appendChild(t);
    }}
    angle+=sweep;
  }});
  const ct=document.createElementNS('http://www.w3.org/2000/svg','text');
  ct.setAttribute('x',cx);ct.setAttribute('y',cy+5);
  ct.setAttribute('text-anchor','middle');ct.setAttribute('font-size','13');
  ct.setAttribute('fill','#555');ct.setAttribute('font-weight','600');
  ct.textContent=total+'隻';svg.appendChild(ct);
}}

drawDonut('us',US_STATS.strong,US_STATS.watch,US_STATS.hold);
drawDonut('hk',HK_STATS.strong,HK_STATS.watch,HK_STATS.hold);
drawDonut('uk',UK_STATS.strong,UK_STATS.watch,UK_STATS.hold);
drawDonut('cn',CN_STATS.strong,CN_STATS.watch,CN_STATS.hold);

const RATING_COLOR={{strong:'#1D9E75',watch:'#378ADD',hold:'#EF9F27'}};
const MKT_STYLE={{
  US:{{bg:'#E6F1FB',text:'#185FA5'}},
  HK:{{bg:'#E1F5EE',text:'#0F6E56'}},
  UK:{{bg:'#FAEEDA',text:'#854F0B'}},
  CN:{{bg:'#FDECEA',text:'#B71C1C'}},
}};
const svg=document.getElementById('bar');
const W=900,H=240,padL=36,padB=46,padT=10,padR=10;
const chartW=W-padL-padR,chartH=H-padB-padT;
const scores=TOP10.map(s=>s.score);
const minScore=Math.max(40,Math.min(...scores)-5);
const maxScore=Math.min(100,Math.max(...scores)+5);
const barW=chartW/TOP10.length,gap=barW*0.22;

[...Array(6)].map((_,i)=>Math.round(minScore+i*(maxScore-minScore)/5)).forEach(v=>{{
  const y=padT+chartH-((v-minScore)/(maxScore-minScore))*chartH;
  const line=document.createElementNS('http://www.w3.org/2000/svg','line');
  line.setAttribute('x1',padL);line.setAttribute('x2',W-padR);
  line.setAttribute('y1',y);line.setAttribute('y2',y);
  line.setAttribute('stroke',v===Math.round(minScore)?'#ccc':'#eee');line.setAttribute('stroke-width','1');
  svg.appendChild(line);
  const t=document.createElementNS('http://www.w3.org/2000/svg','text');
  t.setAttribute('x',padL-4);t.setAttribute('y',y+4);
  t.setAttribute('text-anchor','end');t.setAttribute('font-size','10');
  t.setAttribute('fill','#999');t.textContent=v;svg.appendChild(t);
}});

TOP10.forEach((s,i)=>{{
  const barH=((s.score-minScore)/(maxScore-minScore))*chartH;
  const x=padL+i*barW+gap/2,y=padT+chartH-barH,w=barW-gap;
  const rect=document.createElementNS('http://www.w3.org/2000/svg','rect');
  rect.setAttribute('x',x);rect.setAttribute('y',y);
  rect.setAttribute('width',w);rect.setAttribute('height',barH);
  rect.setAttribute('fill',RATING_COLOR[s.rating]);rect.setAttribute('rx','3');
  svg.appendChild(rect);
  const ts=document.createElementNS('http://www.w3.org/2000/svg','text');
  ts.setAttribute('x',x+w/2);ts.setAttribute('y',y-4);
  ts.setAttribute('text-anchor','middle');ts.setAttribute('font-size','10');
  ts.setAttribute('fill','#555');ts.setAttribute('font-weight','600');
  ts.textContent=s.score;svg.appendChild(ts);
  const tl=document.createElementNS('http://www.w3.org/2000/svg','text');
  tl.setAttribute('x',x+w/2);tl.setAttribute('y',H-padB+14);
  tl.setAttribute('text-anchor','middle');tl.setAttribute('font-size','10');
  tl.setAttribute('fill','#777');tl.textContent=s.label;svg.appendChild(tl);
  const ms=MKT_STYLE[s.mkt];
  const bw=24,bh=13,bx=x+w/2-bw/2,by=H-padB+18;
  const brect=document.createElementNS('http://www.w3.org/2000/svg','rect');
  brect.setAttribute('x',bx);brect.setAttribute('y',by);
  brect.setAttribute('width',bw);brect.setAttribute('height',bh);
  brect.setAttribute('fill',ms.bg);brect.setAttribute('rx','3');
  svg.appendChild(brect);
  const bt=document.createElementNS('http://www.w3.org/2000/svg','text');
  bt.setAttribute('x',bx+bw/2);bt.setAttribute('y',by+9);
  bt.setAttribute('text-anchor','middle');bt.setAttribute('font-size','9');
  bt.setAttribute('fill',ms.text);bt.setAttribute('font-weight','600');
  bt.textContent=s.mkt;svg.appendChild(bt);
}});


function renderSparkline(p){{
  const divPts=p.div_pts, pxPts=p.px_pts||[];
  const hasPx=pxPts.length>1;
  const W=200,H=90,pL=4,pR=4,pT=10,pB=10;
  const cW=W-pL-pR,cH=H-pT-pB;
  const ff='system-ui,-apple-system,sans-serif';
  const DIV_COL='#1D9E75',PX_COL='#378ADD';
  const FSIZE=10,GAP=13;

  function norm(pts,mn,rng){{return pts.map(d=>pT+cH*(1-(d.v-mn)/rng));}}
  function separate(a,b,mn,mx){{
    if(Math.abs(a-b)>=GAP)return [a,b];
    const mid=(a+b)/2;let lo=mid-GAP/2,hi=mid+GAP/2;
    if(lo<mn){{lo=mn;hi=lo+GAP;}}if(hi>mx){{hi=mx;lo=hi-GAP;}}
    return a<=b?[lo,hi]:[hi,lo];
  }}

  const dvals=divPts.map(d=>d.v);
  const dmn=Math.min(...dvals),dmx=Math.max(...dvals),drng=dmx-dmn||dmx*0.3||0.1;
  const dxs=divPts.map((_,i)=>pL+i*cW/(divPts.length-1));
  const dys=norm(divPts,dmn,drng);
  let dpath=`M${{dxs[0].toFixed(1)}},${{dys[0].toFixed(1)}}`;
  for(let i=1;i<dxs.length;i++) dpath+=` L${{dxs[i].toFixed(1)}},${{dys[i].toFixed(1)}}`;

  let pSvg='';
  if(hasPx){{
    const pvals=pxPts.map(d=>d.v);
    const pmn=Math.min(...pvals),pmx=Math.max(...pvals),prng=pmx-pmn||pmx*0.3||0.1;
    const pxs=pxPts.map((_,i)=>pL+i*cW/(pxPts.length-1));
    const pys=norm(pxPts,pmn,prng);
    let ppath=`M${{pxs[0].toFixed(1)}},${{pys[0].toFixed(1)}}`;
    for(let i=1;i<pxs.length;i++) ppath+=` L${{pxs[i].toFixed(1)}},${{pys[i].toFixed(1)}}`;
    const paPath=ppath+` L${{pxs[pxs.length-1].toFixed(1)}},${{H-pB}} L${{pxs[0].toFixed(1)}},${{H-pB}} Z`;

    // labels: dot position → separate if overlap
    const rd0=dys[0]+FSIZE*0.35, rp0=pys[0]+FSIZE*0.35;
    const rdL=dys[dys.length-1]+FSIZE*0.35, rpL=pys[pys.length-1]+FSIZE*0.35;
    const [ld0,lp0]=separate(rd0,rp0,pT+FSIZE,H-pB);
    const [ldL,lpL]=separate(rdL,rpL,pT+FSIZE,H-pB);
    const d0v=dvals[0].toFixed(2),dLv=dvals[dvals.length-1].toFixed(2);
    const p0v=pvals[0].toFixed(2),pLv=pvals[pvals.length-1].toFixed(2);

    // trend descriptions
    const dpct=(dvals[dvals.length-1]-dvals[0])/dvals[0]*100;
    const dcuts=divPts.filter((d,i)=>i>0&&d.v<divPts[i-1].v*0.95).length;
    const dtLbl=dcuts>0?'\u26a0 \u66fe\u7d93\u6e1b\u606f':dpct>5?'\u2191 \u6301\u7e8c\u589e\u9577':'\u2192 \u7a69\u5b9a\u6d3e\u606f';
    const dtCls=dcuts>0?'cut':dpct>5?'grow':'flat';
    const dtBkg={{grow:'#E1F5EE',cut:'#FEF3F2',flat:'#FAEEDA'}}[dtCls];
    const dtCol={{grow:'#0F6E56',cut:'#B42318',flat:'#854F0B'}}[dtCls];
    const ppct=(pvals[pvals.length-1]-pvals[0])/pvals[0]*100;
    const ptLbl=ppct>10?'\u2191 \u80a1\u50f9\u4e0a\u5347':ppct<-10?'\u2193 \u80a1\u50f9\u4e0b\u8dcc':'\u2192 \u80a1\u50f9\u6a6b\u884c';
    const ptCls=ppct>10?'up':ppct<-10?'down':'side';
    const ptBkg={{up:'#E6F1FB',down:'#FEF3F2',side:'#F1EFE8'}}[ptCls];
    const ptCol={{up:'#185FA5',down:'#B42318',side:'#5F5E5A'}}[ptCls];
    const psign=ppct>=0?'+':'';
    const DTXT={{grow:{{hk:'股息 ↑ 持續增長',cn:'股息 ↑ 持续增长',en:'Div ↑ Growing'}},cut:{{hk:'股息 ⚠ 曾經減息',cn:'股息 ⚠ 曾经减息',en:'Div ⚠ Cut'}},flat:{{hk:'股息 → 稔定派息',cn:'股息 → 稳定派息',en:'Div → Stable'}}}};
    const PTXT={{up:{{hk:'股價 ↑ 上升',cn:'股价 ↑ 上升',en:'Price ↑ Rising'}},down:{{hk:'股價 ↓ 下跌',cn:'股价 ↓ 下跌',en:'Price ↓ Falling'}},side:{{hk:'股價 → 橫行',cn:'股价 → 横行',en:'Price → Sideways'}}}};

    pSvg=`
    <path d="${{paPath}}" fill="${{PX_COL}}" fill-opacity="0.07"/>
    <path d="${{ppath}}" stroke="${{PX_COL}}" stroke-width="1.5" fill="none" stroke-linejoin="round" stroke-linecap="round" stroke-dasharray="3 2"/>
    <circle cx="${{pxs[0].toFixed(1)}}" cy="${{pys[0].toFixed(1)}}" r="2.5" fill="${{PX_COL}}"/>
    <circle cx="${{pxs[pxs.length-1].toFixed(1)}}" cy="${{pys[pys.length-1].toFixed(1)}}" r="2.5" fill="${{PX_COL}}"/>
    <text x="${{(dxs[0]-6).toFixed(1)}}" y="${{ld0.toFixed(1)}}" font-size="${{FSIZE}}" font-family="${{ff}}" fill="${{DIV_COL}}" font-weight="600" text-anchor="end">${{d0v}}</text>
    <text x="${{(pxs[0]-6).toFixed(1)}}" y="${{lp0.toFixed(1)}}" font-size="${{FSIZE}}" font-family="${{ff}}" fill="${{PX_COL}}" font-weight="600" text-anchor="end">${{p0v}}</text>
    <text x="${{(dxs[dxs.length-1]+6).toFixed(1)}}" y="${{ldL.toFixed(1)}}" font-size="${{FSIZE}}" font-family="${{ff}}" fill="${{DIV_COL}}" font-weight="600" text-anchor="start">${{dLv}}</text>
    <text x="${{(pxs[pxs.length-1]+6).toFixed(1)}}" y="${{lpL.toFixed(1)}}" font-size="${{FSIZE}}" font-family="${{ff}}" fill="${{PX_COL}}" font-weight="600" text-anchor="start">${{pLv}}</text>`;

    return `<div class="div-spark">
    <div class="div-spark-legend">
      <span class="div-spark-leg" data-zh-hk="股息" data-zh-cn="股息" data-en="Div"><span class="div-spark-leg-line" style="background:${{DIV_COL}}"></span>\u80a1\u606f</span>
      <span class="div-spark-leg" data-zh-hk="股價" data-zh-cn="股价" data-en="Price"><span class="div-spark-leg-line" style="background:${{PX_COL}};opacity:.7"></span>\u80a1\u50f9</span>
      <span style="font-size:9px;color:#aaa;margin-left:auto">${{divPts[0].y}} \u2192 ${{divPts[divPts.length-1].y}}</span>
    </div>
    <svg viewBox="0 0 ${{W}} ${{H}}" style="display:block;width:100%;height:${{H}}px;overflow:visible" aria-hidden="true">
      ${{pSvg}}
      <path d="${{dpath}}" stroke="${{DIV_COL}}" stroke-width="2" fill="none" stroke-linejoin="round" stroke-linecap="round"/>
      <circle cx="${{dxs[0].toFixed(1)}}" cy="${{dys[0].toFixed(1)}}" r="3" fill="${{DIV_COL}}"/>
      <circle cx="${{dxs[dxs.length-1].toFixed(1)}}" cy="${{dys[dys.length-1].toFixed(1)}}" r="3" fill="${{DIV_COL}}"/>
    </svg>
    <div class="div-spark-desc">
      <div class="div-spark-desc-box" style="background:${{dtBkg}}">
        <div class="div-spark-desc-lbl" data-zh-hk="${{DTXT[dtCls].hk}}" data-zh-cn="${{DTXT[dtCls].cn}}" data-en="${{DTXT[dtCls].en}}" style="color:${{dtCol}}">${{DTXT[dtCls].hk}}</div>
        <div class="div-spark-desc-val" style="color:${{dtCol}}">${{dvals[0].toFixed(2)}} \u2192 ${{dvals[dvals.length-1].toFixed(2)}}</div>
      </div>
      <div class="div-spark-desc-box" style="background:${{ptBkg}}">
        <div class="div-spark-desc-lbl" data-zh-hk="${{PTXT[ptCls].hk}}" data-zh-cn="${{PTXT[ptCls].cn}}" data-en="${{PTXT[ptCls].en}}" style="color:${{ptCol}}">${{PTXT[ptCls].hk}}</div>
        <div class="div-spark-desc-val" style="color:${{ptCol}}">${{psign}}${{ppct.toFixed(0)}}%</div>
      </div>
    </div>
  </div>`;
  }}

  // fallback: div only (no price data)
  const col=p.div_trend==='grow'?DIV_COL:p.div_trend==='cut'?'#B71C1C':'#BA7517';
  const badgeTxt=p.div_trend==='grow'?'\u2191 \u6301\u7e8c\u589e\u9577':p.div_trend==='cut'?'\u26a0 \u66fe\u7d93\u6e1b\u606f':'\u2192 \u7a69\u5b9a\u6d3e\u606f';
  const badgeCls=p.div_trend==='grow'?'div-trend-grow':p.div_trend==='cut'?'div-trend-cut':'div-trend-flat';
  return `<div class="div-spark">
    <div class="div-spark-row"><span class="div-trend-badge ${{badgeCls}}">${{badgeTxt}}</span></div>
    <svg viewBox="0 0 ${{W}} ${{H}}" style="display:block;width:100%;height:${{H}}px" aria-hidden="true">
      <path d="${{dpath}}" stroke="${{col}}" stroke-width="2" fill="none" stroke-linejoin="round" stroke-linecap="round"/>
      <circle cx="${{dxs[0].toFixed(1)}}" cy="${{dys[0].toFixed(1)}}" r="3" fill="${{col}}"/>
      <circle cx="${{dxs[dxs.length-1].toFixed(1)}}" cy="${{dys[dys.length-1].toFixed(1)}}" r="3" fill="${{col}}"/>
    </svg>
  </div>`;
}}

const RATING_LABEL={{strong:'🟢🟢 強力買入',watch:'🟢 值得關注',hold:'⚖️ 觀望'}};
const RATING_CLASS={{strong:'r-strong',watch:'r-watch',hold:'r-hold'}};
const MKT_BADGE={{
  HK:'<span class="badge badge-hk">HK</span>',
  US:'<span class="badge badge-us">US</span>',
  UK:'<span class="badge badge-uk">UK</span>',
  CN:'<span class="badge badge-cn">CN</span>',
}};
document.getElementById('picksGrid').innerHTML=PICKS.map(p=>`
  <div class="pick-card${{p.rating==='strong'?' top':''}}">
    <div class="pick-header">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:4px;margin-bottom:4px">
        <div style="display:flex;align-items:center;gap:5px">
          <span style="font-size:14px;font-weight:600;color:#222;white-space:nowrap">${{p.ticker}}</span>
          ${{MKT_BADGE[p.mkt]}}
        </div>
        <span class="${{RATING_CLASS[p.rating]}}" style="white-space:nowrap;flex-shrink:0" data-rating="${{p.rating}}">${{RATING_LABEL[p.rating]}}</span>
      </div>
      <div class="pick-name">${{p.name}}</div>
    </div>
    <div class="pick-stats">
      <div class="pick-stat"><div class="pick-stat-label" data-zh-hk="股息率" data-zh-cn="股息率" data-en="Yield">股息率</div><div class="pick-stat-val">${{p.yield_}}%</div></div>
      <div class="pick-stat"><div class="pick-stat-label">PE</div><div class="pick-stat-val">${{p.pe}}x</div></div>
      <div class="pick-stat"><div class="pick-stat-label">P/B</div><div class="pick-stat-val">${{p.pb}}</div></div>
    </div>
    <div class="pick-track">
      <div class="pick-track-item"><div class="pick-track-label" data-zh-hk="推介日期" data-zh-cn="推介日期" data-en="Picked On">推介日期</div><div class="pick-track-val">${{p.first_date}}</div></div>
      <div class="pick-track-item"><div class="pick-track-label" data-zh-hk="漲跌" data-zh-cn="涨跌" data-en="Change">漲跌</div><div class="pick-track-val" style="color:${{p.chg_color}}">${{p.chg_text}}</div></div>
      <div class="pick-track-item"><div class="pick-track-label" data-zh-hk="推介價" data-zh-cn="推介价" data-en="Then">推介價</div><div class="pick-track-val">${{p.first_price}}</div></div>
      <div class="pick-track-item"><div class="pick-track-label" data-zh-hk="現價" data-zh-cn="现价" data-en="Now">現價</div><div class="pick-track-val">${{p.price}}</div></div>
    </div>
    <div class="pros-cons">
      <div class="pros"><div class="pros-label">✅ 優點</div><div class="pros-text" data-zh-hk="${{p.pro_zh_hk}}" data-zh-cn="${{p.pro_zh_cn}}" data-en="${{p.pro_en}}">${{p.pro_zh_hk}}</div></div>
      <div class="cons"><div class="cons-label">⚠️ 缺點</div><div class="cons-text" data-zh-hk="${{p.con_zh_hk}}" data-zh-cn="${{p.con_zh_cn}}" data-en="${{p.con_en}}">${{p.con_zh_hk}}</div></div>
    </div>
    <div class="score-track"><div class="score-fill" style="width:${{p.score}}%"></div></div>
    <div class="score-row"><span>評分</span><span>${{p.score}}/100</span></div>
    ${{p.div_pts&&p.div_pts.length>1?renderSparkline(p):''}}
  </div>
`).join('');

document.getElementById('avoidGrid').innerHTML=AVOID.map(p=>`
  <div class="pick-card avoid">
    <div class="pick-header">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:4px;margin-bottom:4px">
        <div style="display:flex;align-items:center;gap:5px">
          <span style="font-size:14px;font-weight:600;color:#222;white-space:nowrap">${{p.ticker}}</span>
          ${{MKT_BADGE[p.mkt]}}
        </div>
        <span class="r-avoid" style="white-space:nowrap;flex-shrink:0" data-zh-hk="🔴 高危" data-zh-cn="🔴 高危" data-en="🔴 High Risk">🔴 高危</span>
      </div>
      <div class="pick-name">${{p.name}}</div>
    </div>
    <div class="pick-stats">
      <div class="pick-stat"><div class="pick-stat-label" data-zh-hk="股息率" data-zh-cn="股息率" data-en="Yield">股息率</div><div class="pick-stat-val">${{p.yield_}}%</div></div>
      <div class="pick-stat"><div class="pick-stat-label">PE</div><div class="pick-stat-val">${{p.pe}}x</div></div>
      <div class="pick-stat"><div class="pick-stat-label">P/B</div><div class="pick-stat-val">${{p.pb}}</div></div>
    </div>
    <div class="pick-track">
      <div class="pick-track-item"><div class="pick-track-label" data-zh-hk="避開日期" data-zh-cn="避开日期" data-en="Flagged On">避開日期</div><div class="pick-track-val">${{p.first_date}}</div></div>
      <div class="pick-track-item"><div class="pick-track-label" data-zh-hk="漲跌" data-zh-cn="涨跌" data-en="Change">漲跌</div><div class="pick-track-val" style="color:${{p.chg_color}}">${{p.chg_text}}</div></div>
      <div class="pick-track-item"><div class="pick-track-label" data-zh-hk="避開價" data-zh-cn="避开价" data-en="Then">避開價</div><div class="pick-track-val">${{p.first_price}}</div></div>
      <div class="pick-track-item"><div class="pick-track-label" data-zh-hk="現價" data-zh-cn="现价" data-en="Now">現價</div><div class="pick-track-val">${{p.price}}</div></div>
    </div>
    <div class="risk-box">
      <div class="risk-label" data-zh-hk="⚠️ 高危原因" data-zh-cn="⚠️ 高危原因" data-en="⚠️ Risk Factors">⚠️ 高危原因</div>
      <div class="risk-text" data-zh-hk="${{p.risk_zh_hk}}" data-zh-cn="${{p.risk_zh_cn}}" data-en="${{p.risk_en}}">${{p.risk_zh_hk}}</div>
    </div>
    <div class="score-track"><div class="score-fill avoid" style="width:${{p.score}}%"></div></div>
    <div class="score-row"><span>評分</span><span>${{p.score}}/100</span></div>
  </div>
`).join('');
</script>

</body>
</html>"""
    html = html.replace("</body>", "<script>" + SETLANG_JS + "</script></body>")
    return html

# ── 主程式 ────────────────────────────────────────────────
def main():
    print("=" * 50)
    print(f"generate_html.py  ({datetime.date.today()})")
    print("=" * 50)

    print("\n📂 讀取 Excel 數據庫...")
    stocks = load_excel_data()

    if not stocks:
        print("❌ 找不到任何股票數據，請先執行 daily_importer_global_v5.py")
        return

    print(f"\n✅ 合計 {len(stocks)} 隻股票")

    print("\n📈 讀取股息歷史...")
    div_hist = load_dividend_history()

    stats = get_market_stats(stocks)
    for mkt, s in stats.items():
        print(f"   {mkt}: {s['total']}隻，強力買入 {s['strong']}，值得關注 {s['watch']}，觀望 {s['hold']}，均分 {s['avg']}")

    report_date = str(datetime.date.today())
    html = generate_html(stocks, stats, report_date, div_hist=div_hist)

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    # 先寫到暫存檔，寫完整個內容後再一次性替換正式檔案 ——
    # 避免瀏覽器（或自動重新整理工具）在 write() 進行中途，讀到只寫了一半、被截斷的檔案
    tmp_path = OUTPUT_FILE + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        f.write(html)
    os.replace(tmp_path, OUTPUT_FILE)

    top5 = ', '.join(s['ticker'] for s in stocks[:5])
    print(f"\n✅ 已生成：{OUTPUT_FILE}")
    print(f"   前5名：{top5}")
    print(f"\n👉 下一步：上傳 {OUTPUT_FILE} 到 GitHub")

if __name__ == "__main__":
    main()
