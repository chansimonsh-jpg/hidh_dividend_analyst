#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
screener_global.py
==================
全球股息優質股發現器 — 配合 batch_importer_global_v5.py 使用

功能：
  1. 自動從指數 / ETF 建立候選池（UK ~150，HK ~120，US ~600）
  2. 預篩條件：息率、市值、有無股息記錄
  3. 排除已在 batch_importer 清單的股票
  4. 對新候選股跑完整 5 維度評分（與 batch 完全一致）
  5. 輸出「發現報告」Excel — 新高分股清單供人手審核

建議運行頻率：每週一次（唔係每日）
運行時間估計：約 15–30 分鐘（視網絡速度及候選數量）

使用：
  python screener_global.py

輸出：
  screener_report_YYYYMMDD.xlsx  — 發現報告
  logs/screener_YYYYMMDD.log     — 運行日誌
"""

# ── 標準庫 ──────────────────────────────────────────────
import datetime
import logging
import os
import sys
import time
import warnings
try:
    import akshare as ak
    _AKSHARE_OK = True
except ImportError:
    _AKSHARE_OK = False

# ── 第三方庫 ─────────────────────────────────────────────
import numpy as np
import pandas as pd
import yfinance as yf
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# =========================================================
# ⚙️  CONFIG — 按需調整
# =========================================================
TODAY       = datetime.date.today()
OUTPUT_FILE = f"screener_report_{TODAY:%Y%m%d}.xlsx"

# 預篩門檻
MIN_YIELD = {          # 最低股息率（%）
    "UK": 3.0,
    "HK": 4.0,
    "US": 2.0,
    "CN": 3.0,
}
MIN_MKTCAP = {         # 最低市值（原幣，B = 十億）
    "UK": 500e6,       # £500M
    "HK": 5e9,         # HK$5B
    "US": 2e9,         # US$2B
    "CN": 20e9,        # ¥200億
}
MIN_SCORE_REPORT = 50  # 只報告總分 ≥ 此值的新股（降低可看更多）
SLEEP_BETWEEN   = 0.4  # 每次 yfinance 請求間隔（秒），避免被限速

# 無風險利率（與 batch_importer_global_v5.py 保持一致）
RISK_FREE_RATES = {
    "UK": 4.20,
    "US": 4.30,
    "HK": 3.80,
    "CN": 2.30,
}

# ── 現有清單（與 batch_importer TICKERS_CONFIG 一致）──────
# 這些股票已在 batch 追蹤，screener 會自動排除
EXISTING_TICKERS = {
    # UK
    "LGEN.L","ABDN.L","MNG.L","IMB.L","BATS.L","UKW.L","ORIT.L","TRIG.L",
    "BSIF.L","LAND.L","AV.L","HSBA.L","BARC.L","NWG.L","LLOY.L","SDR.L",
    "INF.L","OSB.L","BLND.L","BBOX.L","HMSO.L","PHP.L","BYG.L","UU.L",
    "SVT.L","NG.L","SSE.L","HFEL.L","MYI.L","MRCH.L","CTY.L","AAIF.L",
    "BP.L","SHEL.L","AAL.L","RIO.L","GSK.L","AZN.L","ULVR.L","BWY.L",
    "TW.L","PSN.L","VOD.L","BT-A.L","BA.L","PSON.L","WPP.L","OCDO.L","SMIN.L",
    # HK
    "0005.HK","2388.HK","1398.HK","0939.HK","3988.HK","1288.HK","0023.HK",
    "0823.HK","0016.HK","0012.HK","0101.HK","1113.HK","0778.HK","2778.HK",
    "0405.HK","0548.HK","0177.HK","0066.HK","1038.HK","0151.HK","2319.HK",
    "0291.HK","1929.HK","0270.HK","0659.HK","1997.HK","0694.HK","0960.HK",
    "0358.HK","2600.HK","0083.HK","0857.HK","0386.HK","1088.HK","1171.HK",
    "0003.HK","0006.HK","0002.HK","0836.HK","0762.HK","0941.HK","0728.HK",
    "6823.HK","0087.HK","0019.HK","1336.HK","2318.HK","0966.HK","0002.HK",
    # US
    "O","AMT","PLD","SPG","VTR","WELL","NNN","STAG","VICI",
    "NEE","DUK","SO","D","AEP","XEL","WEC","ES",
    "JPM","BAC","WFC","C","USB","PRU","MET",
    "XOM","CVX","COP","EOG","PSX",
    "JNJ","ABBV","MRK","PFE",
    "KO","PEP","PG","MO","PM","T","VZ",
    "SCHD","VYM","HDV","JEPI","JEPQ",
    # CN
    "600036.SS","601398.SS","601288.SS","601939.SS","601988.SS",
    "601328.SS","600028.SS","601857.SS","600050.SS","601318.SS",
    "601601.SS","601088.SS","600519.SS","600886.SS","601186.SS",
    "601800.SS","601390.SS","601668.SS","601225.SS","601666.SS",
    "600188.SS","000858.SZ","600887.SS","600900.SS","000002.SZ",
    "000333.SZ","600048.SS","000651.SZ","002304.SZ","600276.SS",
}

# =========================================================
# 靜態備用基本面數據（yfinance 缺失時兜底，與 batch 保持同步）
# =========================================================
STATIC_FALLBACK = {
    # UK
    "LGEN.L":  {"trailingEps":0.32,"payoutRatio":0.62,"ebitda":2.8e9,"totalDebt":8.5e9,"totalCash":2.0e9,"interestExpense":-0.35e9},
    "ABDN.L":  {"trailingEps":0.15,"payoutRatio":0.85,"ebitda":0.5e9,"totalDebt":1.8e9,"totalCash":1.5e9},
    "HSBA.L":  {"trailingEps":1.05,"payoutRatio":0.50},
    "BARC.L":  {"trailingEps":0.38,"payoutRatio":0.30},
    "LLOY.L":  {"trailingEps":0.072,"payoutRatio":0.38},
    "NWG.L":   {"trailingEps":0.52,"payoutRatio":0.35},
    "IMB.L":   {"trailingEps":2.85,"payoutRatio":0.68,"ebitda":3.8e9,"totalDebt":11.5e9,"totalCash":1.5e9,"interestExpense":-0.42e9},
    "BATS.L":  {"trailingEps":3.20,"payoutRatio":0.72,"ebitda":12.5e9,"totalDebt":42.0e9,"totalCash":3.8e9,"interestExpense":-1.8e9},
    "SHEL.L":  {"trailingEps":2.85,"payoutRatio":0.38,"ebitda":52.0e9,"totalDebt":65.0e9,"totalCash":32.0e9,"interestExpense":-2.1e9},
    "BP.L":    {"trailingEps":0.52,"payoutRatio":0.42,"ebitda":28.5e9,"totalDebt":52.0e9,"totalCash":18.0e9,"interestExpense":-1.9e9},
    "RIO.L":   {"trailingEps":6.20,"payoutRatio":0.60,"ebitda":18.5e9,"totalDebt":12.0e9,"totalCash":8.5e9,"interestExpense":-0.45e9},
    "GSK.L":   {"trailingEps":1.28,"payoutRatio":0.48,"ebitda":8.5e9,"totalDebt":18.0e9,"totalCash":4.2e9,"interestExpense":-0.62e9},
    "AZN.L":   {"trailingEps":3.20,"payoutRatio":0.65,"ebitda":12.5e9,"totalDebt":22.0e9,"totalCash":5.8e9,"interestExpense":-0.85e9},
    "NG.L":    {"trailingEps":0.68,"payoutRatio":0.70,"ebitda":4.5e9,"totalDebt":42.0e9,"totalCash":1.8e9,"interestExpense":-1.5e9},
    "SSE.L":   {"trailingEps":1.20,"payoutRatio":0.65,"ebitda":2.8e9,"totalDebt":18.0e9,"totalCash":0.85e9,"interestExpense":-0.72e9},
    "ULVR.L":  {"trailingEps":2.85,"payoutRatio":0.62,"ebitda":10.5e9,"totalDebt":25.0e9,"totalCash":4.5e9,"interestExpense":-0.75e9},
    "VOD.L":   {"trailingEps":0.08,"payoutRatio":0.75,"ebitda":14.5e9,"totalDebt":52.0e9,"totalCash":5.5e9,"interestExpense":-2.1e9},
    # HK
    "0005.HK": {"trailingEps":7.20,"payoutRatio":0.60,"priceToBook":1.10},
    "0941.HK": {"trailingEps":4.20,"payoutRatio":0.52,"priceToBook":1.50},
    "0939.HK": {"trailingEps":1.20,"payoutRatio":0.28,"priceToBook":0.55},
    "1398.HK": {"trailingEps":0.95,"payoutRatio":0.30,"priceToBook":0.60},
    "3988.HK": {"trailingEps":0.72,"payoutRatio":0.32,"priceToBook":0.50},
    "0823.HK": {"trailingEps":2.85,"payoutRatio":0.92,"priceToBook":0.62},
    "0016.HK": {"trailingEps":12.50,"payoutRatio":0.35,"priceToBook":0.35},
    "0001.HK": {"trailingEps":5.20,"payoutRatio":0.50,"priceToBook":0.45},
    "0002.HK": {"trailingEps":7.20,"payoutRatio":0.60,"priceToBook":1.10},
    "0003.HK": {"trailingEps":1.85,"payoutRatio":0.65,"priceToBook":1.80},
    "0006.HK": {"trailingEps":2.20,"payoutRatio":0.58,"priceToBook":1.20},
    "0011.HK": {"trailingEps":8.50,"payoutRatio":0.45,"priceToBook":1.05},
    # US
    "O":    {"trailingEps":1.40,"payoutRatio":0.72,"ebitda":3.5e9,"totalDebt":20.0e9,"totalCash":0.8e9,"interestExpense":-0.85e9},
    "JPM":  {"trailingEps":18.50,"payoutRatio":0.25,"priceToBook":1.90},
    "BAC":  {"trailingEps":3.20,"payoutRatio":0.30,"priceToBook":1.05},
    "JNJ":  {"trailingEps":8.50,"payoutRatio":0.45,"ebitda":28.0e9,"totalDebt":28.0e9,"totalCash":18.0e9,"interestExpense":-0.85e9},
    "KO":   {"trailingEps":2.85,"payoutRatio":0.65,"ebitda":13.0e9,"totalDebt":35.0e9,"totalCash":9.5e9,"interestExpense":-0.95e9},
    "PEP":  {"trailingEps":8.20,"payoutRatio":0.68,"ebitda":16.0e9,"totalDebt":42.0e9,"totalCash":8.5e9,"interestExpense":-1.1e9},
    "PG":   {"trailingEps":6.20,"payoutRatio":0.58,"ebitda":22.0e9,"totalDebt":28.0e9,"totalCash":7.5e9,"interestExpense":-0.75e9},
    "XOM":  {"trailingEps":8.90,"payoutRatio":0.40,"ebitda":55.0e9,"totalDebt":38.0e9,"totalCash":22.0e9,"interestExpense":-1.2e9},
    "CVX":  {"trailingEps":10.20,"payoutRatio":0.42,"ebitda":38.0e9,"totalDebt":28.0e9,"totalCash":8.0e9,"interestExpense":-0.65e9},
    "MO":   {"trailingEps":4.60,"payoutRatio":0.78,"ebitda":12.0e9,"totalDebt":28.0e9,"totalCash":2.0e9,"interestExpense":-1.1e9,"priceToBook":None},
    "PM":   {"trailingEps":6.50,"payoutRatio":0.88,"ebitda":18.0e9,"totalDebt":48.0e9,"totalCash":3.5e9,"interestExpense":-1.8e9,"priceToBook":None},
    "ABBV": {"trailingEps":6.20,"payoutRatio":0.52,"ebitda":25.0e9,"totalDebt":65.0e9,"totalCash":5.5e9,"interestExpense":-2.2e9,"priceToBook":None},
    "T":    {"trailingEps":2.25,"payoutRatio":0.50,"ebitda":45.0e9,"totalDebt":142.0e9,"totalCash":2.5e9,"interestExpense":-6.2e9},
    "VZ":   {"trailingEps":4.20,"payoutRatio":0.62,"ebitda":48.0e9,"totalDebt":148.0e9,"totalCash":2.0e9,"interestExpense":-5.8e9},
}

FILL_FIELDS_STATIC = [
    "trailingEps","payoutRatio","freeCashflow","ebitda",
    "totalDebt","priceToBook","currentRatio","ebit",
    "operatingIncome","totalCash","interestExpense",
    "earningsGrowth","sharesOutstanding",
]

def apply_static_fallback(info: dict, ticker: str) -> dict:
    static = STATIC_FALLBACK.get(ticker)
    if not static:
        return info
    filled = []
    for f in FILL_FIELDS_STATIC:
        if f == "priceToBook" and "priceToBook" in static:
            info[f] = static[f]  # 可以是 None（負股東權益，跳過評分）
        elif info.get(f) is None and static.get(f) is not None:
            info[f] = static[f]
            filled.append(f)
    if filled:
        log(f"    📋 靜態補充: {filled}")
    if not info.get("ebit") and not info.get("operatingIncome"):
        eb = info.get("ebitda") or 0
        if eb > 0:
            info["ebit"] = eb * 0.6
    return info

# =========================================================
# LOGGING
# =========================================================
def _setup_logger():
    os.makedirs("logs", exist_ok=True)
    log_path = os.path.join("logs", f"screener_{TODAY:%Y%m%d}.log")
    logger = logging.getLogger("screener")
    logger.setLevel(logging.DEBUG)
    if logger.handlers:
        logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s  %(message)s", datefmt="%H:%M:%S")
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(fmt)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger, log_path

_log, _log_path = _setup_logger()

def log(msg: str = ""):
    _log.info(msg)

# =========================================================
# 工具函數（與 batch_importer 完全一致）
# =========================================================
def safe(v, default=0.0):
    if v is None:
        return default
    try:
        f = float(v)
        return default if f != f else f
    except Exception:
        return default

def get_status(total):
    if   total >= 75: return "🟢🟢 強力買入"
    elif total >= 60: return "🟢 值得關注"
    elif total >= 45: return "⚖️  觀望"
    elif total >= 30: return "🟡 偏弱"
    else:             return "🔴 避開"

# =========================================================
# 候選池建立 — 三市場
# =========================================================

# ── 完整 FTSE 100 成分股靜態清單（2025年版，每年更新一次即可）────────────
_FTSE100_STATIC = [
    # 能源
    "SHEL.L","BP.L",
    # 礦業
    "RIO.L","AAL.L","GLEN.L","FRES.L","ANTO.L",
    # 製藥/醫療
    "AZN.L","GSK.L","HIK.L",
    # 消費品
    "ULVR.L","DGE.L","ABF.L","IMB.L","BATS.L","RKT.L",
    # 銀行/保險/金融
    "HSBA.L","LLOY.L","BARC.L","NWG.L","AV.L","LGEN.L","PRU.L",
    "ABDN.L","SDR.L","MNG.L","OSB.L","III.L",
    # 電訊
    "VOD.L","BT-A.L",
    # 基建/公用
    "NG.L","SSE.L","UU.L","SVT.L","CNA.L",
    # 工業/航太
    "BA.L","RR.L","WEIR.L","IMI.L","EXPN.L","SMIN.L",
    "DCC.L","SN.L",
    # 地產/REITs
    "LAND.L","BLND.L","BBOX.L","PHP.L","SGRO.L","HMSO.L",
    # 科技/媒體/出版
    "RELX.L","PSON.L","WPP.L","AUTO.L","INF.L",
    # 零售
    "TSCO.L","SBRY.L","MKS.L","NXT.L","JD.L",
    # 建築/材料/地產開發
    "CRH.L","BWY.L","PSN.L","TW.L","CRDA.L","BYG.L",
    # 其他 FTSE 100 藍籌
    "CPG.L","EMG.L","FCIT.L","HL.L","OCDO.L",
    "RTO.L","MNDI.L","SMDS.L","ADM.L","BKNG.L",
    # 額外高息/板塊補充（FTSE 250 高息）
    "HFEL.L","CTY.L","MYI.L","MRCH.L","AAIF.L",
    "UKW.L","ORIT.L","TRIG.L","BSIF.L",
]

# ── 完整 HSI + 高息港股靜態清單（恒指 82 成分 + 高息藍籌擴充）────────────
_HSI_STATIC = [
    # 恒生指數主要成分（2025）
    "0001.HK","0002.HK","0003.HK","0005.HK","0006.HK","0011.HK","0012.HK",
    "0016.HK","0017.HK","0019.HK","0023.HK","0027.HK","0066.HK","0083.HK",
    "0101.HK","0151.HK","0175.HK","0177.HK","0267.HK","0270.HK","0291.HK",
    "0316.HK","0358.HK","0386.HK","0388.HK","0405.HK","0489.HK","0548.HK",
    "0659.HK","0669.HK","0688.HK","0694.HK","0700.HK","0728.HK","0762.HK",
    "0778.HK","0823.HK","0836.HK","0857.HK","0868.HK","0881.HK","0883.HK",
    "0939.HK","0941.HK","0960.HK","0966.HK","1038.HK","1044.HK","1088.HK",
    "1109.HK","1113.HK","1171.HK","1288.HK","1336.HK","1398.HK","1810.HK",
    "1929.HK","1997.HK","2018.HK","2318.HK","2319.HK","2382.HK","2388.HK",
    "2600.HK","2628.HK","2778.HK","3328.HK","3988.HK","6823.HK",
    # 高息藍籌擴充（唔係恒指成分但常見於高息組合）
    "0087.HK","0019.HK","0008.HK","0010.HK","0013.HK","0014.HK","0020.HK",
    "0025.HK","0041.HK","0045.HK","0050.HK","0054.HK","0069.HK","0080.HK",
    "0123.HK","0135.HK","0144.HK","0163.HK","0168.HK","0187.HK","0189.HK",
    "0199.HK","0215.HK","0216.HK","0242.HK","0257.HK","0288.HK","0293.HK",
    "0302.HK","0303.HK","0322.HK","0341.HK","0392.HK","0410.HK","0435.HK",
    "0440.HK","0444.HK","0460.HK","0467.HK","0480.HK","0490.HK","0503.HK",
    "0511.HK","0520.HK","0551.HK","0552.HK","0560.HK","0563.HK","0566.HK",
    "0570.HK","0576.HK","0579.HK","0590.HK","0604.HK","0606.HK","0612.HK",
    "0636.HK","0639.HK","0644.HK","0656.HK","0663.HK","0668.HK","0670.HK",
    "0673.HK","0683.HK","0685.HK","0696.HK","0697.HK","0699.HK","0708.HK",
    "0716.HK","0719.HK","0753.HK","0754.HK","0788.HK","0806.HK","0813.HK",
    "0817.HK","0836.HK","0845.HK","0851.HK","0853.HK","0854.HK","0861.HK",
    "0867.HK","0874.HK","0880.HK","0884.HK","0885.HK","0886.HK","0902.HK",
    "0914.HK","0916.HK","0917.HK","0921.HK","0925.HK","0934.HK","0945.HK",
    "0956.HK","0968.HK","0980.HK","0981.HK","0992.HK","0998.HK","1003.HK",
    "1038.HK","1045.HK","1048.HK","1053.HK","1055.HK","1060.HK","1062.HK",
    "1072.HK","1079.HK","1083.HK","1093.HK","1094.HK","1099.HK","1101.HK",
    "1112.HK","1114.HK","1117.HK","1118.HK","1137.HK","1138.HK","1139.HK",
    "1143.HK","1148.HK","1157.HK","1163.HK","1168.HK","1177.HK","1179.HK",
    "1186.HK","1193.HK","1194.HK","1199.HK","1200.HK","1205.HK","1207.HK",
    "1211.HK","1216.HK","1218.HK","1221.HK","1224.HK","1230.HK","1233.HK",
]

def get_ftse100() -> list:
    """
    取得 FTSE 100 成分股。
    優先：GitHub CSV（穩定，每季更新）
    備用：內嵌完整靜態清單
    """
    import requests as _req, io as _io
    log("  📥 UK: 拉取 FTSE 100 成分股...")
    # ── 方法 1：GitHub datasets CSV ───────────────────────
    # （注：FTSE100 GitHub CSV 暫無穩定公開源，直接用靜態清單）
    # ── 方法 2：完整靜態清單（100 隻 + 候補）─────────────
    tickers = [t for t in _FTSE100_STATIC if t]  # 過濾空字串
    log(f"  ✅ UK 靜態清單: {len(tickers)} 隻")
    return tickers

def get_hsi_components() -> list:
    """
    取得 HSI + 高息港股候選。
    使用內嵌完整靜態清單（恒指 82 成分 + 擴充高息藍籌）
    """
    import requests as _req
    log("  📥 HK: 載入 HSI + 高息港股清單...")
    tickers = list(dict.fromkeys(_HSI_STATIC))   # 去重保序
    log(f"  ✅ HK 靜態清單: {len(tickers)} 隻")
    return tickers

def get_sp500() -> list:
    """
    取得 S&P 500 成分股。
    優先：GitHub datasets CSV（503 隻，穩定可取）
    備用：Dividend Aristocrats + 高息股靜態清單
    """
    import requests as _req, io as _io
    log("  📥 US: 拉取 S&P 500 成分股（GitHub datasets）...")
    try:
        url = ("https://raw.githubusercontent.com/datasets/"
               "s-and-p-500-companies/main/data/constituents.csv")
        r = _req.get(url, timeout=12)
        r.raise_for_status()
        df = pd.read_csv(_io.StringIO(r.text))
        tickers = [str(t).strip().replace(".", "-")
                   for t in df["Symbol"].tolist() if pd.notna(t)]
        # 補充高息股／REITs 唔係 S&P 成分但重要
        extras = [
            "JEPI","JEPQ","MAIN","ARCC","HTGC","GAIN",
            "OHI","MPW","WPC","EPR","LTC","IIPR",
        ]
        tickers += [t for t in extras if t not in tickers]
        log(f"  ✅ S&P 500 + 補充: {len(tickers)} 隻")
        return tickers
    except Exception as e:
        log(f"  ⚠️  GitHub CSV 失敗: {e}，使用 Dividend Aristocrats 備用清單")
        return [
            "ABBV","ABT","ADM","AFL","AOS","APD","ATO","BDX","BEN","BRO",
            "CAH","CAT","CB","CINF","CL","CLX","CTAS","CVX","D","DOV",
            "ECL","ED","EMR","ESS","FAST","GD","GPC","GWW","HRL","HSY",
            "IBM","ITW","JNJ","KMB","KO","LIN","LOW","MCD","MDT","MKC",
            "MMC","MMM","MO","NEE","NUE","O","PEP","PG","PPG","RTX",
            "SHW","SWK","SYY","T","TROW","TGT","VZ","WMT","XOM",
            "SCHD","VYM","HDV","JEPI","JEPQ","DLR","MAIN","ARCC",
            "STAG","VICI","WPC","OHI","MPW","EPR","LTC",
        ]

def get_etf_holdings_yf(etf: str) -> list:
    """
    用 yfinance 嘗試取得 ETF 持倉。
    yfinance 1.4+ 已移除 get_holdings()，改用 funds_data。
    如失敗靜默返回空清單，由 build_candidate_pool 容錯。
    """
    try:
        tk = yf.Ticker(etf)
        # yfinance 1.4+ API
        fd = tk.get_funds_data()
        if fd and hasattr(fd, "top_holdings") and fd.top_holdings is not None:
            return list(fd.top_holdings.index)
    except Exception:
        pass
    return []

def get_csi300() -> list:
    """
    取得滬深 300 成分股候選池。
    優先 AKShare index_stock_cons，失敗用靜態清單。
    返回格式：["600036.SS", "000001.SZ", ...]
    """
    log("  📥 CN: 拉取滬深 300 成分股...")
    if _AKSHARE_OK:
        try:
            df = ak.index_stock_cons(symbol="000300")
            code_col = next((c for c in df.columns
                             if any(k in str(c) for k in ["代码","code","Code"])), df.columns[0])
            tickers = []
            for c in df[code_col].astype(str).str.zfill(6):
                tickers.append(c + (".SS" if c.startswith(("6","9")) else ".SZ"))
            log(f"  ✅ CSI 300（AKShare）: {len(tickers)} 隻")
            return tickers
        except Exception as e:
            log(f"  ⚠️  AKShare CSI300 失敗: {e}，使用靜態清單")

    # 靜態備用：60 隻滬深 300 高息藍籌
    log("  ✅ CN 靜態清單（60 隻高息 A股）")
    return [
        # 銀行
        "600036.SS","601398.SS","601288.SS","601939.SS","601988.SS",
        "601328.SS","600016.SS","601166.SS","600015.SS","002142.SZ",
        # 保險
        "601318.SS","601601.SS","601336.SS",
        # 能源
        "600028.SS","601857.SS","600900.SS","601088.SS","600011.SS",
        "601985.SS","600025.SS","600886.SS",
        # 電訊
        "600050.SS","601728.SS",
        # 基建/公用
        "601390.SS","601800.SS","601186.SS","601668.SS","600548.SS",
        # 煤炭/礦業
        "601225.SS","601666.SS","600188.SS","600019.SS",
        # 地產
        "000002.SZ","600606.SS",
        # 消費/飲料
        "600519.SS","000858.SZ","600887.SS","000568.SZ",
        # 醫藥
        "600276.SS","000538.SZ",
        # 工業
        "600104.SS","000333.SZ","000651.SZ",
        # A股高息 ETF
        "510880.SS","515000.SS","512890.SS",
    ]


def build_candidate_pool() -> dict:
    """
    建立各市場候選池
    返回 {"UK": [...], "HK": [...], "US": [...]}
    """
    log("\n" + "="*55)
    log("📋 第一步：建立候選池")
    log("="*55)

    pool = {}

    # ── UK ──────────────────────────────────────────────────
    uk = set(get_ftse100())
    # 嘗試補充：UK 高息 ETF 持倉（yfinance 1.4+ 可能不支援，靜默失敗）
    for etf in ["HFEL.L", "CTY.L", "MYI.L", "MRCH.L"]:
        holdings = get_etf_holdings_yf(etf)
        if holdings:
            uk.update(t if t.endswith(".L") else t + ".L" for t in holdings)
            log(f"  ➕ {etf} ETF 持倉補充: +{len(holdings)} 隻")
    pool["UK"] = sorted(uk)
    log(f"  🏁 UK 候選池合計: {len(pool['UK'])} 隻")

    # ── HK ──────────────────────────────────────────────────
    hk = set(get_hsi_components())
    pool["HK"] = sorted(hk)
    log(f"  🏁 HK 候選池合計: {len(pool['HK'])} 隻")

    # ── US ──────────────────────────────────────────────────
    us = set(get_sp500())
    # 嘗試補充：US 高息 ETF 持倉（靜默失敗）
    for etf in ["SCHD", "VYM", "HDV"]:
        holdings = get_etf_holdings_yf(etf)
        if holdings:
            us.update(holdings)
            log(f"  ➕ {etf} ETF 持倉補充: +{len(holdings)} 隻")
    pool["US"] = sorted(us)
    log(f"  🏁 US 候選池合計: {len(pool['US'])} 隻")
    
    # ── CN ──────────────────────────────────────────────────
    cn = set(get_csi300())
    pool["CN"] = sorted(cn)
    log(f"  🏁 CN 候選池合計: {len(pool['CN'])} 隻")

    total = sum(len(v) for v in pool.values())
    log(f"  📊 四市場合計: {total} 隻候選")

    return pool

# =========================================================
# 預篩 — 快速過濾不符合條件的候選股
# =========================================================

def prescreen(ticker: str, market: str) -> dict | None:
    """
    快速預篩：只拉 yfinance .info，唔下載完整歷史
    返回 None = 不通過，返回 dict = 通過（含基本資料）
    """
    try:
        info = yf.Ticker(ticker).info or {}
    except Exception as e:
        err = str(e)
        if "404" in err or "delisted" in err.lower() or "No data" in err:
            log(f"    ⏭  {ticker}: 已下市或無資料，跳過")
        else:
            log(f"    ⚠️  {ticker}: yfinance 錯誤 ({err[:60]})")
        return None

    # 必須有股息記錄
    # yfinance dividendYield 格式不一致：
    #   部分返回小數（0.0338 = 3.38%），部分返回已乘 100 的值（3.38）
    #   以 1.0 為分界：> 1.0 視為已係 % 格式，否則 × 100
    _raw_yield = safe(info.get("dividendYield"), 0)
    if _raw_yield > 1.0:
        div_yield = _raw_yield          # 已係 % 形式，直接用
    else:
        div_yield = _raw_yield * 100    # 小數形式，轉 %

    # 合理性上限：>30% 視為數據異常（如 UK 便士問題），改用 dividendRate 估算
    div_rate = safe(info.get("dividendRate") or
                    info.get("trailingAnnualDividendRate"), 0)
    if div_yield > 30:
        price = safe(info.get("currentPrice") or info.get("regularMarketPrice"), 0)
        if market == "UK" and price > 50:
            price /= 100.0
        div_yield = (div_rate / price * 100) if price > 0 and div_rate > 0 else 0.0

    if div_yield <= 0 and div_rate <= 0:
        return None  # 無派息

    # 最低息率
    if div_yield < MIN_YIELD.get(market, 2.0):
        return None

    # 最低市值
    mktcap = safe(info.get("marketCap"), 0)
    if mktcap < MIN_MKTCAP.get(market, 1e9):
        return None

    # 排除 preferred shares / warrant（通常有 "p" 結尾或 "-P" 等）
    if any(ticker.upper().endswith(s) for s in ["-P", "-PRA", "-PRB", ".PF", "W"]):
        return None

    name = info.get("longName") or info.get("shortName") or ticker
    price = safe(info.get("currentPrice") or info.get("regularMarketPrice"), 0)

    # UK Close 換算
    if market == "UK" and price > 50:
        price = price / 100.0

    return {
        "ticker":    ticker,
        "name":      name,
        "market":    market,
        "price":     round(price, 3),
        "yield_pct": round(div_yield, 2),   # 已標準化至 % 單位（如 3.5 = 3.5%）
        "mktcap":    mktcap,
        "info":      info,
    }

# =========================================================
# 評分引擎 — 完整復用 batch_importer 邏輯
# =========================================================

def score_dividend_quality(info, c_yield, y_avg, y_std, df_hist):
    pts = 0.0
    y_buy  = y_avg + 1.5 * y_std
    y_sell = y_avg - 1.5 * y_std
    if   c_yield >= y_buy:  pts += 10
    elif c_yield >= y_avg:  pts += 7
    elif c_yield >= y_sell: pts += 4
    else:                   pts += 1
    pr = safe(info.get("payoutRatio"), -1)
    if pr < 0:
        eps = safe(info.get("trailingEps"), 0)
        div = safe(info.get("dividendRate") or info.get("trailingAnnualDividendRate"), 0)
        pr  = (div / eps) if eps > 0 else -1
    if   0 < pr <= 0.50: pts += 10
    elif 0 < pr <= 0.65: pts += 8
    elif 0 < pr <= 0.80: pts += 5
    elif 0 < pr <= 0.95: pts += 2
    elif pr > 0.95:      pts += 0   # 幾乎全派或超過盈利，明確懲罰
    else:                pts += 2
    fcf      = safe(info.get("freeCashflow"), 0)
    div_rate = safe(info.get("dividendRate") or info.get("trailingAnnualDividendRate"), 0)
    shares   = safe(info.get("sharesOutstanding"), 0)
    div_paid = div_rate * shares if shares > 0 else 0
    if fcf < 0:          pts += 0
    elif fcf > 0 and div_paid > 0:
        fcf_cov = fcf / div_paid
        if   fcf_cov >= 2.0: pts += 8
        elif fcf_cov >= 1.5: pts += 6
        elif fcf_cov >= 1.0: pts += 4
        else:                pts += 1
    else:                pts += 2
    if df_hist is not None and not df_hist.empty:
        divs = df_hist[df_hist["Dividend_Amount"] > 0].copy()
        if len(divs) >= 2:
            divs["Year"] = pd.to_datetime(divs["Date"]).dt.year
            annual = divs.groupby("Year")["Dividend_Amount"].sum()
            cut = any(annual.iloc[i] < annual.iloc[i-1] * 0.90 for i in range(1, len(annual)))
            pts += 0 if cut else 2
        else: pts += 1
    else:   pts += 1
    return round(pts, 2)

def score_valuation(info, c_yield, y_avg, y_std, c_pe, pe_avg, pe_std, rfr=4.3):
    pts = 0.0
    pe_buy  = pe_avg - 1.5 * pe_std
    pe_sell = pe_avg + 1.5 * pe_std
    if   c_pe <= pe_buy:  pts += 8
    elif c_pe <= pe_avg:  pts += 6
    elif c_pe <= pe_sell: pts += 3
    else:                 pts += 0
    pb = safe(info.get("priceToBook"), -1)
    if   pb < 0:    pts += 2
    elif pb <= 0.8: pts += 7
    elif pb <= 1.2: pts += 5
    elif pb <= 2.0: pts += 3
    elif pb <= 3.5: pts += 1
    else:           pts += 0
    spread = c_yield - rfr
    if   spread >= 3.0:  pts += 10
    elif spread >= 2.0:  pts += 8
    elif spread >= 1.0:  pts += 6
    elif spread >= 0.0:  pts += 4
    elif spread >= -1.0: pts += 2
    else:                pts += 0
    return round(pts, 2)

def score_financial_health(info):
    pts = 0.0
    ebitda   = safe(info.get("ebitda"), 0)
    total_d  = safe(info.get("totalDebt"), 0)
    cash     = safe(info.get("totalCash") or info.get("cash"), 0)
    net_debt = total_d - cash
    if ebitda > 0:
        nd_eb = net_debt / ebitda
        if   nd_eb <= 1.0:  pts += 10
        elif nd_eb <= 2.0:  pts += 8
        elif nd_eb <= 3.0:  pts += 6
        elif nd_eb <= 4.5:  pts += 3
        else:               pts += 0
    elif net_debt <= 0: pts += 10
    else:               pts += 2
    ebit    = safe(info.get("ebit") or info.get("operatingIncome"), 0)
    int_exp = abs(safe(info.get("interestExpense"), 0))
    if ebit < 0:                pts += 0
    elif ebit > 0 and int_exp > 0:
        ic = ebit / int_exp
        if   ic >= 8:   pts += 9
        elif ic >= 5:   pts += 7
        elif ic >= 3:   pts += 4
        elif ic >= 1.5: pts += 2
        else:           pts += 0
    else:               pts += 2
    cr = safe(info.get("currentRatio"), -1)
    if   cr < 0:    pts += 1
    elif cr >= 2.0: pts += 6
    elif cr >= 1.5: pts += 5
    elif cr >= 1.0: pts += 3
    else:           pts += 0
    return round(pts, 2)

def score_growth(info, df_hist):
    pts = 0.0
    dgr = None
    if df_hist is not None and not df_hist.empty:
        divs = df_hist[df_hist["Dividend_Amount"] > 0].copy()
        if len(divs) >= 2:
            divs["Year"] = pd.to_datetime(divs["Date"]).dt.year
            annual = divs.groupby("Year")["Dividend_Amount"].sum()
            if len(annual) >= 4:
                dn, do = annual.iloc[-1], annual.iloc[-4]
                dgr = ((dn/do)**(1/3)-1)*100 if do > 0 else 0
            elif len(annual) >= 2:
                n = len(annual)-1
                dn, do = annual.iloc[-1], annual.iloc[0]
                dgr = ((dn/do)**(1/n)-1)*100 if do > 0 else 0
    if dgr is None: dgr = 0
    if   dgr >= 8:  pts += 6
    elif dgr >= 5:  pts += 5
    elif dgr >= 2:  pts += 4
    elif dgr >= 0:  pts += 2
    else:           pts += 0
    eg_raw = info.get("earningsGrowth")
    if eg_raw is None:
        eg_raw = info.get("earningsQuarterlyGrowth")
    if eg_raw is None: pts += 1
    else:
        eg = safe(eg_raw, 0.0)
        if   eg >= 0.10: pts += 4
        elif eg >= 0.03: pts += 3
        elif eg >= 0:    pts += 2
        else:            pts += 0
    return round(pts, 2)

def score_technical(df_hist, info):
    pts = 0.0
    w52l  = safe(info.get("fiftyTwoWeekLow"), 0)
    w52h  = safe(info.get("fiftyTwoWeekHigh"), 0)
    price = safe(info.get("currentPrice") or info.get("regularMarketPrice"), 0)
    if w52l > 0 and w52h > w52l and price > 0:
        pos = (price - w52l) / (w52h - w52l)
        if   pos <= 0.25: pts += 6
        elif pos <= 0.40: pts += 5
        elif pos <= 0.60: pts += 3
        elif pos <= 0.80: pts += 1
        else:             pts += 0
    else: pts += 1
    if df_hist is not None and len(df_hist) >= 15:
        closes = pd.Series(df_hist["Close"].values, dtype=float)
        delta  = closes.diff()
        gain   = delta.clip(lower=0).rolling(14).mean()
        loss   = (-delta.clip(upper=0)).rolling(14).mean()
        rs     = gain / loss.replace(0, np.nan)
        rsi    = (100 - 100 / (1 + rs)).iloc[-1]
        if   rsi <= 30: pts += 4
        elif rsi <= 45: pts += 3
        elif rsi <= 60: pts += 2
        else:           pts += 0
    else: pts += 1
    return round(pts, 2)

def compute_full_score(ticker: str, market: str, info: dict) -> dict | None:
    """
    下載 2 年歷史，計算完整 5 維度評分
    返回結果 dict，或 None（如下載失敗）
    """
    try:
        try:
            df_raw = yf.Ticker(ticker).history(period="2y", auto_adjust=False)
            if df_raw.empty:
                df_raw = yf.Ticker(ticker).history(period="1y", auto_adjust=False)
        except Exception as he:
            err = str(he)
            if "404" in err or "delisted" in err.lower():
                log(f"    ⏭  {ticker}: 歷史數據 404，已下市，跳過")
            else:
                log(f"    ⚠️  {ticker}: 歷史下載失敗 ({err[:60]})")
            return None
        if df_raw.empty:
            log(f"    ⏭  {ticker}: 歷史數據為空，跳過")
            return None

        df_raw = df_raw.reset_index()
        df_raw.columns = [str(c).capitalize() for c in df_raw.columns]
        if "Dividends" in df_raw.columns:
            df_raw.rename(columns={"Dividends": "Dividend_Amount"}, inplace=True)
        else:
            df_raw["Dividend_Amount"] = 0.0
        df_raw["Date"] = pd.to_datetime(df_raw["Date"]).dt.date

        df_hist = df_raw[["Date", "Close", "Dividend_Amount"]].copy()

        # UK 換算
        if market == "UK":
            df_hist["Close"]           = df_hist["Close"] / 100.0
            df_hist["Dividend_Amount"] = df_hist["Dividend_Amount"] / 100.0

        # EPS 驗證
        eps = info.get("trailingEps")
        price = safe(info.get("currentPrice") or info.get("regularMarketPrice"), 0)
        if market == "UK" and price > 50:
            price /= 100.0
        if eps and price > 0:
            implied_pe = price / eps
            if implied_pe > 500 or implied_pe < 0:
                eps = None

        # 息率計算
        rolling_div = df_hist["Dividend_Amount"].rolling(252, min_periods=1).sum()
        hist_yield  = (rolling_div / df_hist["Close"].replace(0, np.nan)) * 100
        hist_yield  = hist_yield.where((hist_yield > 0) & (hist_yield < 30))
        yields = hist_yield.dropna().values
        y_avg = np.mean(yields) if len(yields) > 0 else 5.0
        y_std = max(np.std(yields) if len(yields) > 1 else 1.0, 0.5)

        # 最新息率（優先用 dividendRate）
        div_rate = safe(info.get("dividendRate") or info.get("trailingAnnualDividendRate"), 0)
        last_close = df_hist["Close"].iloc[-1]
        # 優先用 dividendRate / price 計算（最準確），其次用歷史息率均值
        if last_close > 0 and div_rate > 0:
            c_yield = div_rate / last_close * 100
        elif not pd.isna(hist_yield.iloc[-1]):
            c_yield = float(hist_yield.iloc[-1])
        else:
            c_yield = y_avg
        # 合理性：>30% 視為異常，fallback 用歷史均值
        if c_yield > 30:
            c_yield = y_avg

        # EPS fallback
        if not eps or eps <= 0:
            eps = last_close * (y_avg / 100) / 0.8
            if eps <= 0:
                eps = last_close / 12.0

        # PE 計算
        df_hist["PE"] = df_hist["Close"] / eps
        df_hist.loc[df_hist["PE"] > 80, "PE"] = 80
        df_hist.loc[df_hist["PE"] < 0,  "PE"] = 0
        pes = df_hist["PE"].dropna().values
        if len(pes) > 20:
            pes = pes[(pes >= np.percentile(pes, 1)) & (pes <= np.percentile(pes, 99))]
        pe_avg = np.mean(pes) if len(pes) > 0 else 12.0
        pe_std = max(np.std(pes) if len(pes) > 1 else 2.0, 1.5)
        c_pe   = min(max(last_close / eps, 0), 80)

        # 無風險利率
        rfr = RISK_FREE_RATES.get(market, 4.3)

        # ── 靜態備用兜底 ─────────────────────────────────────
        info = apply_static_fallback(dict(info), ticker)

        # ── 合理性修正：payoutRatio > 2.0 視為異常，清除讓評分用估算值 ──
        pr_val = info.get("payoutRatio")
        if pr_val is not None and (pr_val > 2.0 or pr_val < 0):
            log(f"    ⚠️  payoutRatio={pr_val:.3g} 異常，忽略（用派息/EPS 估算）")
            info = dict(info)          # 淺複製，唔改原 dict
            info["payoutRatio"] = None

        # ── 合理性修正：priceToBook < 0（負股東權益）設為 None 跳過 P/B 評分 ──
        pb_val = info.get("priceToBook")
        if pb_val is not None and pb_val < 0:
            log(f"    ⚠️  priceToBook={pb_val:.3g} 為負（負股東權益），跳過 P/B 評分")
            info = dict(info)
            info["priceToBook"] = None

        # 五維度評分
        s_div  = score_dividend_quality(info, c_yield, y_avg, y_std, df_hist)
        s_val  = score_valuation(info, c_yield, y_avg, y_std, c_pe, pe_avg, pe_std, rfr)
        s_fin  = score_financial_health(info)
        s_grow = score_growth(info, df_hist)
        s_tech = score_technical(df_hist, info)
        total  = round(s_div + s_val + s_fin + s_grow + s_tech, 1)

        # 衍生指標
        pb  = safe(info.get("priceToBook"), None)
        cr  = safe(info.get("currentRatio"), None)
        fcf = safe(info.get("freeCashflow"), 0)
        div_total = safe(info.get("dividendRate") or info.get("trailingAnnualDividendRate"), 0)
        shares    = safe(info.get("sharesOutstanding"), 0)
        div_paid  = div_total * shares if shares > 0 else 0
        fcf_cov   = round(fcf/div_paid, 2) if div_paid > 0 and fcf > 0 else None
        ebitda    = safe(info.get("ebitda"), 0)
        net_debt  = safe(info.get("totalDebt"), 0) - safe(info.get("totalCash") or info.get("cash"), 0)
        nd_eb     = round(net_debt/ebitda, 2) if ebitda > 0 else None
        ebit      = safe(info.get("ebit") or info.get("operatingIncome"), 0)
        int_exp   = abs(safe(info.get("interestExpense"), 0))
        ic        = round(ebit/int_exp, 2) if int_exp > 0 and ebit > 0 else None

        # RSI
        rsi = None
        if len(df_hist) >= 15:
            closes = pd.Series(df_hist["Close"].values, dtype=float)
            delta  = closes.diff()
            gain   = delta.clip(lower=0).rolling(14).mean()
            loss   = (-delta.clip(upper=0)).rolling(14).mean()
            rs     = gain / loss.replace(0, np.nan)
            rsi_s  = 100 - 100/(1+rs)
            rsi    = round(float(rsi_s.iloc[-1]), 1) if not rsi_s.empty else None

        # 52W 位置
        w52l  = safe(info.get("fiftyTwoWeekLow"), 0)
        w52h  = safe(info.get("fiftyTwoWeekHigh"), 0)
        w52p  = round((price - w52l) / (w52h - w52l) * 100, 1) if w52h > w52l else None

        # PR
        pr_raw = info.get("payoutRatio")
        if pr_raw:
            pr = round(pr_raw * 100, 1)
        else:
            pr = round(div_rate / eps * 100, 1) if eps > 0 else None

        # DGR
        dgr = None
        divs = df_hist[df_hist["Dividend_Amount"] > 0].copy()
        if len(divs) >= 2:
            divs["Year"] = pd.to_datetime(divs["Date"]).dt.year
            annual = divs.groupby("Year")["Dividend_Amount"].sum()
            if len(annual) >= 4:
                dn, do = annual.iloc[-1], annual.iloc[-4]
                dgr = round(((dn/do)**(1/3)-1)*100, 2) if do > 0 else 0
            elif len(annual) >= 2:
                n = len(annual)-1
                dn, do = annual.iloc[-1], annual.iloc[0]
                dgr = round(((dn/do)**(1/n)-1)*100, 2) if do > 0 else 0

        return {
            "市場":            market,
            "股票代號":        ticker,
            "公司名稱":        info.get("longName") or info.get("shortName") or ticker,
            "現價":            round(last_close, 3),
            "股息率_%":        round(c_yield, 2),
            "息率_5yr均":      round(y_avg, 2),
            "息率_買入線":     round(y_avg + 1.5*y_std, 2),
            "息率_賣出線":     round(y_avg - 1.5*y_std, 2),
            "PE":              round(c_pe, 2),
            "PE_5yr均":        round(pe_avg, 2),
            "Payout_%":        pr,
            "FCF覆蓋":         fcf_cov,
            "Net_Debt/EBITDA": nd_eb,
            "利息覆蓋":        ic,
            "流動比率":        round(cr, 2) if cr else None,
            "P/B":             round(pb, 2) if pb else None,
            "Yield_Spread":    round(c_yield - rfr, 2),
            "DGR_3yr%":        dgr,
            "RSI_14":          rsi,
            "52W位置%":        w52p,
            "S_股息質量_30":   s_div,
            "S_估值_25":       s_val,
            "S_財務健康_25":   s_fin,
            "S_增長_10":       s_grow,
            "S_技術_10":       s_tech,
            "📊 總分_100":     total,
            "📊 評級":         get_status(total),
            "✅ 加入追蹤":     "",   # 留空供人手填 Y，add_to_tracking.py 讀取
        }

    except Exception as e:
        log(f"    ❌ 評分失敗: {e}")
        return None

# =========================================================
# Excel 報告輸出
# =========================================================

def get_score_fill(total):
    if   total >= 75: return PatternFill("solid", fgColor="00B050")
    elif total >= 60: return PatternFill("solid", fgColor="C6EFCE")
    elif total >= 45: return PatternFill("solid", fgColor="FFEB9C")
    elif total >= 30: return PatternFill("solid", fgColor="FFC7CE")
    else:             return PatternFill("solid", fgColor="FF0000")

def get_score_font(total):
    if   total >= 75: return Font(name="Arial", size=10, bold=True, color="FFFFFF")
    elif total >= 60: return Font(name="Arial", size=10, bold=True, color="375623")
    elif total >= 45: return Font(name="Arial", size=10, bold=True, color="7E6000")
    elif total >= 30: return Font(name="Arial", size=10, bold=True, color="9C0006")
    else:             return Font(name="Arial", size=10, bold=True, color="FFFFFF")

def _write_sheet(wb, title, df_in, THIN, HDR_FILL, HDR_FONT, NORM_FONT, SCORE_FILL, SCORE_COLS):
    """將 DataFrame 寫入一個格式化工作表"""
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A2"
    cols = list(df_in.columns)

    # Header row
    for ci, col in enumerate(cols, 1):
        c = ws.cell(1, ci, col)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for ri, row in df_in.iterrows():
        total = safe(row.get("📊 總分_100"), 0)
        rf    = get_score_fill(total)
        rfont = get_score_font(total)
        for ci, col in enumerate(cols, 1):
            val = row[col]
            c   = ws.cell(ri + 2, ci, val)
            c.font   = NORM_FONT
            c.border = THIN
            if col in SCORE_COLS:
                c.fill = SCORE_FILL
            else:
                c.fill = rf
            if col in ("📊 總分_100", "📊 評級"):
                c.fill = rf
                c.font = rfont
                c.alignment = Alignment(horizontal="center")
            if col == "✅ 加入追蹤":
                # 白底綠框，突出可編輯欄位
                c.fill      = PatternFill("solid", fgColor="F2FFF2")
                c.font      = Font(name="Arial", size=11, bold=True, color="375623")
                c.alignment = Alignment(horizontal="center")
                c.border    = Border(
                    left  =Side(style="medium", color="375623"),
                    right =Side(style="medium", color="375623"),
                    top   =Side(style="medium", color="375623"),
                    bottom=Side(style="medium", color="375623"),
                )
            if col in ("股息率_%","息率_5yr均","息率_買入線","息率_賣出線",
                       "Payout_%","Yield_Spread","DGR_3yr%"):
                c.number_format = "0.00"
            if col in ("現價","PE","PE_5yr均","FCF覆蓋","Net_Debt/EBITDA",
                       "利息覆蓋","流動比率","P/B","RSI_14","52W位置%",
                       "S_股息質量_30","S_估值_25","S_財務健康_25","S_增長_10","S_技術_10"):
                c.number_format = "#,##0.00"

    # Column widths
    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ml + 3, 10), 28)
    return ws


def write_report(results: list, watchlist: list,
                 skipped_existing: int, skipped_prescreen: int):
    """輸出 Excel 發現報告（3 分頁：高分發現 / 值得留意 / 摘要）"""
    if not results and not watchlist:
        log("\n⚠️  無新股達到報告門檻，未生成 Excel。")
        return

    wb   = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]   # 刪除預設空白表

    THIN = Border(
        left  =Side(style="thin", color="E0E0E0"),
        right =Side(style="thin", color="E0E0E0"),
        top   =Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    NORM_FONT  = Font(name="Arial", size=10)
    SCORE_FILL = PatternFill("solid", fgColor="E8F4FD")
    SCORE_COLS = {
        "S_股息質量_30","S_估值_25","S_財務健康_25",
        "S_增長_10","S_技術_10","📊 總分_100","📊 評級",
    }
    TRACK_COL = "✅ 加入追蹤"   # 此欄特殊格式：白底、可編輯

    # ── Sheet 1: 達到門檻 ──────────────────────────────────
    if results:
        df1 = pd.DataFrame(results).sort_values(
            "📊 總分_100", ascending=False).reset_index(drop=True)
        _write_sheet(wb, f"新發現 ≥{MIN_SCORE_REPORT}分",
                     df1, THIN, HDR_FILL, HDR_FONT, NORM_FONT, SCORE_FILL, SCORE_COLS)

    # ── Sheet 2: 值得留意（40 至門檻-1 分）────────────────
    if watchlist:
        df2 = pd.DataFrame(watchlist).sort_values(
            "📊 總分_100", ascending=False).reset_index(drop=True)
        _write_sheet(wb, "值得留意 40-49分",
                     df2, THIN, HDR_FILL, HDR_FONT, NORM_FONT, SCORE_FILL, SCORE_COLS)

    # ── Sheet 3: 摘要統計 ──────────────────────────────────
    all_scored = results + watchlist
    df_all     = pd.DataFrame(all_scored) if all_scored else pd.DataFrame()
    ws3        = wb.create_sheet("摘要")
    ws3.freeze_panes = "A2"

    total_scored = len(results) + len(watchlist)
    summary_rows = [
        ["運行日期",                         str(TODAY)],
        ["Log 檔案",                         _log_path],
        ["", ""],
        ["候選池排除（已在追蹤清單）",        skipped_existing],
        ["候選池排除（預篩不通過）",          skipped_prescreen],
        ["完整評分股票數",                   total_scored],
        ["", ""],
        [f"達到報告門檻（≥{MIN_SCORE_REPORT}分）", len(results)],
        ["值得留意（40-49分）",               len(watchlist)],
        ["", ""],
        ["按市場統計（所有評分）", ""],
    ]
    if not df_all.empty:
        by_mkt = (df_all.groupby("市場")["📊 總分_100"]
                  .agg(["count","mean","max"]).reset_index())
        for _, r in by_mkt.iterrows():
            summary_rows.append([
                r["市場"],
                f"{int(r['count'])} 隻，均分 {r['mean']:.1f}，最高 {r['max']:.1f}",
            ])
    summary_rows += [
        ["", ""],
        ["預篩條件", ""],
        ["UK 最低息率",  f"{MIN_YIELD['UK']}%"],
        ["HK 最低息率",  f"{MIN_YIELD['HK']}%"],
        ["US 最低息率",  f"{MIN_YIELD['US']}%"],
        ["UK 最低市值",  f"£{MIN_MKTCAP['UK']/1e6:.0f}M"],
        ["HK 最低市值",  f"HK${MIN_MKTCAP['HK']/1e9:.0f}B"],
        ["US 最低市值",  f"US${MIN_MKTCAP['US']/1e9:.0f}B"],
    ]

    HDR2_FILL = PatternFill("solid", fgColor="2E75B6")
    HDR2_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    ws3.append(["項目", "數值"])
    for ci in (1, 2):
        ws3.cell(1, ci).fill      = HDR2_FILL
        ws3.cell(1, ci).font      = HDR2_FONT
        ws3.cell(1, ci).alignment = Alignment(horizontal="center")

    for row in summary_rows:
        ws3.append(row)
        ri = ws3.max_row
        is_hdr = (row[0] != "" and row[1] == "")
        ws3.cell(ri, 1).font = Font(name="Arial", size=10, bold=is_hdr)
        ws3.cell(ri, 2).font = Font(name="Arial", size=10)

    for col in ws3.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws3.column_dimensions[get_column_letter(col[0].column)].width = min(max(ml + 3, 18), 45)

    wb.save(OUTPUT_FILE)
    log(f"\n✅ 報告已儲存：{OUTPUT_FILE}")
    log(f"   Sheet 1 — 新發現 ≥{MIN_SCORE_REPORT}分:   {len(results)} 隻")
    log(f"   Sheet 2 — 值得留意 40-49分: {len(watchlist)} 隻")
    log(f"   Sheet 3 — 摘要統計")

# =========================================================
# 主程式
# =========================================================
def main():
    log("="*55)
    log(f"🔍 全球股息優質股發現器  ({TODAY})")
    log(f"   Log: {_log_path}")
    log(f"   最低報告分數: {MIN_SCORE_REPORT}")
    log("="*55)

    t_start = time.time()

    # Step 1: 建立候選池
    pool = build_candidate_pool()
    total_candidates = sum(len(v) for v in pool.values())
    log(f"\n📊 候選池合計: {total_candidates} 隻")

    # Step 2: 過濾已在清單的股票
    results        = []
    watchlist      = []   # 40-49 分：值得留意但未達門檻
    skipped_exist  = 0
    skipped_pre    = 0
    scored_count   = 0

    for market, tickers in pool.items():
        rfr = RISK_FREE_RATES.get(market, 4.3)
        log(f"\n{'─'*55}")
        log(f"  [{market}]  候選 {len(tickers)} 隻  RFR={rfr}%")
        log(f"{'─'*55}")

        new_tickers = [t for t in tickers if t not in EXISTING_TICKERS]
        skipped_exist += len(tickers) - len(new_tickers)
        log(f"  排除已追蹤: {len(tickers)-len(new_tickers)} 隻  剩餘新候選: {len(new_tickers)} 隻")

        # Step 3: 預篩
        passed = []
        log(f"  🔎 預篩中...")
        for i, ticker in enumerate(new_tickers, 1):
            time.sleep(SLEEP_BETWEEN)
            result = prescreen(ticker, market)
            if result:
                passed.append(result)
                log(f"    ✅ [{i}/{len(new_tickers)}] {ticker:12s}  息率={result['yield_pct']:.1f}%  市值={result['mktcap']/1e6:.0f}M")
            else:
                skipped_pre += 1

        log(f"  📋 預篩通過: {len(passed)} 隻")

        # Step 4: 完整評分
        if not passed:
            continue
        log(f"  📊 開始完整評分...")
        for i, candidate in enumerate(passed, 1):
            ticker = candidate["ticker"]
            info   = candidate["info"]
            time.sleep(SLEEP_BETWEEN)
            log(f"    [{i}/{len(passed)}] {ticker} ...")
            scored = compute_full_score(ticker, market, info)
            if scored:
                scored_count += 1
                total = scored["📊 總分_100"]
                log(f"      總分: {total}  {scored['📊 評級']}")
                if total >= MIN_SCORE_REPORT:
                    results.append(scored)
                elif total >= 40:
                    watchlist.append(scored)

    # Step 5: 輸出報告
    elapsed = time.time() - t_start
    log(f"\n{'='*55}")
    log(f"✅ 掃描完成  耗時: {elapsed/60:.1f} 分鐘")
    log(f"   已在清單排除:  {skipped_exist} 隻")
    log(f"   預篩不通過:    {skipped_pre} 隻")
    log(f"   完整評分完成:  {scored_count} 隻")
    log(f"   達到報告門檻:  {len(results)} 隻（≥{MIN_SCORE_REPORT}分）")
    log("="*55)

    write_report(results, watchlist, skipped_exist, skipped_pre)
    log(f"\n📋 Log saved: {_log_path}")

if __name__ == "__main__":
    main()
