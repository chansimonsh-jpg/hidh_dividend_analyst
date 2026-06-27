#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
batch_importer_global_v6.py
================================
改動重點（v4 → v5）:
  1. IBAPIFundamentalApp.get_fundamentals_raw() 擴展支援 UK / US / HK 三市場
     （UK 用 LSE/GBP，US 用 SMART/USD，HK 用 SEHK/HKD）
  2. IBKRFundamentalFetcher 新增 get_fundamentals(ticker, market) 三市場通用方法
     （保留 get_hk_fundamentals() 向後兼容）
  3. 新增 check_suspicious_fields() 合理性檢查函數
     檢查項目：payoutRatio > 200%、currentRatio > 50、priceToBook < 0 或 > 100、
     隱含 PE 異常、freeCashflow 數量級與市值差距過大
  4. get_info_with_fallback() 升級為四層架構：
     Layer 1: yfinance
     Layer 2: 合理性檢查，標記異常欄位
     Layer 3: IBKR（三市場通用）— 修正異常欄位 + 補充缺失欄位
     Layer 4: 靜態備用 — 補充 IBKR 仍未能填補的欄位，並覆蓋仍然異常的欄位
  5. 版本號全面更新為 v5
"""

import asyncio
import sys
if sys.version_info >= (3, 10):
    try:
        asyncio.get_event_loop()
    except RuntimeError:
        _loop = asyncio.new_event_loop()
        asyncio.set_event_loop(_loop)

# ── 標準庫 ──────────────────────────────────────────────
import datetime
import os
import socket
import stat
import struct
import threading
import time

# ── 第三方庫 ─────────────────────────────────────────────
import numpy as np
import pandas as pd
import yfinance as yf
import openpyxl
try:
    import akshare as ak
    _AKSHARE_OK = True
except ImportError:
    _AKSHARE_OK = False
    print("⚠️  akshare 未安裝，CN 市場將無法取得數據。請執行：py -3.14 -m pip install akshare")
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# ==========================================
# ⚙️  IBKR Gateway 設置
# ==========================================
IBKR_HOST      = "127.0.0.1"
IBKR_PORT      = 4002   # Gateway=4002, TWS=7497
IBKR_CLIENT_ID = 10
IBKR_TIMEOUT   = 15     # 秒
USE_IBKR       = True   # 設為 False 可跳過 IBKR

# ==========================================
# ⚙️  各市場無風險利率（10年期國債參考值）
#    用於 Yield Spread 及估值評分
#    可每季度手動更新
# ==========================================
RISK_FREE_RATES = {
    "UK": 4.20,   # 英國10年期國債 (Gilt)
    "US": 4.30,   # 美國10年期國債 (Treasury)
    "HK": 3.80,   # 港元利率 (HIBOR/Exchange Fund Bills)
    "CN": 2.30,   # 中國10年期國債（中債）
}

# =========================================================
# IBKR 導入（優先 ib_insync，fallback ibapi）
# =========================================================
def try_import_ibkr():
    """
    嘗試導入 ib_insync（with asyncio patch）
    失敗則嘗試 ibapi，均失敗返回 (None, None)
    """
    try:
        try:
            asyncio.get_event_loop()
        except RuntimeError:
            asyncio.set_event_loop(asyncio.new_event_loop())
        from ib_insync import IB, Stock
        # 壓制 ib_insync 的全局 logger（避免 Error 10358 等訊息直接打印）
        import logging
        logging.getLogger("ib_insync").setLevel(logging.CRITICAL)
        logging.getLogger("ib_insync.client").setLevel(logging.CRITICAL)
        logging.getLogger("ib_insync.wrapper").setLevel(logging.CRITICAL)
        print("  ✅ ib_insync 導入成功")
        return IB, Stock
    except RuntimeError as e:
        print(f"  ⚠️  ib_insync asyncio 問題: {e}")
        try:
            asyncio.set_event_loop(asyncio.new_event_loop())
            from ib_insync import IB, Stock
            import logging
            logging.getLogger("ib_insync").setLevel(logging.CRITICAL)
            logging.getLogger("ib_insync.client").setLevel(logging.CRITICAL)
            logging.getLogger("ib_insync.wrapper").setLevel(logging.CRITICAL)
            print("  ✅ ib_insync 導入成功（asyncio patch 後）")
            return IB, Stock
        except Exception as e2:
            print(f"  ⚠️  ib_insync 仍然失敗: {e2}")
    except ImportError:
        print("  ⚠️  ib_insync 未安裝: pip install ib_insync")
    except Exception as e:
        print(f"  ⚠️  ib_insync 其他錯誤: {e}")

    try:
        from ibapi.client import EClient
        from ibapi.wrapper import EWrapper
        print("  ✅ ibapi (官方) 導入成功，將使用 ibapi 模式")
        return "IBAPI_MODE", "IBAPI_MODE"
    except ImportError:
        print("  ⚠️  ibapi 未安裝: pip install ibapi")
    except Exception as e:
        print(f"  ⚠️  ibapi 錯誤: {e}")

    return None, None

# =========================================================
# IBKR 原生 TCP Socket 實現（完全兼容 Python 3.10+）
# =========================================================
class IBAPIFundamentalApp:
    """
    使用 TCP socket 直接與 IBKR Gateway/TWS 通信
    獲取基本面 XML 數據
    """
    def __init__(self):
        self.fundamental_data = {}
        self._lock = threading.Lock()

    def _recv_all(self, sock, n):
        data = b""
        while len(data) < n:
            chunk = sock.recv(n - len(data))
            if not chunk:
                break
            data += chunk
        return data

    # 市場對應的 IBKR exchange / currency
    MARKET_CONFIG = {
        "HK": {"exchange": b"SEHK",   "currency": b"HKD"},
        "UK": {"exchange": b"LSE",    "currency": b"GBP"},
        "US": {"exchange": b"SMART",  "currency": b"USD"},
    }

    def get_fundamentals_raw(self, ticker_id: str, market: str = "HK") -> str:
        """
        v5: 支援 HK / UK / US 三市場
        UK ticker 去掉 .L 後綴；US ticker 直接使用；HK 去掉 .HK 並去前導零
        """
        cfg = self.MARKET_CONFIG.get(market, self.MARKET_CONFIG["HK"])
        if market == "HK":
            symbol = ticker_id.replace(".HK", "").lstrip("0") or "0"
        elif market == "UK":
            symbol = ticker_id.replace(".L", "")
        else:
            symbol = ticker_id  # US 直接用
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(IBKR_TIMEOUT)
            sock.connect((IBKR_HOST, IBKR_PORT))

            handshake = b"API\x00" + b"v100..176\x00"
            sock.send(struct.pack(">I", len(handshake)) + handshake)

            header = self._recv_all(sock, 4)
            if len(header) < 4:
                sock.close(); return ""
            size = struct.unpack(">I", header)[0]
            self._recv_all(sock, size)  # server version

            start_api = b"71\x002\x00" + str(IBKR_CLIENT_ID).encode() + b"\x00\x00"
            sock.send(struct.pack(">I", len(start_api)) + start_api)

            req_id = 9001
            parts = [
                b"52", b"2", str(req_id).encode(),
                b"0", symbol.encode(), b"STK", b"", b"0", b"", b"",
                cfg["exchange"], cfg["currency"], b"", b"", b"0", b"ReportSnapshot",
            ]
            msg = b"\x00".join(parts) + b"\x00"
            sock.send(struct.pack(">I", len(msg)) + msg)

            xml_data = ""
            sock.settimeout(IBKR_TIMEOUT)
            while True:
                try:
                    header = self._recv_all(sock, 4)
                    if len(header) < 4: break
                    size = struct.unpack(">I", header)[0]
                    if size == 0 or size > 10_000_000: break
                    body = self._recv_all(sock, size)
                    fields = body.split(b"\x00")
                    if fields and fields[0] == b"51":
                        if len(fields) >= 4:
                            xml_data = fields[3].decode("utf-8", errors="ignore")
                            break
                except (socket.timeout, Exception):
                    break

            sock.close()
            return xml_data
        except Exception:
            return ""  # 靜默忽略所有 IBKR 錯誤（含 Error 10358）


# ==========================================
# 📦 HK 靜態備用財務數據（v5：合理性檢查異常時優先使用）
#    來源：公司最新年報 / Wind / Bloomberg 人工整理
# ==========================================
HK_STATIC_FUNDAMENTALS = {
    # ── 銀行 ──────────────────────────────────────────────
    "0939.HK": {"trailingEps": 1.32, "payoutRatio": 0.30, "priceToBook": 0.58,
                "currentRatio": None, "ebitda": None, "totalDebt": None,
                "totalCash": None, "freeCashflow": None, "interestExpense": None,
                "earningsGrowth": 0.05, "sharesOutstanding": 250.01e9},
    "1398.HK": {"trailingEps": 0.94, "payoutRatio": 0.30, "priceToBook": 0.55,
                "currentRatio": None, "ebitda": None, "totalDebt": None,
                "totalCash": None, "freeCashflow": None, "interestExpense": None,
                "earningsGrowth": 0.04, "sharesOutstanding": 356.41e9},
    "3988.HK": {"trailingEps": 0.73, "payoutRatio": 0.30, "priceToBook": 0.48,
                "currentRatio": None, "ebitda": None, "totalDebt": None,
                "totalCash": None, "freeCashflow": None, "interestExpense": None,
                "earningsGrowth": 0.04, "sharesOutstanding": 294.39e9},
    "1288.HK": {"trailingEps": 0.67, "payoutRatio": 0.31, "priceToBook": 0.75,
                "currentRatio": None, "ebitda": None, "totalDebt": None,
                "totalCash": None, "freeCashflow": None, "interestExpense": None,
                "earningsGrowth": 0.06, "sharesOutstanding": 349.99e9},
    "0005.HK": {"trailingEps": 7.20, "payoutRatio": 0.55, "priceToBook": 0.85,
                "currentRatio": None, "ebitda": None, "totalDebt": None,
                "totalCash": None, "freeCashflow": None, "interestExpense": None,
                "earningsGrowth": 0.08, "sharesOutstanding": 20.08e9},
    "2388.HK": {"trailingEps": 2.65, "payoutRatio": 0.62, "priceToBook": 0.90,
                "currentRatio": None, "ebitda": None, "totalDebt": None,
                "totalCash": None, "freeCashflow": None, "interestExpense": None,
                "earningsGrowth": 0.06, "sharesOutstanding": 10.57e9},
    "0023.HK": {"trailingEps": 1.28, "payoutRatio": 0.72, "priceToBook": 0.38,
                "currentRatio": None, "ebitda": None, "totalDebt": None,
                "totalCash": None, "freeCashflow": None, "interestExpense": None,
                "earningsGrowth": -0.05, "sharesOutstanding": 3.24e9},
    # ── 保險 ──────────────────────────────────────────────
    "1336.HK": {"trailingEps": 4.10, "payoutRatio": 0.35, "priceToBook": 0.72,
                "currentRatio": None, "ebitda": None, "totalDebt": None,
                "totalCash": None, "freeCashflow": None, "interestExpense": None,
                "earningsGrowth": 0.08, "sharesOutstanding": 3.0e9},
    "2318.HK": {"trailingEps": 5.86, "payoutRatio": 0.40, "priceToBook": 1.10,
                "currentRatio": None, "ebitda": None, "totalDebt": None,
                "totalCash": None, "freeCashflow": None, "interestExpense": None,
                "earningsGrowth": 0.10, "sharesOutstanding": 18.28e9},
    "0966.HK": {"trailingEps": 1.55, "payoutRatio": 0.32, "priceToBook": 0.45,
                "currentRatio": None, "ebitda": None, "totalDebt": None,
                "totalCash": None, "freeCashflow": None, "interestExpense": None,
                "earningsGrowth": 0.05, "sharesOutstanding": 4.01e9},
    # ── 電訊 ──────────────────────────────────────────────
    "0762.HK": {"trailingEps": 0.78, "payoutRatio": 0.62, "priceToBook": 1.05,
                "currentRatio": 0.85, "ebitda": 96.3e9, "totalDebt": 135.0e9,
                "totalCash": 28.0e9, "freeCashflow": 18.0e9, "interestExpense": 3.5e9,
                "earningsGrowth": 0.04, "sharesOutstanding": 32.09e9},
    "0941.HK": {"trailingEps": 5.78, "payoutRatio": 0.63, "priceToBook": 1.30,
                "currentRatio": 0.95, "ebitda": 350.0e9, "totalDebt": 200.0e9,
                "totalCash": 65.0e9, "freeCashflow": 120.0e9, "interestExpense": 6.0e9,
                "earningsGrowth": 0.08, "sharesOutstanding": 20.43e9},
    "0728.HK": {"trailingEps": 0.55, "payoutRatio": 0.60, "priceToBook": 1.05,
                "currentRatio": 0.80, "ebitda": 110.0e9, "totalDebt": 178.0e9,
                "totalCash": 32.0e9, "freeCashflow": 25.0e9, "interestExpense": 5.5e9,
                "earningsGrowth": 0.05, "sharesOutstanding": 103.7e9},
    "6823.HK": {"trailingEps": 0.68, "payoutRatio": 0.90, "priceToBook": 4.80,
                "currentRatio": 0.60, "ebitda": 7.2e9, "totalDebt": 14.5e9,
                "totalCash": 1.5e9, "freeCashflow": 3.8e9, "interestExpense": 0.65e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 5.02e9},
    # ── 能源 ──────────────────────────────────────────────
    "0857.HK": {"trailingEps": 0.82, "payoutRatio": 0.50, "priceToBook": 0.80,
                "currentRatio": 0.95, "ebitda": 178.0e9, "totalDebt": 290.0e9,
                "totalCash": 45.0e9, "freeCashflow": 55.0e9, "interestExpense": 8.5e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 183.02e9},
    "0386.HK": {"trailingEps": 0.70, "payoutRatio": 0.45, "priceToBook": 0.55,
                "currentRatio": 0.88, "ebitda": 145.0e9, "totalDebt": 320.0e9,
                "totalCash": 55.0e9, "freeCashflow": 48.0e9, "interestExpense": 10.5e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 119.8e9},
    "1088.HK": {"trailingEps": 2.85, "payoutRatio": 0.70, "priceToBook": 0.95,
                "currentRatio": 1.55, "ebitda": 62.0e9, "totalDebt": 38.0e9,
                "totalCash": 22.0e9, "freeCashflow": 32.0e9, "interestExpense": 1.2e9,
                "earningsGrowth": -0.02, "sharesOutstanding": 19.85e9},
    "1171.HK": {"trailingEps": 2.20, "payoutRatio": 0.60, "priceToBook": 0.70,
                "currentRatio": 1.40, "ebitda": 22.0e9, "totalDebt": 28.0e9,
                "totalCash": 8.0e9, "freeCashflow": 12.0e9, "interestExpense": 0.9e9,
                "earningsGrowth": -0.05, "sharesOutstanding": 4.46e9},
    # ── 公用事業 ──────────────────────────────────────────
    "0003.HK": {"trailingEps": 0.52, "payoutRatio": 0.68, "priceToBook": 2.50,
                "currentRatio": 0.75, "ebitda": 9.8e9, "totalDebt": 33.0e9,
                "totalCash": 4.5e9, "freeCashflow": 5.2e9, "interestExpense": 0.95e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 14.73e9},
    "0006.HK": {"trailingEps": 3.80, "payoutRatio": 0.55, "priceToBook": 0.65,
                "currentRatio": 1.20, "ebitda": 12.5e9, "totalDebt": 25.0e9,
                "totalCash": 6.0e9, "freeCashflow": 8.0e9, "interestExpense": 0.85e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 2.41e9},
    "0002.HK": {"trailingEps": 7.20, "payoutRatio": 0.60, "priceToBook": 1.10,
                "currentRatio": 0.95, "ebitda": 21.0e9, "totalDebt": 52.0e9,
                "totalCash": 8.5e9, "freeCashflow": 10.5e9, "interestExpense": 1.8e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 2.63e9},
    "0836.HK": {"trailingEps": 1.45, "payoutRatio": 0.48, "priceToBook": 0.68,
                "currentRatio": 0.75, "ebitda": 25.0e9, "totalDebt": 65.0e9,
                "totalCash": 8.0e9, "freeCashflow": 10.0e9, "interestExpense": 2.2e9,
                "earningsGrowth": 0.05, "sharesOutstanding": 4.16e9},
    # ── 地產 ──────────────────────────────────────────────
    "0016.HK": {"trailingEps": 12.50, "payoutRatio": 0.35, "priceToBook": 0.35,
                "currentRatio": 1.80, "ebitda": 25.0e9, "totalDebt": 75.0e9,
                "totalCash": 15.0e9, "freeCashflow": 8.0e9, "interestExpense": 2.5e9,
                "earningsGrowth": -0.05, "sharesOutstanding": 2.86e9},
    "0001.HK": {"trailingEps": 5.20, "payoutRatio": 0.50, "priceToBook": 0.45,
                "currentRatio": 1.50, "ebitda": 38.0e9, "totalDebt": 120.0e9,
                "totalCash": 28.0e9, "freeCashflow": 18.0e9, "interestExpense": 3.8e9,
                "earningsGrowth": -0.03, "sharesOutstanding": 3.86e9},
    "0012.HK": {"trailingEps": 3.50, "payoutRatio": 0.45, "priceToBook": 0.28,
                "currentRatio": 2.20, "ebitda": 15.0e9, "totalDebt": 55.0e9,
                "totalCash": 12.0e9, "freeCashflow": 5.0e9, "interestExpense": 1.8e9,
                "earningsGrowth": -0.08, "sharesOutstanding": 2.76e9},
    "0083.HK": {"trailingEps": 1.10, "payoutRatio": 0.45, "priceToBook": 0.32,
                "currentRatio": 1.60, "ebitda": 5.0e9, "totalDebt": 18.0e9,
                "totalCash": 4.5e9, "freeCashflow": 2.0e9, "interestExpense": 0.65e9,
                "earningsGrowth": -0.06, "sharesOutstanding": 5.32e9},
    "0101.HK": {"trailingEps": 0.82, "payoutRatio": 0.52, "priceToBook": 0.28,
                "currentRatio": 1.40, "ebitda": 5.8e9, "totalDebt": 28.0e9,
                "totalCash": 5.0e9, "freeCashflow": 2.2e9, "interestExpense": 0.90e9,
                "earningsGrowth": -0.10, "sharesOutstanding": 5.26e9},
    "1113.HK": {"trailingEps": 3.60, "payoutRatio": 0.42, "priceToBook": 0.55,
                "currentRatio": 1.20, "ebitda": 38.0e9, "totalDebt": 135.0e9,
                "totalCash": 12.0e9, "freeCashflow": 12.0e9, "interestExpense": 4.5e9,
                "earningsGrowth": -0.05, "sharesOutstanding": 14.5e9},
    # ── REITs ─────────────────────────────────────────────
    "0823.HK": {"trailingEps": 2.85, "payoutRatio": 0.92, "priceToBook": 0.62,
                "currentRatio": 0.40, "ebitda": 12.5e9, "totalDebt": 55.0e9,
                "totalCash": 3.5e9, "freeCashflow": 10.8e9, "interestExpense": 1.8e9,
                "earningsGrowth": -0.02, "sharesOutstanding": 2.17e9},
    "0778.HK": {"trailingEps": 0.72, "payoutRatio": 0.95, "priceToBook": 0.55,
                "currentRatio": 0.35, "ebitda": 1.8e9, "totalDebt": 8.5e9,
                "totalCash": 0.6e9, "freeCashflow": 1.55e9, "interestExpense": 0.35e9,
                "earningsGrowth": -0.03, "sharesOutstanding": 2.0e9},
    "2778.HK": {"trailingEps": 0.55, "payoutRatio": 0.95, "priceToBook": 0.42,
                "currentRatio": 0.30, "ebitda": 1.5e9, "totalDebt": 10.5e9,
                "totalCash": 0.5e9, "freeCashflow": 1.2e9, "interestExpense": 0.42e9,
                "earningsGrowth": -0.04, "sharesOutstanding": 2.49e9},
    "0405.HK": {"trailingEps": 0.42, "payoutRatio": 0.95, "priceToBook": 0.58,
                "currentRatio": 0.25, "ebitda": 1.1e9, "totalDebt": 6.8e9,
                "totalCash": 0.4e9, "freeCashflow": 0.92e9, "interestExpense": 0.30e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 2.57e9},
    # ── 交通/基建 ─────────────────────────────────────────
    "0548.HK": {"trailingEps": 0.85, "payoutRatio": 0.58, "priceToBook": 0.78,
                "currentRatio": 0.65, "ebitda": 3.2e9, "totalDebt": 12.0e9,
                "totalCash": 2.0e9, "freeCashflow": 2.0e9, "interestExpense": 0.45e9,
                "earningsGrowth": 0.04, "sharesOutstanding": 2.26e9},
    "0177.HK": {"trailingEps": 1.05, "payoutRatio": 0.60, "priceToBook": 0.95,
                "currentRatio": 0.70, "ebitda": 4.8e9, "totalDebt": 16.0e9,
                "totalCash": 3.0e9, "freeCashflow": 2.8e9, "interestExpense": 0.55e9,
                "earningsGrowth": 0.05, "sharesOutstanding": 4.04e9},
    "0066.HK": {"trailingEps": 4.20, "payoutRatio": 0.52, "priceToBook": 1.50,
                "currentRatio": 0.55, "ebitda": 18.0e9, "totalDebt": 65.0e9,
                "totalCash": 8.0e9, "freeCashflow": 8.5e9, "interestExpense": 2.2e9,
                "earningsGrowth": 0.04, "sharesOutstanding": 5.73e9},
    "1038.HK": {"trailingEps": 4.50, "payoutRatio": 0.60, "priceToBook": 0.90,
                "currentRatio": 1.10, "ebitda": 18.5e9, "totalDebt": 48.0e9,
                "totalCash": 8.0e9, "freeCashflow": 10.0e9, "interestExpense": 1.5e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 2.32e9},
    # ── 消費/食品 ─────────────────────────────────────────
    "0151.HK": {"trailingEps": 0.25, "payoutRatio": 0.75, "priceToBook": 2.80,
                "currentRatio": 3.20, "ebitda": 5.5e9, "totalDebt": 4.0e9,
                "totalCash": 8.0e9, "freeCashflow": 3.8e9, "interestExpense": 0.12e9,
                "earningsGrowth": -0.05, "sharesOutstanding": 13.27e9},
    "2319.HK": {"trailingEps": 1.05, "payoutRatio": 0.38, "priceToBook": 2.40,
                "currentRatio": 1.60, "ebitda": 7.8e9, "totalDebt": 22.0e9,
                "totalCash": 6.0e9, "freeCashflow": 4.0e9, "interestExpense": 0.65e9,
                "earningsGrowth": 0.05, "sharesOutstanding": 3.97e9},
    "0291.HK": {"trailingEps": 1.25, "payoutRatio": 0.35, "priceToBook": 2.50,
                "currentRatio": 1.10, "ebitda": 8.5e9, "totalDebt": 18.0e9,
                "totalCash": 5.0e9, "freeCashflow": 4.5e9, "interestExpense": 0.55e9,
                "earningsGrowth": 0.06, "sharesOutstanding": 3.09e9},
    "1929.HK": {"trailingEps": 1.85, "payoutRatio": 0.55, "priceToBook": 1.80,
                "currentRatio": 1.50, "ebitda": 5.2e9, "totalDebt": 12.0e9,
                "totalCash": 3.5e9, "freeCashflow": 2.5e9, "interestExpense": 0.42e9,
                "earningsGrowth": 0.05, "sharesOutstanding": 2.53e9},
    # ── 多元化/工業/金屬 ──────────────────────────────────
    "0270.HK": {"trailingEps": 1.20, "payoutRatio": 0.58, "priceToBook": 0.80,
                "currentRatio": 1.00, "ebitda": 6.5e9, "totalDebt": 20.0e9,
                "totalCash": 4.0e9, "freeCashflow": 3.5e9, "interestExpense": 0.72e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 4.63e9},
    "0659.HK": {"trailingEps": 0.52, "payoutRatio": 0.75, "priceToBook": 0.58,
                "currentRatio": 0.85, "ebitda": 4.8e9, "totalDebt": 22.0e9,
                "totalCash": 3.5e9, "freeCashflow": 2.2e9, "interestExpense": 0.82e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 4.09e9},
    "1997.HK": {"trailingEps": 2.80, "payoutRatio": 0.45, "priceToBook": 0.28,
                "currentRatio": 0.90, "ebitda": 8.0e9, "totalDebt": 38.0e9,
                "totalCash": 5.0e9, "freeCashflow": 3.0e9, "interestExpense": 1.35e9,
                "earningsGrowth": -0.08, "sharesOutstanding": 3.04e9},
    "0694.HK": {"trailingEps": 2.80, "payoutRatio": 0.45, "priceToBook": 0.48,
                "currentRatio": 1.20, "ebitda": 12.0e9, "totalDebt": 35.0e9,
                "totalCash": 3.5e9, "freeCashflow": 5.0e9, "interestExpense": 1.20e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 1.17e9},
    "0960.HK": {"trailingEps": 1.85, "payoutRatio": 0.30, "priceToBook": 0.38,
                "currentRatio": 0.65, "ebitda": 30.0e9, "totalDebt": 185.0e9,
                "totalCash": 18.0e9, "freeCashflow": 5.0e9, "interestExpense": 7.5e9,
                "earningsGrowth": -0.15, "sharesOutstanding": 7.26e9},
    "0358.HK": {"trailingEps": 1.55, "payoutRatio": 0.42, "priceToBook": 0.62,
                "currentRatio": 1.05, "ebitda": 8.5e9, "totalDebt": 30.0e9,
                "totalCash": 5.0e9, "freeCashflow": 3.5e9, "interestExpense": 1.05e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 4.97e9},
    "2600.HK": {"trailingEps": 0.42, "payoutRatio": 0.40, "priceToBook": 0.55,
                "currentRatio": 0.95, "ebitda": 14.0e9, "totalDebt": 55.0e9,
                "totalCash": 10.0e9, "freeCashflow": 4.5e9, "interestExpense": 1.80e9,
                "earningsGrowth": 0.05, "sharesOutstanding": 19.34e9},
    "0087.HK": {"trailingEps": 2.80, "payoutRatio": 0.50, "priceToBook": 0.45,
                "currentRatio": 1.20, "ebitda": 15.0e9, "totalDebt": 42.0e9,
                "totalCash": 6.0e9, "freeCashflow": 6.0e9, "interestExpense": 1.35e9,
                "earningsGrowth": -0.03, "sharesOutstanding": 0.60e9},
    "0019.HK": {"trailingEps": 5.20, "payoutRatio": 0.50, "priceToBook": 0.42,
                "currentRatio": 1.25, "ebitda": 18.0e9, "totalDebt": 52.0e9,
                "totalCash": 7.5e9, "freeCashflow": 7.0e9, "interestExpense": 1.68e9,
                "earningsGrowth": -0.03, "sharesOutstanding": 0.59e9},
}

# ==========================================
# 📦 UK 靜態備用財務數據（v5：合理性檢查異常時優先使用）
#    資產管理/保險/REITs/公用/能源/電訊/醫藥/消費/建房
# ==========================================
UK_STATIC_FUNDAMENTALS = {
    # ── 資產管理 / 保險 ───────────────────────────────────
    "LGEN.L":  {"trailingEps": 0.32, "payoutRatio": 0.62, "priceToBook": 1.10,
                "currentRatio": None, "ebitda": 2.8e9,  "totalDebt": 8.5e9,
                "totalCash": 2.0e9,  "freeCashflow": 1.2e9, "interestExpense": 0.32e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 6.40e9},
    "ABDN.L":  {"trailingEps": 0.15, "payoutRatio": 0.85, "priceToBook": 0.65,
                "currentRatio": None, "ebitda": 0.5e9,  "totalDebt": 1.8e9,
                "totalCash": 1.5e9,  "freeCashflow": 0.3e9, "interestExpense": 0.08e9,
                "earningsGrowth": -0.05, "sharesOutstanding": 1.80e9},
    "MNG.L":   {"trailingEps": 0.22, "payoutRatio": 0.90, "priceToBook": 1.80,
                "currentRatio": None, "ebitda": 0.85e9, "totalDebt": 1.5e9,
                "totalCash": 1.2e9,  "freeCashflow": 0.55e9,"interestExpense": 0.06e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 2.50e9},
    "AV.L":    {"trailingEps": 0.55, "payoutRatio": 0.55, "priceToBook": 1.40,
                "currentRatio": None, "ebitda": 2.2e9,  "totalDebt": 6.5e9,
                "totalCash": 3.5e9,  "freeCashflow": 1.8e9, "interestExpense": 0.28e9,
                "earningsGrowth": 0.06, "sharesOutstanding": 3.90e9},
    "SDR.L":   {"trailingEps": 0.42, "payoutRatio": 0.60, "priceToBook": 1.20,
                "currentRatio": None, "ebitda": 0.65e9, "totalDebt": 1.2e9,
                "totalCash": 0.9e9,  "freeCashflow": 0.45e9,"interestExpense": 0.05e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 0.48e9},
    # ── 銀行 ──────────────────────────────────────────────
    "HSBA.L":  {"trailingEps": 1.05, "payoutRatio": 0.50, "priceToBook": 0.95,
                "currentRatio": None, "ebitda": None,    "totalDebt": None,
                "totalCash": None,   "freeCashflow": None,"interestExpense": None,
                "earningsGrowth": 0.08, "sharesOutstanding": 19.50e9},
    "BARC.L":  {"trailingEps": 0.38, "payoutRatio": 0.30, "priceToBook": 0.55,
                "currentRatio": None, "ebitda": None,    "totalDebt": None,
                "totalCash": None,   "freeCashflow": None,"interestExpense": None,
                "earningsGrowth": 0.10, "sharesOutstanding": 16.20e9},
    "NWG.L":   {"trailingEps": 0.52, "payoutRatio": 0.35, "priceToBook": 0.80,
                "currentRatio": None, "ebitda": None,    "totalDebt": None,
                "totalCash": None,   "freeCashflow": None,"interestExpense": None,
                "earningsGrowth": 0.08, "sharesOutstanding": 8.40e9},
    "LLOY.L":  {"trailingEps": 0.072,"payoutRatio": 0.38, "priceToBook": 0.70,
                "currentRatio": None, "ebitda": None,    "totalDebt": None,
                "totalCash": None,   "freeCashflow": None,"interestExpense": None,
                "earningsGrowth": 0.05, "sharesOutstanding": 57.00e9},
    "OSB.L":   {"trailingEps": 0.95, "payoutRatio": 0.40, "priceToBook": 0.65,
                "currentRatio": None, "ebitda": None,    "totalDebt": None,
                "totalCash": None,   "freeCashflow": None,"interestExpense": None,
                "earningsGrowth": 0.05, "sharesOutstanding": 0.32e9},
    # ── 煙草 ──────────────────────────────────────────────
    "IMB.L":   {"trailingEps": 2.85, "payoutRatio": 0.68, "priceToBook": None,
                "currentRatio": 0.85, "ebitda": 3.8e9,  "totalDebt": 11.5e9,
                "totalCash": 1.5e9,  "freeCashflow": 2.8e9, "interestExpense": 0.55e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 0.89e9},
    "BATS.L":  {"trailingEps": 3.20, "payoutRatio": 0.72, "priceToBook": None,
                "currentRatio": 0.78, "ebitda": 12.5e9, "totalDebt": 42.0e9,
                "totalCash": 3.8e9,  "freeCashflow": 9.5e9, "interestExpense": 1.85e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 2.28e9},
    # ── 可再生能源 REITs / Infrastructure ─────────────────
    "UKW.L":   {"trailingEps": 0.062,"payoutRatio": 1.10, "priceToBook": 1.05,
                "currentRatio": 0.50, "ebitda": 0.32e9, "totalDebt": 1.2e9,
                "totalCash": 0.08e9, "freeCashflow": 0.22e9,"interestExpense": 0.055e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 3.60e9},
    "ORIT.L":  {"trailingEps": 0.045,"payoutRatio": 1.20, "priceToBook": 0.82,
                "currentRatio": 0.40, "ebitda": 0.12e9, "totalDebt": 0.55e9,
                "totalCash": 0.03e9, "freeCashflow": 0.09e9,"interestExpense": 0.022e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 1.55e9},
    "TRIG.L":  {"trailingEps": 0.058,"payoutRatio": 1.10, "priceToBook": 0.88,
                "currentRatio": 0.45, "ebitda": 0.28e9, "totalDebt": 1.0e9,
                "totalCash": 0.06e9, "freeCashflow": 0.20e9,"interestExpense": 0.045e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 2.40e9},
    "BSIF.L":  {"trailingEps": 0.048,"payoutRatio": 1.05, "priceToBook": 0.90,
                "currentRatio": 0.40, "ebitda": 0.10e9, "totalDebt": 0.42e9,
                "totalCash": 0.02e9, "freeCashflow": 0.08e9,"interestExpense": 0.018e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 0.98e9},
    # ── REITs 地產 ────────────────────────────────────────
    "LAND.L":  {"trailingEps": 0.38, "payoutRatio": 0.75, "priceToBook": 0.58,
                "currentRatio": 0.65, "ebitda": 0.62e9, "totalDebt": 3.8e9,
                "totalCash": 0.25e9, "freeCashflow": 0.45e9,"interestExpense": 0.145e9,
                "earningsGrowth": -0.02,"sharesOutstanding": 0.96e9},
    "BLND.L":  {"trailingEps": 0.25, "payoutRatio": 0.72, "priceToBook": 0.52,
                "currentRatio": 0.55, "ebitda": 0.52e9, "totalDebt": 3.2e9,
                "totalCash": 0.20e9, "freeCashflow": 0.38e9,"interestExpense": 0.125e9,
                "earningsGrowth": -0.03,"sharesOutstanding": 1.02e9},
    "BBOX.L":  {"trailingEps": 0.12, "payoutRatio": 0.85, "priceToBook": 0.90,
                "currentRatio": 0.50, "ebitda": 0.38e9, "totalDebt": 2.5e9,
                "totalCash": 0.12e9, "freeCashflow": 0.28e9,"interestExpense": 0.095e9,
                "earningsGrowth": 0.04, "sharesOutstanding": 2.35e9},
    "HMSO.L":  {"trailingEps": 0.08, "payoutRatio": 0.65, "priceToBook": 0.42,
                "currentRatio": 0.45, "ebitda": 0.25e9, "totalDebt": 1.8e9,
                "totalCash": 0.10e9, "freeCashflow": 0.15e9,"interestExpense": 0.075e9,
                "earningsGrowth": -0.05,"sharesOutstanding": 2.85e9},
    "PHP.L":   {"trailingEps": 0.062,"payoutRatio": 0.88, "priceToBook": 1.10,
                "currentRatio": 0.35, "ebitda": 0.18e9, "totalDebt": 1.4e9,
                "totalCash": 0.05e9, "freeCashflow": 0.13e9,"interestExpense": 0.055e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 1.98e9},
    "BYG.L":   {"trailingEps": 0.95, "payoutRatio": 0.42, "priceToBook": 1.80,
                "currentRatio": 1.20, "ebitda": 0.22e9, "totalDebt": 0.58e9,
                "totalCash": 0.08e9, "freeCashflow": 0.16e9,"interestExpense": 0.022e9,
                "earningsGrowth": 0.05, "sharesOutstanding": 0.143e9},
    # ── 公用事業 ──────────────────────────────────────────
    "UU.L":    {"trailingEps": 0.52, "payoutRatio": 0.72, "priceToBook": 2.20,
                "currentRatio": 0.65, "ebitda": 0.88e9, "totalDebt": 8.5e9,
                "totalCash": 0.35e9, "freeCashflow": 0.42e9,"interestExpense": 0.285e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 0.682e9},
    "SVT.L":   {"trailingEps": 1.45, "payoutRatio": 0.68, "priceToBook": 3.50,
                "currentRatio": 0.60, "ebitda": 0.75e9, "totalDebt": 7.2e9,
                "totalCash": 0.28e9, "freeCashflow": 0.35e9,"interestExpense": 0.245e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 0.232e9},
    "NG.L":    {"trailingEps": 0.68, "payoutRatio": 0.70, "priceToBook": 1.85,
                "currentRatio": 0.75, "ebitda": 4.5e9,  "totalDebt": 42.0e9,
                "totalCash": 1.8e9,  "freeCashflow": 1.5e9, "interestExpense": 1.45e9,
                "earningsGrowth": 0.04, "sharesOutstanding": 3.85e9},
    "SSE.L":   {"trailingEps": 1.20, "payoutRatio": 0.65, "priceToBook": 1.60,
                "currentRatio": 0.70, "ebitda": 2.8e9,  "totalDebt": 18.0e9,
                "totalCash": 0.85e9, "freeCashflow": 0.85e9,"interestExpense": 0.62e9,
                "earningsGrowth": 0.05, "sharesOutstanding": 1.10e9},
    # ── 能源 ──────────────────────────────────────────────
    "BP.L":    {"trailingEps": 0.52, "payoutRatio": 0.42, "priceToBook": 1.05,
                "currentRatio": 1.10, "ebitda": 28.5e9, "totalDebt": 52.0e9,
                "totalCash": 18.0e9, "freeCashflow": 12.0e9,"interestExpense": 2.2e9,
                "earningsGrowth": -0.05,"sharesOutstanding": 18.50e9},
    "SHEL.L":  {"trailingEps": 2.85, "payoutRatio": 0.38, "priceToBook": 1.20,
                "currentRatio": 1.15, "ebitda": 52.0e9, "totalDebt": 65.0e9,
                "totalCash": 32.0e9, "freeCashflow": 28.0e9,"interestExpense": 2.8e9,
                "earningsGrowth": -0.03,"sharesOutstanding": 6.50e9},
    # ── 礦業 ──────────────────────────────────────────────
    "AAL.L":   {"trailingEps": 0.85, "payoutRatio": 0.40, "priceToBook": 1.80,
                "currentRatio": 1.50, "ebitda": 5.2e9,  "totalDebt": 12.0e9,
                "totalCash": 4.5e9,  "freeCashflow": 2.8e9, "interestExpense": 0.48e9,
                "earningsGrowth": -0.10,"sharesOutstanding": 1.35e9},
    "RIO.L":   {"trailingEps": 6.20, "payoutRatio": 0.60, "priceToBook": 2.20,
                "currentRatio": 1.80, "ebitda": 18.5e9, "totalDebt": 12.0e9,
                "totalCash": 8.5e9,  "freeCashflow": 10.0e9,"interestExpense": 0.42e9,
                "earningsGrowth": -0.05,"sharesOutstanding": 1.62e9},
    # ── 醫藥 ──────────────────────────────────────────────
    "GSK.L":   {"trailingEps": 1.28, "payoutRatio": 0.48, "priceToBook": 3.80,
                "currentRatio": 1.25, "ebitda": 8.5e9,  "totalDebt": 18.0e9,
                "totalCash": 4.2e9,  "freeCashflow": 5.5e9, "interestExpense": 0.62e9,
                "earningsGrowth": 0.08, "sharesOutstanding": 4.02e9},
    "AZN.L":   {"trailingEps": 3.20, "payoutRatio": 0.65, "priceToBook": 5.50,
                "currentRatio": 1.05, "ebitda": 12.5e9, "totalDebt": 22.0e9,
                "totalCash": 5.8e9,  "freeCashflow": 7.5e9, "interestExpense": 0.75e9,
                "earningsGrowth": 0.15, "sharesOutstanding": 1.58e9},
    # ── 消費品 ────────────────────────────────────────────
    "ULVR.L":  {"trailingEps": 2.85, "payoutRatio": 0.62, "priceToBook": 6.20,
                "currentRatio": 0.85, "ebitda": 10.5e9, "totalDebt": 25.0e9,
                "totalCash": 4.5e9,  "freeCashflow": 7.2e9, "interestExpense": 0.88e9,
                "earningsGrowth": 0.05, "sharesOutstanding": 2.60e9},
    # ── 電訊 ──────────────────────────────────────────────
    "VOD.L":   {"trailingEps": 0.08, "payoutRatio": 0.75, "priceToBook": 0.55,
                "currentRatio": 0.72, "ebitda": 14.5e9, "totalDebt": 52.0e9,
                "totalCash": 5.5e9,  "freeCashflow": 4.5e9, "interestExpense": 2.2e9,
                "earningsGrowth": -0.05,"sharesOutstanding": 26.80e9},
    "BT-A.L":  {"trailingEps": 0.18, "payoutRatio": 0.45, "priceToBook": 1.10,
                "currentRatio": 0.68, "ebitda": 7.8e9,  "totalDebt": 20.0e9,
                "totalCash": 1.8e9,  "freeCashflow": 1.8e9, "interestExpense": 0.88e9,
                "earningsGrowth": 0.03, "sharesOutstanding": 9.90e9},
    # ── 國防 ──────────────────────────────────────────────
    "BA.L":    {"trailingEps": 0.62, "payoutRatio": 0.38, "priceToBook": 4.50,
                "currentRatio": 1.05, "ebitda": 2.8e9,  "totalDebt": 4.5e9,
                "totalCash": 1.8e9,  "freeCashflow": 2.0e9, "interestExpense": 0.18e9,
                "earningsGrowth": 0.12, "sharesOutstanding": 3.15e9},
    # ── 媒體/出版 ─────────────────────────────────────────
    "INF.L":   {"trailingEps": 0.52, "payoutRatio": 0.55, "priceToBook": 2.80,
                "currentRatio": 1.10, "ebitda": 1.2e9,  "totalDebt": 4.2e9,
                "totalCash": 0.85e9, "freeCashflow": 0.75e9,"interestExpense": 0.16e9,
                "earningsGrowth": 0.08, "sharesOutstanding": 1.28e9},
    "PSON.L":  {"trailingEps": 0.68, "payoutRatio": 0.52, "priceToBook": 2.20,
                "currentRatio": 1.25, "ebitda": 0.65e9, "totalDebt": 1.5e9,
                "totalCash": 0.55e9, "freeCashflow": 0.48e9,"interestExpense": 0.06e9,
                "earningsGrowth": 0.05, "sharesOutstanding": 0.78e9},
    "WPP.L":   {"trailingEps": 0.75, "payoutRatio": 0.55, "priceToBook": 1.80,
                "currentRatio": 0.90, "ebitda": 1.8e9,  "totalDebt": 4.8e9,
                "totalCash": 1.5e9,  "freeCashflow": 1.2e9, "interestExpense": 0.22e9,
                "earningsGrowth": 0.02, "sharesOutstanding": 1.02e9},
    # ── 科技/電商 ─────────────────────────────────────────
    "OCDO.L":  {"trailingEps": -0.18,"payoutRatio": None, "priceToBook": 2.50,
                "currentRatio": 1.80, "ebitda": 0.12e9, "totalDebt": 1.8e9,
                "totalCash": 0.65e9, "freeCashflow": -0.15e9,"interestExpense": 0.085e9,
                "earningsGrowth": 0.20, "sharesOutstanding": 0.87e9},
    "SMIN.L":  {"trailingEps": 1.25, "payoutRatio": 0.45, "priceToBook": 2.80,
                "currentRatio": 1.35, "ebitda": 0.85e9, "totalDebt": 1.2e9,
                "totalCash": 0.55e9, "freeCashflow": 0.62e9,"interestExpense": 0.048e9,
                "earningsGrowth": 0.06, "sharesOutstanding": 0.52e9},
    # ── 住宅建築 ──────────────────────────────────────────
    "BWY.L":   {"trailingEps": 2.85, "payoutRatio": 0.40, "priceToBook": 1.20,
                "currentRatio": 3.50, "ebitda": 0.52e9, "totalDebt": 0.25e9,
                "totalCash": 0.32e9, "freeCashflow": 0.38e9,"interestExpense": 0.010e9,
                "earningsGrowth": -0.05,"sharesOutstanding": 0.131e9},
    "TW.L":    {"trailingEps": 0.145,"payoutRatio": 0.55, "priceToBook": 1.35,
                "currentRatio": 4.20, "ebitda": 0.52e9, "totalDebt": 0.18e9,
                "totalCash": 0.28e9, "freeCashflow": 0.40e9,"interestExpense": 0.008e9,
                "earningsGrowth": -0.03,"sharesOutstanding": 3.18e9},
    "PSN.L":   {"trailingEps": 1.20, "payoutRatio": 0.60, "priceToBook": 1.50,
                "currentRatio": 3.80, "ebitda": 0.48e9, "totalDebt": 0.12e9,
                "totalCash": 0.22e9, "freeCashflow": 0.35e9,"interestExpense": 0.006e9,
                "earningsGrowth": -0.08,"sharesOutstanding": 0.316e9},
    # ── 投資信託 ──────────────────────────────────────────
    "HFEL.L":  {"trailingEps": 0.22, "payoutRatio": 0.92, "priceToBook": 0.88,
                "currentRatio": None, "ebitda": None,    "totalDebt": None,
                "totalCash": None,   "freeCashflow": None,"interestExpense": None,
                "earningsGrowth": 0.02, "sharesOutstanding": 0.168e9},
    "MYI.L":   {"trailingEps": 0.38, "payoutRatio": 0.85, "priceToBook": 0.92,
                "currentRatio": None, "ebitda": None,    "totalDebt": None,
                "totalCash": None,   "freeCashflow": None,"interestExpense": None,
                "earningsGrowth": 0.03, "sharesOutstanding": 0.195e9},
    "MRCH.L":  {"trailingEps": 0.32, "payoutRatio": 0.82, "priceToBook": 0.95,
                "currentRatio": None, "ebitda": None,    "totalDebt": None,
                "totalCash": None,   "freeCashflow": None,"interestExpense": None,
                "earningsGrowth": 0.03, "sharesOutstanding": 0.170e9},
    "CTY.L":   {"trailingEps": 0.22, "payoutRatio": 0.88, "priceToBook": 1.05,
                "currentRatio": None, "ebitda": None,    "totalDebt": None,
                "totalCash": None,   "freeCashflow": None,"interestExpense": None,
                "earningsGrowth": 0.03, "sharesOutstanding": 0.388e9},
    "AAIF.L":  {"trailingEps": 0.18, "payoutRatio": 0.90, "priceToBook": 0.82,
                "currentRatio": None, "ebitda": None,    "totalDebt": None,
                "totalCash": None,   "freeCashflow": None,"interestExpense": None,
                "earningsGrowth": 0.02, "sharesOutstanding": 0.322e9},
}

# ==========================================
# 📦 US 靜態備用財務數據（v5：合理性檢查異常時優先使用）
#    REITs / 公用事業 / 銀行 / 能源 / 醫藥 / 消費 / 電訊 / ETF
# ==========================================
US_STATIC_FUNDAMENTALS = {
    # ── REITs ─────────────────────────────────────────────
    "O":    {"trailingEps": 1.40,  "payoutRatio": 0.72, "priceToBook": 1.20,
             "currentRatio": None, "ebitda": 3.5e9,  "totalDebt": 20.0e9,
             "totalCash": 0.8e9,  "freeCashflow": 2.8e9, "interestExpense": 0.90e9,
             "earningsGrowth": 0.05, "sharesOutstanding": 0.87e9},
    "AMT":  {"trailingEps": 4.20,  "payoutRatio": 0.82, "priceToBook": None,
             "currentRatio": None, "ebitda": 5.8e9,  "totalDebt": 38.0e9,
             "totalCash": 1.5e9,  "freeCashflow": 4.2e9, "interestExpense": 1.55e9,
             "earningsGrowth": 0.08, "sharesOutstanding": 0.209e9},
    "PLD":  {"trailingEps": 3.50,  "payoutRatio": 0.68, "priceToBook": 2.00,
             "currentRatio": None, "ebitda": 5.2e9,  "totalDebt": 28.0e9,
             "totalCash": 0.6e9,  "freeCashflow": 3.8e9, "interestExpense": 1.00e9,
             "earningsGrowth": 0.10, "sharesOutstanding": 0.755e9},
    "SPG":  {"trailingEps": 7.80,  "payoutRatio": 0.88, "priceToBook": None,
             "currentRatio": None, "ebitda": 5.0e9,  "totalDebt": 32.0e9,
             "totalCash": 0.8e9,  "freeCashflow": 3.5e9, "interestExpense": 1.25e9,
             "earningsGrowth": 0.06, "sharesOutstanding": 0.328e9},
    "VTR":  {"trailingEps": 0.80,  "payoutRatio": 0.90, "priceToBook": 1.80,
             "currentRatio": None, "ebitda": 1.8e9,  "totalDebt": 12.0e9,
             "totalCash": 0.3e9,  "freeCashflow": 1.2e9, "interestExpense": 0.52e9,
             "earningsGrowth": 0.05, "sharesOutstanding": 0.420e9},
    "WELL": {"trailingEps": 1.50,  "payoutRatio": 0.75, "priceToBook": 3.20,
             "currentRatio": None, "ebitda": 2.8e9,  "totalDebt": 16.0e9,
             "totalCash": 0.5e9,  "freeCashflow": 2.2e9, "interestExpense": 0.62e9,
             "earningsGrowth": 0.12, "sharesOutstanding": 0.428e9},
    "NNN":  {"trailingEps": 2.20,  "payoutRatio": 0.68, "priceToBook": 1.50,
             "currentRatio": None, "ebitda": 0.92e9, "totalDebt": 4.5e9,
             "totalCash": 0.12e9, "freeCashflow": 0.78e9,"interestExpense": 0.18e9,
             "earningsGrowth": 0.04, "sharesOutstanding": 0.185e9},
    "STAG": {"trailingEps": 1.20,  "payoutRatio": 0.78, "priceToBook": 1.40,
             "currentRatio": None, "ebitda": 0.55e9, "totalDebt": 2.8e9,
             "totalCash": 0.08e9, "freeCashflow": 0.42e9,"interestExpense": 0.11e9,
             "earningsGrowth": 0.05, "sharesOutstanding": 0.178e9},
    "VICI": {"trailingEps": 1.85,  "payoutRatio": 0.72, "priceToBook": 1.60,
             "currentRatio": None, "ebitda": 2.8e9,  "totalDebt": 16.5e9,
             "totalCash": 0.4e9,  "freeCashflow": 2.2e9, "interestExpense": 0.68e9,
             "earningsGrowth": 0.06, "sharesOutstanding": 1.06e9},
    # ── 公用事業 ──────────────────────────────────────────
    "NEE":  {"trailingEps": 3.20,  "payoutRatio": 0.62, "priceToBook": 2.50,
             "currentRatio": 0.65, "ebitda": 10.5e9, "totalDebt": 65.0e9,
             "totalCash": 1.8e9,  "freeCashflow": 2.5e9, "interestExpense": 2.45e9,
             "earningsGrowth": 0.08, "sharesOutstanding": 2.05e9},
    "DUK":  {"trailingEps": 5.50,  "payoutRatio": 0.72, "priceToBook": 1.55,
             "currentRatio": 0.55, "ebitda": 8.5e9,  "totalDebt": 62.0e9,
             "totalCash": 0.8e9,  "freeCashflow": 0.5e9, "interestExpense": 2.25e9,
             "earningsGrowth": 0.05, "sharesOutstanding": 0.775e9},
    "SO":   {"trailingEps": 3.80,  "payoutRatio": 0.70, "priceToBook": 2.20,
             "currentRatio": 0.60, "ebitda": 7.2e9,  "totalDebt": 52.0e9,
             "totalCash": 0.6e9,  "freeCashflow": 0.2e9, "interestExpense": 1.95e9,
             "earningsGrowth": 0.06, "sharesOutstanding": 1.09e9},
    "D":    {"trailingEps": 2.85,  "payoutRatio": 0.82, "priceToBook": 1.80,
             "currentRatio": 0.60, "ebitda": 5.5e9,  "totalDebt": 38.0e9,
             "totalCash": 0.5e9,  "freeCashflow": -0.5e9,"interestExpense": 1.55e9,
             "earningsGrowth": 0.04, "sharesOutstanding": 0.840e9},
    "AEP":  {"trailingEps": 5.20,  "payoutRatio": 0.65, "priceToBook": 1.80,
             "currentRatio": 0.55, "ebitda": 5.8e9,  "totalDebt": 42.0e9,
             "totalCash": 0.4e9,  "freeCashflow": 0.1e9, "interestExpense": 1.65e9,
             "earningsGrowth": 0.06, "sharesOutstanding": 0.523e9},
    "XEL":  {"trailingEps": 3.20,  "payoutRatio": 0.65, "priceToBook": 2.10,
             "currentRatio": 0.55, "ebitda": 3.2e9,  "totalDebt": 22.0e9,
             "totalCash": 0.25e9, "freeCashflow": -0.2e9,"interestExpense": 0.82e9,
             "earningsGrowth": 0.06, "sharesOutstanding": 0.549e9},
    "WEC":  {"trailingEps": 4.20,  "payoutRatio": 0.65, "priceToBook": 2.50,
             "currentRatio": 0.55, "ebitda": 2.8e9,  "totalDebt": 18.0e9,
             "totalCash": 0.2e9,  "freeCashflow": 0.3e9, "interestExpense": 0.68e9,
             "earningsGrowth": 0.06, "sharesOutstanding": 0.317e9},
    "ES":   {"trailingEps": 3.50,  "payoutRatio": 0.72, "priceToBook": 1.40,
             "currentRatio": 0.55, "ebitda": 2.5e9,  "totalDebt": 22.0e9,
             "totalCash": 0.3e9,  "freeCashflow": -0.5e9,"interestExpense": 0.82e9,
             "earningsGrowth": 0.04, "sharesOutstanding": 0.348e9},
    # ── 銀行 / 金融 ───────────────────────────────────────
    "JPM":  {"trailingEps": 18.50, "payoutRatio": 0.25, "priceToBook": 1.90,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": 0.12, "sharesOutstanding": 2.87e9},
    "BAC":  {"trailingEps": 3.20,  "payoutRatio": 0.30, "priceToBook": 1.05,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": 0.08, "sharesOutstanding": 7.92e9},
    "WFC":  {"trailingEps": 5.40,  "payoutRatio": 0.30, "priceToBook": 1.15,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": 0.10, "sharesOutstanding": 3.42e9},
    "C":    {"trailingEps": 6.50,  "payoutRatio": 0.28, "priceToBook": 0.62,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": 0.08, "sharesOutstanding": 1.94e9},
    "USB":  {"trailingEps": 3.80,  "payoutRatio": 0.42, "priceToBook": 1.20,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": 0.06, "sharesOutstanding": 1.52e9},
    "PRU":  {"trailingEps": 12.50, "payoutRatio": 0.38, "priceToBook": 0.90,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": 0.08, "sharesOutstanding": 0.370e9},
    "MET":  {"trailingEps": 9.50,  "payoutRatio": 0.30, "priceToBook": 0.85,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": 0.06, "sharesOutstanding": 0.726e9},
    # ── 能源 ──────────────────────────────────────────────
    "XOM":  {"trailingEps": 8.90,  "payoutRatio": 0.40, "priceToBook": 1.90,
             "currentRatio": 1.35, "ebitda": 55.0e9, "totalDebt": 38.0e9,
             "totalCash": 22.0e9, "freeCashflow": 32.0e9,"interestExpense": 1.2e9,
             "earningsGrowth": -0.05,"sharesOutstanding": 4.35e9},
    "CVX":  {"trailingEps": 10.20, "payoutRatio": 0.42, "priceToBook": 1.65,
             "currentRatio": 1.25, "ebitda": 38.0e9, "totalDebt": 28.0e9,
             "totalCash": 8.0e9,  "freeCashflow": 18.0e9,"interestExpense": 0.72e9,
             "earningsGrowth": -0.08,"sharesOutstanding": 1.86e9},
    "COP":  {"trailingEps": 8.20,  "payoutRatio": 0.35, "priceToBook": 2.50,
             "currentRatio": 1.40, "ebitda": 18.0e9, "totalDebt": 18.0e9,
             "totalCash": 5.5e9,  "freeCashflow": 9.5e9, "interestExpense": 0.55e9,
             "earningsGrowth": -0.10,"sharesOutstanding": 1.24e9},
    "EOG":  {"trailingEps": 12.50, "payoutRatio": 0.28, "priceToBook": 2.80,
             "currentRatio": 1.50, "ebitda": 9.5e9,  "totalDebt": 5.0e9,
             "totalCash": 2.5e9,  "freeCashflow": 5.5e9, "interestExpense": 0.22e9,
             "earningsGrowth": -0.05,"sharesOutstanding": 0.598e9},
    "PSX":  {"trailingEps": 9.50,  "payoutRatio": 0.42, "priceToBook": 2.20,
             "currentRatio": 1.20, "ebitda": 5.2e9,  "totalDebt": 12.0e9,
             "totalCash": 2.5e9,  "freeCashflow": 3.5e9, "interestExpense": 0.45e9,
             "earningsGrowth": -0.15,"sharesOutstanding": 0.415e9},
    # ── 醫藥 ──────────────────────────────────────────────
    "JNJ":  {"trailingEps": 8.50,  "payoutRatio": 0.45, "priceToBook": 5.20,
             "currentRatio": 1.45, "ebitda": 28.0e9, "totalDebt": 28.0e9,
             "totalCash": 18.0e9, "freeCashflow": 18.0e9,"interestExpense": 0.88e9,
             "earningsGrowth": 0.06, "sharesOutstanding": 2.40e9},
    "ABBV": {"trailingEps": 6.20,  "payoutRatio": 0.52, "priceToBook": None,
             "currentRatio": 0.95, "ebitda": 25.0e9, "totalDebt": 65.0e9,
             "totalCash": 5.5e9,  "freeCashflow": 18.0e9,"interestExpense": 2.20e9,
             "earningsGrowth": 0.05, "sharesOutstanding": 1.77e9},
    "MRK":  {"trailingEps": 7.80,  "payoutRatio": 0.42, "priceToBook": 5.50,
             "currentRatio": 1.35, "ebitda": 22.0e9, "totalDebt": 35.0e9,
             "totalCash": 8.5e9,  "freeCashflow": 14.0e9,"interestExpense": 1.15e9,
             "earningsGrowth": 0.12, "sharesOutstanding": 2.54e9},
    "PFE":  {"trailingEps": 1.45,  "payoutRatio": 0.80, "priceToBook": 1.65,
             "currentRatio": 1.05, "ebitda": 15.0e9, "totalDebt": 62.0e9,
             "totalCash": 5.0e9,  "freeCashflow": 8.5e9, "interestExpense": 2.45e9,
             "earningsGrowth": -0.30,"sharesOutstanding": 5.68e9},
    # ── 消費品 ────────────────────────────────────────────
    "KO":   {"trailingEps": 2.85,  "payoutRatio": 0.65, "priceToBook": 10.50,
             "currentRatio": 1.10, "ebitda": 13.0e9, "totalDebt": 35.0e9,
             "totalCash": 9.5e9,  "freeCashflow": 9.5e9, "interestExpense": 1.05e9,
             "earningsGrowth": 0.06, "sharesOutstanding": 4.31e9},
    "PEP":  {"trailingEps": 8.20,  "payoutRatio": 0.68, "priceToBook": 12.50,
             "currentRatio": 0.85, "ebitda": 16.0e9, "totalDebt": 42.0e9,
             "totalCash": 8.5e9,  "freeCashflow": 8.5e9, "interestExpense": 1.35e9,
             "earningsGrowth": 0.04, "sharesOutstanding": 1.38e9},
    "PG":   {"trailingEps": 6.20,  "payoutRatio": 0.58, "priceToBook": 8.50,
             "currentRatio": 0.75, "ebitda": 22.0e9, "totalDebt": 28.0e9,
             "totalCash": 7.5e9,  "freeCashflow": 14.0e9,"interestExpense": 0.92e9,
             "earningsGrowth": 0.05, "sharesOutstanding": 2.36e9},
    "MO":   {"trailingEps": 4.60,  "payoutRatio": 0.78, "priceToBook": None,
             "currentRatio": 0.62, "ebitda": 12.0e9, "totalDebt": 28.0e9,
             "totalCash": 2.0e9,  "freeCashflow": 9.5e9, "interestExpense": 1.08e9,
             "earningsGrowth": 0.03, "sharesOutstanding": 1.73e9},
    "PM":   {"trailingEps": 6.50,  "payoutRatio": 0.88, "priceToBook": None,
             "currentRatio": 0.92, "ebitda": 18.0e9, "totalDebt": 48.0e9,
             "totalCash": 3.5e9,  "freeCashflow": 10.0e9,"interestExpense": 1.55e9,
             "earningsGrowth": 0.08, "sharesOutstanding": 1.55e9},
    # ── 電訊 ──────────────────────────────────────────────
    "T":    {"trailingEps": 2.25,  "payoutRatio": 0.50, "priceToBook": 1.05,
             "currentRatio": 0.62, "ebitda": 45.0e9, "totalDebt": 142.0e9,
             "totalCash": 2.5e9,  "freeCashflow": 16.0e9,"interestExpense": 6.20e9,
             "earningsGrowth": 0.03, "sharesOutstanding": 7.15e9},
    "VZ":   {"trailingEps": 4.20,  "payoutRatio": 0.62, "priceToBook": 1.60,
             "currentRatio": 0.68, "ebitda": 48.0e9, "totalDebt": 148.0e9,
             "totalCash": 2.0e9,  "freeCashflow": 18.0e9,"interestExpense": 5.80e9,
             "earningsGrowth": 0.02, "sharesOutstanding": 4.21e9},
    # ── ETF（無財務數據，只補 payoutRatio/sharesOutstanding）─
    "SCHD": {"trailingEps": None,  "payoutRatio": None,  "priceToBook": None,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": None,"sharesOutstanding": None},
    "VYM":  {"trailingEps": None,  "payoutRatio": None,  "priceToBook": None,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": None,"sharesOutstanding": None},
    "HDV":  {"trailingEps": None,  "payoutRatio": None,  "priceToBook": None,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": None,"sharesOutstanding": None},
    "JEPI": {"trailingEps": None,  "payoutRatio": None,  "priceToBook": None,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": None,"sharesOutstanding": None},
    "JEPQ": {"trailingEps": None,  "payoutRatio": None,  "priceToBook": None,
             "currentRatio": None, "ebitda": None,    "totalDebt": None,
             "totalCash": None,   "freeCashflow": None,"interestExpense": None,
             "earningsGrowth": None,"sharesOutstanding": None},
}

# ==========================================
# ⚙️  股票清單
# ==========================================
# =========================================================
# TICKERS_CONFIG — 從外部 JSON 載入（單一真相來源）
# 手動新增：直接編輯 tickers_config.json
# Screener 新增：執行 add_to_tracking.py
# =========================================================
_TICKERS_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "tickers_config.json")
# =========================================================
# 通用日誌函數（模組層級早期使用）
# =========================================================
log = print  # 簡單別名，輸出到 stdout
def _load_tickers_config(path: str) -> dict:
    """
    讀取 tickers_config.json，返回與舊版相容的格式：
    { ticker: (name, market), ... }
    """
    import json as _json
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"找不到 {path}\n"
            f"請先執行 export_tickers_json.py 生成設定檔。"
        )
    with open(path, encoding="utf-8") as f:
        data = _json.load(f)
    tickers = data.get("tickers", {})
    result  = {t: (v["name"], v["market"]) for t, v in tickers.items()}
    meta    = data.get("_meta", {})
    log(f"  📋 tickers_config.json 載入: {len(result)} 隻股票"
        f"  (更新日期: {meta.get('last_updated','?')})")
    by_mkt = {}
    for _, (_, mkt) in result.items():
        by_mkt[mkt] = by_mkt.get(mkt, 0) + 1
    for mkt, cnt in sorted(by_mkt.items()):
        log(f"     {mkt}: {cnt} 隻")
    return result

TICKERS_CONFIG = _load_tickers_config(_TICKERS_JSON)



TODAY_DATE = datetime.date.today()

EXCEL_FILES = {
    "UK": "UK_Dividend_Analysis.xlsx",
    "HK": "HK_Dividend_Analysis.xlsx",
    "US": "US_Dividend_Analysis.xlsx",
    "CN": "CN_Dividend_Analysis.xlsx",
}

def ticker_to_sheet(tid: str) -> str:
    """Ticker → 工作表名稱（與 daily_importer 完全一致，確保兩個腳本互通）"""
    return tid.replace(".L", "").replace(".HK", "").replace("-", "_").replace(".", "_")

STANDARD_COLUMNS = [
    "Date", "Close", "Volume", "Dividend_Amount",
    "Current_Yield_%",
    "Yield_5yr_Avg", "Yield_買入線", "Yield_賣出線",
    "Current_PE",
    "PE_5yr_Avg", "PE_買入線", "PE_賣出線",
    "Company_Name",
]

SCORE_SNAPSHOT_COLS = [
    "Score_日期",
    "Payout_Ratio_%", "FCF_Coverage", "Net_Debt_EBITDA",
    "Interest_Coverage", "Current_Ratio", "PB_Ratio",
    "Yield_Spread_vs_Bond", "DGR_3yr_%", "RSI_14",
    "52W_Position_%",
    "Score_股息質量_30", "Score_估值_25", "Score_財務健康_25",
    "Score_增長潛力_10", "Score_技術面_10", "Score_總分_100",
]

ALL_COLUMNS     = STANDARD_COLUMNS + SCORE_SNAPSHOT_COLS
STANDARD_NCOLS  = len(STANDARD_COLUMNS)      # 13 — cols A-M
SCORE_COL_START = STANDARD_NCOLS + 1         # 14 — col N onwards（與 daily_importer 一致）
SUMMARY_COLS  = [
    "股票代號", "公司名稱", "現價",
    "最新股息率", "5年均息率", "🟢 息率買入線", "🔴 息率賣出線",
    "最新 PE",   "5年均 PE",  "🟢 PE買入線",   "🔴 PE賣出線",
    "Payout_%", "FCF覆蓋", "Net_Debt/EBITDA", "利息覆蓋",
    "流動比率", "P/B",     "Yield_Spread",     "DGR_3yr%",
    "RSI_14",   "52W位置%",
    "S_股息質量", "S_估值", "S_財務健康", "S_增長", "S_技術",
    "📊 總分_100", "📊 綜合診斷",
]

# =========================================================
# IBKRFundamentalFetcher（多模式，靜默 Error 10358）
# =========================================================
class IBKRFundamentalFetcher:
    def __init__(self):
        self.ib        = None
        self.connected = False
        self.mode      = None
        self._ib_app   = IBAPIFundamentalApp()

        IB_cls, Stock_cls = try_import_ibkr()
        self.IB_cls    = IB_cls
        self.Stock_cls = Stock_cls

        if IB_cls == "IBAPI_MODE":
            self.mode = "SOCKET"
        elif IB_cls is not None:
            self.mode = "IB_INSYNC"
        else:
            self.mode = "SOCKET"

    def connect(self) -> bool:
        if not USE_IBKR:
            return False
        if self.mode == "IB_INSYNC":
            return self._connect_ib_insync()
        else:
            return self._test_socket_connection()

    def _connect_ib_insync(self) -> bool:
        try:
            try:
                asyncio.get_event_loop()
            except RuntimeError:
                asyncio.set_event_loop(asyncio.new_event_loop())
            self.ib = self.IB_cls()

            # 靜默 Error 10358（Fundamentals not allowed）及其他非致命錯誤
            def _silent_error(reqId, errorCode, errorString, contract):
                if errorCode not in (10358, 2104, 2106, 2158, 1100, 1101, 1102):
                    pass  # 可按需啟用: print(f"IBKR Error {errorCode}: {errorString}")
            self.ib.errorEvent += _silent_error

            self.ib.connect(IBKR_HOST, IBKR_PORT,
                            clientId=IBKR_CLIENT_ID,
                            timeout=IBKR_TIMEOUT, readonly=True)
            self.connected = self.ib.isConnected()
            if self.connected:
                print(f"  ✅ IBKR ib_insync 已連接 ({IBKR_HOST}:{IBKR_PORT})")
            else:
                print("  ⚠️  ib_insync 連接失敗，改用 SOCKET 模式")
                self.mode = "SOCKET"
                return self._test_socket_connection()
        except Exception as e:
            print(f"  ⚠️  ib_insync 連接錯誤: {e}，改用 SOCKET 模式")
            self.mode = "SOCKET"
            return self._test_socket_connection()
        return self.connected

    def _test_socket_connection(self) -> bool:
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(5)
            result = s.connect_ex((IBKR_HOST, IBKR_PORT))
            s.close()
            if result == 0:
                self.connected = True
                print(f"  ✅ IBKR Gateway 可達 (SOCKET 模式) ({IBKR_HOST}:{IBKR_PORT})")
            else:
                self.connected = False
                print(f"  ⚠️  IBKR Gateway 無法連接 port {IBKR_PORT}")
        except Exception as e:
            self.connected = False
            print(f"  ⚠️  IBKR Gateway 連接測試失敗: {e}")
        return self.connected

    def disconnect(self):
        if self.mode == "IB_INSYNC" and self.ib:
            try:
                self.ib.disconnect()
                print("  IBKR 已斷開")
            except Exception:
                pass
        self.connected = False

    def get_fundamentals(self, ticker_id: str, market: str) -> dict:
        """v5: 三市場通用（HK / UK / US）"""
        if not self.connected:
            return {}
        if self.mode == "IB_INSYNC":
            return self._get_via_ib_insync(ticker_id, market)
        else:
            return self._get_via_socket(ticker_id, market)

    # 保留舊名稱作向後兼容
    def get_hk_fundamentals(self, ticker_id: str) -> dict:
        return self.get_fundamentals(ticker_id, "HK")

    def _get_via_ib_insync(self, ticker_id: str, market: str = "HK") -> dict:
        exchange_map = {"HK": ("SEHK", "HKD"), "UK": ("LSE", "GBP"), "US": ("SMART", "USD")}
        exchange, currency = exchange_map.get(market, ("SMART", "USD"))
        if market == "HK":
            symbol = ticker_id.replace(".HK", "").lstrip("0") or "0"
        elif market == "UK":
            symbol = ticker_id.replace(".L", "")
        else:
            symbol = ticker_id
        try:
            contract = self.Stock_cls(symbol, exchange, currency)
            self.ib.qualifyContracts(contract)
            xml_str  = self.ib.reqFundamentalData(contract, "ReportSnapshot")
            return self._parse_xml(xml_str) if xml_str else {}
        except Exception:
            return {}

    def _get_via_socket(self, ticker_id: str, market: str = "HK") -> dict:
        try:
            xml_str = self._ib_app.get_fundamentals_raw(ticker_id, market)
            return self._parse_xml(xml_str) if xml_str else {}
        except Exception:
            return {}

    def _parse_xml(self, xml_str: str) -> dict:
        if not xml_str:
            return {}
        import xml.etree.ElementTree as ET
        result = {}
        try:
            root = ET.fromstring(xml_str)
            field_map = {
                "EPSExclExtraItems":  ("trailingEps",       1.0,  False),
                "EPS":                ("trailingEps",       1.0,  False),
                "PayoutRatio":        ("payoutRatio",       1.0,  True),
                "BookValuePerShare":  ("bookValuePerShare", 1.0,  False),
                "CurrentRatio":       ("currentRatio",      1.0,  False),
                "EBITDA":             ("ebitda",            1e6,  False),
                "TotalDebt":          ("totalDebt",         1e6,  False),
                "CashAndEquivalents": ("totalCash",         1e6,  False),
                "FreeCashFlow":       ("freeCashflow",      1e6,  False),
                "InterestExpense":    ("interestExpense",   1e6,  False),
                "SharesOutstanding":  ("sharesOutstanding", 1e6,  False),
                "EarningsGrowth":     ("earningsGrowth",    1.0,  True),
            }
            for xml_tag, (py_field, mult, div100) in field_map.items():
                if py_field in result:
                    continue
                for elem in root.iter(xml_tag):
                    try:
                        val = float(elem.text)
                        if div100 and abs(val) > 1:
                            val /= 100.0
                        result[py_field] = val * mult
                        break
                    except Exception:
                        pass
            if "bookValuePerShare" in result and result["bookValuePerShare"] > 0:
                result["_bvps"] = result.pop("bookValuePerShare")
        except Exception:
            pass
        return result


# =========================================================
# 全局 IBKR 實例
# =========================================================
_ibkr_fetcher = IBKRFundamentalFetcher()

# =========================================================
# v5 新增：數據合理性檢查
# =========================================================
def check_suspicious_fields(info: dict, ticker_id: str) -> list:
    """
    檢查 yfinance 返回的數據是否有異常欄位。
    返回異常欄位清單（空清單表示全部正常）。
    """
    suspicious = []

    # ① payoutRatio > 2.0（200%）或 < 0
    pr = info.get("payoutRatio")
    if pr is not None and (pr > 2.0 or pr < 0):
        suspicious.append(("payoutRatio", pr, "超出合理範圍 0~2.0"))

    # ② currentRatio > 50 或 < 0（非銀行/保險股）
    cr = info.get("currentRatio")
    if cr is not None and (cr > 50 or cr < 0):
        suspicious.append(("currentRatio", cr, "超出合理範圍 0~50"))

    # ③ priceToBook < 0 或 > 100
    pb = info.get("priceToBook")
    if pb is not None and (pb < 0 or pb > 100):
        suspicious.append(("priceToBook", pb, "超出合理範圍 0~100"))

    # ④ trailingEps 導致隱含 PE 異常
    eps = info.get("trailingEps")
    raw_price = info.get("currentPrice") or info.get("regularMarketPrice")
    if eps is not None and raw_price and eps != 0:
        # UK 股票：info["currentPrice"] 有時用便士，有時用英鎊
        # 用啟發式判斷：若 price > 50 且 EPS < 5，很可能 price 是便士
        if ticker_id.endswith(".L") and raw_price > 50 and abs(eps) < 10:
            price = raw_price / 100.0  # 統一換算為英鎊
        else:
            price = raw_price
        implied_pe = price / eps
        if eps < 0:
            pass  # 虧損，不標記
        elif implied_pe > 500:
            note = "疑似英鎊/便士單位不一致" if ticker_id.endswith(".L") else "隱含 PE 異常"
            suspicious.append(("trailingEps", eps, f"隱含 PE={implied_pe:.1f}，{note}"))
        elif implied_pe < 0:
            suspicious.append(("trailingEps", eps, f"隱含 PE={implied_pe:.1f} 異常"))

    # ⑤ freeCashflow 數量級與市值相差 > 1000 倍（明顯單位錯誤）
    fcf = info.get("freeCashflow")
    mktcap = info.get("marketCap")
    if fcf is not None and mktcap and mktcap > 0:
        ratio = abs(fcf) / mktcap
        if ratio > 10:
            suspicious.append(("freeCashflow", fcf, f"FCF/市值={ratio:.1f} 疑似單位錯誤"))

    if suspicious:
        for f, v, reason in suspicious:
            print(f"    ⚠️  v5 合理性警告 [{ticker_id}]: {f}={v:.6g} ({reason})")

    return [f for f, _, _ in suspicious]


# =========================================================
# 多層數據源獲取（v5：yfinance → 合理性檢查 → IBKR補救 → 靜態）
# =========================================================
def get_info_with_fallback(ticker_id: str, market: str) -> dict:
    """
    v5 四層數據源：
      Layer 1: yfinance
      Layer 2: 合理性檢查，找出異常欄位
      Layer 3: IBKR（三市場通用）補救異常欄位 + 填補缺失欄位
      Layer 4: 靜態備用（UK / HK / US）補充仍缺失的欄位
    """
    info = {}

    # ── Layer 1: yfinance ─────────────────────────────────
    try:
        info = yf.Ticker(ticker_id).info or {}
        key_ok = [f for f in ["trailingEps", "payoutRatio", "freeCashflow",
                               "ebitda", "totalDebt", "priceToBook", "currentRatio",
                               "ebit", "operatingIncome", "totalCash", "interestExpense"]
                  if info.get(f) is not None]
        if key_ok:
            print(f"    📡 yfinance OK: {key_ok}")
    except Exception as e:
        print(f"    ⚠️  yfinance 失敗: {e}")

    # ── Layer 2: 合理性檢查 ────────────────────────────────
    suspicious_fields = check_suspicious_fields(info, ticker_id)

    # ── 決定使用哪個靜態表 ────────────────────────────────
    static_map = {
        "UK": UK_STATIC_FUNDAMENTALS,
        "HK": HK_STATIC_FUNDAMENTALS,
        "US": US_STATIC_FUNDAMENTALS,
    }
    FILL_FIELDS = ["trailingEps", "payoutRatio", "freeCashflow", "ebitda",
                   "totalDebt", "priceToBook", "currentRatio", "ebit",
                   "operatingIncome", "totalCash", "interestExpense",
                   "earningsGrowth", "sharesOutstanding"]

    missing = [f for f in FILL_FIELDS if not info.get(f)]

    # ── Layer 3: IBKR（三市場通用）──────────────────────
    # 觸發條件：有缺失欄位 OR 有合理性異常欄位
    needs_ibkr = missing or suspicious_fields
    if needs_ibkr and USE_IBKR and _ibkr_fetcher.connected:
        try:
            ibkr_data = _ibkr_fetcher.get_fundamentals(ticker_id, market)
            if ibkr_data:
                # 計算 P/B（IBKR 返回每股帳面值）
                if "_bvps" in ibkr_data:
                    price = safe(info.get("currentPrice") or
                                 info.get("regularMarketPrice"), 0)
                    if price > 0:
                        ibkr_data["priceToBook"] = round(price / ibkr_data["_bvps"], 3)
                    del ibkr_data["_bvps"]

                filled = []
                replaced = []
                for f in FILL_FIELDS:
                    if ibkr_data.get(f) is None:
                        continue
                    if f in suspicious_fields:
                        # 異常欄位：用 IBKR 值覆蓋 yfinance 的異常值
                        old_val = info.get(f)
                        info[f] = ibkr_data[f]
                        replaced.append(f"{f}({old_val:.3g}→{ibkr_data[f]:.3g})")
                    elif not info.get(f):
                        # 缺失欄位：正常填補
                        info[f] = ibkr_data[f]
                        filled.append(f)

                if replaced:
                    print(f"    🔌 IBKR 修正異常: {replaced}")
                if filled:
                    print(f"    🔌 IBKR 補充缺失: {filled}")

                missing = [f for f in FILL_FIELDS if not info.get(f)]
        except Exception:
            pass  # 完全靜默

    # ── Layer 4: 靜態備用（三市場通用）──────────────────
    static = static_map.get(market, {}).get(ticker_id, {})
    if static and missing:
        filled = []
        for f in missing:
            if static.get(f) is not None and not info.get(f):
                info[f] = static[f]
                filled.append(f)
        if filled:
            print(f"    📋 靜態補充({market}): {filled}")
        # 靜態數據也覆蓋仍然異常的欄位（IBKR 未能修正的情況）
        still_suspicious = [f for f in suspicious_fields if f in (static or {})]
        for f in still_suspicious:
            if static.get(f) is not None:
                info[f] = static[f]
                print(f"    📋 靜態覆蓋異常欄位({market}): {f}={static[f]}")

    # ── ebit 估算（ebitda × 0.6）────────────────────────
    if not info.get("ebit") and not info.get("operatingIncome"):
        ebitda_val = safe(info.get("ebitda"), 0)
        if ebitda_val > 0:
            info["ebit"] = ebitda_val * 0.6

    return info


# ==========================================
# 📊 工具函數
# ==========================================
def safe(v, default=0.0):
    if v is None:
        return default
    try:
        f = float(v)
        return default if f != f else f
    except Exception:
        return default


# ==========================================
# 📊 評分系統（五維度）
# ==========================================
def score_dividend_quality(info, c_yield, y_avg, y_std, df_hist):
    """股息質量（滿分 30）"""
    pts = 0.0
    # ① 息率位置（10分）
    y_buy  = y_avg + 1.5 * y_std
    y_sell = y_avg - 1.5 * y_std
    if   c_yield >= y_buy:  pts += 10
    elif c_yield >= y_avg:  pts += 7
    elif c_yield >= y_sell: pts += 4
    else:                   pts += 1
    # ② 派息比率（10分）
    pr = safe(info.get("payoutRatio"), -1)
    if pr < 0:
        eps = safe(info.get("trailingEps"), 0)
        div = safe(info.get("dividendRate") or
                   info.get("trailingAnnualDividendRate"), 0)
        pr  = (div / eps) if eps > 0 else -1
    if   0 < pr <= 0.50: pts += 10
    elif 0 < pr <= 0.65: pts += 8
    elif 0 < pr <= 0.80: pts += 5
    elif 0 < pr <= 0.95: pts += 2
    elif pr > 0.95:      pts += 0
    else:                pts += 5   # 未知，給中等分
    # ③ FCF 覆蓋（8分）
    fcf       = safe(info.get("freeCashflow"), 0)
    div_total = safe(info.get("dividendRate") or
                     info.get("trailingAnnualDividendRate"), 0)
    shares    = safe(info.get("sharesOutstanding"), 0)
    div_paid  = div_total * shares if shares > 0 else 0
    if fcf < 0:
        pts += 0                 # 現金流為負，明確懲罰
    elif fcf > 0 and div_paid > 0:
        fcf_cov = fcf / div_paid
        if   fcf_cov >= 2.0: pts += 8
        elif fcf_cov >= 1.5: pts += 6
        elif fcf_cov >= 1.0: pts += 4
        else:                pts += 1
    else:
        pts += 4  # 未知，給中等分
    # ④ 股息連續性（2分）
    if df_hist is not None and not df_hist.empty:
        divs = df_hist[df_hist["Dividend_Amount"] > 0].copy()
        if len(divs) >= 2:
            divs["Year"] = pd.to_datetime(divs["Date"]).dt.year
            annual = divs.groupby("Year")["Dividend_Amount"].sum()
            cut = any(
                annual.iloc[i] < annual.iloc[i-1] * 0.90
                for i in range(1, len(annual))
            )
            pts += 0 if cut else 2
        else:
            pts += 1
    else:
        pts += 1
    return round(pts, 2)


def score_valuation(info, c_yield, y_avg, y_std, c_pe, pe_avg, pe_std,
                    risk_free_rate=4.3):
    """估值（滿分 25）"""
    pts = 0.0
    # ① PE 位置（8分）
    pe_buy  = pe_avg - 1.5 * pe_std
    pe_sell = pe_avg + 1.5 * pe_std
    if   c_pe <= pe_buy:  pts += 8
    elif c_pe <= pe_avg:  pts += 6
    elif c_pe <= pe_sell: pts += 3
    else:                 pts += 0
    # ② P/B（7分）
    pb = safe(info.get("priceToBook"), -1)
    if   pb <= 0:    pts += 2   # 未知/負值
    elif pb <= 0.8:  pts += 7   # 真正折讓才給滿分
    elif pb <= 1.2:  pts += 5
    elif pb <= 2.0:  pts += 3
    elif pb <= 3.5:  pts += 1
    else:            pts += 0
    # ③ Yield Spread vs 無風險利率（10分）
    spread = c_yield - risk_free_rate
    if   spread >= 3.0:  pts += 10
    elif spread >= 2.0:  pts += 8
    elif spread >= 1.0:  pts += 6
    elif spread >= 0.0:  pts += 4
    elif spread >= -1.0: pts += 2
    else:                pts += 0
    return round(pts, 2)


def score_financial_health(info):
    """財務健康（滿分 25）"""
    pts = 0.0
    ebitda  = safe(info.get("ebitda"), 0)
    total_d = safe(info.get("totalDebt"), 0)
    cash    = safe(info.get("totalCash") or info.get("cash"), 0)
    net_debt = total_d - cash
    # ① Net Debt / EBITDA（10分）
    if ebitda > 0:
        nd_eb = net_debt / ebitda
        if   nd_eb <= 1.0:  pts += 10
        elif nd_eb <= 2.0:  pts += 8
        elif nd_eb <= 3.0:  pts += 6
        elif nd_eb <= 4.5:  pts += 3
        else:               pts += 0
    elif net_debt <= 0:     pts += 10
    else:                   pts += 4
    # ② 利息覆蓋（Interest Coverage）（9分）
    ebit    = safe(info.get("ebit") or info.get("operatingIncome"), 0)
    int_exp = abs(safe(info.get("interestExpense"), 0))
    if ebit < 0:              pts += 0    # 營業虧損，明確懲罰
    elif ebit > 0 and int_exp > 0:
        ic = ebit / int_exp
        if   ic >= 8:   pts += 9
        elif ic >= 5:   pts += 7
        elif ic >= 3:   pts += 4
        elif ic >= 1.5: pts += 2
        else:           pts += 0
    else:
        pts += 4  # 未知，給中等分
    # ③ 流動比率（6分）
    cr = safe(info.get("currentRatio"), -1)
    if   cr < 0:    pts += 3   # 未知（銀行/保險常見）
    elif cr >= 2.0: pts += 6
    elif cr >= 1.5: pts += 5
    elif cr >= 1.0: pts += 3
    else:           pts += 0
    return round(pts, 2)


def score_growth(info, df_hist):
    """增長潛力（滿分 10）"""
    pts = 0.0
    dgr = None
    if df_hist is not None and not df_hist.empty:
        divs = df_hist[df_hist["Dividend_Amount"] > 0].copy()
        if len(divs) >= 2:
            divs["Year"] = pd.to_datetime(divs["Date"]).dt.year
            annual = divs.groupby("Year")["Dividend_Amount"].sum()
            if len(annual) >= 4:
                d_new = annual.iloc[-1]; d_old = annual.iloc[-4]
                dgr   = ((d_new/d_old)**(1/3)-1)*100 if d_old > 0 else 0
            elif len(annual) >= 2:
                d_new = annual.iloc[-1]; d_old = annual.iloc[0]
                n     = len(annual) - 1
                dgr   = ((d_new/d_old)**(1/n)-1)*100 if d_old > 0 else 0
    if dgr is None: dgr = 0
    # ① 股息增長率（6分）
    if   dgr >= 8:  pts += 6
    elif dgr >= 5:  pts += 5
    elif dgr >= 2:  pts += 4
    elif dgr >= 0:  pts += 2
    else:           pts += 0
    # ② 盈利增長（4分）
    eg_raw = info.get("earningsGrowth")
    if eg_raw is None:
        eg_raw = info.get("earningsQuarterlyGrowth")
    if eg_raw is None:
        pts += 2  # 未知，給中等分
    else:
        eg = safe(eg_raw, 0.0)
        if   eg >= 0.10: pts += 4
        elif eg >= 0.03: pts += 3
        elif eg >= 0:    pts += 2
        else:            pts += 0
    return round(pts, 2)


def score_technical(df_hist, info):
    """技術面（滿分 10）"""
    pts = 0.0
    w52_low  = safe(info.get("fiftyTwoWeekLow"), 0)
    w52_high = safe(info.get("fiftyTwoWeekHigh"), 0)
    price    = safe(info.get("currentPrice") or
                    info.get("regularMarketPrice"), 0)
    # ① 52週位置（6分）
    if w52_low > 0 and w52_high > w52_low and price > 0:
        pos = (price - w52_low) / (w52_high - w52_low)
        if   pos <= 0.25: pts += 6
        elif pos <= 0.40: pts += 5
        elif pos <= 0.60: pts += 3
        elif pos <= 0.80: pts += 1
        else:             pts += 0
    else:
        pts += 3
    # ② RSI（4分）
    if df_hist is not None and len(df_hist) >= 15:
        closes = pd.Series(df_hist["Close"].values, dtype=float)
        delta  = closes.diff()
        gain   = delta.clip(lower=0).rolling(14).mean()
        loss   = (-delta.clip(upper=0)).rolling(14).mean()
        rs     = gain / loss.replace(0, np.nan)
        rsi    = (100 - 100 / (1 + rs)).iloc[-1]
        if   rsi <= 30:  pts += 4
        elif rsi <= 45:  pts += 3
        elif rsi <= 60:  pts += 2
        else:            pts += 0
    else:
        pts += 2
    return round(pts, 2)


def compute_score(info, c_yield, y_avg, y_std, c_pe, pe_avg, pe_std,
                  df_hist, risk_free_rate=4.3):
    """整合五維度評分"""
    s_div  = score_dividend_quality(info, c_yield, y_avg, y_std, df_hist)
    s_val  = score_valuation(info, c_yield, y_avg, y_std, c_pe, pe_avg, pe_std,
                             risk_free_rate=risk_free_rate)
    s_fin  = score_financial_health(info)
    s_grow = score_growth(info, df_hist)
    s_tech = score_technical(df_hist, info)
    # 防禦：任何意外 None 改為 0
    s_div  = s_div  if s_div  is not None else 0.0
    s_val  = s_val  if s_val  is not None else 0.0
    s_fin  = s_fin  if s_fin  is not None else 0.0
    s_grow = s_grow if s_grow is not None else 0.0
    s_tech = s_tech if s_tech is not None else 0.0
    total  = s_div + s_val + s_fin + s_grow + s_tech
    return {
        "Score_股息質量_30": s_div,
        "Score_估值_25":     s_val,
        "Score_財務健康_25": s_fin,
        "Score_增長潛力_10": s_grow,
        "Score_技術面_10":   s_tech,
        "Score_總分_100":    round(total, 1),
    }


def get_score_snapshot(info, df_stock, c_yield, y_avg, y_std,
                       c_pe, pe_avg, pe_std, risk_free_rate=4.3):
    """生成評分快照字典"""
    scores = compute_score(info, c_yield, y_avg, y_std, c_pe, pe_avg, pe_std,
                           df_stock, risk_free_rate=risk_free_rate)
    # Payout Ratio
    pr = info.get("payoutRatio")
    if pr is None:
        eps = safe(info.get("trailingEps"), 0)
        div = safe(info.get("dividendRate") or
                   info.get("trailingAnnualDividendRate"), 0)
        pr  = round(div / eps * 100, 2) if eps > 0 else None
    else:
        pr = round(pr * 100, 2)
    # FCF Coverage
    fcf      = safe(info.get("freeCashflow"), 0)
    div_rate = safe(info.get("dividendRate") or
                    info.get("trailingAnnualDividendRate"), 0)
    shares   = safe(info.get("sharesOutstanding"), 0)
    div_paid = div_rate * shares if shares > 0 else 0
    fcf_cov  = round(fcf / div_paid, 2) if div_paid > 0 and fcf > 0 else None
    # Net Debt / EBITDA
    ebitda  = safe(info.get("ebitda"), 0)
    total_d = safe(info.get("totalDebt"), 0)
    cash    = safe(info.get("totalCash") or info.get("cash"), 0)
    net_debt = total_d - cash
    nd_eb   = round(net_debt / ebitda, 2) if ebitda > 0 else None
    # Interest Coverage
    ebit    = safe(info.get("ebit") or info.get("operatingIncome"), 0)
    int_exp = abs(safe(info.get("interestExpense"), 0))
    ic      = round(ebit / int_exp, 2) if int_exp > 0 and ebit > 0 else None
    # Current Ratio / P/B
    cr = safe(info.get("currentRatio"), None)
    pb = safe(info.get("priceToBook"), None)
    # 52W Position
    w52_low  = safe(info.get("fiftyTwoWeekLow"), 0)
    w52_high = safe(info.get("fiftyTwoWeekHigh"), 0)
    price    = safe(info.get("currentPrice") or
                    info.get("regularMarketPrice"), 0)
    w52_pos  = round((price - w52_low) / (w52_high - w52_low) * 100, 1) \
               if w52_high > w52_low else None
    # RSI
    rsi = None
    if df_stock is not None and len(df_stock) >= 15:
        closes = pd.Series(df_stock["Close"].values, dtype=float)
        delta  = closes.diff()
        gain   = delta.clip(lower=0).rolling(14).mean()
        loss   = (-delta.clip(upper=0)).rolling(14).mean()
        rs     = gain / loss.replace(0, np.nan)
        rsi_s  = 100 - 100 / (1 + rs)
        rsi    = round(float(rsi_s.iloc[-1]), 1) if not rsi_s.empty else None
    # DGR 3yr
    dgr = None
    if df_stock is not None and not df_stock.empty and \
            "Dividend_Amount" in df_stock.columns:
        divs = df_stock[df_stock["Dividend_Amount"] > 0].copy()
        if len(divs) >= 2:
            divs["Year"] = pd.to_datetime(divs["Date"]).dt.year
            annual = divs.groupby("Year")["Dividend_Amount"].sum()
            if len(annual) >= 4:
                d_new = annual.iloc[-1]; d_old = annual.iloc[-4]
                dgr   = round(((d_new/d_old)**(1/3)-1)*100, 2) \
                        if d_old > 0 else 0
            elif len(annual) >= 2:
                d_new = annual.iloc[-1]; d_old = annual.iloc[0]
                n     = len(annual) - 1
                dgr   = round(((d_new/d_old)**(1/n)-1)*100, 2) \
                        if d_old > 0 else 0
    return {
        "Score_日期":           TODAY_DATE,
        "Payout_Ratio_%":       pr,
        "FCF_Coverage":         fcf_cov,
        "Net_Debt_EBITDA":      nd_eb,
        "Interest_Coverage":    ic,
        "Current_Ratio":        round(cr, 2) if cr is not None else None,
        "PB_Ratio":             round(pb, 2) if pb is not None else None,
        "Yield_Spread_vs_Bond": round(c_yield - risk_free_rate, 2),  # 動態利率
        "DGR_3yr_%":            dgr,
        "RSI_14":               rsi,
        "52W_Position_%":       w52_pos,
        **scores,
    }


# ==========================================
# 🎨 Excel 樣式函數
# ==========================================
def get_score_fill(total_val):
    if   total_val >= 75: return PatternFill("solid", fgColor="00B050")
    elif total_val >= 60: return PatternFill("solid", fgColor="C6EFCE")
    elif total_val >= 45: return PatternFill("solid", fgColor="FFEB9C")
    elif total_val >= 30: return PatternFill("solid", fgColor="FFC7CE")
    else:                 return PatternFill("solid", fgColor="FF0000")

def get_score_font(total_val):
    if   total_val >= 75: return Font(name="Arial", size=11, bold=True, color="FFFFFF")
    elif total_val >= 60: return Font(name="Arial", size=11, bold=True, color="375623")
    elif total_val >= 45: return Font(name="Arial", size=11, bold=True, color="7E6000")
    elif total_val >= 30: return Font(name="Arial", size=11, bold=True, color="9C0006")
    else:                 return Font(name="Arial", size=11, bold=True, color="FFFFFF")

def get_status(total):
    if   total >= 75: return "🟢🟢 強力買入"
    elif total >= 60: return "🟢 值得關注"
    elif total >= 45: return "⚖️ 觀望"
    elif total >= 30: return "🟡 偏弱"
    else:             return "🔴 避開"

def apply_styles_to_summary(ws, total_col):
    header_fill = PatternFill("solid", fgColor="3F3F3F")
    score_hdr   = PatternFill("solid", fgColor="1F4E79")
    score_fill  = PatternFill("solid", fgColor="E8F4FD")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    normal_font = Font(name="Arial", size=10)
    link_font   = Font(name="Arial", size=10, color="0000FF", underline="single")
    thin_border = Border(
        left=Side(style="thin",  color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin",   color="E0E0E0"),
        bottom=Side(style="thin",color="E0E0E0"),
    )
    def _sn(tid):
        return (tid.replace(".L","").replace(".HK","")
                   .replace("-","_").replace(".","_"))
    for col in range(1, total_col + 1):
        c = ws.cell(1, col)
        c.fill = score_hdr if col >= 22 else header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in range(2, ws.max_row + 1):
        total_val = safe(ws.cell(row, total_col - 1).value, 0)
        row_fill  = get_score_fill(total_val)
        row_font  = get_score_font(total_val)
        tid = str(ws.cell(row, 1).value or "")
        for col in range(1, total_col + 1):
            c = ws.cell(row, col)
            c.border = thin_border
            if col < 22:
                c.fill = row_fill; c.font = normal_font
            else:
                c.fill = score_fill; c.font = normal_font
            if col == total_col - 1:
                c.fill = row_fill; c.font = row_font
                c.number_format = "0.0"
                c.alignment = Alignment(horizontal="center")
            if col == total_col:
                c.fill = row_fill; c.font = row_font
                c.alignment = Alignment(horizontal="center")
            if col == 1 and tid:
                c.hyperlink = f"#'{_sn(tid)}'!A1"
                c.font = link_font
            if col in [4, 5, 6, 7, 13, 17, 18]:  c.number_format = "0.00"
            if col in [3, 8, 9, 10, 11]:          c.number_format = "#,##0.00"
            if col in [22, 23, 24, 25, 26]:        c.number_format = "0.0"


# =========================================================
# CN 市場數據層（AKShare 新浪財經源 — 可突破大陸以外 IP 限制）
# =========================================================
def _cn_code(ticker_id: str) -> str:
    """600036.SS → sh600036，000001.SZ → sz000001（新浪格式）"""
    code = ticker_id.replace(".SS","").replace(".SZ","")
    if ticker_id.endswith(".SS"):
        return "sh" + code
    else:
        return "sz" + code

def get_cn_history(ticker_id: str, period_years: int = 5) -> pd.DataFrame:
    """
    取得 A股歷史日線數據（前復權）— 使用新浪財經源。
    返回含 Date / Close / Volume / Dividend_Amount 的 DataFrame。
    新浪源可在大陸以外 IP 正常訪問。
    """
    if not _AKSHARE_OK:
        return pd.DataFrame()
    sina_code = _cn_code(ticker_id)
    try:
        df = ak.stock_zh_a_daily(symbol=sina_code, adjust="qfq")
        # 欄位：date open high low close volume amount outstanding_share turnover
        df = df.rename(columns={"date": "Date", "close": "Close", "volume": "Volume"})
        df["Date"]            = pd.to_datetime(df["Date"]).dt.date
        df["Close"]           = pd.to_numeric(df["Close"],  errors="coerce")
        df["Volume"]          = pd.to_numeric(df["Volume"], errors="coerce").fillna(0).astype(int)
        df["Dividend_Amount"] = 0.0
        df = df[["Date","Close","Volume","Dividend_Amount"]].copy()
        # 只取最近 period_years 年
        import datetime as _dt
        cutoff = _dt.date.today() - _dt.timedelta(days=period_years*366)
        df = df[df["Date"] >= cutoff].reset_index(drop=True)
        print(f"    ✅ AKShare(新浪) 歷史數據: {len(df)} 行")
        return df
    except Exception as e:
        print(f"    ⚠️  AKShare CN hist 失敗: {e}")
        return pd.DataFrame()

def get_cn_dividends(ticker_id: str) -> pd.DataFrame:
    """
    取得 A股歷史派息記錄。
    A股通常年派一次，除息日在年報後（4–6月）。
    返回 DataFrame with Date / Dividend_Amount（人民幣每股）。
    """
    if not _AKSHARE_OK:
        return pd.DataFrame()
    code = ticker_id.replace(".SS","").replace(".SZ","")
    try:
        df = ak.stock_dividend_cninfo(symbol=code)
        date_col = next((c for c in df.columns
                         if any(k in c for k in ["除息","日期","date"])), None)
        div_col  = next((c for c in df.columns
                         if any(k in c for k in ["每股","股息","派息","dividend"])), None)
        if date_col and div_col:
            raw_divs = pd.to_numeric(df[div_col], errors="coerce").fillna(0)
            result = pd.DataFrame({
                "Date": pd.to_datetime(df[date_col], errors="coerce").dt.date,
                # ⚠️ AKShare stock_dividend_cninfo 返回「角」(0.1元)為單位
                # 需除以 10 換算成「元」，否則息率虛高 10 倍
                "Dividend_Amount": raw_divs / 10.0,
            }).dropna()
            result = result[result["Dividend_Amount"] > 0].reset_index(drop=True)
            print(f"    ✅ AKShare 股息記錄: {len(result)} 筆（已換算至元，÷10）")
            return result
    except Exception as e:
        print(f"    ⚠️  AKShare 股息失敗（可能被封）: {e}")
    return pd.DataFrame()

def get_cn_fundamentals(ticker_id: str) -> dict:
    """
    取得 A股基本面（yfinance .SS/.SZ 格式作補充）。
    AKShare 即時行情在大陸以外 IP 常被封，改用 yfinance 取 52W 高低等數據。
    """
    info = {}
    # yfinance 支援 SS/SZ 格式
    try:
        yf_info = yf.Ticker(ticker_id).info or {}
        for k in ["fiftyTwoWeekHigh","fiftyTwoWeekLow","currentPrice",
                  "regularMarketPrice","sharesOutstanding","earningsGrowth",
                  "trailingEps","payoutRatio","priceToBook","marketCap"]:
            if yf_info.get(k) is not None:
                info[k] = yf_info[k]
    except Exception:
        pass
    return info

# ==========================================
# 🚀 主程式
# ==========================================
print("==========================================")
print(f"🚀 全球三市場評分看盤系統 v6 ── ({TODAY_DATE})")
print("==========================================")
print(f"📌 無風險利率設定: UK={RISK_FREE_RATES['UK']}%  "
      f"US={RISK_FREE_RATES['US']}%  HK={RISK_FREE_RATES['HK']}%  CN={RISK_FREE_RATES['CN']}%")
print("==========================================")

# ── 連接 IBKR（只連一次）────────────────────────────────
if USE_IBKR:
    _ibkr_fetcher.connect()

summary_by_market    = {"UK": [], "HK": [], "US": [], "CN": []}
stock_data_by_market = {"UK": {}, "HK": {}, "US": {}, "CN": {}}
results_by_market     = {"UK": {}, "HK": {}, "US": {}, "CN": {}}   # 新股的完整評分結果（給 rebuild_summary 用）

# =========================================================
# 預掃描：邊個市場已有 Excel？邊隻 ticker 已經有自己嗰個 sheet？
# ── 呢一步係「漸進式更新」嘅關鍵：已存在嘅 sheet 完全唔會
#    再被處理／覆寫，daily_importer 每日 append 落去嗰啲歷史
#    紀錄先可以完整保留。淨係真係未出現過嘅新 ticker 先會被抓取。
# =========================================================
existing_wb     = {}   # market -> 已開啟嘅 openpyxl Workbook（若檔案存在）
existing_sheets = {}   # market -> set(已存在嘅個股 sheet 名)

for market, excel_file in EXCEL_FILES.items():
    sum_name = f"{market} 總覽"
    if os.path.exists(excel_file):
        wb_e = openpyxl.load_workbook(excel_file)
        existing_wb[market]     = wb_e
        existing_sheets[market] = set(wb_e.sheetnames) - {sum_name}
        print(f"📂 [{market}] 已有 {excel_file}，現存 {len(existing_sheets[market])} 隻個股 sheet")
    else:
        existing_wb[market]     = None
        existing_sheets[market] = set()
        print(f"📂 [{market}] 未找到 {excel_file}，將會全新建立")

# 完整 ticker → (ticker, name) 對應表（不論新舊），rebuild_summary 需要全市場視角
smap_by_market = {"UK": {}, "HK": {}, "US": {}, "CN": {}}
for ticker_id, (name, market) in TICKERS_CONFIG.items():
    smap_by_market.setdefault(market, {})[ticker_to_sheet(ticker_id)] = (ticker_id, name)

for ticker_id, (name, market) in TICKERS_CONFIG.items():
    sheet_name = ticker_to_sheet(ticker_id)

    # ⏭ 已經有自己嗰個 sheet＝已追蹤緊嘅舊股，完全唔再處理
    #    （唔再重新拉歷史、唔再洗 yfinance/IBKR 配額，最重要係唔會
    #    觸碰到 daily_importer 逐日 append 落去嗰段歷史）
    if sheet_name in existing_sheets.get(market, set()):
        continue

    print(f"\n📥 [{market}] {ticker_id} ({name}) ... (新股，建立歷史)")

    try:
        if market == "CN":
            # ── CN 市場：AKShare 新浪源（可突破大陸以外 IP 限制）──
            df_stock = get_cn_history(ticker_id, period_years=5)
            if df_stock.empty:
                print("  ❌ CN 無歷史數據，跳過"); continue
            # 合併實際除息記錄（A股年派一次，合併後方可計算滾動息率）
            cn_divs = get_cn_dividends(ticker_id)
            if not cn_divs.empty:
                div_map = {d: v for d, v in zip(cn_divs["Date"], cn_divs["Dividend_Amount"])}
                df_stock["Dividend_Amount"] = df_stock["Date"].map(div_map).fillna(0.0)
            info = get_cn_fundamentals(ticker_id)
        else:
            # ── UK / HK / US：沿用 yfinance ──────────────────
            tk     = yf.Ticker(ticker_id)
            df_raw = tk.history(period="5y")
            if df_raw.empty:
                print("  ⚠️  5年數據為空，嘗試 2年...")
                df_raw = tk.history(period="2y")
                if df_raw.empty:
                    print("  ❌ 無歷史數據，跳過"); continue

            df_raw = df_raw.reset_index()
            df_raw.columns = [str(c).capitalize() for c in df_raw.columns]
            if "Dividends" in df_raw.columns:
                df_raw.rename(columns={"Dividends": "Dividend_Amount"}, inplace=True)
            df_raw["Date"] = pd.to_datetime(df_raw["Date"]).dt.date
            for col in ["Close", "Volume", "Dividend_Amount"]:
                if col not in df_raw.columns:
                    df_raw[col] = 0.0 if col != "Volume" else 0

            df_stock = df_raw[["Date", "Close", "Volume", "Dividend_Amount"]].copy()

            # UK：yfinance history 用便士（GBX），統一換算英鎊（GBP）
            if market == "UK":
                df_stock["Close"]           = df_stock["Close"] / 100.0
                df_stock["Dividend_Amount"] = df_stock["Dividend_Amount"] / 100.0

            info = get_info_with_fallback(ticker_id, market)

        # ── EPS 驗證 ──────────────────────────────────────
        current_eps = None
        try:
            current_eps = info.get("trailingEps", None)
            raw_price   = info.get("currentPrice", 0)
            # UK：info["currentPrice"] 可能是便士，Close 已換算為英鎊
            chk_price   = (raw_price / 100.0
                           if market == "UK" and raw_price > 50
                           else raw_price)
            if current_eps and chk_price > 0:
                temp_pe = chk_price / current_eps
                if temp_pe > 80 or temp_pe < 2:
                    current_eps = None
        except Exception:
            pass

        # ── Current Yield % ───────────────────────────────
        # A股年派一次用 400 日滾動，其他市場 252 日
        rolling_window = 400 if market == "CN" else 252
        rolling_div = df_stock["Dividend_Amount"].rolling(rolling_window, min_periods=1).sum()
        hist_yield  = (rolling_div / df_stock["Close"].replace(0, np.nan)) * 100
        hist_yield  = hist_yield.where((hist_yield > 0) & (hist_yield < 30))
        df_stock["Current_Yield_%"] = hist_yield

        # 最後一行：優先用過去 365 日實際除息記錄重新計算
        # 比 dividendRate 更可靠：自動反映削息/加息，唔受格式問題影響
        last_close  = df_stock["Close"].iloc[-1]
        last_date   = df_stock.index[-1] if hasattr(df_stock.index[-1], "year")                       else pd.Timestamp(df_stock["Date"].iloc[-1] if "Date" in df_stock.columns else df_stock.index[-1])
        cutoff      = last_date - pd.Timedelta(days=365)
        recent_divs = df_stock["Dividend_Amount"][df_stock.index >= cutoff]                       if hasattr(df_stock.index[-1], "year")                       else df_stock["Dividend_Amount"].tail(252)
        annual_div_actual = recent_divs.sum()

        if annual_div_actual > 0 and last_close > 0:
            c_yield_current = round(annual_div_actual / last_close * 100, 4)
            print(f"  ✅ 實際除息計算息率: {c_yield_current:.2f}%  (年化股息={annual_div_actual:.4f})")
        else:
            # 後備：dividendRate（格式可能不一致，加合理性上限）
            div_rate = safe(info.get("dividendRate") or info.get("trailingAnnualDividendRate"), 0)
            # UK: dividendRate 有時係便士，需檢查
            if market == "UK" and div_rate > last_close * 0.5 and div_rate > 1:
                div_rate /= 100.0   # 疑似便士，換算英鎊
            c_yield_current = round(div_rate / last_close * 100, 4) if last_close > 0 and div_rate > 0 else 0.0
            if c_yield_current > 0:
                print(f"  ℹ️  除息記錄不足，使用 dividendRate 後備: {c_yield_current:.2f}%")

        # 合理性上限：>30% 視為異常（如除息記錄包含特別股息），改用 hist_yield 均值
        if c_yield_current > 30:
            c_yield_current = float(hist_yield.dropna().mean()) if hist_yield.dropna().any() else 0.0
            print(f"  ⚠️  息率 >30% 異常，改用歷史均值: {c_yield_current:.2f}%")

        if c_yield_current > 0:
            df_stock.at[df_stock.index[-1], "Current_Yield_%"] = c_yield_current

        yields = hist_yield.dropna().values
        y_avg  = np.mean(yields) if len(yields) > 0 else 5.0
        y_std  = np.std(yields)  if len(yields) > 1 else 1.0
        if y_std == 0: y_std = 0.5

        df_stock["Yield_5yr_Avg"]  = y_avg
        df_stock["Yield_買入線"]   = y_avg + 1.5 * y_std
        df_stock["Yield_賣出線"]   = y_avg - 1.5 * y_std

        # ── EPS Fallback（基於股息反推）──────────────────
        if not current_eps or current_eps <= 0 or np.isnan(current_eps):
            safe_div    = df_stock["Close"].iloc[-1] * (y_avg / 100)
            current_eps = safe_div / 0.8
            if current_eps <= 0:
                current_eps = df_stock["Close"].iloc[-1] / 12.0
            print(f"  ℹ️  EPS fallback = {current_eps:.4f} (基於股息反推)")

        # ── Current PE ────────────────────────────────────
        df_stock["Current_PE"] = df_stock["Close"] / current_eps
        df_stock.loc[df_stock["Current_PE"] > 80, "Current_PE"] = 80
        df_stock.loc[df_stock["Current_PE"] < 0,  "Current_PE"] = 0

        pes = df_stock["Current_PE"].dropna().values
        pe_filtered = (
            pes[(pes >= np.percentile(pes, 1)) & (pes <= np.percentile(pes, 99))]
            if len(pes) > 20 else pes
        )
        pe_avg = np.mean(pe_filtered) if len(pe_filtered) > 0 else 12.0
        pe_std = np.std(pe_filtered)  if len(pe_filtered) > 1 else 2.0
        if pe_std == 0: pe_std = 1.5

        df_stock["PE_5yr_Avg"]  = pe_avg
        df_stock["PE_買入線"]   = pe_avg - 1.5 * pe_std
        df_stock["PE_賣出線"]   = pe_avg + 1.5 * pe_std

        # ── 評分（傳入市場對應利率）───────────────────────
        rfr     = RISK_FREE_RATES.get(market, 4.3)
        c_yield = df_stock["Current_Yield_%"].iloc[-1]
        if c_yield is None or (isinstance(c_yield, float) and np.isnan(c_yield)):
            c_yield = 0.0
        c_pe    = df_stock["Current_PE"].iloc[-1]
        snap    = get_score_snapshot(info, df_stock, c_yield, y_avg, y_std,
                                     c_pe, pe_avg, pe_std,
                                     risk_free_rate=rfr)

        # ── 公司名稱 ──────────────────────────────────────
        df_stock["Company_Name"] = None
        df_stock.at[df_stock.index[-1], "Company_Name"] = name

        # ── 寫入 df_stock ─────────────────────────────────
        for col in SCORE_SNAPSHOT_COLS:
            df_stock[col] = None
        last_idx = df_stock.index[-1]
        for col in SCORE_SNAPSHOT_COLS:
            df_stock.at[last_idx, col] = snap.get(col)

        stock_data_by_market[market][sheet_name] = df_stock[ALL_COLUMNS].copy()

        results_by_market[market][sheet_name] = {
            "snap": snap, "y_avg": y_avg, "y_std": y_std,
            "pe_avg": pe_avg, "pe_std": pe_std, "name": name,
            "close": df_stock["Close"].iloc[-1], "yield": c_yield, "pe": c_pe,
        }

        y_buy   = y_avg + 1.5 * y_std
        y_sell  = y_avg - 1.5 * y_std
        pe_buy  = pe_avg - 1.5 * pe_std
        pe_sell = pe_avg + 1.5 * pe_std
        total   = snap["Score_總分_100"]

        summary_by_market[market].append({
            "股票代號": ticker_id, "公司名稱": name,
            "現價":     round(df_stock["Close"].iloc[-1], 2),
            "最新股息率": round(c_yield, 2), "5年均息率": round(y_avg, 2),
            "🟢 息率買入線": round(y_buy, 2),  "🔴 息率賣出線": round(y_sell, 2),
            "最新 PE":  round(c_pe, 2),    "5年均 PE":  round(pe_avg, 2),
            "🟢 PE買入線":  round(pe_buy, 2), "🔴 PE賣出線": round(pe_sell, 2),
            "Payout_%":        snap["Payout_Ratio_%"],
            "FCF覆蓋":         snap["FCF_Coverage"],
            "Net_Debt/EBITDA": snap["Net_Debt_EBITDA"],
            "利息覆蓋":        snap["Interest_Coverage"],
            "流動比率":        snap["Current_Ratio"],
            "P/B":             snap["PB_Ratio"],
            "Yield_Spread":    snap["Yield_Spread_vs_Bond"],
            "DGR_3yr%":        snap["DGR_3yr_%"],
            "RSI_14":          snap["RSI_14"],
            "52W位置%":        snap["52W_Position_%"],
            "S_股息質量":      snap["Score_股息質量_30"],
            "S_估值":          snap["Score_估值_25"],
            "S_財務健康":      snap["Score_財務健康_25"],
            "S_增長":          snap["Score_增長潛力_10"],
            "S_技術":          snap["Score_技術面_10"],
            "📊 總分_100":     total,
            "📊 綜合診斷":     get_status(total),
        })
        print(f"  ✅ 總分: {total}  ({get_status(total)})")

    except Exception as e:
        import traceback
        print(f"  ❌ 處理 {ticker_id} 時出錯: {e}")
        print(f"     {traceback.format_exc()}")

# ── 斷開 IBKR ────────────────────────────────────────────
if USE_IBKR:
    _ibkr_fetcher.disconnect()

# ==========================================
# 💾 寫入 3 個 Excel 文件
# ==========================================
print("\n💾 寫入3個Excel文件...")

def write_df_to_new_sheet(wb, sheet_name, df):
    """將 DataFrame 寫成一個全新嘅 sheet（唔會觸碰任何現存 sheet/工作簿其他內容）。
    行為貼近 pandas .to_excel()：NaN → 空白格。"""
    ws = wb.create_sheet(title=sheet_name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        vals = [None if (isinstance(v, float) and v != v) else v for v in row]
        ws.append(vals)
    return ws


def style_one_stock_sheet(ws, sn):
    """個股 sheet 美化＋圖表。只喺呢個 sheet 第一次建立時叫用一次，
    現存（已經美化過）嗰啲 sheet 唔會經過呢個函數，所以唔會被重新處理。"""
    score_fill  = PatternFill("solid", fgColor="E8F4FD")
    header_fill = PatternFill("solid", fgColor="3F3F3F")
    score_hdr   = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    normal_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin",  color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin",   color="E0E0E0"),
        bottom=Side(style="thin",color="E0E0E0"),
    )
    ncols = len(ALL_COLUMNS)

    ws.freeze_panes = "A2"
    ws.views.sheetView[0].showGridLines = True
    for col in range(1, ncols + 1):
        c = ws.cell(1, col)
        c.fill = score_hdr if col > len(STANDARD_COLUMNS) else header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in range(2, ws.max_row + 1):
        for col in range(1, ncols + 1):
            c = ws.cell(row, col)
            c.font = normal_font; c.border = thin_border
            if col > len(STANDARD_COLUMNS) and c.value is not None:
                c.fill = score_fill
        ws.cell(row, 2).number_format = "#,##0.00"
        ws.cell(row, 3).number_format = "#,##0"
        for col in [5, 6, 7, 8]:    ws.cell(row, col).number_format = '0.00"%"'
        for col in [9, 10, 11, 12]: ws.cell(row, col).number_format = "#,##0.00"
    # 圖表
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    ch1 = LineChart()
    ch1.title = f"{sn} 5年歷史股息率"
    ch1.y_axis.title = "股息率 (%)"; ch1.width = 17; ch1.height = 10
    ch1.add_data(Reference(ws, min_col=5, max_col=8,
                           min_row=1, max_row=ws.max_row),
                 titles_from_data=True)
    ch1.set_categories(cats); ch1.x_axis.tickLblSkip = 120
    for i, col_color in enumerate(["4F81BD","7F7F7F","27AE60","C0392B"]):
        ch1.series[i].graphicalProperties.line.solidFill = col_color
    ws.add_chart(ch1, "O2")

    ch2 = LineChart()
    ch2.title = f"{sn} 5年歷史 PE"
    ch2.y_axis.title = "PE 倍數"; ch2.width = 17; ch2.height = 10
    ch2.add_data(Reference(ws, min_col=9, max_col=12,
                           min_row=1, max_row=ws.max_row),
                 titles_from_data=True)
    ch2.set_categories(cats); ch2.x_axis.tickLblSkip = 120
    for i, col_color in enumerate(["8E44AD","7F7F7F","27AE60","C0392B"]):
        ch2.series[i].graphicalProperties.line.solidFill = col_color
    ws.add_chart(ch2, "O18")

    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[
            get_column_letter(col[0].column)].width = min(max(ml + 3, 10), 22)


def rebuild_summary(wb, market, smap, results):
    """重建總覽分頁：呢次新處理嘅股票用 results 嗰份新鮮計算結果；
    冇處理過（即係已經追蹤緊嘅舊股）就讀返佢自己 sheet 最後一行嘅快照 ──
    同 daily_importer.py 嘅 rebuild_summary() 一模一樣嘅邏輯，
    保證兩個腳本對「總覽」嘅理解永遠一致。"""
    sum_name = f"{market} 總覽"

    if sum_name in wb.sheetnames: del wb[sum_name]
    ws = wb.create_sheet(sum_name); ws.freeze_panes = "A2"

    THIN = Border(left=Side(style="thin", color="E0E0E0"), right=Side(style="thin", color="E0E0E0"),
                  top=Side(style="thin", color="E0E0E0"), bottom=Side(style="thin", color="E0E0E0"))
    HDR_FILL   = PatternFill("solid", fgColor="3F3F3F")
    SCORE_HDR  = PatternFill("solid", fgColor="1F4E79")
    SCORE_FILL = PatternFill("solid", fgColor="E8F4FD")
    HDR_FONT   = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    NORM_FONT  = Font(name="Arial", size=10)
    LINK_FONT  = Font(name="Arial", size=10, color="0000FF", underline="single")
    ncols = len(SUMMARY_COLS)

    ws.append(SUMMARY_COLS)
    for col in range(1, ncols + 1):
        c = ws.cell(1, col)
        c.fill = SCORE_HDR if col >= 22 else HDR_FILL
        c.font = HDR_FONT; c.alignment = Alignment(horizontal="center", wrap_text=True)

    rows = []
    for sn, (ticker, smap_name) in smap.items():
        r = results.get(sn)
        if r:
            snap = r["snap"]
            y_avg = r["y_avg"]; y_std = r["y_std"]
            pe_avg = r["pe_avg"]; pe_std = r["pe_std"]
            total = safe(snap.get("Score_總分_100"), 0)
            name = r["name"] if r.get("name") else smap_name
            rows.append((sn, ticker, name, r["close"], r["yield"], y_avg,
                         y_avg + 1.5 * y_std, y_avg - 1.5 * y_std,
                         r["pe"], pe_avg,
                         max(0, pe_avg - 1.5 * pe_std), pe_avg + 1.5 * pe_std,
                         snap, total))
        else:
            if sn not in wb.sheetnames: continue
            ws_s = wb[sn]; last = ws_s.max_row
            if last < 2: continue
            snap = {SCORE_SNAPSHOT_COLS[i]: ws_s.cell(last, SCORE_COL_START + i).value
                    for i in range(len(SCORE_SNAPSHOT_COLS))}
            total = safe(snap.get("Score_總分_100"), 0)
            rows.append((sn, ticker, smap_name,
                safe(ws_s.cell(last, 2).value, 0), safe(ws_s.cell(last, 6).value, 0),
                safe(ws_s.cell(last, 7).value, 0), safe(ws_s.cell(last, 8).value, 0), safe(ws_s.cell(last, 9).value, 0),
                safe(ws_s.cell(last, 10).value, 0), safe(ws_s.cell(last, 11).value, 0),
                safe(ws_s.cell(last, 12).value, 0), safe(ws_s.cell(last, 13).value, 0),
                snap, total))

    rows.sort(key=lambda x: x[13], reverse=True)

    for sn, ticker, company, close, cy, ya, yb, ys, pe, pa, pb2, ps, snap, total in rows:
        ws.append([
            ticker, company, round(close, 2),
            round(cy, 2), round(ya, 2), round(yb, 2), round(ys, 2),
            round(pe, 2), round(pa, 2), round(pb2, 2), round(ps, 2),
            snap.get("Payout_Ratio_%"), snap.get("FCF_Coverage"),
            snap.get("Net_Debt_EBITDA"), snap.get("Interest_Coverage"),
            snap.get("Current_Ratio"), snap.get("PB_Ratio"),
            snap.get("Yield_Spread_vs_Bond"), snap.get("DGR_3yr_%"),
            snap.get("RSI_14"), snap.get("52W_Position_%"),
            snap.get("Score_股息質量_30"), snap.get("Score_估值_25"),
            snap.get("Score_財務健康_25"), snap.get("Score_增長潛力_10"),
            snap.get("Score_技術面_10"), total, get_status(total),
        ])
        ridx = ws.max_row
        rf = get_score_fill(total); rfont = get_score_font(total)

        for col in range(1, ncols + 1):
            c = ws.cell(ridx, col); c.border = THIN
            if col >= 22: c.fill = SCORE_FILL; c.font = NORM_FONT
            else:         c.fill = rf;         c.font = NORM_FONT
            if col == ncols - 1:
                c.fill = rf; c.font = rfont; c.number_format = "0.0"
                c.alignment = Alignment(horizontal="center")
            if col == ncols:
                c.fill = rf; c.font = rfont
                c.alignment = Alignment(horizontal="center")
            if col == 1:
                c.hyperlink = f"#'{sn}'!A1"; c.font = LINK_FONT
            if col in [4, 5, 6, 7, 12, 17, 18]: c.number_format = "0.00"
            if col in [3, 8, 9, 10, 11]:        c.number_format = "#,##0.00"
            if col in [22, 23, 24, 25, 26]:     c.number_format = "0.0"

    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ml + 3, 10), 22)

    # 總覽移至最前
    sheets = wb._sheets
    try:
        idx = next(i for i, s in enumerate(sheets) if s.title == sum_name)
        sheets.insert(0, sheets.pop(idx))
    except Exception:
        pass
    print(f"  📊 總覽已重建: {sum_name}（共 {len(rows)} 隻，含 {len(results)} 隻新股）")


for market, excel_file in EXCEL_FILES.items():
    smap         = smap_by_market.get(market, {})
    stock_sheets = stock_data_by_market[market]
    results      = results_by_market[market]

    if not smap:
        print(f"  ⚠️  {market} 沒有任何追蹤股票，跳過"); continue

    file_exists = existing_wb.get(market) is not None

    if not stock_sheets:
        if file_exists:
            print(f"  ℹ️  {market}: 沒有新股票需要加入，{excel_file} 維持不變（歷史完整保留）")
        else:
            print(f"  ⚠️  {market}: 無數據（檔案不存在亦無新股可建立），跳過")
        continue

    if file_exists:
        wb = existing_wb[market]
        try: os.chmod(excel_file, stat.S_IWRITE)
        except Exception: pass
        print(f"\n  📁 {excel_file}：新增 {len(stock_sheets)} 隻新股票"
              f"（現存 {len(existing_sheets.get(market, set()))} 隻 sheet 完全不變）...")
    else:
        wb = openpyxl.Workbook()
        print(f"\n  📁 {excel_file}：全新建立（{len(stock_sheets)} 隻股票）...")

    print(f"  🎨 建立並美化 {len(stock_sheets)} 個新 sheet...")
    for sn, df in stock_sheets.items():
        ws_new = write_df_to_new_sheet(wb, sn, df)
        style_one_stock_sheet(ws_new, sn)

    if not file_exists and "Sheet" in wb.sheetnames:
        del wb["Sheet"]   # openpyxl 新工作簿預設帶嘅空白分頁

    rebuild_summary(wb, market, smap, results)

    wb.save(excel_file)
    try: os.chmod(excel_file, stat.S_IREAD)
    except Exception: pass
    print(f"  ✅ {excel_file} 完成！")

n_files = len([f for f in EXCEL_FILES.values() if os.path.exists(f)])
print(f"\n🎯 全部完成！共生成 {n_files} 個文件：")
for market, f in EXCEL_FILES.items():
    status = "✅" if os.path.exists(f) else "⏭ (無數據)"
    print(f"   {status} {market}: {f}")
