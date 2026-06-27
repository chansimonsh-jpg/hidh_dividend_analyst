#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
daily_importer_global_v6.py
============================
對應 batch_importer_global_v6.py 的每日更新器
  - 讀取 UK/HK/US 三個獨立 Excel 文件
  - 每日更新最新行情 + 重算評分快照
  - v5 新增：四層數據源架構（yfinance → 合理性檢查 → IBKR補救 → 靜態備用）
  - IBKR 擴展支援三市場（HK/UK/US），不再限於 HK 專用
  - 完整繼承 v5 評分邏輯（含市場獨立無風險利率）
  - 靜默處理 IBKR Error 10358
  - 完整 5 級顏色總覽
  - 每次更新後重建總覽分頁
"""

import asyncio, sys, datetime, os, stat, time, shutil, socket, struct
import threading, warnings, logging
try:
    import akshare as ak
    _AKSHARE_OK = True
except ImportError:
    _AKSHARE_OK = False
import numpy as np, pandas as pd, yfinance as yf, openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from yfinance import ticker

warnings.filterwarnings("ignore")

# =========================================================
# LOGGING — 同時輸出至 terminal 及 logs/daily_YYYYMMDD.log
# =========================================================
def _setup_logger():
    os.makedirs("logs", exist_ok=True)
    log_path = os.path.join("logs", f"daily_{datetime.date.today():%Y%m%d}.log")
    logger = logging.getLogger("daily_importer")
    logger.setLevel(logging.DEBUG)
    if logger.handlers:
        logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s  %(message)s", datefmt="%H:%M:%S")
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.DEBUG)
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger, log_path

_log, _log_path = _setup_logger()

def log(msg: str = ""):
    """Drop-in for log() — writes to terminal + log file."""
    _log.info(msg)

if sys.version_info >= (3, 10):
    try:
        asyncio.get_event_loop()
    except RuntimeError:
        asyncio.set_event_loop(asyncio.new_event_loop())

# =========================================================
# CONFIG
# =========================================================
EXCEL_FILES = {
    "UK": "UK_Dividend_Analysis.xlsx",
    "HK": "HK_Dividend_Analysis.xlsx",
    "US": "US_Dividend_Analysis.xlsx",
    "CN": "CN_Dividend_Analysis.xlsx",
}
SUMMARY_SHEET_NAME = "{market} 總覽"   # e.g. "UK 總覽"
BACKUP_FOLDER  = "x:\\backup"
MAX_SAVE_RETRY = 3

RISK_FREE_RATES = {
    "UK": 4.20,
    "US": 4.30,
    "HK": 3.80,
    "CN": 2.30,
}

IBKR_HOST      = "127.0.0.1"
IBKR_PORT      = 4002
IBKR_CLIENT_ID = 10
IBKR_TIMEOUT   = 15
USE_IBKR       = True

STANDARD_NCOLS = 13   # cols A-M
SCORE_COL_START = 14  # col N onwards

SCORE_SNAPSHOT_COLS = [
    "Score_日期",
    "Payout_Ratio_%", "FCF_Coverage", "Net_Debt_EBITDA",
    "Interest_Coverage", "Current_Ratio", "PB_Ratio",
    "Yield_Spread_vs_Bond", "DGR_3yr_%", "RSI_14", "52W_Position_%",
    "Score_股息質量_30", "Score_估值_25", "Score_財務健康_25",
    "Score_增長潛力_10", "Score_技術面_10", "Score_總分_100",
]

SUMMARY_COLS = [
    "股票代號", "公司名稱", "現價",
    "最新股息率", "5年均息率", "🟢 息率買入線", "🔴 息率賣出線",
    "最新 PE",   "5年均 PE",  "🟢 PE買入線",   "🔴 PE賣出線",
    "Payout_%", "FCF覆蓋", "Net_Debt/EBITDA", "利息覆蓋",
    "流動比率", "P/B",     "Yield_Spread",     "DGR_3yr%",
    "RSI_14",   "52W位置%",
    "S_股息質量", "S_估值", "S_財務健康", "S_增長", "S_技術",
    "📊 總分_100", "📊 綜合診斷",
]

# =========================================================
# STATIC FUNDAMENTALS (copied from batch v5)
# =========================================================
HK_STATIC_FUNDAMENTALS = {
    "0939.HK": {"trailingEps":1.32,"payoutRatio":0.30,"priceToBook":0.58,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.05,"sharesOutstanding":250.01e9},
    "1398.HK": {"trailingEps":0.94,"payoutRatio":0.30,"priceToBook":0.55,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.04,"sharesOutstanding":356.41e9},
    "3988.HK": {"trailingEps":0.73,"payoutRatio":0.30,"priceToBook":0.48,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.04,"sharesOutstanding":294.39e9},
    "1288.HK": {"trailingEps":0.67,"payoutRatio":0.31,"priceToBook":0.75,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.06,"sharesOutstanding":349.99e9},
    "0005.HK": {"trailingEps":7.20,"payoutRatio":0.55,"priceToBook":0.85,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.08,"sharesOutstanding":20.08e9},
    "2388.HK": {"trailingEps":2.65,"payoutRatio":0.62,"priceToBook":0.90,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.06,"sharesOutstanding":10.57e9},
    "0023.HK": {"trailingEps":1.28,"payoutRatio":0.72,"priceToBook":0.38,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":-0.05,"sharesOutstanding":3.24e9},
    "1336.HK": {"trailingEps":4.10,"payoutRatio":0.35,"priceToBook":0.72,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.08,"sharesOutstanding":3.0e9},
    "2318.HK": {"trailingEps":5.86,"payoutRatio":0.40,"priceToBook":1.10,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.10,"sharesOutstanding":18.28e9},
    "0966.HK": {"trailingEps":1.55,"payoutRatio":0.32,"priceToBook":0.45,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.05,"sharesOutstanding":4.01e9},
    "0762.HK": {"trailingEps":0.78,"payoutRatio":0.62,"priceToBook":1.05,"currentRatio":0.85,"ebitda":96.3e9,"totalDebt":135.0e9,"totalCash":28.0e9,"freeCashflow":18.0e9,"interestExpense":3.5e9,"earningsGrowth":0.04,"sharesOutstanding":32.09e9},
    "0941.HK": {"trailingEps":5.78,"payoutRatio":0.63,"priceToBook":1.30,"currentRatio":0.95,"ebitda":350.0e9,"totalDebt":200.0e9,"totalCash":65.0e9,"freeCashflow":120.0e9,"interestExpense":6.0e9,"earningsGrowth":0.08,"sharesOutstanding":20.43e9},
    "0728.HK": {"trailingEps":0.55,"payoutRatio":0.60,"priceToBook":1.05,"currentRatio":0.80,"ebitda":110.0e9,"totalDebt":178.0e9,"totalCash":32.0e9,"freeCashflow":25.0e9,"interestExpense":5.5e9,"earningsGrowth":0.05,"sharesOutstanding":103.7e9},
    "6823.HK": {"trailingEps":0.68,"payoutRatio":0.90,"priceToBook":4.80,"currentRatio":0.60,"ebitda":7.2e9,"totalDebt":14.5e9,"totalCash":1.5e9,"freeCashflow":3.8e9,"interestExpense":0.65e9,"earningsGrowth":0.02,"sharesOutstanding":5.02e9},
    "0857.HK": {"trailingEps":0.82,"payoutRatio":0.50,"priceToBook":0.80,"currentRatio":0.95,"ebitda":178.0e9,"totalDebt":290.0e9,"totalCash":45.0e9,"freeCashflow":55.0e9,"interestExpense":8.5e9,"earningsGrowth":0.03,"sharesOutstanding":183.02e9},
    "0386.HK": {"trailingEps":0.70,"payoutRatio":0.45,"priceToBook":0.55,"currentRatio":0.88,"ebitda":145.0e9,"totalDebt":320.0e9,"totalCash":55.0e9,"freeCashflow":48.0e9,"interestExpense":10.5e9,"earningsGrowth":0.02,"sharesOutstanding":119.8e9},
    "1088.HK": {"trailingEps":2.85,"payoutRatio":0.70,"priceToBook":0.95,"currentRatio":1.55,"ebitda":62.0e9,"totalDebt":38.0e9,"totalCash":22.0e9,"freeCashflow":32.0e9,"interestExpense":1.2e9,"earningsGrowth":-0.02,"sharesOutstanding":19.85e9},
    "1171.HK": {"trailingEps":2.20,"payoutRatio":0.60,"priceToBook":0.70,"currentRatio":1.40,"ebitda":22.0e9,"totalDebt":28.0e9,"totalCash":8.0e9,"freeCashflow":12.0e9,"interestExpense":0.9e9,"earningsGrowth":-0.05,"sharesOutstanding":4.46e9},
    "0003.HK": {"trailingEps":0.52,"payoutRatio":0.68,"priceToBook":2.50,"currentRatio":0.75,"ebitda":9.8e9,"totalDebt":33.0e9,"totalCash":4.5e9,"freeCashflow":5.2e9,"interestExpense":0.95e9,"earningsGrowth":0.03,"sharesOutstanding":14.73e9},
    "0006.HK": {"trailingEps":3.80,"payoutRatio":0.55,"priceToBook":0.65,"currentRatio":1.20,"ebitda":12.5e9,"totalDebt":25.0e9,"totalCash":6.0e9,"freeCashflow":8.0e9,"interestExpense":0.85e9,"earningsGrowth":0.02,"sharesOutstanding":2.41e9},
    "0002.HK": {"trailingEps":7.20,"payoutRatio":0.60,"priceToBook":1.10,"currentRatio":0.95,"ebitda":21.0e9,"totalDebt":52.0e9,"totalCash":8.5e9,"freeCashflow":10.5e9,"interestExpense":1.8e9,"earningsGrowth":0.03,"sharesOutstanding":2.63e9},
    "0836.HK": {"trailingEps":1.45,"payoutRatio":0.48,"priceToBook":0.68,"currentRatio":0.75,"ebitda":25.0e9,"totalDebt":65.0e9,"totalCash":8.0e9,"freeCashflow":10.0e9,"interestExpense":2.2e9,"earningsGrowth":0.05,"sharesOutstanding":4.16e9},
    "0016.HK": {"trailingEps":12.50,"payoutRatio":0.35,"priceToBook":0.35,"currentRatio":1.80,"ebitda":25.0e9,"totalDebt":75.0e9,"totalCash":15.0e9,"freeCashflow":8.0e9,"interestExpense":2.5e9,"earningsGrowth":-0.05,"sharesOutstanding":2.86e9},
    "0001.HK": {"trailingEps":5.20,"payoutRatio":0.50,"priceToBook":0.45,"currentRatio":1.50,"ebitda":38.0e9,"totalDebt":120.0e9,"totalCash":28.0e9,"freeCashflow":18.0e9,"interestExpense":3.8e9,"earningsGrowth":-0.03,"sharesOutstanding":3.86e9},
    "0012.HK": {"trailingEps":3.50,"payoutRatio":0.45,"priceToBook":0.28,"currentRatio":2.20,"ebitda":15.0e9,"totalDebt":55.0e9,"totalCash":12.0e9,"freeCashflow":5.0e9,"interestExpense":1.8e9,"earningsGrowth":-0.08,"sharesOutstanding":2.76e9},
    "0083.HK": {"trailingEps":1.10,"payoutRatio":0.45,"priceToBook":0.32,"currentRatio":1.60,"ebitda":5.0e9,"totalDebt":18.0e9,"totalCash":4.5e9,"freeCashflow":2.0e9,"interestExpense":0.65e9,"earningsGrowth":-0.06,"sharesOutstanding":5.32e9},
    "0101.HK": {"trailingEps":0.82,"payoutRatio":0.52,"priceToBook":0.28,"currentRatio":1.40,"ebitda":5.8e9,"totalDebt":28.0e9,"totalCash":5.0e9,"freeCashflow":2.2e9,"interestExpense":0.90e9,"earningsGrowth":-0.10,"sharesOutstanding":5.26e9},
    "1113.HK": {"trailingEps":3.60,"payoutRatio":0.42,"priceToBook":0.55,"currentRatio":1.20,"ebitda":38.0e9,"totalDebt":135.0e9,"totalCash":12.0e9,"freeCashflow":12.0e9,"interestExpense":4.5e9,"earningsGrowth":-0.05,"sharesOutstanding":14.5e9},
    "0823.HK": {"trailingEps":2.85,"payoutRatio":0.92,"priceToBook":0.62,"currentRatio":0.40,"ebitda":12.5e9,"totalDebt":55.0e9,"totalCash":3.5e9,"freeCashflow":10.8e9,"interestExpense":1.8e9,"earningsGrowth":-0.02,"sharesOutstanding":2.17e9},
    "0778.HK": {"trailingEps":0.72,"payoutRatio":0.95,"priceToBook":0.55,"currentRatio":0.35,"ebitda":1.8e9,"totalDebt":8.5e9,"totalCash":0.6e9,"freeCashflow":1.55e9,"interestExpense":0.35e9,"earningsGrowth":-0.03,"sharesOutstanding":2.0e9},
    "2778.HK": {"trailingEps":0.55,"payoutRatio":0.95,"priceToBook":0.42,"currentRatio":0.30,"ebitda":1.5e9,"totalDebt":10.5e9,"totalCash":0.5e9,"freeCashflow":1.2e9,"interestExpense":0.42e9,"earningsGrowth":-0.04,"sharesOutstanding":2.49e9},
    "0405.HK": {"trailingEps":0.42,"payoutRatio":0.95,"priceToBook":0.58,"currentRatio":0.25,"ebitda":1.1e9,"totalDebt":6.8e9,"totalCash":0.4e9,"freeCashflow":0.92e9,"interestExpense":0.30e9,"earningsGrowth":0.02,"sharesOutstanding":2.57e9},
    "0548.HK": {"trailingEps":0.85,"payoutRatio":0.58,"priceToBook":0.78,"currentRatio":0.65,"ebitda":3.2e9,"totalDebt":12.0e9,"totalCash":2.0e9,"freeCashflow":2.0e9,"interestExpense":0.45e9,"earningsGrowth":0.04,"sharesOutstanding":2.26e9},
    "0177.HK": {"trailingEps":1.05,"payoutRatio":0.60,"priceToBook":0.95,"currentRatio":0.70,"ebitda":4.8e9,"totalDebt":16.0e9,"totalCash":3.0e9,"freeCashflow":2.8e9,"interestExpense":0.55e9,"earningsGrowth":0.05,"sharesOutstanding":4.04e9},
    "0066.HK": {"trailingEps":4.20,"payoutRatio":0.52,"priceToBook":1.50,"currentRatio":0.55,"ebitda":18.0e9,"totalDebt":65.0e9,"totalCash":8.0e9,"freeCashflow":8.5e9,"interestExpense":2.2e9,"earningsGrowth":0.04,"sharesOutstanding":5.73e9},
    "1038.HK": {"trailingEps":4.50,"payoutRatio":0.60,"priceToBook":0.90,"currentRatio":1.10,"ebitda":18.5e9,"totalDebt":48.0e9,"totalCash":8.0e9,"freeCashflow":10.0e9,"interestExpense":1.5e9,"earningsGrowth":0.03,"sharesOutstanding":2.32e9},
    "0151.HK": {"trailingEps":0.25,"payoutRatio":0.75,"priceToBook":2.80,"currentRatio":3.20,"ebitda":5.5e9,"totalDebt":4.0e9,"totalCash":8.0e9,"freeCashflow":3.8e9,"interestExpense":0.12e9,"earningsGrowth":-0.05,"sharesOutstanding":13.27e9},
    "2319.HK": {"trailingEps":1.05,"payoutRatio":0.38,"priceToBook":2.40,"currentRatio":1.60,"ebitda":7.8e9,"totalDebt":22.0e9,"totalCash":6.0e9,"freeCashflow":4.0e9,"interestExpense":0.65e9,"earningsGrowth":0.05,"sharesOutstanding":3.97e9},
    "0291.HK": {"trailingEps":1.25,"payoutRatio":0.35,"priceToBook":2.50,"currentRatio":1.10,"ebitda":8.5e9,"totalDebt":18.0e9,"totalCash":5.0e9,"freeCashflow":4.5e9,"interestExpense":0.55e9,"earningsGrowth":0.06,"sharesOutstanding":3.09e9},
    "1929.HK": {"trailingEps":1.85,"payoutRatio":0.55,"priceToBook":1.80,"currentRatio":1.50,"ebitda":5.2e9,"totalDebt":12.0e9,"totalCash":3.5e9,"freeCashflow":2.5e9,"interestExpense":0.42e9,"earningsGrowth":0.05,"sharesOutstanding":2.53e9},
    "0270.HK": {"trailingEps":1.20,"payoutRatio":0.58,"priceToBook":0.80,"currentRatio":1.00,"ebitda":6.5e9,"totalDebt":20.0e9,"totalCash":4.0e9,"freeCashflow":3.5e9,"interestExpense":0.72e9,"earningsGrowth":0.03,"sharesOutstanding":4.63e9},
    "0659.HK": {"trailingEps":0.52,"payoutRatio":0.75,"priceToBook":0.58,"currentRatio":0.85,"ebitda":4.8e9,"totalDebt":22.0e9,"totalCash":3.5e9,"freeCashflow":2.2e9,"interestExpense":0.82e9,"earningsGrowth":0.02,"sharesOutstanding":4.09e9},
    "1997.HK": {"trailingEps":2.80,"payoutRatio":0.45,"priceToBook":0.28,"currentRatio":0.90,"ebitda":8.0e9,"totalDebt":38.0e9,"totalCash":5.0e9,"freeCashflow":3.0e9,"interestExpense":1.35e9,"earningsGrowth":-0.08,"sharesOutstanding":3.04e9},
    "0694.HK": {"trailingEps":2.80,"payoutRatio":0.45,"priceToBook":0.48,"currentRatio":1.20,"ebitda":12.0e9,"totalDebt":35.0e9,"totalCash":3.5e9,"freeCashflow":5.0e9,"interestExpense":1.20e9,"earningsGrowth":0.03,"sharesOutstanding":1.17e9},
    "0960.HK": {"trailingEps":1.85,"payoutRatio":0.30,"priceToBook":0.38,"currentRatio":0.65,"ebitda":30.0e9,"totalDebt":185.0e9,"totalCash":18.0e9,"freeCashflow":5.0e9,"interestExpense":7.5e9,"earningsGrowth":-0.15,"sharesOutstanding":7.26e9},
    "0358.HK": {"trailingEps":1.55,"payoutRatio":0.42,"priceToBook":0.62,"currentRatio":1.05,"ebitda":8.5e9,"totalDebt":30.0e9,"totalCash":5.0e9,"freeCashflow":3.5e9,"interestExpense":1.05e9,"earningsGrowth":0.03,"sharesOutstanding":4.97e9},
    "2600.HK": {"trailingEps":0.42,"payoutRatio":0.40,"priceToBook":0.55,"currentRatio":0.95,"ebitda":14.0e9,"totalDebt":55.0e9,"totalCash":10.0e9,"freeCashflow":4.5e9,"interestExpense":1.80e9,"earningsGrowth":0.05,"sharesOutstanding":19.34e9},
    "0087.HK": {"trailingEps":2.80,"payoutRatio":0.50,"priceToBook":0.45,"currentRatio":1.20,"ebitda":15.0e9,"totalDebt":42.0e9,"totalCash":6.0e9,"freeCashflow":6.0e9,"interestExpense":1.35e9,"earningsGrowth":-0.03,"sharesOutstanding":0.60e9},
    "0019.HK": {"trailingEps":5.20,"payoutRatio":0.50,"priceToBook":0.42,"currentRatio":1.25,"ebitda":18.0e9,"totalDebt":52.0e9,"totalCash":7.5e9,"freeCashflow":7.0e9,"interestExpense":1.68e9,"earningsGrowth":-0.03,"sharesOutstanding":0.59e9},
}

UK_STATIC_FUNDAMENTALS = {
    "LGEN.L": {"trailingEps":0.32,"payoutRatio":0.62,"priceToBook":1.10,"currentRatio":None,"ebitda":2.8e9,"totalDebt":8.5e9,"totalCash":2.0e9,"freeCashflow":1.2e9,"interestExpense":0.32e9,"earningsGrowth":0.03,"sharesOutstanding":6.40e9},
    "ABDN.L": {"trailingEps":0.15,"payoutRatio":0.85,"priceToBook":0.65,"currentRatio":None,"ebitda":0.5e9,"totalDebt":1.8e9,"totalCash":1.5e9,"freeCashflow":0.3e9,"interestExpense":0.08e9,"earningsGrowth":-0.05,"sharesOutstanding":1.80e9},
    "MNG.L":  {"trailingEps":0.22,"payoutRatio":0.90,"priceToBook":1.80,"currentRatio":None,"ebitda":0.85e9,"totalDebt":1.5e9,"totalCash":1.2e9,"freeCashflow":0.55e9,"interestExpense":0.06e9,"earningsGrowth":0.02,"sharesOutstanding":2.50e9},
    "AV.L":   {"trailingEps":0.55,"payoutRatio":0.55,"priceToBook":1.40,"currentRatio":None,"ebitda":2.2e9,"totalDebt":6.5e9,"totalCash":3.5e9,"freeCashflow":1.8e9,"interestExpense":0.28e9,"earningsGrowth":0.06,"sharesOutstanding":3.90e9},
    "SDR.L":  {"trailingEps":0.42,"payoutRatio":0.60,"priceToBook":1.20,"currentRatio":None,"ebitda":0.65e9,"totalDebt":1.2e9,"totalCash":0.9e9,"freeCashflow":0.45e9,"interestExpense":0.05e9,"earningsGrowth":0.02,"sharesOutstanding":0.48e9},
    "HSBA.L": {"trailingEps":1.05,"payoutRatio":0.50,"priceToBook":0.95,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.08,"sharesOutstanding":19.50e9},
    "BARC.L": {"trailingEps":0.38,"payoutRatio":0.30,"priceToBook":0.55,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.10,"sharesOutstanding":16.20e9},
    "NWG.L":  {"trailingEps":0.52,"payoutRatio":0.35,"priceToBook":0.80,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.08,"sharesOutstanding":8.40e9},
    "LLOY.L": {"trailingEps":0.072,"payoutRatio":0.38,"priceToBook":0.70,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.05,"sharesOutstanding":57.00e9},
    "OSB.L":  {"trailingEps":0.95,"payoutRatio":0.40,"priceToBook":0.65,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.05,"sharesOutstanding":0.32e9},
    "IMB.L":  {"trailingEps":2.85,"payoutRatio":0.68,"priceToBook":None,"currentRatio":0.85,"ebitda":3.8e9,"totalDebt":11.5e9,"totalCash":1.5e9,"freeCashflow":2.8e9,"interestExpense":0.55e9,"earningsGrowth":0.03,"sharesOutstanding":0.89e9},
    "BATS.L": {"trailingEps":3.20,"payoutRatio":0.72,"priceToBook":None,"currentRatio":0.78,"ebitda":12.5e9,"totalDebt":42.0e9,"totalCash":3.8e9,"freeCashflow":9.5e9,"interestExpense":1.85e9,"earningsGrowth":0.02,"sharesOutstanding":2.28e9},
    "UKW.L":  {"trailingEps":0.062,"payoutRatio":1.10,"priceToBook":1.05,"currentRatio":0.50,"ebitda":0.32e9,"totalDebt":1.2e9,"totalCash":0.08e9,"freeCashflow":0.22e9,"interestExpense":0.055e9,"earningsGrowth":0.03,"sharesOutstanding":3.60e9},
    "ORIT.L": {"trailingEps":0.045,"payoutRatio":1.20,"priceToBook":0.82,"currentRatio":0.40,"ebitda":0.12e9,"totalDebt":0.55e9,"totalCash":0.03e9,"freeCashflow":0.09e9,"interestExpense":0.022e9,"earningsGrowth":0.02,"sharesOutstanding":1.55e9},
    "TRIG.L": {"trailingEps":0.058,"payoutRatio":1.10,"priceToBook":0.88,"currentRatio":0.45,"ebitda":0.28e9,"totalDebt":1.0e9,"totalCash":0.06e9,"freeCashflow":0.20e9,"interestExpense":0.045e9,"earningsGrowth":0.02,"sharesOutstanding":2.40e9},
    "BSIF.L": {"trailingEps":0.048,"payoutRatio":1.05,"priceToBook":0.90,"currentRatio":0.40,"ebitda":0.10e9,"totalDebt":0.42e9,"totalCash":0.02e9,"freeCashflow":0.08e9,"interestExpense":0.018e9,"earningsGrowth":0.02,"sharesOutstanding":0.98e9},
    "LAND.L": {"trailingEps":0.38,"payoutRatio":0.75,"priceToBook":0.58,"currentRatio":0.65,"ebitda":0.62e9,"totalDebt":3.8e9,"totalCash":0.25e9,"freeCashflow":0.45e9,"interestExpense":0.145e9,"earningsGrowth":-0.02,"sharesOutstanding":0.96e9},
    "BLND.L": {"trailingEps":0.25,"payoutRatio":0.72,"priceToBook":0.52,"currentRatio":0.55,"ebitda":0.52e9,"totalDebt":3.2e9,"totalCash":0.20e9,"freeCashflow":0.38e9,"interestExpense":0.125e9,"earningsGrowth":-0.03,"sharesOutstanding":1.02e9},
    "BBOX.L": {"trailingEps":0.12,"payoutRatio":0.85,"priceToBook":0.90,"currentRatio":0.50,"ebitda":0.38e9,"totalDebt":2.5e9,"totalCash":0.12e9,"freeCashflow":0.28e9,"interestExpense":0.095e9,"earningsGrowth":0.04,"sharesOutstanding":2.35e9},
    "HMSO.L": {"trailingEps":0.08,"payoutRatio":0.65,"priceToBook":0.42,"currentRatio":0.45,"ebitda":0.25e9,"totalDebt":1.8e9,"totalCash":0.10e9,"freeCashflow":0.15e9,"interestExpense":0.075e9,"earningsGrowth":-0.05,"sharesOutstanding":2.85e9},
    "PHP.L":  {"trailingEps":0.062,"payoutRatio":0.88,"priceToBook":1.10,"currentRatio":0.35,"ebitda":0.18e9,"totalDebt":1.4e9,"totalCash":0.05e9,"freeCashflow":0.13e9,"interestExpense":0.055e9,"earningsGrowth":0.03,"sharesOutstanding":1.98e9},
    "BYG.L":  {"trailingEps":0.95,"payoutRatio":0.42,"priceToBook":1.80,"currentRatio":1.20,"ebitda":0.22e9,"totalDebt":0.58e9,"totalCash":0.08e9,"freeCashflow":0.16e9,"interestExpense":0.022e9,"earningsGrowth":0.05,"sharesOutstanding":0.143e9},
    "UU.L":   {"trailingEps":0.52,"payoutRatio":0.72,"priceToBook":2.20,"currentRatio":0.65,"ebitda":0.88e9,"totalDebt":8.5e9,"totalCash":0.35e9,"freeCashflow":0.42e9,"interestExpense":0.285e9,"earningsGrowth":0.03,"sharesOutstanding":0.682e9},
    "SVT.L":  {"trailingEps":1.45,"payoutRatio":0.68,"priceToBook":3.50,"currentRatio":0.60,"ebitda":0.75e9,"totalDebt":7.2e9,"totalCash":0.28e9,"freeCashflow":0.35e9,"interestExpense":0.245e9,"earningsGrowth":0.03,"sharesOutstanding":0.232e9},
    "NG.L":   {"trailingEps":0.68,"payoutRatio":0.70,"priceToBook":1.85,"currentRatio":0.75,"ebitda":4.5e9,"totalDebt":42.0e9,"totalCash":1.8e9,"freeCashflow":1.5e9,"interestExpense":1.45e9,"earningsGrowth":0.04,"sharesOutstanding":3.85e9},
    "SSE.L":  {"trailingEps":1.20,"payoutRatio":0.65,"priceToBook":1.60,"currentRatio":0.70,"ebitda":2.8e9,"totalDebt":18.0e9,"totalCash":0.85e9,"freeCashflow":0.85e9,"interestExpense":0.62e9,"earningsGrowth":0.05,"sharesOutstanding":1.10e9},
    "BP.L":   {"trailingEps":0.52,"payoutRatio":0.42,"priceToBook":1.05,"currentRatio":1.10,"ebitda":28.5e9,"totalDebt":52.0e9,"totalCash":18.0e9,"freeCashflow":12.0e9,"interestExpense":2.2e9,"earningsGrowth":-0.05,"sharesOutstanding":18.50e9},
    "SHEL.L": {"trailingEps":2.85,"payoutRatio":0.38,"priceToBook":1.20,"currentRatio":1.15,"ebitda":52.0e9,"totalDebt":65.0e9,"totalCash":32.0e9,"freeCashflow":28.0e9,"interestExpense":2.8e9,"earningsGrowth":-0.03,"sharesOutstanding":6.50e9},
    "AAL.L":  {"trailingEps":0.85,"payoutRatio":0.40,"priceToBook":1.80,"currentRatio":1.50,"ebitda":5.2e9,"totalDebt":12.0e9,"totalCash":4.5e9,"freeCashflow":2.8e9,"interestExpense":0.48e9,"earningsGrowth":-0.10,"sharesOutstanding":1.35e9},
    "RIO.L":  {"trailingEps":6.20,"payoutRatio":0.60,"priceToBook":2.20,"currentRatio":1.80,"ebitda":18.5e9,"totalDebt":12.0e9,"totalCash":8.5e9,"freeCashflow":10.0e9,"interestExpense":0.42e9,"earningsGrowth":-0.05,"sharesOutstanding":1.62e9},
    "GSK.L":  {"trailingEps":1.28,"payoutRatio":0.48,"priceToBook":3.80,"currentRatio":1.25,"ebitda":8.5e9,"totalDebt":18.0e9,"totalCash":4.2e9,"freeCashflow":5.5e9,"interestExpense":0.62e9,"earningsGrowth":0.08,"sharesOutstanding":4.02e9},
    "AZN.L":  {"trailingEps":3.20,"payoutRatio":0.65,"priceToBook":5.50,"currentRatio":1.05,"ebitda":12.5e9,"totalDebt":22.0e9,"totalCash":5.8e9,"freeCashflow":7.5e9,"interestExpense":0.75e9,"earningsGrowth":0.15,"sharesOutstanding":1.58e9},
    "ULVR.L": {"trailingEps":2.85,"payoutRatio":0.62,"priceToBook":6.20,"currentRatio":0.85,"ebitda":10.5e9,"totalDebt":25.0e9,"totalCash":4.5e9,"freeCashflow":7.2e9,"interestExpense":0.88e9,"earningsGrowth":0.05,"sharesOutstanding":2.60e9},
    "VOD.L":  {"trailingEps":0.08,"payoutRatio":0.75,"priceToBook":0.55,"currentRatio":0.72,"ebitda":14.5e9,"totalDebt":52.0e9,"totalCash":5.5e9,"freeCashflow":4.5e9,"interestExpense":2.2e9,"earningsGrowth":-0.05,"sharesOutstanding":26.80e9},
    "BT-A.L": {"trailingEps":0.18,"payoutRatio":0.45,"priceToBook":1.10,"currentRatio":0.68,"ebitda":7.8e9,"totalDebt":20.0e9,"totalCash":1.8e9,"freeCashflow":1.8e9,"interestExpense":0.88e9,"earningsGrowth":0.03,"sharesOutstanding":9.90e9},
    "BA.L":   {"trailingEps":0.62,"payoutRatio":0.38,"priceToBook":4.50,"currentRatio":1.05,"ebitda":2.8e9,"totalDebt":4.5e9,"totalCash":1.8e9,"freeCashflow":2.0e9,"interestExpense":0.18e9,"earningsGrowth":0.12,"sharesOutstanding":3.15e9},
    "INF.L":  {"trailingEps":0.52,"payoutRatio":0.55,"priceToBook":2.80,"currentRatio":1.10,"ebitda":1.2e9,"totalDebt":4.2e9,"totalCash":0.85e9,"freeCashflow":0.75e9,"interestExpense":0.16e9,"earningsGrowth":0.08,"sharesOutstanding":1.28e9},
    "PSON.L": {"trailingEps":0.68,"payoutRatio":0.52,"priceToBook":2.20,"currentRatio":1.25,"ebitda":0.65e9,"totalDebt":1.5e9,"totalCash":0.55e9,"freeCashflow":0.48e9,"interestExpense":0.06e9,"earningsGrowth":0.05,"sharesOutstanding":0.78e9},
    "WPP.L":  {"trailingEps":0.75,"payoutRatio":0.55,"priceToBook":1.80,"currentRatio":0.90,"ebitda":1.8e9,"totalDebt":4.8e9,"totalCash":1.5e9,"freeCashflow":1.2e9,"interestExpense":0.22e9,"earningsGrowth":0.02,"sharesOutstanding":1.02e9},
    "OCDO.L": {"trailingEps":-0.18,"payoutRatio":None,"priceToBook":2.50,"currentRatio":1.80,"ebitda":0.12e9,"totalDebt":1.8e9,"totalCash":0.65e9,"freeCashflow":-0.15e9,"interestExpense":0.085e9,"earningsGrowth":0.20,"sharesOutstanding":0.87e9},
    "SMIN.L": {"trailingEps":1.25,"payoutRatio":0.45,"priceToBook":2.80,"currentRatio":1.35,"ebitda":0.85e9,"totalDebt":1.2e9,"totalCash":0.55e9,"freeCashflow":0.62e9,"interestExpense":0.048e9,"earningsGrowth":0.06,"sharesOutstanding":0.52e9},
    "BWY.L":  {"trailingEps":2.85,"payoutRatio":0.40,"priceToBook":1.20,"currentRatio":3.50,"ebitda":0.52e9,"totalDebt":0.25e9,"totalCash":0.32e9,"freeCashflow":0.38e9,"interestExpense":0.010e9,"earningsGrowth":-0.05,"sharesOutstanding":0.131e9},
    "TW.L":   {"trailingEps":0.145,"payoutRatio":0.55,"priceToBook":1.35,"currentRatio":4.20,"ebitda":0.52e9,"totalDebt":0.18e9,"totalCash":0.28e9,"freeCashflow":0.40e9,"interestExpense":0.008e9,"earningsGrowth":-0.03,"sharesOutstanding":3.18e9},
    "PSN.L":  {"trailingEps":1.20,"payoutRatio":0.60,"priceToBook":1.50,"currentRatio":3.80,"ebitda":0.48e9,"totalDebt":0.12e9,"totalCash":0.22e9,"freeCashflow":0.35e9,"interestExpense":0.006e9,"earningsGrowth":-0.08,"sharesOutstanding":0.316e9},
    "HFEL.L": {"trailingEps":0.22,"payoutRatio":0.92,"priceToBook":0.88,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.02,"sharesOutstanding":0.168e9},
    "MYI.L":  {"trailingEps":0.38,"payoutRatio":0.85,"priceToBook":0.92,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.03,"sharesOutstanding":0.195e9},
    "MRCH.L": {"trailingEps":0.32,"payoutRatio":0.82,"priceToBook":0.95,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.03,"sharesOutstanding":0.170e9},
    "CTY.L":  {"trailingEps":0.22,"payoutRatio":0.88,"priceToBook":1.05,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.03,"sharesOutstanding":0.388e9},
    "AAIF.L": {"trailingEps":0.18,"payoutRatio":0.90,"priceToBook":0.82,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.02,"sharesOutstanding":0.322e9},
}

US_STATIC_FUNDAMENTALS = {
    "O":    {"trailingEps":1.40,"payoutRatio":0.72,"priceToBook":1.20,"currentRatio":None,"ebitda":3.5e9,"totalDebt":20.0e9,"totalCash":0.8e9,"freeCashflow":2.8e9,"interestExpense":0.90e9,"earningsGrowth":0.05,"sharesOutstanding":0.87e9},
    "AMT":  {"trailingEps":4.20,"payoutRatio":0.82,"priceToBook":None,"currentRatio":None,"ebitda":5.8e9,"totalDebt":38.0e9,"totalCash":1.5e9,"freeCashflow":4.2e9,"interestExpense":1.55e9,"earningsGrowth":0.08,"sharesOutstanding":0.209e9},
    "PLD":  {"trailingEps":3.50,"payoutRatio":0.68,"priceToBook":2.00,"currentRatio":None,"ebitda":5.2e9,"totalDebt":28.0e9,"totalCash":0.6e9,"freeCashflow":3.8e9,"interestExpense":1.00e9,"earningsGrowth":0.10,"sharesOutstanding":0.755e9},
    "SPG":  {"trailingEps":7.80,"payoutRatio":0.88,"priceToBook":None,"currentRatio":None,"ebitda":5.0e9,"totalDebt":32.0e9,"totalCash":0.8e9,"freeCashflow":3.5e9,"interestExpense":1.25e9,"earningsGrowth":0.06,"sharesOutstanding":0.328e9},
    "VTR":  {"trailingEps":0.80,"payoutRatio":0.90,"priceToBook":1.80,"currentRatio":None,"ebitda":1.8e9,"totalDebt":12.0e9,"totalCash":0.3e9,"freeCashflow":1.2e9,"interestExpense":0.52e9,"earningsGrowth":0.05,"sharesOutstanding":0.420e9},
    "WELL": {"trailingEps":1.50,"payoutRatio":0.75,"priceToBook":3.20,"currentRatio":None,"ebitda":2.8e9,"totalDebt":16.0e9,"totalCash":0.5e9,"freeCashflow":2.2e9,"interestExpense":0.62e9,"earningsGrowth":0.12,"sharesOutstanding":0.428e9},
    "NNN":  {"trailingEps":2.20,"payoutRatio":0.68,"priceToBook":1.50,"currentRatio":None,"ebitda":0.92e9,"totalDebt":4.5e9,"totalCash":0.12e9,"freeCashflow":0.78e9,"interestExpense":0.18e9,"earningsGrowth":0.04,"sharesOutstanding":0.185e9},
    "STAG": {"trailingEps":1.20,"payoutRatio":0.78,"priceToBook":1.40,"currentRatio":None,"ebitda":0.55e9,"totalDebt":2.8e9,"totalCash":0.08e9,"freeCashflow":0.42e9,"interestExpense":0.11e9,"earningsGrowth":0.05,"sharesOutstanding":0.178e9},
    "VICI": {"trailingEps":1.85,"payoutRatio":0.72,"priceToBook":1.60,"currentRatio":None,"ebitda":2.8e9,"totalDebt":16.5e9,"totalCash":0.4e9,"freeCashflow":2.2e9,"interestExpense":0.68e9,"earningsGrowth":0.06,"sharesOutstanding":1.06e9},
    "NEE":  {"trailingEps":3.20,"payoutRatio":0.62,"priceToBook":2.50,"currentRatio":0.65,"ebitda":10.5e9,"totalDebt":65.0e9,"totalCash":1.8e9,"freeCashflow":2.5e9,"interestExpense":2.45e9,"earningsGrowth":0.08,"sharesOutstanding":2.05e9},
    "DUK":  {"trailingEps":5.50,"payoutRatio":0.72,"priceToBook":1.55,"currentRatio":0.55,"ebitda":8.5e9,"totalDebt":62.0e9,"totalCash":0.8e9,"freeCashflow":0.5e9,"interestExpense":2.25e9,"earningsGrowth":0.05,"sharesOutstanding":0.775e9},
    "SO":   {"trailingEps":3.80,"payoutRatio":0.70,"priceToBook":2.20,"currentRatio":0.60,"ebitda":7.2e9,"totalDebt":52.0e9,"totalCash":0.6e9,"freeCashflow":0.2e9,"interestExpense":1.95e9,"earningsGrowth":0.06,"sharesOutstanding":1.09e9},
    "D":    {"trailingEps":2.85,"payoutRatio":0.82,"priceToBook":1.80,"currentRatio":0.60,"ebitda":5.5e9,"totalDebt":38.0e9,"totalCash":0.5e9,"freeCashflow":-0.5e9,"interestExpense":1.55e9,"earningsGrowth":0.04,"sharesOutstanding":0.840e9},
    "AEP":  {"trailingEps":5.20,"payoutRatio":0.65,"priceToBook":1.80,"currentRatio":0.55,"ebitda":5.8e9,"totalDebt":42.0e9,"totalCash":0.4e9,"freeCashflow":0.1e9,"interestExpense":1.65e9,"earningsGrowth":0.06,"sharesOutstanding":0.523e9},
    "XEL":  {"trailingEps":3.20,"payoutRatio":0.65,"priceToBook":2.10,"currentRatio":0.55,"ebitda":3.2e9,"totalDebt":22.0e9,"totalCash":0.25e9,"freeCashflow":-0.2e9,"interestExpense":0.82e9,"earningsGrowth":0.06,"sharesOutstanding":0.549e9},
    "WEC":  {"trailingEps":4.20,"payoutRatio":0.65,"priceToBook":2.50,"currentRatio":0.55,"ebitda":2.8e9,"totalDebt":18.0e9,"totalCash":0.2e9,"freeCashflow":0.3e9,"interestExpense":0.68e9,"earningsGrowth":0.06,"sharesOutstanding":0.317e9},
    "ES":   {"trailingEps":3.50,"payoutRatio":0.72,"priceToBook":1.40,"currentRatio":0.55,"ebitda":2.5e9,"totalDebt":22.0e9,"totalCash":0.3e9,"freeCashflow":-0.5e9,"interestExpense":0.82e9,"earningsGrowth":0.04,"sharesOutstanding":0.348e9},
    "JPM":  {"trailingEps":18.50,"payoutRatio":0.25,"priceToBook":1.90,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.12,"sharesOutstanding":2.87e9},
    "BAC":  {"trailingEps":3.20,"payoutRatio":0.30,"priceToBook":1.05,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.08,"sharesOutstanding":7.92e9},
    "WFC":  {"trailingEps":5.40,"payoutRatio":0.30,"priceToBook":1.15,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.10,"sharesOutstanding":3.42e9},
    "C":    {"trailingEps":6.50,"payoutRatio":0.28,"priceToBook":0.62,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.08,"sharesOutstanding":1.94e9},
    "USB":  {"trailingEps":3.80,"payoutRatio":0.42,"priceToBook":1.20,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.06,"sharesOutstanding":1.52e9},
    "PRU":  {"trailingEps":12.50,"payoutRatio":0.38,"priceToBook":0.90,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.08,"sharesOutstanding":0.370e9},
    "MET":  {"trailingEps":9.50,"payoutRatio":0.30,"priceToBook":0.85,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":0.06,"sharesOutstanding":0.726e9},
    "XOM":  {"trailingEps":8.90,"payoutRatio":0.40,"priceToBook":1.90,"currentRatio":1.35,"ebitda":55.0e9,"totalDebt":38.0e9,"totalCash":22.0e9,"freeCashflow":32.0e9,"interestExpense":1.2e9,"earningsGrowth":-0.05,"sharesOutstanding":4.35e9},
    "CVX":  {"trailingEps":10.20,"payoutRatio":0.42,"priceToBook":1.65,"currentRatio":1.25,"ebitda":38.0e9,"totalDebt":28.0e9,"totalCash":8.0e9,"freeCashflow":18.0e9,"interestExpense":0.72e9,"earningsGrowth":-0.08,"sharesOutstanding":1.86e9},
    "COP":  {"trailingEps":8.20,"payoutRatio":0.35,"priceToBook":2.50,"currentRatio":1.40,"ebitda":18.0e9,"totalDebt":18.0e9,"totalCash":5.5e9,"freeCashflow":9.5e9,"interestExpense":0.55e9,"earningsGrowth":-0.10,"sharesOutstanding":1.24e9},
    "EOG":  {"trailingEps":12.50,"payoutRatio":0.28,"priceToBook":2.80,"currentRatio":1.50,"ebitda":9.5e9,"totalDebt":5.0e9,"totalCash":2.5e9,"freeCashflow":5.5e9,"interestExpense":0.22e9,"earningsGrowth":-0.05,"sharesOutstanding":0.598e9},
    "PSX":  {"trailingEps":9.50,"payoutRatio":0.42,"priceToBook":2.20,"currentRatio":1.20,"ebitda":5.2e9,"totalDebt":12.0e9,"totalCash":2.5e9,"freeCashflow":3.5e9,"interestExpense":0.45e9,"earningsGrowth":-0.15,"sharesOutstanding":0.415e9},
    "JNJ":  {"trailingEps":8.50,"payoutRatio":0.45,"priceToBook":5.20,"currentRatio":1.45,"ebitda":28.0e9,"totalDebt":28.0e9,"totalCash":18.0e9,"freeCashflow":18.0e9,"interestExpense":0.88e9,"earningsGrowth":0.06,"sharesOutstanding":2.40e9},
    "ABBV": {"trailingEps":6.20,"payoutRatio":0.52,"priceToBook":None,"currentRatio":0.95,"ebitda":25.0e9,"totalDebt":65.0e9,"totalCash":5.5e9,"freeCashflow":18.0e9,"interestExpense":2.20e9,"earningsGrowth":0.05,"sharesOutstanding":1.77e9},
    "MRK":  {"trailingEps":7.80,"payoutRatio":0.42,"priceToBook":5.50,"currentRatio":1.35,"ebitda":22.0e9,"totalDebt":35.0e9,"totalCash":8.5e9,"freeCashflow":14.0e9,"interestExpense":1.15e9,"earningsGrowth":0.12,"sharesOutstanding":2.54e9},
    "PFE":  {"trailingEps":1.45,"payoutRatio":0.80,"priceToBook":1.65,"currentRatio":1.05,"ebitda":15.0e9,"totalDebt":62.0e9,"totalCash":5.0e9,"freeCashflow":8.5e9,"interestExpense":2.45e9,"earningsGrowth":-0.30,"sharesOutstanding":5.68e9},
    "KO":   {"trailingEps":2.85,"payoutRatio":0.65,"priceToBook":10.50,"currentRatio":1.10,"ebitda":13.0e9,"totalDebt":35.0e9,"totalCash":9.5e9,"freeCashflow":9.5e9,"interestExpense":1.05e9,"earningsGrowth":0.06,"sharesOutstanding":4.31e9},
    "PEP":  {"trailingEps":8.20,"payoutRatio":0.68,"priceToBook":12.50,"currentRatio":0.85,"ebitda":16.0e9,"totalDebt":42.0e9,"totalCash":8.5e9,"freeCashflow":8.5e9,"interestExpense":1.35e9,"earningsGrowth":0.04,"sharesOutstanding":1.38e9},
    "PG":   {"trailingEps":6.20,"payoutRatio":0.58,"priceToBook":8.50,"currentRatio":0.75,"ebitda":22.0e9,"totalDebt":28.0e9,"totalCash":7.5e9,"freeCashflow":14.0e9,"interestExpense":0.92e9,"earningsGrowth":0.05,"sharesOutstanding":2.36e9},
    "MO":   {"trailingEps":4.60,"payoutRatio":0.78,"priceToBook":None,"currentRatio":0.62,"ebitda":12.0e9,"totalDebt":28.0e9,"totalCash":2.0e9,"freeCashflow":9.5e9,"interestExpense":1.08e9,"earningsGrowth":0.03,"sharesOutstanding":1.73e9},
    "PM":   {"trailingEps":6.50,"payoutRatio":0.88,"priceToBook":None,"currentRatio":0.92,"ebitda":18.0e9,"totalDebt":48.0e9,"totalCash":3.5e9,"freeCashflow":10.0e9,"interestExpense":1.55e9,"earningsGrowth":0.08,"sharesOutstanding":1.55e9},
    "T":    {"trailingEps":2.25,"payoutRatio":0.50,"priceToBook":1.05,"currentRatio":0.62,"ebitda":45.0e9,"totalDebt":142.0e9,"totalCash":2.5e9,"freeCashflow":16.0e9,"interestExpense":6.20e9,"earningsGrowth":0.03,"sharesOutstanding":7.15e9},
    "VZ":   {"trailingEps":4.20,"payoutRatio":0.62,"priceToBook":1.60,"currentRatio":0.68,"ebitda":48.0e9,"totalDebt":148.0e9,"totalCash":2.0e9,"freeCashflow":18.0e9,"interestExpense":5.80e9,"earningsGrowth":0.02,"sharesOutstanding":4.21e9},
    "SCHD": {"trailingEps":None,"payoutRatio":None,"priceToBook":None,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":None,"sharesOutstanding":None},
    "VYM":  {"trailingEps":None,"payoutRatio":None,"priceToBook":None,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":None,"sharesOutstanding":None},
    "HDV":  {"trailingEps":None,"payoutRatio":None,"priceToBook":None,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":None,"sharesOutstanding":None},
    "JEPI": {"trailingEps":None,"payoutRatio":None,"priceToBook":None,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":None,"sharesOutstanding":None},
    "JEPQ": {"trailingEps":None,"payoutRatio":None,"priceToBook":None,"currentRatio":None,"ebitda":None,"totalDebt":None,"totalCash":None,"freeCashflow":None,"interestExpense":None,"earningsGrowth":None,"sharesOutstanding":None},
}

STATIC_MAP = {"UK": UK_STATIC_FUNDAMENTALS, "HK": HK_STATIC_FUNDAMENTALS, "US": US_STATIC_FUNDAMENTALS}

# =========================================================
# HELPERS
# =========================================================
def safe(v, default=0.0):
    if v is None: return default
    try:
        f = float(v)
        return default if f != f else f
    except: return default

def ticker_to_sheet(tid):
    return tid.replace(".L","").replace(".HK","").replace("-","_").replace(".","_")

def get_status(total):
    if   total >= 75: return "🟢🟢 強力買入"
    elif total >= 60: return "🟢 值得關注"
    elif total >= 45: return "⚖️ 觀望"
    elif total >= 30: return "🟡 偏弱"
    else:             return "🔴 避開"

def get_score_fill(total):
    if   total >= 75: return PatternFill("solid", fgColor="00B050")
    elif total >= 60: return PatternFill("solid", fgColor="C6EFCE")
    elif total >= 45: return PatternFill("solid", fgColor="FFEB9C")
    elif total >= 30: return PatternFill("solid", fgColor="FFC7CE")
    else:             return PatternFill("solid", fgColor="FF0000")

def get_score_font(total):
    if   total >= 75: return Font(name="Arial",size=11,bold=True,color="FFFFFF")
    elif total >= 60: return Font(name="Arial",size=11,bold=True,color="375623")
    elif total >= 45: return Font(name="Arial",size=11,bold=True,color="7E6000")
    elif total >= 30: return Font(name="Arial",size=11,bold=True,color="9C0006")
    else:             return Font(name="Arial",size=11,bold=True,color="FFFFFF")

# =========================================================
# IBKR SOCKET（v5 新增：daily 亦支援三市場）
# =========================================================
class IBAPIFundamentalApp:
    MARKET_CONFIG = {
        "HK": {"exchange": b"SEHK",  "currency": b"HKD"},
        "UK": {"exchange": b"LSE",   "currency": b"GBP"},
        "US": {"exchange": b"SMART", "currency": b"USD"},
    }
    def __init__(self): pass

    def _recv_all(self, sock, n):
        data = b""
        while len(data) < n:
            chunk = sock.recv(n - len(data))
            if not chunk: break
            data += chunk
        return data

    def get_fundamentals_raw(self, ticker_id: str, market: str = "HK") -> str:
        cfg = self.MARKET_CONFIG.get(market, self.MARKET_CONFIG["HK"])
        if market == "HK":   symbol = ticker_id.replace(".HK","").lstrip("0") or "0"
        elif market == "UK": symbol = ticker_id.replace(".L","")
        else:                symbol = ticker_id
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(IBKR_TIMEOUT)
            sock.connect((IBKR_HOST, IBKR_PORT))
            sock.send(struct.pack(">I", 14) + b"API\x00v100..176\x00")
            sz = struct.unpack(">I", self._recv_all(sock, 4))[0]
            self._recv_all(sock, sz)
            sock.send(struct.pack(">I", len(b"71\x002\x001\x00\x00")) + b"71\x002\x001\x00\x00")
            parts = [b"52",b"2",b"9001",b"0",symbol.encode(),b"STK",b"",b"0",b"",b"",
                     cfg["exchange"],cfg["currency"],b"",b"",b"0",b"ReportSnapshot"]
            msg = b"\x00".join(parts) + b"\x00"
            sock.send(struct.pack(">I", len(msg)) + msg)
            xml_data = ""
            while True:
                try:
                    h = self._recv_all(sock, 4)
                    if len(h) < 4: break
                    sz2 = struct.unpack(">I", h)[0]
                    if sz2 == 0 or sz2 > 10_000_000: break
                    body = self._recv_all(sock, sz2)
                    fields = body.split(b"\x00")
                    if fields and fields[0] == b"51" and len(fields) >= 4:
                        xml_data = fields[3].decode("utf-8", errors="ignore"); break
                except Exception: break
            sock.close(); return xml_data
        except Exception: return ""

    def get_last_price(self, ticker_id: str, market: str = "HK") -> float:
        """用 IBKR reqMktData 取最後成交/收市價，失敗回傳 0.0"""
        cfg = self.MARKET_CONFIG.get(market, self.MARKET_CONFIG["HK"])

        if market == "HK":
            raw = ticker_id.replace(".HK", "")
            symbol = raw   # 保留原樣，e.g. "2600", "0005"
        elif market == "UK":
            symbol = ticker_id.replace(".L", "")
        else:
            symbol = ticker_id

        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(IBKR_TIMEOUT)
            sock.connect((IBKR_HOST, IBKR_PORT))
            sock.send(struct.pack(">I", 14) + b"API\x00v100..176\x00")
            sz = struct.unpack(">I", self._recv_all(sock, 4))[0]
            self._recv_all(sock, sz)
            sock.send(struct.pack(">I", len(b"71\x002\x001\x00\x00")) + b"71\x002\x001\x00\x00")

            parts = [b"1", b"11", b"9001",
                    b"0", symbol.encode(), b"STK",
                    b"", b"0", b"", b"", cfg["exchange"], b"", cfg["currency"], b"", b"",
                    b"0", b"", b"1",   # snapshot=1
                    b"0", b""]
            msg = b"\x00".join(parts) + b"\x00"
            sock.send(struct.pack(">I", len(msg)) + msg)

            last_price = 0.0
            deadline = time.time() + IBKR_TIMEOUT
            while time.time() < deadline:
                try:
                    h = self._recv_all(sock, 4)
                    if len(h) < 4: break
                    sz2 = struct.unpack(">I", h)[0]
                    if sz2 == 0 or sz2 > 1_000_000: break
                    body = self._recv_all(sock, sz2)
                    fields = body.split(b"\x00")
                    if not fields: continue
                    if fields[0] == b"1" and len(fields) >= 5:
                        try:
                            tick_type = int(fields[3])
                            price_val = float(fields[4])
                            if tick_type in (4, 9, 75) and price_val > 0:
                                last_price = price_val
                                if tick_type in (9, 75):
                                    break
                        except Exception: pass
                    if fields[0] == b"57": break   # tickSnapshotEnd
                except Exception: break

            try:
                cmsg = b"\x00".join([b"2", b"1", b"9001"]) + b"\x00"
                sock.send(struct.pack(">I", len(cmsg)) + cmsg)
            except Exception: pass
            sock.close()
            return last_price
        except Exception:
            return 0.0


def _parse_ibkr_xml(xml_str: str, price: float = 0) -> dict:
    if not xml_str: return {}
    import xml.etree.ElementTree as ET
    result = {}
    field_map = {
        "EPSExclExtraItems": ("trailingEps",       1.0,  False),
        "EPS":               ("trailingEps",       1.0,  False),
        "PayoutRatio":       ("payoutRatio",       1.0,  True),
        "BookValuePerShare": ("_bvps",             1.0,  False),
        "CurrentRatio":      ("currentRatio",      1.0,  False),
        "EBITDA":            ("ebitda",            1e6,  False),
        "TotalDebt":         ("totalDebt",         1e6,  False),
        "CashAndEquivalents":("totalCash",         1e6,  False),
        "FreeCashFlow":      ("freeCashflow",      1e6,  False),
        "InterestExpense":   ("interestExpense",   1e6,  False),
        "SharesOutstanding": ("sharesOutstanding", 1e6,  False),
        "EarningsGrowth":    ("earningsGrowth",    1.0,  True),
    }
    try:
        root = ET.fromstring(xml_str)
        for xml_tag, (py_field, mult, div100) in field_map.items():
            if py_field in result: continue
            for elem in root.iter(xml_tag):
                try:
                    val = float(elem.text)
                    if div100 and abs(val) > 1: val /= 100.0
                    result[py_field] = val * mult; break
                except Exception: pass
        if "_bvps" in result and result["_bvps"] > 0 and price > 0:
            result["priceToBook"] = round(price / result["_bvps"], 3)
            del result["_bvps"]
    except Exception: pass
    return result

_ibkr_app = IBAPIFundamentalApp()

def _get_ibkr_data(ticker_id: str, market: str, price: float = 0) -> dict:
    try:
        xml = _ibkr_app.get_fundamentals_raw(ticker_id, market)
        return _parse_ibkr_xml(xml, price) if xml else {}
    except Exception: return {}

# =========================================================
# v5 合理性檢查
# =========================================================
def check_suspicious_fields(info: dict, ticker_id: str) -> list:
    suspicious = []
    pr = info.get("payoutRatio")
    if pr is not None and (pr > 2.0 or pr < 0):
        suspicious.append("payoutRatio")
        log(f"    ⚠️  v5 合理性警告 [{ticker_id}]: payoutRatio={pr:.3g} 超出範圍")
    cr = info.get("currentRatio")
    if cr is not None and (cr > 50 or cr < 0):
        suspicious.append("currentRatio")
        log(f"    ⚠️  v5 合理性警告 [{ticker_id}]: currentRatio={cr:.3g} 超出範圍")
    pb = info.get("priceToBook")
    if pb is not None and (pb < 0 or pb > 100):
        suspicious.append("priceToBook")
        log(f"    ⚠️  v5 合理性警告 [{ticker_id}]: priceToBook={pb:.3g} 超出範圍")
    eps = info.get("trailingEps")
    raw_price = info.get("currentPrice") or info.get("regularMarketPrice")
    if eps is not None and raw_price and eps != 0:
        if ticker_id.endswith(".L") and raw_price > 50 and abs(eps) < 10:
            price = raw_price / 100.0
        else:
            price = raw_price
        if eps < 0:
            pass
        else:
            implied_pe = price / eps
            if implied_pe > 500:
                suspicious.append("trailingEps")
                note = "疑似英鎊/便士單位不一致" if ticker_id.endswith(".L") else "隱含 PE 異常"
                log(f"    ⚠️  v5 合理性警告 [{ticker_id}]: trailingEps={eps:.6g} (隱含 PE={implied_pe:.1f}，{note})")
    return suspicious

# =========================================================
# INFO WITH FALLBACK（v5：四層架構）
# =========================================================

def get_info_with_fallback(ticker_id, market):
    # Layer 1: yfinance
    info = {}
    try:
        info = yf.Ticker(ticker_id).info or {}
    except Exception as e:
        log(f"    ⚠️  yfinance: {e}")
    FILL_FIELDS = ["trailingEps","payoutRatio","freeCashflow","ebitda",
                   "totalDebt","priceToBook","currentRatio","ebit",
                   "operatingIncome","totalCash","interestExpense",
                   "earningsGrowth","sharesOutstanding"]
    # Layer 2: 合理性檢查
    suspicious = check_suspicious_fields(info, ticker_id)
    # ✅ 修正：用 is None 判斷，唔好用 not（避免 0.0 被誤判為缺失）
    missing = [f for f in FILL_FIELDS if info.get(f) is None]
    # Layer 3: IBKR（三市場通用）— 修正異常 + 補缺失
    if (missing or suspicious) and USE_IBKR:
        price = safe(info.get("currentPrice") or info.get("regularMarketPrice"), 0)
        ibkr_data = _get_ibkr_data(ticker_id, market, price)
        if ibkr_data:
            replaced, filled = [], []
            for f in FILL_FIELDS:
                if ibkr_data.get(f) is None: continue
                if f in suspicious:
                    old = info.get(f)
                    info[f] = ibkr_data[f]
                    replaced.append(f"{f}({old:.3g}→{ibkr_data[f]:.3g})")
                elif info.get(f) is None:
                    info[f] = ibkr_data[f]; filled.append(f)
            if replaced: log(f"    🔌 IBKR 修正異常: {replaced}")
            if filled:   log(f"    🔌 IBKR 補充缺失: {filled}")
            # ✅ 重新計算 missing（用 is None）
            missing = [f for f in FILL_FIELDS if info.get(f) is None]
    # Layer 4: 靜態備用（Static Fundamentals）
    static = STATIC_MAP.get(market, {}).get(ticker_id, {})
    if static and missing:
        filled = []
        for f in missing:
            # ✅ 修正：static 值係 None 就唔補（靜態字典本身有 None 值）
            if static.get(f) is not None and info.get(f) is None:
                info[f] = static[f]; filled.append(f)
        if filled: log(f"    📋 靜態補充({market}): {filled}")
    # 靜態覆蓋仍異常欄位（IBKR 未能修正）
    for f in suspicious:
        if static.get(f) is not None:
            info[f] = static[f]
            log(f"    📋 靜態覆蓋異常({market}): {f}={static[f]}")
    if not info.get("ebit") and not info.get("operatingIncome"):
        ev = safe(info.get("ebitda"), 0)
        if ev > 0: info["ebit"] = ev * 0.6
    return info


# =========================================================
# SCORING (identical to batch v4)
# =========================================================
def score_dividend_quality(info, c_yield, y_avg, y_std, df_hist):
    pts = 0.0
    y_buy = y_avg + 1.5*y_std; y_sell = y_avg - 1.5*y_std
    if   c_yield >= y_buy:  pts += 10
    elif c_yield >= y_avg:  pts += 7
    elif c_yield >= y_sell: pts += 4
    else:                   pts += 1
    pr = safe(info.get("payoutRatio"), -1)
    if pr < 0:
        eps = safe(info.get("trailingEps"),0); div = safe(info.get("dividendRate") or info.get("trailingAnnualDividendRate"),0)
        pr = (div/eps) if eps>0 else -1
    if   pr > 1.0:       pts += 0   # 派息超過盈利，明確懲罰
    elif 0<pr<=0.50:     pts += 10
    elif 0<pr<=0.65:     pts += 8
    elif 0<pr<=0.80:     pts += 5
    elif 0<pr<=0.95:     pts += 2
    elif 0<pr<=1.0:      pts += 0   # 幾乎全派，風險高
    else:                pts += 2   # 數據缺失：給低分
    fcf=safe(info.get("freeCashflow"),0); div_r=safe(info.get("dividendRate") or info.get("trailingAnnualDividendRate"),0)
    sh=safe(info.get("sharesOutstanding"),0); dp=div_r*sh if sh>0 else 0
    if fcf>0 and dp>0:
        fc=fcf/dp
        if fc>=2.0: pts+=8
        elif fc>=1.5: pts+=6
        elif fc>=1.0: pts+=4
        else: pts+=1
    elif fcf<0: pts+=0  # FCF 為負，明確懲罰
    else: pts+=2        # 數據缺失：給低分
    if df_hist is not None and not df_hist.empty:
        divs=df_hist[df_hist["Dividend_Amount"]>0].copy()
        if len(divs)>=2:
            divs["Year"]=pd.to_datetime(divs["Date"]).dt.year
            annual=divs.groupby("Year")["Dividend_Amount"].sum()
            cut=any(annual.iloc[i]<annual.iloc[i-1]*0.90 for i in range(1,len(annual)))
            pts+=0 if cut else 2
        else: pts+=1
    else: pts+=1
    return round(pts,2)

def score_valuation(info, c_yield, y_avg, y_std, c_pe, pe_avg, pe_std, rfr=4.3):
    pts=0.0
    pe_buy=pe_avg-1.5*pe_std; pe_sell=pe_avg+1.5*pe_std
    if   c_pe<=pe_buy:  pts+=8
    elif c_pe<=pe_avg:  pts+=6
    elif c_pe<=pe_sell: pts+=3
    pb=safe(info.get("priceToBook"),-1)
    if   pb<=0:   pts+=2   # 數據缺失：給低分
    elif pb<=0.8: pts+=7   # 真正折讓才給滿分
    elif pb<=1.2: pts+=5
    elif pb<=2.0: pts+=3
    elif pb<=3.5: pts+=1
    else:         pts+=0
    sp=c_yield-rfr
    if   sp>=3.0: pts+=10
    elif sp>=2.0: pts+=8
    elif sp>=1.0: pts+=6
    elif sp>=0.0: pts+=4
    elif sp>=-1.0:pts+=2
    return round(pts,2)

def score_financial_health(info):
    pts=0.0
    ebitda=safe(info.get("ebitda"),0); td=safe(info.get("totalDebt"),0)
    cash=safe(info.get("totalCash") or info.get("cash"),0); nd=td-cash
    if ebitda>0:
        r=nd/ebitda
        if   r<=1.0: pts+=10
        elif r<=2.0: pts+=8
        elif r<=3.0: pts+=6
        elif r<=4.5: pts+=3
    elif nd<=0: pts+=10    # 淨現金，財務最穩健
    else: pts+=2           # 有負債但無數據：給低分
    ebit=safe(info.get("ebit") or info.get("operatingIncome"),0)
    ie=abs(safe(info.get("interestExpense"),0))
    if ebit>0 and ie>0:
        ic=ebit/ie
        if   ic>=8:   pts+=9
        elif ic>=5:   pts+=7
        elif ic>=3:   pts+=4
        elif ic>=1.5: pts+=2
    elif ebit<0: pts+=0    # 營業虧損，明確懲罰
    else: pts+=2           # 數據缺失：給低分
    cr=safe(info.get("currentRatio"),-1)
    if   cr<0:    pts+=1   # 數據缺失：給低分
    elif cr>=2.0: pts+=6
    elif cr>=1.5: pts+=5
    elif cr>=1.0: pts+=3
    return round(pts,2)

def score_growth(info, df_hist):
    pts=0.0; dgr=None
    if df_hist is not None and not df_hist.empty:
        divs=df_hist[df_hist["Dividend_Amount"]>0].copy()
        if len(divs)>=2:
            divs["Year"]=pd.to_datetime(divs["Date"]).dt.year
            annual=divs.groupby("Year")["Dividend_Amount"].sum()
            if len(annual)>=4:
                dn,do=annual.iloc[-1],annual.iloc[-4]
                dgr=((dn/do)**(1/3)-1)*100 if do>0 else 0
            elif len(annual)>=2:
                n=len(annual)-1; dn,do=annual.iloc[-1],annual.iloc[0]
                dgr=((dn/do)**(1/n)-1)*100 if do>0 else 0
    if dgr is None: dgr=0
    if   dgr>=8: pts+=6
    elif dgr>=5: pts+=5
    elif dgr>=2: pts+=4
    elif dgr>=0: pts+=2
    eg=info.get("earningsGrowth")
    if eg is None: eg=info.get("earningsQuarterlyGrowth")
    if eg is None: pts+=1  # 數據缺失：給低分
    else:
        eg=safe(eg,0)
        if   eg>=0.10: pts+=4
        elif eg>=0.03: pts+=3
        elif eg>=0:    pts+=2
    return round(pts,2)

def score_technical(df_hist, info):
    pts=0.0
    w52l=safe(info.get("fiftyTwoWeekLow"),0); w52h=safe(info.get("fiftyTwoWeekHigh"),0)
    price=safe(info.get("currentPrice") or info.get("regularMarketPrice"),0)
    if w52l>0 and w52h>w52l and price>0:
        pos=(price-w52l)/(w52h-w52l)
        if   pos<=0.25: pts+=6
        elif pos<=0.40: pts+=5
        elif pos<=0.60: pts+=3
        elif pos<=0.80: pts+=1
    else: pts+=1  # 數據缺失：給低分
    if df_hist is not None and len(df_hist)>=15:
        closes=pd.Series(df_hist["Close"].values,dtype=float)
        delta=closes.diff(); gain=delta.clip(lower=0).rolling(14).mean()
        loss=(-delta.clip(upper=0)).rolling(14).mean()
        rs=gain/loss.replace(0,np.nan); rsi=(100-100/(1+rs)).iloc[-1]
        if   rsi<=30: pts+=4
        elif rsi<=45: pts+=3
        elif rsi<=60: pts+=2
    else: pts+=1  # 數據缺失：給低分
    return round(pts,2)

def compute_score(info, c_yield, y_avg, y_std, c_pe, pe_avg, pe_std, df_hist, rfr=4.3):
    s1=score_dividend_quality(info,c_yield,y_avg,y_std,df_hist)
    s2=score_valuation(info,c_yield,y_avg,y_std,c_pe,pe_avg,pe_std,rfr)
    s3=score_financial_health(info)
    s4=score_growth(info,df_hist)
    s5=score_technical(df_hist,info)
    return {"Score_股息質量_30":s1,"Score_估值_25":s2,"Score_財務健康_25":s3,
            "Score_增長潛力_10":s4,"Score_技術面_10":s5,"Score_總分_100":round(s1+s2+s3+s4+s5,1)}

def build_score_snapshot(info, df_hist, c_yield, y_avg, y_std, c_pe, pe_avg, pe_std, rfr=4.3):
    scores=compute_score(info,c_yield,y_avg,y_std,c_pe,pe_avg,pe_std,df_hist,rfr)
    pr=info.get("payoutRatio")
    if pr is None:
        eps=safe(info.get("trailingEps"),0); div=safe(info.get("dividendRate") or info.get("trailingAnnualDividendRate"),0)
        pr=round(div/eps*100,2) if eps>0 else None
    else: pr=round(pr*100,2)
    fcf=safe(info.get("freeCashflow"),0); dr=safe(info.get("dividendRate") or info.get("trailingAnnualDividendRate"),0)
    sh=safe(info.get("sharesOutstanding"),0); dp=dr*sh if sh>0 else 0
    fcf_cov=round(fcf/dp,2) if dp>0 and fcf>0 else None
    ebitda=safe(info.get("ebitda"),0); td=safe(info.get("totalDebt"),0)
    cash=safe(info.get("totalCash") or info.get("cash"),0)
    nd_eb=round((td-cash)/ebitda,2) if ebitda>0 else None
    ebit=safe(info.get("ebit") or info.get("operatingIncome"),0)
    ie=abs(safe(info.get("interestExpense"),0))
    ic=round(ebit/ie,2) if ie>0 and ebit>0 else None
    cr=safe(info.get("currentRatio"),None); pb=safe(info.get("priceToBook"),None)
    w52l=safe(info.get("fiftyTwoWeekLow"),0); w52h=safe(info.get("fiftyTwoWeekHigh"),0)
    price=safe(info.get("currentPrice") or info.get("regularMarketPrice"),0)
    w52p=round((price-w52l)/(w52h-w52l)*100,1) if w52h>w52l else None
    rsi=None
    if df_hist is not None and len(df_hist)>=15:
        closes=pd.Series(df_hist["Close"].values,dtype=float); delta=closes.diff()
        gain=delta.clip(lower=0).rolling(14).mean(); loss=(-delta.clip(upper=0)).rolling(14).mean()
        rs=gain/loss.replace(0,np.nan); rsi_s=100-100/(1+rs)
        rsi=round(float(rsi_s.iloc[-1]),1) if not rsi_s.empty else None
    dgr=None
    if df_hist is not None and not df_hist.empty and "Dividend_Amount" in df_hist.columns:
        divs=df_hist[df_hist["Dividend_Amount"]>0].copy()
        if len(divs)>=2:
            divs["Year"]=pd.to_datetime(divs["Date"]).dt.year
            annual=divs.groupby("Year")["Dividend_Amount"].sum()
            if len(annual)>=4:
                dn,do=annual.iloc[-1],annual.iloc[-4]
                dgr=round(((dn/do)**(1/3)-1)*100,2) if do>0 else 0
            elif len(annual)>=2:
                n=len(annual)-1; dn,do=annual.iloc[-1],annual.iloc[0]
                dgr=round(((dn/do)**(1/n)-1)*100,2) if do>0 else 0
    return {
        "Score_日期":datetime.date.today(),"Payout_Ratio_%":pr,"FCF_Coverage":fcf_cov,
        "Net_Debt_EBITDA":nd_eb,"Interest_Coverage":ic,
        "Current_Ratio":round(cr,2) if cr is not None else None,
        "PB_Ratio":round(pb,2) if pb is not None else None,
        "Yield_Spread_vs_Bond":round(c_yield-rfr,2),"DGR_3yr_%":dgr,"RSI_14":rsi,
        "52W_Position_%":w52p,**scores,
    }

# =========================================================
# BACKUP & SAVE
# =========================================================
def create_backup(filepath):
    if not os.path.exists(filepath): return
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    ts=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dst=os.path.join(BACKUP_FOLDER, f"backup_{os.path.basename(filepath)}_{ts}.xlsx")
    shutil.copy2(filepath, dst)
    log(f"  Backup: {dst}")

def safe_save(wb, filepath):
    tmp=filepath.replace(".xlsx","._saving.xlsx")
    try:
        wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True
        wb.save(tmp)
    except Exception as e:
        log(f"  SAVE ERROR (write): {e}"); return False
    try: wb.close()
    except: pass
    for retry in range(MAX_SAVE_RETRY):
        try:
            shutil.move(tmp, filepath)
            os.chmod(filepath, stat.S_IREAD)
            log(f"  Saved: {filepath}"); return True
        except PermissionError as e:
            log(f"  Locked — close Excel. Retry {retry+1}/{MAX_SAVE_RETRY}"); time.sleep(4)
        except Exception as e:
            log(f"  SAVE ERROR: {e}"); time.sleep(2)
    log(f"  Data safe in: {tmp}"); return False

# =========================================================
# DATE ROW
# =========================================================
def find_date_row(ws, target):
    for row in range(2, ws.max_row+1):
        v=ws.cell(row,1).value
        if isinstance(v, datetime.datetime): v=v.date()
        if isinstance(v, pd.Timestamp):      v=v.date()
        if v==target: return row
    return None

def set_cells(ws, row, vals):
    for col,val in vals.items(): ws.cell(row,col).value=val

# =========================================================
# BUILD MARKET MAP FROM WORKBOOK
# =========================================================
def build_market_map(wb, market):
    """Returns dict: sheet_name -> (full_ticker, company_name)"""
    smap={}
    sum_name=f"{market} 總覽"
    if sum_name not in wb.sheetnames: return smap
    ws=wb[sum_name]
    for row in range(2, ws.max_row+1):
        tv=ws.cell(row,1).value
        if not tv: continue
        tv=str(tv)
        nv=ws.cell(row,2).value          # col B = 公司名稱
        name=str(nv) if nv else tv       # fallback: ticker
        sn=ticker_to_sheet(tv)
        smap[sn]=(tv, name)
    return smap

# =========================================================
# READ HISTORY FROM SHEET
# =========================================================
def read_sheet_history(ws):
    rows=[]
    for r in range(2, ws.max_row+1):
        dv=ws.cell(r,1).value
        if isinstance(dv, datetime.datetime): dv=dv.date()
        rows.append({
            "Date":dv,
            "Close":safe(ws.cell(r,2).value, np.nan),
            "Dividend_Amount":safe(ws.cell(r,4).value, 0.0),
        })
    return pd.DataFrame(rows)

# =========================================================
# UPDATE ONE STOCK
# =========================================================
def update_stock(ws, ticker, market):
    log(f"  [{market}] {ticker}")

    # ── 取最新行情 ──────────────────────────────────────────
    if market == "CN" and _AKSHARE_OK:
        # A股：AKShare 新浪源（可突破大陸以外 IP 限制）
        try:
            code = ticker.replace(".SS","").replace(".SZ","")
            sina_code = ("sh" if ticker.endswith(".SS") else "sz") + code
            df_ak = ak.stock_zh_a_daily(symbol=sina_code, adjust="qfq")
            df_ak = df_ak.rename(columns={"date":"Date","close":"Close","volume":"Volume"})
            df_ak["Date"]   = pd.to_datetime(df_ak["Date"]).dt.date
            df_ak["Close"]  = pd.to_numeric(df_ak["Close"],  errors="coerce")
            df_ak["Volume"] = pd.to_numeric(df_ak["Volume"], errors="coerce").fillna(0).astype(int)
            # 只取最近 15 日
            df_ak = df_ak.tail(15).copy()
            # DatetimeIndex 格式與 yfinance 一致，避免 reset_index() 重複插入 Date 欄
            df_ak_out = df_ak.drop(columns=["Date"]).copy()
            df_ak_out.index = pd.to_datetime(df_ak["Date"].values)
            df_ak_out.index.name = "Date"
            df_new = df_ak_out
            log(f"    AKShare(新浪) CN 數據: {len(df_ak)} 行")
        except Exception as e:
            log(f"    ⚠️  AKShare CN 失敗: {e}，嘗試 yfinance...")
            try:
                df_new = yf.Ticker(ticker).history(period="10d")
            except Exception as e2:
                log(f"    ❌ yfinance CN 亦失敗: {e2}")
                df_new = pd.DataFrame()
    else:
        try:
            df_new = yf.Ticker(ticker).history(period="10d")
        except Exception as e:
            log(f"    ⚠️  yfinance history() 失敗: {e}")
            df_new = pd.DataFrame()
    ld = lc = lv = ldiv = None
    if not df_new.empty:
        df_new = df_new.reset_index()
        df_new["Date"] = pd.to_datetime(df_new["Date"]).dt.date
        df_new = df_new.dropna(subset=["Close"])
        if not df_new.empty:
            if market == "UK":
                df_new["Close"]     = df_new["Close"]     / 100.0
                df_new["Dividends"] = df_new.get("Dividends", pd.Series([0]*len(df_new))) / 100.0
            latest = df_new.iloc[-1]
            ld   = latest["Date"]
            lc   = round(float(latest["Close"]), 4)
            lv   = int(latest.get("Volume", 0) or 0)
            ldiv = float(latest.get("Dividends", 0) or 0)
    # ── IBKR 補充收市價 ──
    if lc is None or (isinstance(lc, float) and np.isnan(lc)):
        reason = "yfinance 無數據" if lc is None else "Close=nan"
        log(f"    ⚠️  {reason}，嘗試 IBKR 補充收市價...")
        if USE_IBKR:
            ibkr_price = _ibkr_app.get_last_price(ticker, market)
            if ibkr_price > 0:
                if market == "UK": ibkr_price /= 100.0
                lc   = round(ibkr_price, 4)
                ld   = ld or datetime.date.today()
                lv   = lv or 0
                ldiv = ldiv or 0.0
                log(f"    🔌 IBKR 補充收市價: {ld}  Close={lc}")
            else:
                log(f"    ❌ IBKR 亦無法取得收市價，跳過")
                return None
        else:
            log("    No data"); return None
    try:
        log(f"    {ld}  Close={lc}")
        # ── 讀取工作表歷史數據 ──
        df_hist = read_sheet_history(ws)
        # ── 取基本面資料（四層架構）──
        info = get_info_with_fallback(ticker, market)
        # ── 公司名稱 ──
        name = info.get("longName") or info.get("shortName") or ""
        # ── 股息 ──
        if ldiv is None: ldiv = 0.0
        # ── 股息歷史統計 ──
        rfr = RISK_FREE_RATES.get(market, 4.3)
        # 計算現價（UK 已除 100）
        price = lc
        # ── 計算現時息率（優先用歷史除息記錄，唔依賴 dividendRate 格式）──
        # 方法：讀取工作表過去最多 252 行的 Dividend_Amount，加總作年化股息
        div_rows = []
        max_r    = ws.max_row
        # A股年派一次用 400 行，其他市場 252 行
        lookback = 400 if market == "CN" else 252
        start_r  = max(2, max_r - (lookback - 1))
        for r in range(start_r, max_r + 1):
            dv = safe(ws.cell(r, 4).value, 0.0)   # col D = Dividend_Amount
            if dv > 0:
                div_rows.append(dv)
        annual_div_actual = sum(div_rows)

        if annual_div_actual > 0 and price > 0:
            c_yield = round(annual_div_actual / price * 100, 4)
            log(f"    息率（實際除息）: {c_yield:.2f}%  年化股息={annual_div_actual:.4f}")
        else:
            # 後備：dividendRate（加 UK 單位合理性檢查）
            div_rate = safe(info.get("dividendRate") or info.get("trailingAnnualDividendRate"), 0)
            if market == "UK" and div_rate > price * 0.5 and div_rate > 1:
                div_rate /= 100.0   # 疑似便士，換算英鎊
            c_yield = round(div_rate / price * 100, 4) if price > 0 and div_rate > 0 else 0.0
            if c_yield > 0:
                log(f"    息率（dividendRate 後備）: {c_yield:.2f}%")

        # 合理性上限：>30% 視為異常
        if c_yield > 30:
            log(f"    ⚠️  息率 {c_yield:.1f}% >30% 異常，改用歷史均值")
            c_yield = 0.0   # 下方 y_avg 會補回合理值
        # 從工作表讀取歷史息率統計
        y_vals = []
        pe_vals = []
        for r in range(2, ws.max_row + 1):
            yv = ws.cell(r, 6).value   # col F = 息率
            pv = ws.cell(r, 10).value  # col J = PE
            if yv is not None:
                try: y_vals.append(float(yv))
                except: pass
            if pv is not None:
                try: pe_vals.append(float(pv))
                except: pass
        y_avg  = float(np.mean(y_vals))  if y_vals  else c_yield
        y_std  = float(np.std(y_vals))   if len(y_vals) > 1  else 0.5
        pe_avg = float(np.mean(pe_vals)) if pe_vals else 15.0
        pe_std = float(np.std(pe_vals))  if len(pe_vals) > 1 else 3.0
        # 計算現時 PE
        eps = safe(info.get("trailingEps"), 0)
        if market == "UK" and eps > 0 and price > 0:
            implied_pe = price / eps
            if implied_pe > 500:           # EPS 係便士單位
                eps = eps / 100.0
        c_pe = round(price / eps, 2) if eps > 0 and price > 0 else 0.0
        # ── 評分快照 ──
        snap = build_score_snapshot(
            info, df_hist, c_yield, y_avg, y_std, c_pe, pe_avg, pe_std, rfr
        )
        # ── 寫入或新增日期行 ──
        existing = find_date_row(ws, ld)
        if existing:
            row = existing
        else:
            row = ws.max_row + 1
        # 計算 5年均息率 / PE 統計（寫入 col E–M）
        y_buy  = round(y_avg + 1.5 * y_std, 4)
        y_sell = round(y_avg - 1.5 * y_std, 4)
        pe_buy  = round(max(0, pe_avg - 1.5 * pe_std), 2)
        pe_sell = round(pe_avg + 1.5 * pe_std, 2)
        set_cells(ws, row, {
            1:  ld,           # A: 日期
            2:  lc,           # B: 收市價
            3:  lv,           # C: 成交量
            4:  ldiv,         # D: 股息
            5:  round(c_yield, 4),   # E: 現時息率
            6:  round(y_avg,  4),    # F: 5年均息率
            7:  round(y_buy,  4),    # G: 買入線
            8:  round(y_sell, 4),    # H: 賣出線
            9:  round(c_pe,   2),    # I: 現時PE
            10: round(pe_avg, 2),    # J: 5年均PE
            11: round(pe_buy, 2),    # K: PE買入線
            12: round(pe_sell,2),    # L: PE賣出線
            13: name,                # M: 公司名稱
        })
        # 寫入評分快照 (col N onwards)
        for i, col_name in enumerate(SCORE_SNAPSHOT_COLS):
            ws.cell(row, SCORE_COL_START + i).value = snap.get(col_name)
        return {
            "close": lc,
            "yield": c_yield,
            "y_avg": y_avg,
            "y_std": y_std,
            "pe":    c_pe,
            "pe_avg":pe_avg,
            "pe_std":pe_std,
            "snap":  snap,
            "name":  name,
        }
    except Exception as e:
        import traceback
        log(f"    ERROR: {e}\n{traceback.format_exc()}")
        return None

# =========================================================
# REBUILD SUMMARY SHEET
# =========================================================
def rebuild_summary(wb, market, smap, results):
    sum_name=f"{market} 總覽"

    if sum_name in wb.sheetnames: del wb[sum_name]
    ws=wb.create_sheet(sum_name); ws.freeze_panes="A2"

    THIN=Border(left=Side(style="thin",color="E0E0E0"),right=Side(style="thin",color="E0E0E0"),
                top=Side(style="thin",color="E0E0E0"),bottom=Side(style="thin",color="E0E0E0"))
    HDR_FILL=PatternFill("solid",fgColor="3F3F3F")
    SCORE_HDR=PatternFill("solid",fgColor="1F4E79")
    SCORE_FILL=PatternFill("solid",fgColor="E8F4FD")
    HDR_FONT=Font(name="Arial",size=11,bold=True,color="FFFFFF")
    NORM_FONT=Font(name="Arial",size=10)
    LINK_FONT=Font(name="Arial",size=10,color="0000FF",underline="single")
    ncols=len(SUMMARY_COLS)

    ws.append(SUMMARY_COLS)
    for col in range(1,ncols+1):
        c=ws.cell(1,col)
        c.fill=SCORE_HDR if col>=22 else HDR_FILL
        c.font=HDR_FONT; c.alignment=Alignment(horizontal="center",wrap_text=True)

    rows=[]
    for sn,(ticker,smap_name) in smap.items():
        r=results.get(sn)
        if r:
            snap=r["snap"]
            y_avg=r["y_avg"]; y_std=r["y_std"]
            pe_avg=r["pe_avg"]; pe_std=r["pe_std"]
            total=safe(snap.get("Score_總分_100"),0)
            # yfinance name 優先；若為空則用 smap 中的舊名稱
            name = r["name"] if r.get("name") else smap_name
            rows.append((sn,ticker,name,r["close"],r["yield"],y_avg,
                         y_avg+1.5*y_std,y_avg-1.5*y_std,
                         r["pe"],pe_avg,
                         max(0,pe_avg-1.5*pe_std),pe_avg+1.5*pe_std,
                         snap,total))
        else:
            if sn not in wb.sheetnames: continue
            ws_s=wb[sn]; last=ws_s.max_row
            if last<2: continue
            snap={SCORE_SNAPSHOT_COLS[i]:ws_s.cell(last,SCORE_COL_START+i).value for i in range(len(SCORE_SNAPSHOT_COLS))}
            total=safe(snap.get("Score_總分_100"),0)
            rows.append((sn,ticker,smap_name,
                safe(ws_s.cell(last,2).value,0),safe(ws_s.cell(last,6).value,0),
                safe(ws_s.cell(last,7).value,0),safe(ws_s.cell(last,8).value,0),safe(ws_s.cell(last,9).value,0),
                safe(ws_s.cell(last,10).value,0),safe(ws_s.cell(last,11).value,0),
                safe(ws_s.cell(last,12).value,0),safe(ws_s.cell(last,13).value,0),
                snap,total))

    rows.sort(key=lambda x:x[13],reverse=True)

    for sn,ticker,company,close,cy,ya,yb,ys,pe,pa,pb2,ps,snap,total in rows:
        ws.append([
            ticker,company,round(close,2),
            round(cy,2),round(ya,2),round(yb,2),round(ys,2),
            round(pe,2),round(pa,2),round(pb2,2),round(ps,2),
            snap.get("Payout_Ratio_%"),snap.get("FCF_Coverage"),
            snap.get("Net_Debt_EBITDA"),snap.get("Interest_Coverage"),
            snap.get("Current_Ratio"),snap.get("PB_Ratio"),
            snap.get("Yield_Spread_vs_Bond"),snap.get("DGR_3yr_%"),
            snap.get("RSI_14"),snap.get("52W_Position_%"),
            snap.get("Score_股息質量_30"),snap.get("Score_估值_25"),
            snap.get("Score_財務健康_25"),snap.get("Score_增長潛力_10"),
            snap.get("Score_技術面_10"),total,get_status(total),
        ])
        ridx=ws.max_row
        rf=get_score_fill(total); rfont=get_score_font(total)

        for col in range(1,ncols+1):
            c=ws.cell(ridx,col); c.border=THIN
            if col>=22: c.fill=SCORE_FILL; c.font=NORM_FONT
            else:       c.fill=rf;         c.font=NORM_FONT
            if col==ncols-1:   # 總分
                c.fill=rf; c.font=rfont; c.number_format="0.0"
                c.alignment=Alignment(horizontal="center")
            if col==ncols:     # 評級
                c.fill=rf; c.font=rfont
                c.alignment=Alignment(horizontal="center")
            if col==1:
                c.hyperlink=f"#'{sn}'!A1"; c.font=LINK_FONT
            if col in [4,5,6,7,12,17,18]: c.number_format="0.00"
            if col in [3,8,9,10,11]:      c.number_format="#,##0.00"
            if col in [22,23,24,25,26]:   c.number_format="0.0"

    for col in ws.columns:
        ml=max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(ml+3,10),22)

    # Move to front
    sheets=wb._sheets
    try:
        idx=next(i for i,s in enumerate(sheets) if s.title==sum_name)
        sheets.insert(0,sheets.pop(idx))
    except: pass
    log(f"  Summary rebuilt: {sum_name}")

# =========================================================
# MAIN
# =========================================================
def main():
    log("="*50)
    log(f"GLOBAL DIVIDEND DAILY IMPORTER v6  ({datetime.date.today()})")
    log(f"RFR: UK={RISK_FREE_RATES['UK']}%  US={RISK_FREE_RATES['US']}%  HK={RISK_FREE_RATES['HK']}%")
    log(f"Log file: {_log_path}")
    log("="*50)

    for market, filepath in EXCEL_FILES.items():
        log(f"\n{'─'*50}")
        log(f"  {market}  →  {filepath}")
        log(f"{'─'*50}")

        if not os.path.exists(filepath):
            log(f"  File not found — run batch first."); continue

        create_backup(filepath)
        try: os.chmod(filepath, stat.S_IREAD|stat.S_IWRITE)
        except Exception as e: log(f"  Warning: {e}")

        wb=load_workbook(filepath)
        smap=build_market_map(wb, market)
        log(f"  Stocks: {len(smap)}")

        results={}
        sum_name=f"{market} 總覽"
        for sn in list(wb.sheetnames):
            if sn==sum_name: continue
            if sn not in smap: continue
            ws=wb[sn]; ws.freeze_panes="A2"
            ticker, _name = smap[sn]
            r=update_stock(ws, ticker, market)
            if r: results[sn]=r
            # ── 返回總覽連結（每個個股分頁第一行最後一格）──
            back_col = ws.max_column + 1
            bc = ws.cell(1, back_col)
            bc.value     = f"⬅ 返回 {sum_name}"
            bc.hyperlink  = f"#'{sum_name}'!A1"
            bc.font       = Font(name="Arial", size=10, bold=True,
                                 color="0563C1", underline="single")
            bc.alignment  = Alignment(horizontal="center")

        rebuild_summary(wb, market, smap, results)
        success=safe_save(wb, filepath)
        log(f"  {'✅ Done' if success else '❌ Save failed'}: {filepath}")

    log("\n"+"="*50)
    log("DAILY IMPORT COMPLETED")
    log(f"Log saved: {_log_path}")
    log("="*50)

if __name__=="__main__":
    main()
