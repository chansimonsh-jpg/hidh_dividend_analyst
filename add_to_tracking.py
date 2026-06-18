#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
add_to_tracking.py
==================
從 Screener 報告 Excel 讀取你勾選的股票，寫入 tickers_config.json。
之後直接跑 batch_importer_global_v6.py 即可包含新股。

使用流程：
  1. 跑 screener_global.py → 生成 screener_report_YYYYMMDD.xlsx
  2. 打開 Excel，在「✅ 加入追蹤」欄填入 Y（大小寫均可）
  3. 儲存 Excel
  4. 執行本腳本：python add_to_tracking.py
  5. 確認後跑 batch_importer_global_v6.py

可選參數：
  --report   指定報告檔案路徑（預設自動找最新的 screener_report_*.xlsx）
  --json     指定 tickers_config.json 路徑（預設同目錄）
  --dry-run  只顯示將要加入的股票，不實際寫入

範例：
  python add_to_tracking.py
  python add_to_tracking.py --report screener_report_20260607.xlsx
  python add_to_tracking.py --dry-run
"""

import argparse
import datetime
import glob
import json
import os
import sys


# ── 設定 ──────────────────────────────────────────────────────────────────────
DEFAULT_JSON   = "tickers_config.json"
TRACK_COL_NAME = "✅ 加入追蹤"    # Screener Excel 的勾選欄標題
TICKER_COL     = "股票代號"
NAME_COL       = "公司名稱"
MARKET_COL     = "市場"
SCORE_COL      = "📊 總分_100"
RATING_COL     = "📊 評級"
YIELD_COL      = "股息率_%"

# 接受的「加入」標記值（大小寫不敏感）
YES_VALUES = {"y", "yes", "是", "✓", "✅", "1", "true", "加"}


def find_latest_report() -> str | None:
    """自動找最新的 screener_report_*.xlsx"""
    files = sorted(glob.glob("screener_report_*.xlsx"), reverse=True)
    return files[0] if files else None


def load_report(path: str) -> list[dict]:
    """
    讀取 Screener Excel，提取所有勾選了「加入追蹤」的行。
    同時掃描「新發現」和「值得留意」兩個分頁。
    """
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要安裝 openpyxl：pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    selected = []

    target_sheets = [s for s in wb.sheetnames
                     if "新發現" in s or "值得留意" in s]
    if not target_sheets:
        # fallback: try all non-摘要 sheets
        target_sheets = [s for s in wb.sheetnames if s != "摘要"]

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        headers = [str(ws.cell(1, c).value or "").strip()
                   for c in range(1, ws.max_column + 1)]

        # Find required column indices
        def col_idx(name):
            try:
                return headers.index(name)
            except ValueError:
                return None

        i_track  = col_idx(TRACK_COL_NAME)
        i_ticker = col_idx(TICKER_COL)
        i_name   = col_idx(NAME_COL)
        i_market = col_idx(MARKET_COL)
        i_score  = col_idx(SCORE_COL)
        i_rating = col_idx(RATING_COL)
        i_yield  = col_idx(YIELD_COL)

        if i_track is None or i_ticker is None:
            print(f"  ⚠️  分頁「{sheet_name}」找不到必要欄位，跳過")
            continue

        for row in range(2, ws.max_row + 1):
            track_val = str(ws.cell(row, i_track + 1).value or "").strip().lower()
            if track_val not in YES_VALUES:
                continue

            ticker = str(ws.cell(row, i_ticker + 1).value or "").strip()
            if not ticker:
                continue

            name   = str(ws.cell(row, i_name   + 1).value or ticker).strip() if i_name   is not None else ticker
            market = str(ws.cell(row, i_market  + 1).value or "").strip()    if i_market is not None else ""
            score  = ws.cell(row, i_score  + 1).value if i_score  is not None else None
            rating = ws.cell(row, i_rating + 1).value if i_rating is not None else None
            yield_ = ws.cell(row, i_yield  + 1).value if i_yield  is not None else None

            selected.append({
                "ticker": ticker,
                "name":   name,
                "market": market,
                "score":  score,
                "rating": rating,
                "yield":  yield_,
                "sheet":  sheet_name,
            })

    return selected


def load_json(path: str) -> dict:
    """載入現有 tickers_config.json"""
    if not os.path.exists(path):
        print(f"❌ 找不到 {path}")
        print("   請先執行 export_tickers_json.py 生成設定檔")
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_json(data: dict, path: str):
    """儲存 tickers_config.json（更新 last_updated）"""
    data["_meta"]["last_updated"] = str(datetime.date.today())
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def main():
    parser = argparse.ArgumentParser(description="從 Screener 報告加入追蹤股票")
    parser.add_argument("--report",  default=None,         help="Screener Excel 路徑")
    parser.add_argument("--json",    default=DEFAULT_JSON,  help="tickers_config.json 路徑")
    parser.add_argument("--dry-run", action="store_true",   help="只顯示，不寫入")
    args = parser.parse_args()

    # ── 找報告檔案 ─────────────────────────────────────────────────────────
    report_path = args.report
    if not report_path:
        report_path = find_latest_report()
        if not report_path:
            print("❌ 找不到 screener_report_*.xlsx")
            print("   請指定：python add_to_tracking.py --report screener_report_YYYYMMDD.xlsx")
            sys.exit(1)
        print(f"📄 自動選取最新報告：{report_path}")
    else:
        if not os.path.exists(report_path):
            print(f"❌ 檔案不存在：{report_path}")
            sys.exit(1)

    # ── 讀報告 ─────────────────────────────────────────────────────────────
    print(f"\n🔍 讀取報告：{report_path}")
    selected = load_report(report_path)

    if not selected:
        print("\n⚠️  未找到任何勾選股票。")
        print(f"   請在 Excel「{TRACK_COL_NAME}」欄填入 Y 後儲存再執行。")
        sys.exit(0)

    print(f"\n✅ 發現 {len(selected)} 隻已勾選股票：")
    print(f"  {'股票代號':<12} {'市場':<5} {'評分':>6}  {'息率':>6}  {'評級':<15} 公司名稱")
    print("  " + "─" * 72)
    for s in selected:
        score_str = f"{s['score']:.1f}" if isinstance(s['score'], (int, float)) else "─"
        yield_str = f"{s['yield']:.1f}%" if isinstance(s['yield'], (int, float)) else "─"
        print(f"  {s['ticker']:<12} {s['market']:<5} {score_str:>6}  {yield_str:>6}  "
              f"{str(s['rating'] or ''):<15} {s['name']}")

    # ── 載入 JSON，找出新增 vs 重複 ────────────────────────────────────────
    data    = load_json(args.json)
    current = data.get("tickers", {})

    new_stocks = [s for s in selected if s["ticker"] not in current]
    dup_stocks = [s for s in selected if s["ticker"] in current]

    if dup_stocks:
        print(f"\n⏭  已在追蹤清單（跳過）：{', '.join(s['ticker'] for s in dup_stocks)}")

    if not new_stocks:
        print("\n⚠️  所有勾選股票已在追蹤清單，無需更新。")
        sys.exit(0)

    # ── Dry run ────────────────────────────────────────────────────────────
    if args.dry_run:
        print(f"\n🔍 [Dry Run] 以下 {len(new_stocks)} 隻將被加入（未實際寫入）：")
        for s in new_stocks:
            print(f"   {s['ticker']:<12} {s['market']:<5} {s['name']}")
        print("\n   移除 --dry-run 參數以實際寫入。")
        sys.exit(0)

    # ── 確認 ───────────────────────────────────────────────────────────────
    print(f"\n📝 將加入 {len(new_stocks)} 隻新股至 {args.json}：")
    for s in new_stocks:
        print(f"   + {s['ticker']:<12} ({s['market']})  {s['name']}")

    confirm = input("\n確認加入？[Y/n] ").strip().lower()
    if confirm not in ("", "y", "yes", "是"):
        print("已取消。")
        sys.exit(0)

    # ── 寫入 JSON ──────────────────────────────────────────────────────────
    today = str(datetime.date.today())
    added = []
    for s in new_stocks:
        # 自動補全 market（若 Screener 報告有欄位）
        market = s["market"]
        if not market:
            t = s["ticker"]
            if t.endswith(".HK"):             market = "HK"
            elif t.endswith(".L"):            market = "UK"
            elif t.endswith((".SS",".SZ")):   market = "CN"
            else:                             market = "US"

        current[s["ticker"]] = {
            "name":       s["name"],
            "market":     market,
            "added_date": today,
            "source":     "screener",          # 標記來源
            "score_at_add": s["score"],        # 記錄加入時的評分
            "yield_at_add": s["yield"],        # 記錄加入時的息率
        }
        added.append(s["ticker"])

    data["tickers"] = current
    save_json(data, args.json)

    print(f"\n✅ 已成功加入 {len(added)} 隻股票：{', '.join(added)}")
    print(f"   {args.json} 現共 {len(current)} 隻股票")
    print(f"\n👉 下一步：執行 batch_importer_global_v6.py 建立新股的完整數據")


if __name__ == "__main__":
    main()
